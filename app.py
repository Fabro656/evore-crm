# =============================================================
# EVORE CRM — v12 (Lotes, Permisos por módulo, Notificaciones,
#                  Diagnóstico, Recetas BOM, Materias Primas, Reservas)
# =============================================================

from flask import Flask, render_template, redirect, url_for, flash, request, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_login import (LoginManager, UserMixin, login_user,
                         logout_user, login_required, current_user)
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta, date as date_type
import os, json, secrets
from functools import wraps
from jinja2 import DictLoader

app = Flask(__name__)
_secret_key = os.environ.get('SECRET_KEY', 'evore-crm-stable-fallback-key-2026-xK9mP')
if _secret_key == 'evore-crm-stable-fallback-key-2026-xK9mP':
    print('INFO: Using fallback SECRET_KEY. Set SECRET_KEY in Railway Variables for better security.')
app.config['SECRET_KEY'] = _secret_key
_db_url = os.environ.get('DATABASE_URL', 'sqlite:///crm.db')
if _db_url.startswith('postgres://'): _db_url = _db_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = _db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = False
# Flask-Mail (optional — set MAIL_* vars in Railway to enable email)
app.config['MAIL_SERVER']   = os.environ.get('MAIL_SERVER', '')
app.config['MAIL_PORT']     = int(os.environ.get('MAIL_PORT', 587))
app.config['MAIL_USE_TLS']  = os.environ.get('MAIL_USE_TLS', '1') == '1'
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME', '')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD', '')
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_USERNAME', 'noreply@evore.us')

db = SQLAlchemy(app)

# Flask-Mail — graceful degradation if not installed/configured
try:
    from flask_mail import Mail, Message as MailMessage
    _mail = Mail(app)
    _mail_ok = bool(app.config['MAIL_SERVER'])
except ImportError:
    _mail = None; _mail_ok = False

def _send_email(to, subject, body):
    if not _mail_ok or not _mail: return
    try:
        with app.app_context():
            msg = MailMessage(subject, recipients=[to], body=body)
            _mail.send(msg)
    except Exception as e:
        print(f'Email error: {e}')
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Inicia sesión para continuar.'
login_manager.login_message_category = 'warning'

@app.after_request
def security_headers(response):
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    return response

_MODULOS_TODOS = ['clientes','ventas','cotizaciones','tareas','calendario',
                  'notas','inventario','produccion','gastos','reportes']
_MODULOS_ROL = {
    'admin':      _MODULOS_TODOS,
    'vendedor':   ['clientes','ventas','cotizaciones','tareas','calendario','notas'],
    'produccion': ['inventario','produccion','gastos','notas','calendario','tareas'],
    'contador':   ['gastos','reportes','produccion','notas'],
    'usuario':    ['tareas','notas','calendario'],
}

def _modulos_user(user):
    if not user or not user.is_authenticated: return []
    if user.rol == 'admin': return _MODULOS_TODOS
    try:
        custom = json.loads(user.modulos_permitidos or '[]')
        if custom: return custom
    except: pass
    return _MODULOS_ROL.get(user.rol, ['tareas','notas'])

def requiere_modulo(modulo):
    def decorator(f):
        @wraps(f)
        def wrapped(*a, **kw):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))
            if current_user.rol == 'admin' or modulo in _modulos_user(current_user):
                return f(*a, **kw)
            flash('No tienes acceso a este módulo.', 'danger')
            return redirect(url_for('dashboard'))
        return wrapped
    return decorator

@app.context_processor
def inject_globals():
    modulos = _modulos_user(current_user) if current_user.is_authenticated else []
    notif_count = 0
    if current_user.is_authenticated:
        try:
            notif_count = Notificacion.query.filter_by(
                usuario_id=current_user.id, leida=False).count()
        except: pass
    return {'now': datetime.utcnow(), 'modulos_user': modulos, 'notif_count': notif_count}

@app.template_filter('cop')
def cop(value):
    try: return '$ {:,.0f}'.format(float(value or 0)).replace(',','.')
    except: return '$ 0'

@app.template_filter('moneda')
def moneda(value):
    try: return '${:,.2f}'.format(float(value or 0))
    except: return '$0.00'

@app.template_filter('moneda0')
def moneda0(value):
    try: return '${:,.0f}'.format(float(value or 0))
    except: return '$0'

# =============================================================
# MODELOS
# =============================================================

class User(UserMixin, db.Model):
    __tablename__ = 'users'
    id                  = db.Column(db.Integer, primary_key=True)
    nombre              = db.Column(db.String(100), nullable=False)
    email               = db.Column(db.String(120), unique=True, nullable=False)
    password_hash       = db.Column(db.String(256), nullable=False)
    rol                 = db.Column(db.String(20), default='usuario')
    activo              = db.Column(db.Boolean, default=True)
    modulos_permitidos  = db.Column(db.Text, default='[]')   # JSON list
    creado_en           = db.Column(db.DateTime, default=datetime.utcnow)
    def set_password(self, p):   self.password_hash = generate_password_hash(p)
    def check_password(self, p): return check_password_hash(self.password_hash, p)

class ContactoCliente(db.Model):
    __tablename__ = 'contactos_cliente'
    id         = db.Column(db.Integer, primary_key=True)
    cliente_id = db.Column(db.Integer, db.ForeignKey('clientes.id'), nullable=False)
    nombre     = db.Column(db.String(100), nullable=False)
    cargo      = db.Column(db.String(100))
    email      = db.Column(db.String(120))
    telefono   = db.Column(db.String(20))

class Cliente(db.Model):
    __tablename__ = 'clientes'
    id              = db.Column(db.Integer, primary_key=True)
    nombre          = db.Column(db.String(100), nullable=False)
    empresa         = db.Column(db.String(100))
    nit             = db.Column(db.String(30))
    estado_relacion = db.Column(db.String(30), default='prospecto')
    dir_comercial   = db.Column(db.Text)
    dir_entrega     = db.Column(db.Text)
    notas           = db.Column(db.Text)
    estado          = db.Column(db.String(20), default='activo')
    creado_en       = db.Column(db.DateTime, default=datetime.utcnow)
    actualizado_en  = db.Column(db.DateTime, default=datetime.utcnow)
    contactos       = db.relationship('ContactoCliente', backref='cliente_rel', lazy=True, cascade='all, delete-orphan')
    ventas          = db.relationship('Venta', backref='cliente', lazy=True)

class OrdenCompra(db.Model):
    __tablename__ = 'ordenes_compra'
    id                      = db.Column(db.Integer, primary_key=True)
    numero                  = db.Column(db.String(20))               # OC-YYYY-NNN
    proveedor_id            = db.Column(db.Integer, db.ForeignKey('proveedores.id'), nullable=True)
    cotizacion_id           = db.Column(db.Integer, db.ForeignKey('cotizaciones_proveedor.id'), nullable=True)
    transportista_id        = db.Column(db.Integer, db.ForeignKey('proveedores.id'), nullable=True)
    estado                  = db.Column(db.String(30), default='borrador')  # borrador, enviada, recibida, cancelada
    fecha_emision           = db.Column(db.Date)
    fecha_esperada          = db.Column(db.Date)   # fecha entrega esperada (calculada desde cotización)
    fecha_estimada_pago     = db.Column(db.Date)   # fecha estimada de pago al proveedor
    fecha_estimada_recogida = db.Column(db.Date)   # fecha estimada para recoger con transportista
    subtotal                = db.Column(db.Float, default=0)
    iva                     = db.Column(db.Float, default=0)
    total                   = db.Column(db.Float, default=0)
    notas                   = db.Column(db.Text)
    creado_por              = db.Column(db.Integer, db.ForeignKey('users.id'))
    creado_en               = db.Column(db.DateTime, default=datetime.utcnow)
    proveedor               = db.relationship('Proveedor', foreign_keys=[proveedor_id])
    transportista           = db.relationship('Proveedor', foreign_keys=[transportista_id])
    cotizacion_ref          = db.relationship('CotizacionProveedor', foreign_keys=[cotizacion_id])
    items                   = db.relationship('OrdenCompraItem', backref='orden', lazy=True, cascade='all, delete-orphan')

class OrdenCompraItem(db.Model):
    __tablename__ = 'ordenes_compra_items'
    id              = db.Column(db.Integer, primary_key=True)
    orden_id        = db.Column(db.Integer, db.ForeignKey('ordenes_compra.id'), nullable=False)
    cotizacion_id   = db.Column(db.Integer, db.ForeignKey('cotizaciones_proveedor.id'), nullable=True)
    nombre_item     = db.Column(db.String(200), nullable=False)
    descripcion     = db.Column(db.Text)
    cantidad        = db.Column(db.Float, default=1)
    unidad          = db.Column(db.String(30), default='unidades')
    precio_unit     = db.Column(db.Float, default=0)
    subtotal        = db.Column(db.Float, default=0)

class Proveedor(db.Model):
    __tablename__ = 'proveedores'
    id         = db.Column(db.Integer, primary_key=True)
    nombre     = db.Column(db.String(100), nullable=False)
    empresa    = db.Column(db.String(100))
    nit        = db.Column(db.String(30))
    email      = db.Column(db.String(200))
    telefono   = db.Column(db.String(50))
    direccion  = db.Column(db.Text)
    categoria  = db.Column(db.String(100))
    tipo       = db.Column(db.String(20), default='proveedor')  # proveedor, transportista, ambos
    notas      = db.Column(db.Text)
    activo     = db.Column(db.Boolean, default=True)
    creado_en  = db.Column(db.DateTime, default=datetime.utcnow)

class VentaProducto(db.Model):
    __tablename__ = 'venta_productos'
    id          = db.Column(db.Integer, primary_key=True)
    venta_id    = db.Column(db.Integer, db.ForeignKey('ventas.id'), nullable=False)
    producto_id = db.Column(db.Integer, db.ForeignKey('productos.id'), nullable=True)
    nombre_prod = db.Column(db.String(200))
    cantidad    = db.Column(db.Float, default=1)
    precio_unit = db.Column(db.Float, default=0)
    subtotal    = db.Column(db.Float, default=0)

class Venta(db.Model):
    __tablename__ = 'ventas'
    id                  = db.Column(db.Integer, primary_key=True)
    titulo              = db.Column(db.String(200), nullable=False)
    cliente_id          = db.Column(db.Integer, db.ForeignKey('clientes.id'), nullable=True)
    subtotal            = db.Column(db.Float, default=0)
    iva                 = db.Column(db.Float, default=0)
    total               = db.Column(db.Float, default=0)
    porcentaje_anticipo = db.Column(db.Float, default=0)
    monto_anticipo      = db.Column(db.Float, default=0)
    saldo               = db.Column(db.Float, default=0)
    estado              = db.Column(db.String(30), default='prospecto')
    fecha_anticipo      = db.Column(db.Date)
    dias_entrega        = db.Column(db.Integer, default=30)
    fecha_entrega_est   = db.Column(db.Date)
    notas               = db.Column(db.Text)
    creado_en           = db.Column(db.DateTime, default=datetime.utcnow)
    creado_por          = db.Column(db.Integer, db.ForeignKey('users.id'))
    numero              = db.Column(db.String(20))               # v12.3 VNT-YYYY-NNN
    cliente_informado_en= db.Column(db.DateTime, nullable=True)  # v12.2
    entregado_en        = db.Column(db.DateTime, nullable=True)   # v12.2
    items               = db.relationship('VentaProducto', backref='venta', lazy=True, cascade='all, delete-orphan')
    ordenes_produccion  = db.relationship('OrdenProduccion', foreign_keys='OrdenProduccion.venta_id', lazy=True)

class TareaAsignado(db.Model):
    __tablename__ = 'tarea_asignados'
    id       = db.Column(db.Integer, primary_key=True)
    tarea_id = db.Column(db.Integer, db.ForeignKey('tareas.id'), nullable=False)
    user_id  = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    user     = db.relationship('User', foreign_keys=[user_id])

class TareaComentario(db.Model):
    __tablename__ = 'tarea_comentarios'
    id        = db.Column(db.Integer, primary_key=True)
    tarea_id  = db.Column(db.Integer, db.ForeignKey('tareas.id'), nullable=False)
    autor_id  = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    mensaje   = db.Column(db.Text, nullable=False)
    creado_en = db.Column(db.DateTime, default=datetime.utcnow)
    autor     = db.relationship('User', foreign_keys=[autor_id])

class Tarea(db.Model):
    __tablename__ = 'tareas'
    id                = db.Column(db.Integer, primary_key=True)
    titulo            = db.Column(db.String(200), nullable=False)
    descripcion       = db.Column(db.Text)
    estado            = db.Column(db.String(20), default='pendiente')
    prioridad         = db.Column(db.String(10), default='media')
    fecha_vencimiento = db.Column(db.Date)
    asignado_a        = db.Column(db.Integer, db.ForeignKey('users.id'))
    creado_por        = db.Column(db.Integer, db.ForeignKey('users.id'))
    creado_en         = db.Column(db.DateTime, default=datetime.utcnow)
    cotizacion_id     = db.Column(db.Integer, db.ForeignKey('cotizaciones.id'), nullable=True)
    tarea_tipo        = db.Column(db.String(50), nullable=True)   # comprar_materias, verificar_abono
    tarea_pareja_id   = db.Column(db.Integer, db.ForeignKey('tareas.id'), nullable=True)
    asignado_user     = db.relationship('User', foreign_keys=[asignado_a], backref='tareas_asignadas')
    creador           = db.relationship('User', foreign_keys=[creado_por])
    asignados         = db.relationship('TareaAsignado', backref='tarea', lazy=True, cascade='all, delete-orphan')
    comentarios       = db.relationship('TareaComentario', backref='tarea', lazy=True, cascade='all, delete-orphan', order_by='TareaComentario.creado_en')

class Producto(db.Model):
    __tablename__ = 'productos'
    id           = db.Column(db.Integer, primary_key=True)
    nombre       = db.Column(db.String(200), nullable=False)
    descripcion  = db.Column(db.Text)
    sku          = db.Column(db.String(50))
    nso          = db.Column(db.String(50))
    precio       = db.Column(db.Float, default=0)
    costo        = db.Column(db.Float, default=0)
    stock        = db.Column(db.Integer, default=0)
    stock_minimo    = db.Column(db.Integer, default=5)
    categoria       = db.Column(db.String(100))
    activo          = db.Column(db.Boolean, default=True)
    fecha_caducidad = db.Column(db.Date, nullable=True)
    creado_en       = db.Column(db.DateTime, default=datetime.utcnow)
    venta_items     = db.relationship('VentaProducto', backref='producto', lazy=True)

class CompraMateria(db.Model):
    __tablename__ = 'compras_materia'
    id              = db.Column(db.Integer, primary_key=True)
    producto_id     = db.Column(db.Integer, db.ForeignKey('productos.id'), nullable=True)
    materia_id      = db.Column(db.Integer, db.ForeignKey('materias_primas.id'), nullable=True)
    nombre_item     = db.Column(db.String(200), nullable=False)
    tipo_compra     = db.Column(db.String(50), default='insumo')  # materia_prima, producto_terminado, insumo, otro
    unidad          = db.Column(db.String(30), default='unidades')
    proveedor       = db.Column(db.String(200))
    proveedor_id    = db.Column(db.Integer, db.ForeignKey('proveedores.id'), nullable=True)
    fecha           = db.Column(db.Date, nullable=False)
    nro_factura     = db.Column(db.String(100))
    cantidad        = db.Column(db.Float, default=1)
    costo_producto  = db.Column(db.Float, default=0)
    impuestos       = db.Column(db.Float, default=0)
    transporte      = db.Column(db.Float, default=0)
    costo_total     = db.Column(db.Float, default=0)
    precio_unitario = db.Column(db.Float, default=0)
    tiene_caducidad = db.Column(db.Boolean, default=False)
    fecha_caducidad = db.Column(db.Date, nullable=True)
    notas           = db.Column(db.Text)
    creado_por      = db.Column(db.Integer, db.ForeignKey('users.id'))
    creado_en       = db.Column(db.DateTime, default=datetime.utcnow)
    producto        = db.relationship('Producto', foreign_keys=[producto_id])
    materia         = db.relationship('MateriaPrima', foreign_keys=[materia_id])
    proveedor_rel   = db.relationship('Proveedor', foreign_keys=[proveedor_id])

class CotizacionProveedor(db.Model):
    __tablename__ = 'cotizaciones_proveedor'
    id                    = db.Column(db.Integer, primary_key=True)
    numero                = db.Column(db.String(20))               # CP-YYYY-NNN
    proveedor_id          = db.Column(db.Integer, db.ForeignKey('proveedores.id'), nullable=True)
    tipo_cotizacion       = db.Column(db.String(20), default='general')  # granel, general
    tipo_producto_servicio= db.Column(db.String(100))  # ej: materia prima, servicio logístico, empaque
    nombre_producto       = db.Column(db.String(200), nullable=False)
    descripcion           = db.Column(db.Text)
    sku                   = db.Column(db.String(50))
    precio_unitario       = db.Column(db.Float, default=0)
    unidades_minimas      = db.Column(db.Integer, default=1)
    unidad                = db.Column(db.String(30), default='unidades')
    plazo_entrega         = db.Column(db.String(100))   # texto descriptivo legacy
    plazo_entrega_dias    = db.Column(db.Integer, default=0)   # días hábiles (structured)
    condiciones_pago      = db.Column(db.String(200))   # texto legacy
    condicion_pago_tipo   = db.Column(db.String(30), default='contado')  # contado, credito, anticipo_saldo, consignacion
    condicion_pago_dias   = db.Column(db.Integer, default=0)  # días de crédito
    anticipo_porcentaje   = db.Column(db.Float, default=0)    # % anticipo si aplica
    fecha_cotizacion      = db.Column(db.Date)
    vigencia              = db.Column(db.Date)
    estado                = db.Column(db.String(20), default='vigente')  # vigente, vencida, en_revision
    notas                 = db.Column(db.Text)
    calendario_integrado  = db.Column(db.Boolean, default=False)  # ya se agendó entrega al calendario
    creado_por            = db.Column(db.Integer, db.ForeignKey('users.id'))
    creado_en             = db.Column(db.DateTime, default=datetime.utcnow)
    proveedor             = db.relationship('Proveedor', foreign_keys=[proveedor_id])

class CotizacionGranel(db.Model):
    __tablename__ = 'cotizaciones_granel'
    id               = db.Column(db.Integer, primary_key=True)
    producto_id      = db.Column(db.Integer, db.ForeignKey('productos.id'), nullable=True)
    nombre_producto  = db.Column(db.String(200), nullable=False)
    sku              = db.Column(db.String(50))
    nso              = db.Column(db.String(50))
    proveedor        = db.Column(db.String(200))
    precio_unitario  = db.Column(db.Float, default=0)
    unidades_minimas = db.Column(db.Integer, default=1)
    fecha_cotizacion = db.Column(db.Date)
    vigencia         = db.Column(db.Date)
    estado           = db.Column(db.String(20), default='vigente')
    notas            = db.Column(db.Text)
    creado_por       = db.Column(db.Integer, db.ForeignKey('users.id'))
    creado_en        = db.Column(db.DateTime, default=datetime.utcnow)
    producto         = db.relationship('Producto', foreign_keys=[producto_id])

class DocumentoLegal(db.Model):
    __tablename__ = 'documentos_legales'
    id                = db.Column(db.Integer, primary_key=True)
    tipo              = db.Column(db.String(50), nullable=False)
    # tipos: permiso_sanitario, registro_invima, nso, contrato, licencia, certificado, otro
    titulo            = db.Column(db.String(200), nullable=False)
    numero            = db.Column(db.String(100))   # número del documento
    entidad           = db.Column(db.String(200))   # entidad emisora (INVIMA, cámara de comercio, etc.)
    descripcion       = db.Column(db.Text)
    estado            = db.Column(db.String(20), default='vigente')  # vigente, vencido, en_tramite, suspendido
    fecha_emision     = db.Column(db.Date)
    fecha_vencimiento = db.Column(db.Date)
    recordatorio_dias = db.Column(db.Integer, default=30)  # días antes de vencer para alertar
    archivo_url       = db.Column(db.String(500))
    notas             = db.Column(db.Text)
    activo            = db.Column(db.Boolean, default=True)
    creado_por        = db.Column(db.Integer, db.ForeignKey('users.id'))
    creado_en         = db.Column(db.DateTime, default=datetime.utcnow)

class AsientoContable(db.Model):
    __tablename__ = 'asientos_contables'
    id          = db.Column(db.Integer, primary_key=True)
    numero      = db.Column(db.String(20))     # AC-YYYY-NNN
    fecha       = db.Column(db.Date, nullable=False)
    descripcion = db.Column(db.String(300), nullable=False)
    tipo        = db.Column(db.String(30), default='manual')  # manual, venta, compra, gasto
    referencia  = db.Column(db.String(100))
    debe        = db.Column(db.Float, default=0)
    haber       = db.Column(db.Float, default=0)
    cuenta_debe = db.Column(db.String(100))
    cuenta_haber= db.Column(db.String(100))
    notas       = db.Column(db.Text)
    creado_por  = db.Column(db.Integer, db.ForeignKey('users.id'))
    creado_en   = db.Column(db.DateTime, default=datetime.utcnow)

class ReglaTributaria(db.Model):
    __tablename__ = 'reglas_tributarias'
    id               = db.Column(db.Integer, primary_key=True)
    nombre           = db.Column(db.String(100), nullable=False)
    descripcion      = db.Column(db.Text)
    porcentaje       = db.Column(db.Float, default=0)
    aplica_a         = db.Column(db.String(30), default='ventas')  # ventas, ingresos, profit, proveedor_producto, proveedor_granel
    proveedor_nombre = db.Column(db.String(200))
    activo           = db.Column(db.Boolean, default=True)
    creado_en        = db.Column(db.DateTime, default=datetime.utcnow)

class GastoOperativo(db.Model):
    __tablename__ = 'gastos_operativos'
    id           = db.Column(db.Integer, primary_key=True)
    fecha        = db.Column(db.Date, nullable=False)
    tipo         = db.Column(db.String(50), nullable=False)
    tipo_custom  = db.Column(db.String(100))
    descripcion  = db.Column(db.String(200))
    monto        = db.Column(db.Float, default=0, nullable=False)
    recurrencia  = db.Column(db.String(10), default='unico')  # unico, mensual
    es_plantilla = db.Column(db.Boolean, default=False)
    notas        = db.Column(db.Text)
    creado_por   = db.Column(db.Integer, db.ForeignKey('users.id'))
    creado_en    = db.Column(db.DateTime, default=datetime.utcnow)

class Nota(db.Model):
    __tablename__ = 'notas'
    id             = db.Column(db.Integer, primary_key=True)
    titulo         = db.Column(db.String(200))
    contenido      = db.Column(db.Text, nullable=False)
    cliente_id     = db.Column(db.Integer, db.ForeignKey('clientes.id'), nullable=True)
    producto_id    = db.Column(db.Integer, db.ForeignKey('productos.id'), nullable=True)
    modulo         = db.Column(db.String(50))   # ventas, produccion, inventario, gastos, tareas, otro
    fecha_revision = db.Column(db.Date, nullable=True)
    creado_por     = db.Column(db.Integer, db.ForeignKey('users.id'))
    creado_en      = db.Column(db.DateTime, default=datetime.utcnow)
    actualizado_en = db.Column(db.DateTime, default=datetime.utcnow)
    cliente        = db.relationship('Cliente', foreign_keys=[cliente_id])
    producto       = db.relationship('Producto', foreign_keys=[producto_id])
    autor          = db.relationship('User', foreign_keys=[creado_por])

class Actividad(db.Model):
    __tablename__ = 'actividades'
    id          = db.Column(db.Integer, primary_key=True)
    tipo        = db.Column(db.String(20))   # crear, editar, eliminar, completar
    entidad     = db.Column(db.String(50))   # cliente, venta, tarea, nota...
    entidad_id  = db.Column(db.Integer)
    descripcion = db.Column(db.String(300))
    usuario_id  = db.Column(db.Integer, db.ForeignKey('users.id'))
    creado_en   = db.Column(db.DateTime, default=datetime.utcnow)
    usuario     = db.relationship('User', foreign_keys=[usuario_id])

class ConfigEmpresa(db.Model):
    __tablename__ = 'config_empresa'
    id        = db.Column(db.Integer, primary_key=True)
    nombre    = db.Column(db.String(200), default='Evore')
    nit       = db.Column(db.String(30))
    direccion = db.Column(db.Text)
    telefono  = db.Column(db.String(30))
    email     = db.Column(db.String(120))
    ciudad    = db.Column(db.String(100))
    sitio_web = db.Column(db.String(200))

class Evento(db.Model):
    __tablename__ = 'eventos'
    id          = db.Column(db.Integer, primary_key=True)
    titulo      = db.Column(db.String(200), nullable=False)
    tipo        = db.Column(db.String(20), default='recordatorio')  # cita, reunion, recordatorio
    fecha       = db.Column(db.Date, nullable=False)
    hora_inicio = db.Column(db.String(5))
    hora_fin    = db.Column(db.String(5))
    descripcion = db.Column(db.Text)
    usuario_id  = db.Column(db.Integer, db.ForeignKey('users.id'))
    creado_en   = db.Column(db.DateTime, default=datetime.utcnow)
    usuario     = db.relationship('User', foreign_keys=[usuario_id])

class CotizacionItem(db.Model):
    __tablename__ = 'cotizacion_items'
    id            = db.Column(db.Integer, primary_key=True)
    cotizacion_id = db.Column(db.Integer, db.ForeignKey('cotizaciones.id'), nullable=False)
    producto_id   = db.Column(db.Integer, db.ForeignKey('productos.id'), nullable=True)
    nombre_prod   = db.Column(db.String(200))
    cantidad      = db.Column(db.Float, default=1)
    precio_unit   = db.Column(db.Float, default=0)
    subtotal      = db.Column(db.Float, default=0)

class Cotizacion(db.Model):
    __tablename__ = 'cotizaciones'
    id                  = db.Column(db.Integer, primary_key=True)
    numero              = db.Column(db.String(20))
    titulo              = db.Column(db.String(200), nullable=False)
    cliente_id          = db.Column(db.Integer, db.ForeignKey('clientes.id'), nullable=True)
    subtotal            = db.Column(db.Float, default=0)
    iva                 = db.Column(db.Float, default=0)
    total               = db.Column(db.Float, default=0)
    porcentaje_anticipo = db.Column(db.Float, default=50)
    monto_anticipo      = db.Column(db.Float, default=0)
    saldo               = db.Column(db.Float, default=0)
    estado              = db.Column(db.String(30), default='borrador')  # borrador, enviada, aprobada, confirmacion_orden
    fecha_emision       = db.Column(db.Date, default=date_type.today)
    fecha_validez       = db.Column(db.Date)
    dias_entrega        = db.Column(db.Integer, default=30)
    fecha_entrega_est   = db.Column(db.Date)
    condiciones_pago    = db.Column(db.Text)
    notas               = db.Column(db.Text)
    creado_en           = db.Column(db.DateTime, default=datetime.utcnow)
    creado_por          = db.Column(db.Integer, db.ForeignKey('users.id'))
    items               = db.relationship('CotizacionItem', backref='cotizacion', lazy=True, cascade='all, delete-orphan')
    cliente             = db.relationship('Cliente', foreign_keys=[cliente_id])

class LoteProducto(db.Model):
    __tablename__ = 'lotes_producto'
    id                  = db.Column(db.Integer, primary_key=True)
    producto_id         = db.Column(db.Integer, db.ForeignKey('productos.id'), nullable=False)
    numero_lote         = db.Column(db.String(80), nullable=False)
    nso                 = db.Column(db.String(80))
    fecha_produccion    = db.Column(db.Date)
    fecha_vencimiento   = db.Column(db.Date)
    unidades_producidas = db.Column(db.Float, default=0)
    unidades_restantes  = db.Column(db.Float, default=0)
    notas               = db.Column(db.Text)
    creado_por          = db.Column(db.Integer, db.ForeignKey('users.id'))
    creado_en           = db.Column(db.DateTime, default=datetime.utcnow)
    producto            = db.relationship('Producto', foreign_keys=[producto_id], backref='lotes')

class MateriaPrima(db.Model):
    __tablename__ = 'materias_primas'
    id               = db.Column(db.Integer, primary_key=True)
    nombre           = db.Column(db.String(200), nullable=False)
    descripcion      = db.Column(db.Text)
    unidad           = db.Column(db.String(30), default='unidades')  # kg, g, litros, ml, unidades
    stock_disponible = db.Column(db.Float, default=0)
    stock_reservado  = db.Column(db.Float, default=0)
    stock_minimo     = db.Column(db.Float, default=0)
    costo_unitario   = db.Column(db.Float, default=0)
    categoria        = db.Column(db.String(100))
    proveedor        = db.Column(db.String(200))
    proveedor_id     = db.Column(db.Integer, db.ForeignKey('proveedores.id'), nullable=True)
    producto_id      = db.Column(db.Integer, db.ForeignKey('productos.id'), nullable=True)
    activo           = db.Column(db.Boolean, default=True)
    creado_en        = db.Column(db.DateTime, default=datetime.utcnow)
    producto         = db.relationship('Producto', foreign_keys=[producto_id])
    proveedor_rel    = db.relationship('Proveedor', foreign_keys=[proveedor_id])

class RecetaProducto(db.Model):
    __tablename__ = 'recetas_producto'
    id               = db.Column(db.Integer, primary_key=True)
    producto_id      = db.Column(db.Integer, db.ForeignKey('productos.id'), nullable=False)
    unidades_produce = db.Column(db.Integer, default=1)    # cuántas unidades produce esta receta
    descripcion      = db.Column(db.Text)
    activo           = db.Column(db.Boolean, default=True)
    creado_en        = db.Column(db.DateTime, default=datetime.utcnow)
    producto         = db.relationship('Producto', foreign_keys=[producto_id], backref='receta')
    items            = db.relationship('RecetaItem', backref='receta', lazy=True,
                                       cascade='all, delete-orphan')

class RecetaItem(db.Model):
    __tablename__ = 'receta_items'
    id               = db.Column(db.Integer, primary_key=True)
    receta_id        = db.Column(db.Integer, db.ForeignKey('recetas_producto.id'), nullable=False)
    materia_prima_id = db.Column(db.Integer, db.ForeignKey('materias_primas.id'), nullable=False)
    cantidad_por_unidad = db.Column(db.Float, default=0)
    materia          = db.relationship('MateriaPrima', foreign_keys=[materia_prima_id])

class ReservaProduccion(db.Model):
    __tablename__ = 'reservas_produccion'
    id               = db.Column(db.Integer, primary_key=True)
    materia_prima_id = db.Column(db.Integer, db.ForeignKey('materias_primas.id'), nullable=False)
    cantidad         = db.Column(db.Float, default=0)
    estado           = db.Column(db.String(20), default='reservado')  # reservado, usado, cancelado
    producto_id      = db.Column(db.Integer, db.ForeignKey('productos.id'), nullable=True)
    lote_id          = db.Column(db.Integer, db.ForeignKey('lotes_producto.id'), nullable=True)
    notas            = db.Column(db.Text)
    creado_por       = db.Column(db.Integer, db.ForeignKey('users.id'))
    creado_en        = db.Column(db.DateTime, default=datetime.utcnow)
    venta_id         = db.Column(db.Integer, db.ForeignKey('ventas.id'), nullable=True)
    materia          = db.relationship('MateriaPrima', foreign_keys=[materia_prima_id])
    producto         = db.relationship('Producto', foreign_keys=[producto_id])
    venta            = db.relationship('Venta', foreign_keys=[venta_id])

class OrdenProduccion(db.Model):
    __tablename__ = 'ordenes_produccion'
    id                = db.Column(db.Integer, primary_key=True)
    cotizacion_id     = db.Column(db.Integer, db.ForeignKey('cotizaciones.id'), nullable=True)
    venta_id          = db.Column(db.Integer, db.ForeignKey('ventas.id'), nullable=True)
    producto_id       = db.Column(db.Integer, db.ForeignKey('productos.id'), nullable=False)
    cantidad_total    = db.Column(db.Float, default=0)   # total requerido
    cantidad_stock    = db.Column(db.Float, default=0)   # ya en stock al crear
    cantidad_producir = db.Column(db.Float, default=0)   # diferencia a producir
    numero_lote       = db.Column(db.String(80))
    estado            = db.Column(db.String(30), default='en_produccion')
    # pendiente_materiales | en_produccion | completado
    notas             = db.Column(db.Text)
    creado_por        = db.Column(db.Integer, db.ForeignKey('users.id'))
    creado_en         = db.Column(db.DateTime, default=datetime.utcnow)
    completado_en     = db.Column(db.DateTime, nullable=True)
    fecha_inicio_real = db.Column(db.Date, nullable=True)      # v14 Gantt
    fecha_fin_estimada= db.Column(db.Date, nullable=True)      # v14 Gantt
    producto          = db.relationship('Producto', foreign_keys=[producto_id])
    cotizacion        = db.relationship('Cotizacion', foreign_keys=[cotizacion_id])
    venta             = db.relationship('Venta', foreign_keys=[venta_id])

class Notificacion(db.Model):
    __tablename__ = 'notificaciones'
    id         = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    tipo       = db.Column(db.String(40), default='info')  # tarea_asignada, alerta_stock, info
    titulo     = db.Column(db.String(200), nullable=False)
    mensaje    = db.Column(db.Text)
    url        = db.Column(db.String(300))
    leida      = db.Column(db.Boolean, default=False)
    creado_en  = db.Column(db.DateTime, default=datetime.utcnow)
    usuario    = db.relationship('User', foreign_keys=[usuario_id])

@login_manager.user_loader
def load_user(uid): return User.query.get(int(uid))

# =============================================================
# TEMPLATES
# =============================================================

_CSS = """{% raw %}<style>
:root{--sb:#1a1f36;--ac:#5e72e4;--bg:#f4f6fb}
body{background:var(--bg);font-family:'Segoe UI',sans-serif}
#sb{position:fixed;top:0;left:0;height:100vh;width:252px;background:var(--sb);display:flex;flex-direction:column;z-index:1000}
.sb-brand{padding:1.3rem 1.2rem .85rem;color:#fff;font-size:1.3rem;font-weight:700;border-bottom:1px solid rgba(255,255,255,.08);letter-spacing:1px}
.sb-brand span{color:var(--ac)}
.sb-sec{font-size:.67rem;font-weight:700;letter-spacing:1.5px;text-transform:uppercase;color:rgba(255,255,255,.3);padding:.65rem 1.2rem .2rem}
.sb-nav .nav-link{color:#a8b0d3;padding:.48rem 1.2rem;border-radius:8px;margin:.05rem .65rem;display:flex;align-items:center;gap:.65rem;font-size:.87rem;transition:all .2s}
.sb-nav .nav-link:hover{background:rgba(255,255,255,.07);color:#fff}
.sb-nav .nav-link.active{background:var(--ac);color:#fff}
.sb-nav .nav-link i{font-size:1rem;width:19px}
.sb-foot{padding:.85rem 1.2rem;border-top:1px solid rgba(255,255,255,.08);color:#a8b0d3;font-size:.82rem;margin-top:auto}
.u-name{color:#fff;font-weight:600}
.u-rol{font-size:.67rem;padding:2px 8px;border-radius:20px;background:rgba(94,114,228,.3);color:var(--ac)}
#main{margin-left:252px;min-height:100vh}
.topbar{background:#fff;padding:.68rem 1.4rem;border-bottom:1px solid #e8ecf0;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:100;box-shadow:0 1px 4px rgba(0,0,0,.05)}
.pg-title{font-size:1.1rem;font-weight:600;color:#1a1f36;margin:0}
.content{padding:1.4rem}
.sc{background:#fff;border-radius:12px;padding:1.2rem 1.4rem;box-shadow:0 2px 8px rgba(0,0,0,.06);transition:transform .2s}
.sc:hover{transform:translateY(-2px);box-shadow:0 4px 16px rgba(0,0,0,.1)}
.si{width:46px;height:46px;border-radius:11px;display:flex;align-items:center;justify-content:center;font-size:1.3rem}
.sv{font-size:1.7rem;font-weight:700;color:#1a1f36}
.sl{color:#8898aa;font-size:.82rem}
.tc{background:#fff;border-radius:12px;box-shadow:0 2px 8px rgba(0,0,0,.06);overflow:hidden}
.ch{background:#fff;border-bottom:1px solid #f0f0f0;padding:.85rem 1.4rem;font-weight:600;color:#1a1f36}
.table{margin:0}
.table th{font-size:.71rem;font-weight:700;text-transform:uppercase;letter-spacing:.5px;color:#8898aa;border-bottom:1px solid #f0f0f0;padding:.65rem 1rem}
.table td{padding:.65rem 1rem;vertical-align:middle;border-bottom:1px solid #f8f9fa;color:#525f7f}
.table tbody tr:last-child td{border-bottom:none}
.table tbody tr:hover{background:#f8f9fe}
.b{padding:3px 10px;border-radius:20px;font-size:.72rem;font-weight:600}
.b-activo,.b-ganado,.b-completada,.b-baja,.b-vip,.b-vigente,.b-aprobada{background:#d4edda;color:#155724}
.b-inactivo,.b-perdido,.b-alta,.b-vencida{background:#f8d7da;color:#721c24}
.b-prospecto,.b-pendiente,.b-media,.b-anticipo_pagado,.b-en_revision,.b-enviada{background:#fff3cd;color:#856404}
.b-negociacion,.b-en_progreso,.b-cliente_activo,.b-confirmacion_orden{background:#cce5ff;color:#004085}
.b-borrador{background:#e9ecef;color:#495057}
.b-unico{background:#e8f4fd;color:#0c5460}
.b-mensual{background:#d1ecf1;color:#0c5460}
.b-cita{background:#fce4ec;color:#880e4f}
.b-reunion{background:#e8eaf6;color:#283593}
.b-recordatorio{background:#fff8e1;color:#f57f17}
.ev-tarea{background:#fff4e5;color:#fb6340;border-left:3px solid #fb6340}
.ev-evento{background:#e8eeff;color:#5e72e4;border-left:3px solid #5e72e4}
.ev-nota{background:#f3e8ff;color:#8965e0;border-left:3px solid #8965e0}
.ev-caducidad{background:#fde8e8;color:#f5365c;border-left:3px solid #f5365c}
.ev-venta{background:#e3f9ee;color:#2dce89;border-left:3px solid #2dce89}
.cal-day{border:1px solid #f0f2ff;min-height:110px;vertical-align:top;padding:.3rem;cursor:pointer;transition:background .15s}
.cal-day:hover{background:#f8f9fe}
.cal-day.today{background:#eef0ff}
.cal-day.other-month{background:#fafbff;opacity:.6}
.cal-ev{font-size:.7rem;padding:2px 5px;border-radius:4px;margin-bottom:2px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;cursor:pointer}
.fc{background:#fff;border-radius:12px;padding:1.8rem;box-shadow:0 2px 8px rgba(0,0,0,.06);max-width:820px}
.form-label{font-weight:600;font-size:.875rem;color:#525f7f}
.form-control,.form-select{border:1.5px solid #e9ecef;border-radius:8px;padding:.5rem .75rem;font-size:.9rem;transition:border-color .2s}
.form-control:focus,.form-select:focus{border-color:var(--ac);box-shadow:0 0 0 3px rgba(94,114,228,.15)}
.btn-primary{background:var(--ac);border-color:var(--ac)}
.btn-primary:hover{background:#4a5bd4;border-color:#4a5bd4}
.alert{border-radius:10px;border:none}
.prod-row{background:#f8f9fe;border-radius:8px;padding:.75rem;margin-bottom:.5rem;border:1px solid #e8ecf0}
.totales-box{background:#f8f9fe;border-radius:10px;padding:1rem 1.4rem;border:1px solid #e8ecf0}
.chat-bubble{background:#f0f2ff;border-radius:12px 12px 12px 2px;padding:.65rem 1rem;margin-bottom:.5rem;max-width:80%}
.chat-bubble.mine{background:#5e72e4;color:#fff;border-radius:12px 12px 2px 12px;margin-left:auto}
.chat-bubble.mine .chat-meta{color:rgba(255,255,255,.7)}
.chat-meta{font-size:.72rem;color:#8898aa;margin-top:.25rem}
@media(max-width:768px){#sb{width:54px}#sb .nav-link span,.sb-brand .bt,.sb-sec,.ui{display:none}#main{margin-left:54px}}
/* Notificaciones */
.notif-btn{position:relative;background:none;border:none;color:rgba(255,255,255,.7);font-size:1.1rem;padding:.3rem .45rem;border-radius:8px;cursor:pointer;transition:all .2s}
.notif-btn:hover{color:#fff;background:rgba(255,255,255,.1)}
.notif-badge{position:absolute;top:0;right:0;background:#f5365c;color:#fff;font-size:.6rem;font-weight:700;min-width:16px;height:16px;border-radius:8px;display:flex;align-items:center;justify-content:center;padding:0 3px;line-height:1}
.notif-dd{position:fixed;bottom:70px;left:0;width:300px;max-height:400px;overflow-y:auto;background:#fff;border-radius:12px;box-shadow:0 -4px 30px rgba(0,0,0,.18);z-index:9999}
.notif-dd-head{padding:.75rem 1rem;font-size:.8rem;font-weight:700;color:#1a1f36;border-bottom:1px solid #f0f2ff;display:flex;justify-content:space-between;align-items:center}
.notif-item{padding:.65rem 1rem;border-bottom:1px solid #f8f9fe;cursor:pointer;transition:background .15s}
.notif-item:hover{background:#f8f9fe}
.notif-item.unread{background:#f0f4ff}
.notif-item .ni-title{font-size:.82rem;font-weight:600;color:#1a1f36}
.notif-item .ni-msg{font-size:.75rem;color:#525f7f;margin-top:2px}
.notif-item .ni-time{font-size:.68rem;color:#adb5bd}
/* Diagnóstico */
#diagBtn{position:fixed;bottom:28px;right:28px;z-index:2000;width:52px;height:52px;border-radius:50%;background:linear-gradient(135deg,#5e72e4,#825ee4);border:none;color:#fff;font-size:1.3rem;box-shadow:0 4px 20px rgba(94,114,228,.4);cursor:pointer;display:flex;align-items:center;justify-content:center;transition:all .25s}
#diagBtn:hover{transform:scale(1.1);box-shadow:0 6px 28px rgba(94,114,228,.55)}
#diagPanel{position:fixed;bottom:90px;right:28px;width:360px;background:#fff;border-radius:16px;box-shadow:0 8px 40px rgba(0,0,0,.18);z-index:1999;overflow:hidden;display:none}
#diagPanel .dp-head{padding:.9rem 1.2rem;background:linear-gradient(135deg,#1a1f36,#2d3561);color:#fff;font-weight:700;font-size:.9rem;display:flex;justify-content:space-between}
#diagPanel .dp-body{max-height:420px;overflow-y:auto;padding:.75rem 1rem}
.diag-item{padding:.5rem .75rem;border-radius:8px;margin-bottom:.4rem;font-size:.82rem;display:flex;gap:.6rem;align-items:flex-start}
.diag-rojo{background:#fde8e8;color:#f5365c}
.diag-amarillo{background:#fff4e5;color:#fb6340}
.diag-verde{background:#e3f9ee;color:#2dce89}
/* Materias primas / lotes */
.mp-row{background:#f8f9fe;border-radius:8px;padding:.6rem .9rem;margin-bottom:.4rem;border-left:3px solid #5e72e4}
.lote-row{background:#f8f9fe;border-radius:8px;padding:.6rem .9rem;margin-bottom:.4rem;border-left:3px solid #2dce89}
.lote-venc{border-left-color:#f5365c!important}
.lote-warn{border-left-color:#fb6340!important}
.stock-bar-wrap{background:#e8ecf5;border-radius:4px;height:7px;overflow:hidden;flex:1}
.stock-bar{height:100%;border-radius:4px;background:#2dce89}
.stock-bar.warn{background:#fb6340}.stock-bar.bad{background:#f5365c}
</style>{% endraw %}"""

_CDN = """<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
<link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css" rel="stylesheet">"""
_BSJ = '<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>'

T = {}

T['base.html'] = """<!DOCTYPE html>
<html lang="es"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{% block title %}Evore CRM{% endblock %}</title>
""" + _CDN + _CSS + """
</head><body>
<nav id="sb">
  <div class="sb-brand" style="padding:.85rem 1rem .7rem"><a href="/" style="text-decoration:none;color:inherit"><div style="overflow:hidden;line-height:0"><svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 474.45 119.52" width="188" height="47" style="filter:brightness(0) invert(1);opacity:.88;display:block"><path d="M163.87,76.22V43.3c0-13-1.37-15.81-1.37-15.81h45.27v4.07s-4.09-2.22-23.45-2.22H167.87V54.21h30.9v4.07S196.13,56,183.41,56H167.87V90.18H187C206.4,90.18,210.5,88,210.5,88V92h-48S163.87,89.26,163.87,76.22Z"/><path d="M215.13,27.49h7s-.28,2.77,4.72,15.81l17.91,46.33L263.21,43.3c5.18-13,3.82-15.81,3.82-15.81h5.82s-2.46,2.68-7.46,15.16L245.76,92h-4.27l-19-49C217.58,30.17,215.13,27.49,215.13,27.49Z"/><path d="M274.07,59.67c0-18.5,14.45-33.39,32.36-33.39s32.36,14.89,32.36,33.39-14.46,33.38-32.36,33.38S274.07,78.16,274.07,59.67Zm60.81,0c0-17.48-12.73-31.63-28.45-31.63S278,42.19,278,59.67s12.72,31.62,28.45,31.62S334.88,77.15,334.88,59.67Z"/><path d="M346.1,76.22V43.3c0-13-1.36-15.81-1.36-15.81h28.63c11,0,19.91,9.34,19.91,20.8s-8.73,20.63-19.55,20.81L381,76.5C393.55,89.35,397.64,92,397.64,92h-8.36S388,89.35,375.73,76.87l-7.54-7.77H350.1v7.12c0,13,1.36,15.81,1.36,15.81h-6.72S346.1,89.26,346.1,76.22Zm43.27-27.93c0-10.35-7.18-18.95-16-18.95H350.1V67.25h23.27C382.19,67.25,389.37,58.65,389.37,48.29Z"/><path d="M403.73,76.22V43.3c0-13-1.37-15.81-1.37-15.81h45.27v4.07s-4.09-2.22-23.45-2.22H407.72V54.21h30.91v4.07S436,56,423.27,56H407.72V90.18H426.9c19.36,0,23.45-2.21,23.45-2.21V92h-48S403.73,89.26,403.73,76.22Z"/><path d="M132.51,80.73a88.84,88.84,0,0,0-10.58-4.88s-1.06-.37-1-.67c.16-.91,2.08-1.17,7.25-8.06,5.35-7.14,6.47-10.49,7.48-13.42.86-2.49,1.79-8.27,1.29-8.87s-6.86-1-12.4.22c-5.31,1.2-8.35,2.62-8.65,2.28-.48-.53.71-4.69.83-8.52a36.63,36.63,0,0,0-.3-8.65c-.47-3-1.11-5.7-2-5.95s-3.76.62-5.26,1.27a62.18,62.18,0,0,0-7.26,4.28c-2.19,1.52-4.91,4.13-5.68,4.29s-2.1-4.53-5.19-9.9a64.94,64.94,0,0,0-4.29-6.84C85.44,15.47,83.05,12,81.26,12h0c-1.79,0-4.17,3.45-5.48,5.29a63.75,63.75,0,0,0-4.28,6.84c-3.1,5.37-4.34,10.07-5.19,9.9s-3.56-2.86-5.75-4.37a60.26,60.26,0,0,0-7.2-4.2c-1.5-.65-4.37-1.53-5.26-1.27s-1.53,3-2,5.95a36.08,36.08,0,0,0-.31,8.65c.12,3.83,1.32,8,.84,8.52-.3.34-3.35-1.08-8.66-2.28-5.53-1.24-11.82-.92-12.4-.22s.44,6.38,1.29,8.87c1,2.93,2.13,6.28,7.49,13.42,5.17,6.89,7.08,7.15,7.24,8.06.06.3-1,.67-1,.67A88.84,88.84,0,0,0,30,80.73c-2.76,1.7-5.9,4.24-5.9,5.54,0,1.59,2.67,3.48,6.14,5.78s11.55,5.88,20.12,5.07c9.65-.91,15.56-5.42,16.37-5.07.37.16.08,3.94,3.88,7.86,3.34,3.43,8,7.5,10.65,7.5s7.31-4.07,10.65-7.5c3.8-3.92,3.51-7.7,3.88-7.86.81-.35,6.73,4.16,16.37,5.07,8.58.81,16.61-2.75,20.13-5.07s6.14-4.19,6.13-5.78C138.41,85,135.27,82.43,132.51,80.73ZM120.37,48.38c6.29-2.14,13.77-2.27,14.29-1.59s-.51,11.81-13.15,24.48c-8.57,8.59-16.05,9.66-22.83,10.46a49.69,49.69,0,0,1-12.32-.54,26.06,26.06,0,0,0,8.55-9c3.31-5.9,4.74-10.88,10.5-15.29S114.07,50.51,120.37,48.38ZM104.31,30.92c5.53-3.75,8.42-4.45,8.92-4.12s1.29,3.49,1.25,9.2a88,88,0,0,1-1.28,12.24s-1.74.95-6.5,4.28a38.24,38.24,0,0,0-7.13,6.11,56.1,56.1,0,0,0-.32-12A73.79,73.79,0,0,0,97,37.13S98.78,34.67,104.31,30.92Zm-33-1.39c3-6,8.24-14.66,9.91-14.66s7,8.68,9.91,14.66a59.24,59.24,0,0,1,6.13,25,32.38,32.38,0,0,1-7.47,20c-3.88,4.88-8.13,6.14-8.57,6.14s-4.69-1.26-8.57-6.14a32.38,32.38,0,0,1-7.47-20A59.24,59.24,0,0,1,71.35,29.53ZM49.29,26.8c.5-.33,3.4.37,8.92,4.12s7.36,6.21,7.36,6.21a72.86,72.86,0,0,0-2.3,9.53,56.1,56.1,0,0,0-.32,12,38.24,38.24,0,0,0-7.13-6.11c-4.76-3.33-6.5-4.28-6.5-4.28A88,88,0,0,1,48,36C48,30.29,48.77,27.14,49.29,26.8ZM41,71.27C28.37,58.6,27.35,47.47,27.86,46.79s8-.55,14.3,1.59,9.19,4.14,15,8.54,7.2,9.39,10.5,15.29a26.06,26.06,0,0,0,8.55,9,49.69,49.69,0,0,1-12.32.54C57.06,80.93,49.58,79.86,41,71.27ZM62.77,91.11a31.75,31.75,0,0,1-17.52,3.76c-4.91-.33-8.86-2.14-12.74-4.16s-5.77-3.82-5.77-4.54,3.57-3.23,6.92-4.81a85.68,85.68,0,0,1,9-3.93c.81-.28,1-.05,2.33.69a30.76,30.76,0,0,0,9.68,4.43,107,107,0,0,0,14,2S66.75,89.14,62.77,91.11Zm25.1,9.17c-4.37,4.22-6,4.41-6.61,4.41s-2.24-.19-6.61-4.41-6.23-8.18-4.86-12.1a5.72,5.72,0,0,1,1.6-2.69c2.73-2.62,9.86-2.33,9.86-2.33h0s7.12-.29,9.85,2.33a5.72,5.72,0,0,1,1.6,2.69C94.1,92.1,92.23,96.06,87.87,100.28ZM130,90.71c-3.88,2-7.83,3.83-12.74,4.16a31.75,31.75,0,0,1-17.52-3.76c-4-2-6-6.58-6-6.58a107,107,0,0,0,14-2,30.63,30.63,0,0,0,9.68-4.43c1.39-.74,1.52-1,2.34-.69a86.72,86.72,0,0,1,9,3.93c3.36,1.58,6.92,4.1,6.92,4.81S134,88.63,130,90.71Z"/></svg><span class="bt" style="font-size:.65rem;letter-spacing:1px;opacity:.55;white-space:nowrap"> CRM</span></div></a></div>
  <div class="sb-nav py-2" style="overflow-y:auto;flex:1">
    <div class="sb-sec">Principal</div>
    <a href="{{ url_for('dashboard') }}" class="nav-link {% if request.endpoint=='dashboard' %}active{% endif %}">
      <i class="bi bi-grid-1x2-fill"></i><span>Dashboard</span></a>
    {% set m=modulos_user %}
    {% if 'clientes' in m or 'ventas' in m or 'cotizaciones' in m or 'tareas' in m or 'calendario' in m or 'notas' in m %}
    <div class="sb-sec">Comercial</div>
    {% if 'clientes' in m %}<a href="{{ url_for('clientes') }}" class="nav-link {% if 'cliente' in request.endpoint %}active{% endif %}"><i class="bi bi-people-fill"></i><span>Clientes</span></a>{% endif %}
    {% if 'ventas' in m %}<a href="{{ url_for('ventas') }}" class="nav-link {% if 'venta' in request.endpoint %}active{% endif %}"><i class="bi bi-graph-up-arrow"></i><span>Ventas</span></a>{% endif %}
    {% if 'cotizaciones' in m %}<a href="{{ url_for('cotizaciones') }}" class="nav-link {% if request.endpoint in ['cotizaciones','cotizacion_nueva','cotizacion_ver','cotizacion_editar'] %}active{% endif %}"><i class="bi bi-file-earmark-text-fill"></i><span>Cotizaciones</span></a>{% endif %}
    {% if 'cotizaciones' in m %}<a href="{{ url_for('granel') }}" class="nav-link {% if 'granel' in request.endpoint %}active{% endif %}" style="padding-left:2.4rem;font-size:.82rem"><i class="bi bi-building"></i><span>Cot. Granel</span></a>{% endif %}
    {% if 'tareas' in m %}<a href="{{ url_for('tareas') }}" class="nav-link {% if 'tarea' in request.endpoint %}active{% endif %}"><i class="bi bi-check2-square"></i><span>Tareas</span></a>{% endif %}
    {% if 'calendario' in m %}<a href="{{ url_for('calendario') }}" class="nav-link {% if request.endpoint=='calendario' or 'evento' in request.endpoint %}active{% endif %}"><i class="bi bi-calendar3"></i><span>Calendario</span></a>{% endif %}
    {% if 'notas' in m %}<a href="{{ url_for('notas') }}" class="nav-link {% if 'nota' in request.endpoint %}active{% endif %}"><i class="bi bi-sticky-fill"></i><span>Notas</span></a>{% endif %}
    {% endif %}
    {% if 'clientes' in m %}
    <div class="sb-sec">Proveedores</div>
    <a href="{{ url_for('proveedores') }}" class="nav-link {% if request.endpoint in ['proveedores','proveedor_nuevo','proveedor_editar'] %}active{% endif %}"><i class="bi bi-truck"></i><span>Proveedores</span></a>
    <a href="{{ url_for('proveedores', tipo_f='transportista') }}" class="nav-link {% if request.endpoint in ['proveedores'] and request.args.get('tipo_f')=='transportista' %}active{% endif %}" style="padding-left:2.4rem;font-size:.82rem"><i class="bi bi-truck-front"></i><span>Transportistas</span></a>
    <a href="{{ url_for('cotizaciones_proveedor') }}" class="nav-link {% if 'cotizacion_proveedor' in request.endpoint or 'cotizaciones_proveedor' in request.endpoint %}active{% endif %}"><i class="bi bi-file-earmark-check"></i><span>Cotizaciones</span></a>
    <a href="{{ url_for('cotizaciones_proveedor', tipo='granel') }}" class="nav-link" style="padding-left:2.4rem;font-size:.82rem"><i class="bi bi-building"></i><span>Granel</span></a>
    <a href="{{ url_for('cotizaciones_proveedor', tipo='general') }}" class="nav-link" style="padding-left:2.4rem;font-size:.82rem"><i class="bi bi-box"></i><span>General</span></a>
    <a href="{{ url_for('ordenes_compra') }}" class="nav-link {% if 'orden_compra' in request.endpoint %}active{% endif %}" style="padding-left:2.4rem;font-size:.82rem"><i class="bi bi-cart-check"></i><span>Órd. Compra</span></a>
    {% endif %}
    {% if 'inventario' in m or 'produccion' in m %}
    <div class="sb-sec">Operaciones</div>
    {% if 'inventario' in m %}<a href="{{ url_for('inventario') }}" class="nav-link {% if 'inventario' in request.endpoint or 'producto' in request.endpoint or 'lote' in request.endpoint %}active{% endif %}"><i class="bi bi-box-seam-fill"></i><span>Inventario</span></a>{% endif %}
    {% if 'inventario' in m %}<a href="{{ url_for('inventario_ingresos') }}" class="nav-link {% if request.endpoint=='inventario_ingresos' %}active{% endif %}" style="padding-left:2.4rem;font-size:.82rem"><i class="bi bi-plus-square"></i><span>Multi-Ingreso</span></a>{% endif %}
    {% if 'produccion' in m %}<a href="{{ url_for('produccion_index') }}" class="nav-link {% if 'produccion' in request.endpoint or 'compra' in request.endpoint or 'materia' in request.endpoint or 'receta' in request.endpoint or 'reserva' in request.endpoint %}active{% endif %}"><i class="bi bi-gear-fill"></i><span>Producción</span></a>{% endif %}
    {% if 'produccion' in m %}<a href="{{ url_for('ordenes_produccion') }}" class="nav-link {% if request.endpoint=='ordenes_produccion' or request.endpoint=='orden_completar' %}active{% endif %}" style="padding-left:2.4rem;font-size:.82rem"><i class="bi bi-list-check"></i><span>Órdenes</span></a>{% endif %}
    {% if 'produccion' in m %}<a href="{{ url_for('gantt') }}" class="nav-link {% if request.endpoint=='gantt' %}active{% endif %}" style="padding-left:2.4rem;font-size:.82rem"><i class="bi bi-bar-chart-steps"></i><span>Gantt</span></a>{% endif %}
    {% endif %}
    <div class="sb-sec">Legal</div>
    <a href="{{ url_for('legal_index') }}" class="nav-link {% if 'legal' in request.endpoint %}active{% endif %}"><i class="bi bi-shield-check"></i><span>Documentos legales</span></a>
    <div class="sb-sec">Finanzas</div>
    <a href="{{ url_for('contable_index') }}" class="nav-link {% if 'contable' in request.endpoint %}active{% endif %}"><i class="bi bi-calculator"></i><span>Contabilidad</span></a>
    {% if 'contable' in request.endpoint %}
    <a href="{{ url_for('contable_ingresos') }}" class="nav-link {% if request.endpoint=='contable_ingresos' %}active{% endif %}" style="padding-left:2.4rem;font-size:.82rem"><i class="bi bi-graph-up"></i><span>Ingresos</span></a>
    <a href="{{ url_for('contable_egresos') }}" class="nav-link {% if request.endpoint=='contable_egresos' %}active{% endif %}" style="padding-left:2.4rem;font-size:.82rem"><i class="bi bi-graph-down"></i><span>Egresos</span></a>
    <a href="{{ url_for('contable_libro_diario') }}" class="nav-link {% if request.endpoint=='contable_libro_diario' %}active{% endif %}" style="padding-left:2.4rem;font-size:.82rem"><i class="bi bi-journal-text"></i><span>Libro diario</span></a>
    {% endif %}
    {% if 'gastos' in m %}<a href="{{ url_for('gastos') }}" class="nav-link {% if 'gasto' in request.endpoint %}active{% endif %}"><i class="bi bi-receipt"></i><span>Gastos</span></a>{% endif %}
    <a href="{{ url_for('impuestos') }}" class="nav-link {% if 'impuesto' in request.endpoint %}active{% endif %}"><i class="bi bi-percent"></i><span>Reglas tributarias</span></a>
    {% if 'reportes' in m %}<a href="{{ url_for('reportes') }}" class="nav-link {% if 'reporte' in request.endpoint %}active{% endif %}"><i class="bi bi-bar-chart-fill"></i><span>Reportes</span></a>{% endif %}
    <div class="sb-sec">Herramientas</div>
    <a href="{{ url_for('buscador') }}" class="nav-link {% if request.endpoint=='buscador' %}active{% endif %}"><i class="bi bi-search"></i><span>Buscador</span></a>
    {% if current_user.rol == 'admin' %}
    <div class="sb-sec">Admin</div>
    <a href="{{ url_for('admin_usuarios') }}" class="nav-link {% if 'admin_usuario' in request.endpoint %}active{% endif %}"><i class="bi bi-shield-person-fill"></i><span>Usuarios</span></a>
    <a href="{{ url_for('actividad') }}" class="nav-link {% if request.endpoint=='actividad' %}active{% endif %}"><i class="bi bi-clock-history"></i><span>Actividad</span></a>
    <a href="{{ url_for('admin_config') }}" class="nav-link {% if request.endpoint=='admin_config' %}active{% endif %}"><i class="bi bi-gear"></i><span>Empresa</span></a>
    {% endif %}
  </div>
  <div class="sb-foot">
    <div class="d-flex align-items-center justify-content-between mb-1">
      <div class="d-flex align-items-center gap-2">
        <div class="rounded-circle bg-primary d-flex align-items-center justify-content-center text-white fw-bold"
             style="width:31px;height:31px;font-size:.8rem;flex-shrink:0">{{ current_user.nombre[0].upper() }}</div>
        <div class="ui"><div class="u-name">{{ current_user.nombre }}</div>
          <span class="u-rol">{{ current_user.rol }}</span></div>
      </div>
      <div>
        <button class="notif-btn" id="notifBtn" onclick="toggleNotif(event)" title="Notificaciones">
          <i class="bi bi-bell-fill"></i>
          {% if notif_count > 0 %}<span class="notif-badge">{{ notif_count if notif_count < 10 else '9+' }}</span>{% endif %}
        </button>
      </div>
    </div>
    <a href="{{ url_for('perfil') }}" class="nav-link mt-1"><i class="bi bi-person-gear"></i><span>Mi perfil</span></a>
    <a href="{{ url_for('logout') }}" class="nav-link mt-1 text-danger"><i class="bi bi-box-arrow-right"></i><span>Salir</span></a>
  </div>
</nav>
<div id="main">
  <div class="topbar">
    <h1 class="pg-title">{% block page_title %}{% endblock %}</h1>
    <div class="d-flex align-items-center gap-3">
      {% block topbar_actions %}{% endblock %}
      <span class="text-muted" style="font-size:.82rem"><i class="bi bi-calendar3"></i> {{ now.strftime('%d %b %Y') }}</span>
    </div>
  </div>
  <div class="content">
    {% with messages = get_flashed_messages(with_categories=true) %}
    {% if messages %}{% for cat, msg in messages %}
    <div class="alert alert-{{ cat }} alert-dismissible fade show">{{ msg }}
      <button type="button" class="btn-close" data-bs-dismiss="alert"></button></div>
    {% endfor %}{% endif %}{% endwith %}
    {% block content %}{% endblock %}
  </div>
</div>
<!-- Panel notificaciones flotante -->
<div class="notif-dd" id="notifDd" style="display:none">
  <div class="notif-dd-head"><span><i class="bi bi-bell me-1"></i>Notificaciones</span>
    <form method="POST" action="{{ url_for('notificaciones_marcar_todas') }}" style="display:inline;margin:0">
      <button type="submit" style="font-size:.75rem;color:#5e72e4;text-decoration:none;background:none;border:none;padding:0;cursor:pointer" onclick="document.getElementById('notifDd').style.display='none'">Marcar leídas</button>
    </form>
  </div>
  <div id="notifList"><div class="p-3 text-center text-muted" style="font-size:.8rem">Cargando...</div></div>
  <div style="padding:.5rem 1rem;text-align:center;border-top:1px solid #f0f2ff">
    <a href="{{ url_for('notificaciones') }}" style="font-size:.78rem;color:#5e72e4;text-decoration:none">Ver todas</a>
  </div>
</div>
<!-- Botón diagnóstico flotante -->
<!-- Botón buscador flotante -->
<button id="diagBtn" onclick="window.location='/buscador'" title="Buscador global"><i class="bi bi-search"></i></button>
""" + _BSJ + """
<script>
// Notificaciones
function toggleNotif(e){
  e.stopPropagation();
  var dd=document.getElementById('notifDd');
  if(dd.style.display==='block'){dd.style.display='none';return;}
  dd.style.display='block';
  fetch('/notificaciones/recientes').then(r=>r.json()).then(data=>{
    var html='';
    if(data.length===0){html='<div class="p-3 text-center text-muted" style="font-size:.8rem">Sin notificaciones</div>';}
    else{data.forEach(n=>{html+='<div class="notif-item'+(n.leida?'':' unread')+'" onclick="window.location=\\''+n.url+'\\'">'
      +'<div class="ni-title">'+n.titulo+'</div>'
      +'<div class="ni-msg">'+n.mensaje+'</div>'
      +'<div class="ni-time">'+n.tiempo+'</div></div>';});}
    document.getElementById('notifList').innerHTML=html;
  }).catch(()=>{});
}
document.addEventListener('click',function(e){
  var dd=document.getElementById('notifDd');
  if(dd&&!dd.contains(e.target)&&e.target.id!=='notifBtn')dd.style.display='none';
});
// Buscador: handled by direct link
</script>
<script>
/* ── Auto-título global ─────────────────────────────────────────
   setupAutoTitulo('#idTitulo', ['#campo1','#campo2',...])
   Rellena el campo título uniendo los valores de los campos fuente
   con ", ". Si el usuario edita el título manualmente, deja de
   rellenarlo automáticamente.
   ─────────────────────────────────────────────────────────────── */
function setupAutoTitulo(tituloSel, sourceIds){
  var tEl = document.querySelector(tituloSel);
  if(!tEl) return;
  function build(){
    if(tEl._editadoManual) return;
    var partes = [];
    sourceIds.forEach(function(sid){
      var el = document.querySelector(sid);
      if(!el) return;
      var v = '';
      if(el.tagName === 'SELECT'){
        var o = el.options[el.selectedIndex];
        v = (o && o.value) ? o.text : '';
      } else {
        v = (el.value || '').trim();
      }
      if(el.type === 'date' && v){
        try{
          var d = new Date(v + 'T00:00:00');
          v = d.toLocaleDateString('es-CO',{day:'2-digit',month:'2-digit',year:'numeric'});
        }catch(e){}
      }
      if(v) partes.push(v);
    });
    if(partes.length) tEl.value = partes.join(', ');
  }
  tEl.addEventListener('input', function(){ this._editadoManual = true; });
  sourceIds.forEach(function(sid){
    var el = document.querySelector(sid);
    if(el){ el.addEventListener('change', build); el.addEventListener('input', build); }
  });
  build();
}
</script>
{% block scripts %}{% endblock %}
</body></html>"""

T['login.html'] = """<!DOCTYPE html>
<html lang="es"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Evore CRM — Login</title>""" + _CDN + """
<style>
body{background:linear-gradient(135deg,#1a1f36,#2d3561);min-height:100vh;display:flex;align-items:center;justify-content:center;font-family:'Segoe UI',sans-serif}
.card{border-radius:20px;padding:2.3rem;width:100%;max-width:420px;box-shadow:0 20px 60px rgba(0,0,0,.3);border:none}
.brand{font-size:2rem;font-weight:800;color:#1a1f36}.brand span{color:#5e72e4}
.form-control{border:1.5px solid #e9ecef;border-radius:10px;padding:.62rem 1rem}
.form-control:focus{border-color:#5e72e4;box-shadow:0 0 0 3px rgba(94,114,228,.15)}
.igt{background:#f8f9fe;border:1.5px solid #e9ecef;border-right:none;color:#8898aa;border-radius:10px 0 0 10px}
.input-group .form-control{border-left:none;border-radius:0 10px 10px 0}
.btn-lg{background:linear-gradient(135deg,#5e72e4,#4a5bd4);border:none;color:#fff;padding:.75rem;border-radius:10px;font-weight:600;transition:all .3s;width:100%}
.btn-lg:hover{transform:translateY(-1px);box-shadow:0 4px 15px rgba(94,114,228,.4);color:#fff}
</style></head><body>
<div class="card bg-white">
  <div class="text-center mb-4">
    <svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 341.94 261.01" width="170" height="130" style="display:block;margin:0 auto"><path fill="#1d1d1b" stroke="#1d1d1b" stroke-width="0" d="M28.41,217.71V184.78C28.41,171.75,27,169,27,169H72.31V173s-4.09-2.22-23.45-2.22H32.41V195.7h30.9v4.07s-2.64-2.32-15.36-2.32H32.41v34.22H51.58c19.36,0,23.45-2.22,23.45-2.22v4.07H27S28.41,230.75,28.41,217.71Z"/><path fill="#1d1d1b" stroke="#1d1d1b" stroke-width="0" d="M79.67,169h7s-.28,2.78,4.72,15.81l17.91,46.33,18.45-46.33c5.18-13,3.82-15.81,3.82-15.81h5.81s-2.45,2.68-7.45,15.17L110.3,233.52H106l-19-49C82.12,171.65,79.67,169,79.67,169Z"/><path fill="#1d1d1b" stroke="#1d1d1b" stroke-width="0" d="M138.61,201.15c0-18.49,14.45-33.38,32.36-33.38s32.36,14.89,32.36,33.38S188.87,234.54,171,234.54,138.61,219.65,138.61,201.15Zm60.81,0c0-17.47-12.73-31.62-28.45-31.62s-28.45,14.15-28.45,31.62,12.72,31.63,28.45,31.63S199.42,218.63,199.42,201.15Z"/><path fill="#1d1d1b" stroke="#1d1d1b" stroke-width="0" d="M210.64,217.71V184.78c0-13-1.36-15.81-1.36-15.81h28.63c11,0,19.9,9.34,19.9,20.81s-8.72,20.62-19.54,20.81l7.27,7.39c12.55,12.86,16.64,15.54,16.64,15.54h-8.36s-1.28-2.68-13.55-15.17l-7.54-7.76H214.64v7.12c0,13,1.36,15.81,1.36,15.81h-6.72S210.64,230.75,210.64,217.71Zm43.27-27.93c0-10.36-7.18-19-16-19H214.64v37.92h23.27C246.73,208.74,253.91,200.14,253.91,189.78Z"/><path fill="#1d1d1b" stroke="#1d1d1b" stroke-width="0" d="M268.26,217.71V184.78c0-13-1.36-15.81-1.36-15.81h45.27V173s-4.09-2.22-23.45-2.22H272.26V195.7h30.91v4.07s-2.64-2.32-15.36-2.32H272.26v34.22h19.18c19.36,0,23.45-2.22,23.45-2.22v4.07h-48S268.26,230.75,268.26,217.71Z"/><path fill="#1d1d1b" d="M239.09,111.71C234.61,109,225,105.23,225,105.23s-1.42-.5-1.34-.9c.21-1.2,2.76-1.54,9.63-10.7,7.11-9.49,8.6-13.94,9.94-17.84,1.14-3.3,2.38-11,1.72-11.79s-9.12-1.35-16.48.3c-7.06,1.58-11.1,3.47-11.5,3-.64-.71.95-6.23,1.11-11.33.17-5.28.32-6.75-.41-11.49-.61-4-1.47-7.57-2.65-7.91s-5,.82-7,1.69A81.41,81.41,0,0,0,198.41,44c-2.91,2-6.52,5.48-7.55,5.69s-2.79-6-6.9-13.15a86,86,0,0,0-5.69-9.09c-1.74-2.45-4.91-7-7.28-7h0c-2.38,0-5.54,4.59-7.28,7A84.29,84.29,0,0,0,158,36.52c-4.12,7.13-5.76,13.38-6.9,13.15s-4.73-3.8-7.64-5.8a79.94,79.94,0,0,0-9.56-5.58c-2-.87-5.81-2-7-1.69s-2,3.94-2.65,7.91c-.73,4.74-.58,6.21-.41,11.49.16,5.1,1.74,10.62,1.11,11.33-.4.44-4.44-1.45-11.5-3C106.08,62.65,97.72,63.08,97,64s.57,8.49,1.71,11.79c1.35,3.9,2.83,8.35,10,17.84,6.87,9.16,9.41,9.5,9.63,10.7.07.4-1.35.9-1.35.9s-9.57,3.72-14.06,6.48C99.17,114,95,117.34,95,119.08c0,2.11,3.54,4.62,8.14,7.67s15.36,7.83,26.75,6.75c12.82-1.22,20.68-7.21,21.76-6.74.49.21.11,5.24,5.16,10.44,4.43,4.56,10.58,10,14.15,10s9.72-5.42,14.15-10c5-5.2,4.67-10.23,5.16-10.44,1.07-.47,8.93,5.52,21.76,6.74,11.39,1.08,22.07-3.65,26.74-6.75s8.17-5.56,8.15-7.67C246.93,117.34,242.76,114,239.09,111.71ZM223,68.71c8.36-2.84,18.31-3,19-2.1s-.68,15.69-17.47,32.53C213.08,110.56,203.15,112,194.13,113c-7.48.88-16.37-.72-16.37-.72a34.59,34.59,0,0,0,11.36-11.93c4.39-7.85,6.29-14.46,13.95-20.31S214.59,71.55,223,68.71ZM201.61,45.52c7.35-5,11.19-5.92,11.86-5.48s1.71,4.63,1.66,12.23a115.69,115.69,0,0,1-1.71,16.27s-2.31,1.26-8.63,5.68a50.76,50.76,0,0,0-9.48,8.13,74.94,74.94,0,0,0-.42-15.92,99,99,0,0,0-3-12.66S194.26,50.51,201.61,45.52Zm-43.8-1.86c3.92-7.94,10.94-19.47,13.17-19.47s9.25,11.53,13.17,19.47c4.74,9.58,8.13,21.9,8.14,33.28,0,9.88-4,19.13-9.93,26.62-5.15,6.47-10.79,8.16-11.38,8.16s-6.23-1.69-11.39-8.16c-6-7.49-9.94-16.74-9.92-26.62C149.68,65.56,153.07,53.24,157.81,43.66ZM128.49,40c.67-.44,4.51.49,11.86,5.48s9.77,8.25,9.77,8.25a99,99,0,0,0-3.05,12.66,74.94,74.94,0,0,0-.42,15.92,50.76,50.76,0,0,0-9.48-8.13c-6.32-4.42-8.63-5.68-8.63-5.68a115.69,115.69,0,0,1-1.71-16.27C126.78,44.67,127.8,40.5,128.49,40Zm-11,59.1C100.69,82.3,99.34,67.51,100,66.61s10.63-.74,19,2.1,12.22,5.51,19.88,11.37,9.56,12.46,14,20.31a34.59,34.59,0,0,0,11.36,11.93s-8.89,1.6-16.37.72C138.81,112,128.88,110.56,117.49,99.14Zm28.92,26.37c-5.29,2.62-12.37,5.73-23.29,5-6.53-.44-11.77-2.84-16.93-5.52-5.33-2.76-7.66-5.09-7.66-6s4.73-4.29,9.19-6.39a112.32,112.32,0,0,1,12-5.22c1.08-.37,1.26-.07,3.11.91a40.92,40.92,0,0,0,12.86,5.9c6.85,1.73,18.66,2.63,18.66,2.63S151.7,122.88,146.41,125.51Zm33.35,12.19c-5.8,5.61-8,5.86-8.78,5.86s-3-.25-8.78-5.86-8.28-10.87-6.47-16.09a7.73,7.73,0,0,1,2.13-3.58c3.63-3.47,13.1-3.09,13.1-3.09h0s9.47-.38,13.1,3.09a7.73,7.73,0,0,1,2.13,3.58C188,126.83,185.56,132.09,179.76,137.7Zm56-12.72c-5.16,2.68-10.4,5.08-16.93,5.52-10.92.74-18-2.37-23.29-5s-7.92-8.74-7.92-8.74,11.81-.9,18.66-2.63a40.92,40.92,0,0,0,12.86-5.9c1.84-1,2-1.28,3.11-.91a112.32,112.32,0,0,1,12,5.22c4.46,2.1,9.19,5.45,9.19,6.39S241.1,122.22,235.77,125Z"/></svg>
    <p class="text-muted mb-0" style="font-size:.75rem;letter-spacing:3px;margin-top:.5rem">SISTEMA DE GESTIÓN</p>
  </div>
  {% with messages = get_flashed_messages(with_categories=true) %}
  {% if messages %}{% for cat, msg in messages %}
  <div class="alert alert-{{ cat }} py-2 mb-3" style="border-radius:10px;border:none;font-size:.875rem">{{ msg }}</div>
  {% endfor %}{% endif %}{% endwith %}
  <form method="POST">
    <div class="mb-3"><label class="form-label fw-semibold" style="font-size:.875rem;color:#525f7f">Correo</label>
      <div class="input-group"><span class="input-group-text igt"><i class="bi bi-envelope"></i></span>
        <input type="email" name="email" class="form-control" required autofocus></div></div>
    <div class="mb-4"><label class="form-label fw-semibold" style="font-size:.875rem;color:#525f7f">Contraseña</label>
      <div class="input-group"><span class="input-group-text igt"><i class="bi bi-lock"></i></span>
        <input type="password" name="password" class="form-control" required></div></div>
    <button type="submit" class="btn btn-lg"><i class="bi bi-box-arrow-in-right me-2"></i>Iniciar Sesión</button>
  </form>
</div>""" + _BSJ + """</body></html>"""

T['dashboard.html'] = """{% extends 'base.html' %}
{% block title %}Dashboard{% endblock %}{% block page_title %}Dashboard{% endblock %}
{% block topbar_actions %}
<button id="btnM" class="btn btn-sm btn-outline-secondary" onclick="toggleMoneda()">
  <i class="bi bi-currency-exchange me-1"></i><span id="lblM">Ver en USD</span></button>
<small class="text-muted" id="tasaInfo"></small>
{% endblock %}
{% block content %}
<div class="row g-3 mb-3">
  <div class="col-6 col-lg-3"><div class="sc"><div class="d-flex justify-content-between align-items-start">
    <div><div class="sv">{{ total_clientes }}</div><div class="sl">Clientes activos</div></div>
    <div class="si" style="background:#e8eeff"><i class="bi bi-people-fill" style="color:#5e72e4"></i></div>
  </div></div></div>
  <div class="col-6 col-lg-3"><div class="sc"><div class="d-flex justify-content-between align-items-start">
    <div><div class="sv">{{ ventas_ganadas }}</div><div class="sl">Ventas ganadas</div></div>
    <div class="si" style="background:#e3f9ee"><i class="bi bi-graph-up-arrow" style="color:#2dce89"></i></div>
  </div></div></div>
  <div class="col-6 col-lg-3"><div class="sc"><div class="d-flex justify-content-between align-items-start">
    <div><div class="sv">{{ tareas_pendientes }}</div><div class="sl">Tareas pendientes</div></div>
    <div class="si" style="background:#fff4e5"><i class="bi bi-check2-square" style="color:#fb6340"></i></div>
  </div></div></div>
  <div class="col-6 col-lg-3"><div class="sc"><div class="d-flex justify-content-between align-items-start">
    <div><div class="sv valor-cop" data-cop="{{ ingresos_totales }}">$ {{ '{:,.0f}'.format(ingresos_totales).replace(',','.') }}</div>
      <div class="sl">Ingresos COP</div></div>
    <div class="si" style="background:#fce8ff"><i class="bi bi-currency-dollar" style="color:#c300ff"></i></div>
  </div></div></div>
</div>
<div class="row g-3 mb-3">
  <div class="col-6 col-lg-3"><div class="sc"><div class="d-flex justify-content-between align-items-start">
    <div><div class="sv valor-cop" data-cop="{{ gastos_totales }}">$ {{ '{:,.0f}'.format(gastos_totales).replace(',','.') }}</div>
      <div class="sl">Gastos operativos</div></div>
    <div class="si" style="background:#ffeaea"><i class="bi bi-receipt" style="color:#e74c3c"></i></div>
  </div></div></div>
  <div class="col-6 col-lg-3"><div class="sc"><div class="d-flex justify-content-between align-items-start">
    <div><div class="sv valor-cop" data-cop="{{ balance }}">$ {{ '{:,.0f}'.format(balance).replace(',','.') }}</div>
      <div class="sl">Utilidad bruta</div></div>
    <div class="si" style="background:#e8fff3"><i class="bi bi-bar-chart-fill" style="color:#27ae60"></i></div>
  </div></div></div>
  <div class="col-6 col-lg-3"><div class="sc"><div class="d-flex justify-content-between align-items-start">
    <div><div class="sv valor-cop" data-cop="{{ impuestos_estimados }}">$ {{ '{:,.0f}'.format(impuestos_estimados).replace(',','.') }}</div>
      <div class="sl">Impuestos estimados</div></div>
    <div class="si" style="background:#fce4ec"><i class="bi bi-percent" style="color:#e83e8c"></i></div>
  </div></div></div>
  <div class="col-6 col-lg-3"><div class="sc" style="border:2px solid {{ '#2dce89' if saldo_neto >= 0 else '#f5365c' }}"><div class="d-flex justify-content-between align-items-start">
    <div><div class="sv valor-cop" data-cop="{{ saldo_neto }}" style="color:{{ '#2dce89' if saldo_neto >= 0 else '#f5365c' }}">$ {{ '{:,.0f}'.format(saldo_neto).replace(',','.') }}</div>
      <div class="sl">Saldo neto real</div></div>
    <div class="si" style="background:{{ '#e8fff3' if saldo_neto >= 0 else '#ffeaea' }}"><i class="bi bi-cash-stack" style="color:{{ '#2dce89' if saldo_neto >= 0 else '#f5365c' }}"></i></div>
  </div></div></div>
</div>
<div class="row g-3 mb-4">
  <div class="col-6 col-lg-3"><div class="sc"><div class="d-flex justify-content-between align-items-start">
    <div><div class="sv valor-cop" data-cop="{{ saldo_pendiente }}">$ {{ '{:,.0f}'.format(saldo_pendiente).replace(',','.') }}</div>
      <div class="sl">Saldo por cobrar</div></div>
    <div class="si" style="background:#fff8e1"><i class="bi bi-hourglass-split" style="color:#f39c12"></i></div>
  </div></div></div>
  <div class="col-6 col-lg-3"><div class="sc"><div class="d-flex justify-content-between align-items-start">
    <div><div class="sv">{{ productos_bajo_stock }}</div><div class="sl">Stock bajo</div></div>
    <div class="si" style="background:#ffeaea"><i class="bi bi-exclamation-triangle" style="color:#e74c3c"></i></div>
  </div></div></div>
  <div class="col-6 col-lg-3"></div>
  <div class="col-6 col-lg-3"><div class="sc" style="cursor:pointer" onclick="window.location='{{ url_for('contable_index') }}'">
    <div class="d-flex align-items-center gap-2">
      <i class="bi bi-calculator" style="font-size:1.3rem;color:#5e72e4"></i>
      <div><div style="font-size:.8rem;font-weight:600;color:#1a1f36">Ver contabilidad detallada</div>
        <div style="font-size:.72rem;color:#8898aa">Desglose de impuestos + utilidad neta</div></div>
    </div>
  </div></div>
</div>
<div class="row g-4">
  <div class="col-lg-6"><div class="tc">
    <div class="ch d-flex justify-content-between align-items-center">
      <span><i class="bi bi-check2-square me-2 text-warning"></i>Tareas pendientes</span>
      <a href="{{ url_for('tarea_nueva') }}" class="btn btn-sm btn-primary"><i class="bi bi-plus"></i></a></div>
    {% if tareas_recientes %}<table class="table"><tbody>
    {% for t in tareas_recientes %}<tr>
      <td><a href="{{ url_for('tarea_ver', id=t.id) }}" class="fw-semibold text-decoration-none" style="color:#1a1f36">{{ t.titulo }}</a>
        <div><small class="text-muted">{{ t.asignado_user.nombre if t.asignado_user else '' }}</small></div></td>
      <td><span class="b b-{{ t.prioridad }}">{{ t.prioridad.title() }}</span></td>
    </tr>{% endfor %}</tbody></table>
    {% else %}<div class="text-center text-muted py-4"><i class="bi bi-check2-all" style="font-size:2rem"></i>
      <p class="mt-2 mb-0">Sin pendientes</p></div>{% endif %}
  </div></div>
  <div class="col-lg-6"><div class="tc">
    <div class="ch d-flex justify-content-between align-items-center">
      <span><i class="bi bi-graph-up me-2 text-success"></i>Ventas recientes</span>
      <a href="{{ url_for('venta_nueva') }}" class="btn btn-sm btn-primary"><i class="bi bi-plus"></i></a></div>
    {% if ventas_recientes %}<table class="table"><tbody>
    {% for v in ventas_recientes %}<tr>
      <td><div class="fw-semibold" style="color:#1a1f36">{{ v.titulo }}</div>
        <small class="text-muted">{{ v.cliente.empresa or v.cliente.nombre if v.cliente else '—' }}</small></td>
      <td class="fw-semibold valor-cop" data-cop="{{ v.total }}">$ {{ '{:,.0f}'.format(v.total).replace(',','.') }}</td>
      <td><span class="b b-{{ v.estado }}">{{ v.estado.replace('_',' ').title() }}</span></td>
    </tr>{% endfor %}</tbody></table>
    {% else %}<div class="text-center text-muted py-4"><i class="bi bi-graph-up" style="font-size:2rem"></i>
      <p class="mt-2 mb-0">Sin ventas</p></div>{% endif %}
  </div></div>
</div>
{% if actividades_recientes %}
<div class="tc mt-4"><div class="ch d-flex justify-content-between align-items-center">
  <span><i class="bi bi-clock-history me-2 text-secondary"></i>Actividad reciente</span>
  {% if current_user.rol=='admin' %}<a href="{{ url_for('actividad') }}" class="btn btn-sm btn-outline-secondary">Ver todo</a>{% endif %}
</div>
<table class="table"><tbody>{% for a in actividades_recientes %}<tr>
  <td style="width:130px"><small class="text-muted">{{ a.creado_en.strftime('%d/%m %H:%M') }}</small></td>
  <td style="width:80px"><span class="badge {% if a.tipo=='crear' %}bg-success{% elif a.tipo=='editar' %}bg-primary{% elif a.tipo=='eliminar' %}bg-danger{% elif a.tipo=='completar' %}bg-info{% else %}bg-secondary{% endif %}">{{ a.tipo.title() }}</span></td>
  <td style="font-size:.88rem">{{ a.descripcion }}</td>
  <td style="width:100px"><small class="text-muted">{{ a.usuario.nombre if a.usuario else '' }}</small></td>
</tr>{% endfor %}</tbody></table>
</div>{% endif %}
{% endblock %}
{% block scripts %}<script>
let tasaCOP=null,enUSD=false;
async function cargarTasa(){
  try{const r=await fetch('https://open.er-api.com/v6/latest/USD');
    const d=await r.json();tasaCOP=d.rates.COP;
    document.getElementById('tasaInfo').textContent='1 USD = $'+tasaCOP.toFixed(0).replace(/\B(?=(\d{3})+(?!\d))/g,'.')+ ' COP';
  }catch(e){}
}
function toggleMoneda(){
  if(!tasaCOP)return;enUSD=!enUSD;
  document.getElementById('lblM').textContent=enUSD?'Ver en COP':'Ver en USD';
  document.querySelectorAll('.valor-cop').forEach(el=>{
    const c=parseFloat(el.dataset.cop)||0;
    el.textContent=enUSD?'USD '+(c/tasaCOP).toLocaleString('en-US',{maximumFractionDigits:0}):'$ '+c.toLocaleString('es-CO',{maximumFractionDigits:0});
  });
}
cargarTasa();
</script>{% endblock %}"""

T['clientes/index.html'] = """{% extends 'base.html' %}
{% block title %}Clientes{% endblock %}{% block page_title %}Clientes{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('cliente_nuevo') }}" class="btn btn-primary btn-sm"><i class="bi bi-plus-lg me-1"></i>Nuevo</a>{% endblock %}
{% block content %}
<div class="tc mb-3"><div class="p-3"><form method="GET" class="row g-2 align-items-end">
  <div class="col-sm-4"><input type="text" name="buscar" class="form-control form-control-sm" placeholder="Nombre, empresa, NIT..." value="{{ busqueda }}"></div>
  <div class="col-sm-3"><select name="estado_rel" class="form-select form-select-sm">
    <option value="">Todas las relaciones</option>
    {% for er in ['prospecto','negociacion','cliente_activo','vip','inactivo','perdido'] %}
    <option value="{{ er }}" {% if estado_rel_f==er %}selected{% endif %}>{{ er.replace('_',' ').title() }}</option>{% endfor %}
  </select></div>
  <div class="col-auto">
    <button type="submit" class="btn btn-primary btn-sm"><i class="bi bi-search"></i></button>
    <a href="{{ url_for('clientes') }}" class="btn btn-outline-secondary btn-sm">Limpiar</a>
  </div>
</form></div></div>
<div class="tc"><div class="ch"><i class="bi bi-people-fill me-2"></i>{{ items|length }} cliente(s)</div>
{% if items %}<div class="table-responsive"><table class="table">
  <thead><tr><th>Empresa</th><th>NIT</th><th>Relación</th><th>Contacto principal</th><th>Alta</th><th></th></tr></thead>
  <tbody>{% for c in items %}<tr>
    <td><a href="{{ url_for('cliente_ver', id=c.id) }}" class="fw-semibold text-decoration-none" style="color:#1a1f36">{{ c.empresa or c.nombre }}</a>
      {% if c.empresa %}<div><small class="text-muted">{{ c.nombre }}</small></div>{% endif %}</td>
    <td><small class="text-muted">{{ c.nit or '—' }}</small></td>
    <td><span class="b b-{{ c.estado_relacion }}">{{ c.estado_relacion.replace('_',' ').title() }}</span></td>
    <td>{% if c.contactos %}<small>{{ c.contactos[0].nombre }}{% if c.contactos[0].cargo %} · {{ c.contactos[0].cargo }}{% endif %}</small>{% else %}—{% endif %}</td>
    <td><small class="text-muted">{{ c.creado_en.strftime('%d/%m/%Y') }}</small></td>
    <td><div class="d-flex gap-1">
      <a href="{{ url_for('cliente_ver', id=c.id) }}" class="btn btn-sm btn-outline-primary"><i class="bi bi-eye"></i></a>
      <a href="{{ url_for('cliente_editar', id=c.id) }}" class="btn btn-sm btn-outline-secondary"><i class="bi bi-pencil"></i></a>
      <form method="POST" action="{{ url_for('cliente_eliminar', id=c.id) }}" onsubmit="return confirm('¿Eliminar?')">
        <button class="btn btn-sm btn-outline-danger"><i class="bi bi-trash"></i></button></form>
    </div></td>
  </tr>{% endfor %}</tbody>
</table></div>
{% else %}<div class="text-center text-muted py-5"><i class="bi bi-people" style="font-size:3rem"></i>
  <p class="mt-3">Sin clientes.</p><a href="{{ url_for('cliente_nuevo') }}" class="btn btn-primary">Agregar</a></div>
{% endif %}</div>{% endblock %}"""

T['clientes/form.html'] = """{% extends 'base.html' %}
{% block title %}{{ titulo }}{% endblock %}{% block page_title %}{{ titulo }}{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('clientes') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Volver</a>{% endblock %}
{% block content %}<div class="fc" style="max-width:900px"><form method="POST">
<h6 class="text-muted mb-3 text-uppercase" style="letter-spacing:1px;font-size:.75rem">Información de la empresa</h6>
<div class="row g-3 mb-4">
  <div class="col-md-6"><label class="form-label">Empresa *</label>
    <input type="text" name="empresa" class="form-control" value="{{ obj.empresa if obj else '' }}" required></div>
  <div class="col-md-3"><label class="form-label">NIT</label>
    <input type="text" name="nit" class="form-control" placeholder="900.123.456-7" value="{{ obj.nit if obj else '' }}"></div>
  <div class="col-md-3"><label class="form-label">Estado relación</label>
    <select name="estado_relacion" class="form-select">
      {% for er,lbl in [('prospecto','Prospecto'),('negociacion','Negociación'),('cliente_activo','Cliente Activo'),('vip','VIP'),('inactivo','Inactivo'),('perdido','Perdido')] %}
      <option value="{{ er }}" {% if obj and obj.estado_relacion==er %}selected{% elif not obj and er=='prospecto' %}selected{% endif %}>{{ lbl }}</option>{% endfor %}
    </select></div>
  <div class="col-md-6"><label class="form-label">Dirección Cámara de Comercio</label>
    <input type="text" name="dir_comercial" class="form-control" value="{{ obj.dir_comercial if obj else '' }}"></div>
  <div class="col-md-6"><label class="form-label">Dirección de entrega</label>
    <input type="text" name="dir_entrega" class="form-control" value="{{ obj.dir_entrega if obj else '' }}"></div>
  <div class="col-12"><label class="form-label">Notas</label>
    <textarea name="notas" class="form-control" rows="2">{{ obj.notas if obj else '' }}</textarea></div>
</div><hr class="my-3">
<div class="d-flex justify-content-between align-items-center mb-3">
  <h6 class="text-muted mb-0 text-uppercase" style="letter-spacing:1px;font-size:.75rem">Contactos</h6>
  <button type="button" class="btn btn-sm btn-outline-primary" onclick="addContacto()"><i class="bi bi-plus-lg me-1"></i>Agregar</button>
</div>
<div id="contactosContainer">
  {% if obj and obj.contactos %}{% for c in obj.contactos %}
  <div class="prod-row mb-2"><div class="row g-2 align-items-end">
    <div class="col-md-3"><label class="form-label mb-1" style="font-size:.8rem">Nombre *</label>
      <input type="text" name="c_nombre[]" class="form-control form-control-sm" value="{{ c.nombre }}" required></div>
    <div class="col-md-2"><label class="form-label mb-1" style="font-size:.8rem">Cargo</label>
      <input type="text" name="c_cargo[]" class="form-control form-control-sm" value="{{ c.cargo or '' }}"></div>
    <div class="col-md-3"><label class="form-label mb-1" style="font-size:.8rem">Email</label>
      <input type="email" name="c_email[]" class="form-control form-control-sm" value="{{ c.email or '' }}"></div>
    <div class="col-md-3"><label class="form-label mb-1" style="font-size:.8rem">Teléfono</label>
      <input type="text" name="c_telefono[]" class="form-control form-control-sm" value="{{ c.telefono or '' }}"></div>
    <div class="col-md-1"><button type="button" class="btn btn-sm btn-outline-danger w-100" onclick="this.closest('.prod-row').remove()"><i class="bi bi-trash"></i></button></div>
  </div></div>{% endfor %}{% endif %}
</div>
<div class="d-flex gap-2 mt-4">
  <button type="submit" class="btn btn-primary"><i class="bi bi-check-lg me-1"></i>{{ 'Actualizar' if obj else 'Crear Cliente' }}</button>
  <a href="{{ url_for('clientes') }}" class="btn btn-outline-secondary">Cancelar</a>
</div></form></div>{% endblock %}
{% block scripts %}<script>
function addContacto(){
  document.getElementById('contactosContainer').insertAdjacentHTML('beforeend',`<div class="prod-row mb-2"><div class="row g-2 align-items-end">
    <div class="col-md-3"><label class="form-label mb-1" style="font-size:.8rem">Nombre *</label><input type="text" name="c_nombre[]" class="form-control form-control-sm" required></div>
    <div class="col-md-2"><label class="form-label mb-1" style="font-size:.8rem">Cargo</label><input type="text" name="c_cargo[]" class="form-control form-control-sm"></div>
    <div class="col-md-3"><label class="form-label mb-1" style="font-size:.8rem">Email</label><input type="email" name="c_email[]" class="form-control form-control-sm"></div>
    <div class="col-md-3"><label class="form-label mb-1" style="font-size:.8rem">Teléfono</label><input type="text" name="c_telefono[]" class="form-control form-control-sm"></div>
    <div class="col-md-1"><button type="button" class="btn btn-sm btn-outline-danger w-100" onclick="this.closest('.prod-row').remove()"><i class="bi bi-trash"></i></button></div>
  </div></div>`);
}
</script>{% endblock %}"""

T['clientes/ver.html'] = """{% extends 'base.html' %}
{% block title %}{{ obj.empresa or obj.nombre }}{% endblock %}
{% block page_title %}{{ obj.empresa or obj.nombre }}{% endblock %}
{% block topbar_actions %}
<a href="{{ url_for('cliente_editar', id=obj.id) }}" class="btn btn-primary btn-sm"><i class="bi bi-pencil me-1"></i>Editar</a>
<a href="{{ url_for('clientes') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Volver</a>
{% endblock %}
{% block content %}<div class="row g-4">
  <div class="col-lg-4">
    <div class="fc mb-3">
      <div class="d-flex align-items-center gap-3 mb-3">
        <div class="rounded-circle d-flex align-items-center justify-content-center text-white fw-bold"
             style="width:50px;height:50px;background:#5e72e4;font-size:1.2rem;flex-shrink:0">{{ (obj.empresa or obj.nombre)[0].upper() }}</div>
        <div><h5 class="mb-1">{{ obj.empresa or obj.nombre }}</h5>
          <span class="b b-{{ obj.estado_relacion }}">{{ obj.estado_relacion.replace('_',' ').title() }}</span></div>
      </div><hr>
      <dl class="row mb-0" style="font-size:.88rem">
        {% if obj.empresa %}<dt class="col-5 text-muted">Rep. Legal</dt><dd class="col-7">{{ obj.nombre }}</dd>{% endif %}
        <dt class="col-5 text-muted">NIT</dt><dd class="col-7">{{ obj.nit or '—' }}</dd>
        <dt class="col-5 text-muted">Dir. Comercial</dt><dd class="col-7">{{ obj.dir_comercial or '—' }}</dd>
        <dt class="col-5 text-muted">Dir. Entrega</dt><dd class="col-7">{{ obj.dir_entrega or '—' }}</dd>
      </dl>
      {% if obj.notas %}<hr><p style="font-size:.88rem">{{ obj.notas }}</p>{% endif %}
    </div>
    {% if obj.contactos %}<div class="fc">
      <h6 class="text-muted text-uppercase mb-3" style="font-size:.72rem;letter-spacing:1px">Contactos</h6>
      {% for c in obj.contactos %}
      <div class="d-flex gap-2 mb-3">
        <div class="rounded-circle d-flex align-items-center justify-content-center text-white fw-bold"
             style="width:34px;height:34px;background:#8898aa;font-size:.8rem;flex-shrink:0">{{ c.nombre[0].upper() }}</div>
        <div style="font-size:.88rem">
          <div class="fw-semibold">{{ c.nombre }}{% if c.cargo %} <span class="text-muted fw-normal">· {{ c.cargo }}</span>{% endif %}</div>
          {% if c.email %}<div><a href="mailto:{{ c.email }}" class="text-decoration-none">{{ c.email }}</a></div>{% endif %}
          {% if c.telefono %}<div class="text-muted">{{ c.telefono }}</div>{% endif %}
        </div>
      </div>{% endfor %}
    </div>{% endif %}
  </div>
  <div class="col-lg-8"><div class="tc">
    <div class="ch d-flex justify-content-between">
      <span><i class="bi bi-graph-up me-2"></i>Ventas</span>
      <a href="{{ url_for('venta_nueva') }}" class="btn btn-sm btn-primary"><i class="bi bi-plus"></i></a></div>
    {% if obj.ventas %}<table class="table">
      <thead><tr><th>N°</th><th>Título</th><th>Total</th><th>Saldo</th><th>Pago</th><th>Producción</th></tr></thead>
      <tbody>{% for v in obj.ventas %}<tr>
        <td><small class="fw-semibold text-muted">{{ v.numero or '—' }}</small></td>
        <td class="fw-semibold">{{ v.titulo }}</td>
        <td>$ {{ '{:,.0f}'.format(v.total).replace(',','.') }}</td>
        <td class="{{ 'text-danger fw-semibold' if v.saldo > 0 else '' }}">$ {{ '{:,.0f}'.format(v.saldo).replace(',','.') }}</td>
        <td><span class="b b-{{ v.estado }}" style="font-size:.75rem">{{ v.estado.replace('_',' ').title() }}</span></td>
        <td>
          {% set ops = v.ordenes_produccion %}
          {% if ops %}
            {% set comp = ops|selectattr('estado','equalto','completado')|list|length %}
            {% set total_ops = ops|length %}
            {% if comp == total_ops %}
              <span class="badge bg-success"><i class="bi bi-check-all me-1"></i>Completo</span>
            {% else %}
              <span class="badge bg-primary">{{ comp }}/{{ total_ops }} ords.</span>
            {% endif %}
          {% elif v.estado in ['anticipo_pagado','ganado'] %}
            <span class="badge bg-secondary">Sin prod.</span>
          {% else %}
            <span class="text-muted" style="font-size:.8rem">—</span>
          {% endif %}
        </td>
      </tr>{% endfor %}</tbody>
    </table>
    {% else %}<div class="text-center text-muted py-4"><i class="bi bi-graph-up" style="font-size:2rem"></i>
      <p class="mt-2 mb-0">Sin ventas</p></div>{% endif %}
  </div></div>
</div>{% endblock %}"""

T['ordenes_compra/index.html'] = """{% extends 'base.html' %}
{% block title %}Órdenes de Compra{% endblock %}{% block page_title %}Órdenes de Compra{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('orden_compra_nueva') }}" class="btn btn-primary btn-sm"><i class="bi bi-plus-lg me-1"></i>Nueva Orden</a>{% endblock %}
{% block content %}
<div class="mb-3 d-flex gap-2 flex-wrap">
  {% for est,lbl in [('','Todas'),('borrador','Borrador'),('enviada','Enviada'),('recibida','Recibida'),('cancelada','Cancelada')] %}
  <a href="{{ url_for('ordenes_compra', estado=est) }}" class="btn btn-sm {{ 'btn-primary' if estado_f==est else 'btn-outline-secondary' }}">{{ lbl }}</a>{% endfor %}
</div>
<div class="tc"><div class="ch"><i class="bi bi-cart-check me-2"></i>{{ items|length }} orden(es) de compra</div>
{% if items %}<div class="table-responsive"><table class="table">
  <thead><tr><th>N°</th><th>Proveedor</th><th>Cotización</th><th>Estado</th><th>Emisión</th><th>Entrega esp.</th><th>Est. pago</th><th>Transportista</th><th>Total</th><th></th></tr></thead>
  <tbody>{% for oc in items %}<tr>
    <td class="fw-semibold" style="color:#1a1f36">{{ oc.numero or '—' }}</td>
    <td>{{ oc.proveedor.empresa or oc.proveedor.nombre if oc.proveedor else '—' }}</td>
    <td>{% if oc.cotizacion_ref %}<small class="text-muted">{{ oc.cotizacion_ref.numero }}</small>{% else %}—{% endif %}</td>
    <td>
      {% if oc.estado=='borrador' %}<span class="badge bg-secondary">Borrador</span>
      {% elif oc.estado=='enviada' %}<span class="badge bg-primary">Enviada</span>
      {% elif oc.estado=='recibida' %}<span class="badge bg-success">Recibida</span>
      {% else %}<span class="badge bg-danger">Cancelada</span>{% endif %}
    </td>
    <td><small>{{ oc.fecha_emision.strftime('%d/%m/%Y') if oc.fecha_emision else '—' }}</small></td>
    <td>{% if oc.fecha_esperada %}<small class="{{ 'text-danger fw-semibold' if oc.fecha_esperada < now.date() and oc.estado != 'recibida' else '' }}">{{ oc.fecha_esperada.strftime('%d/%m/%Y') }}</small>{% else %}—{% endif %}</td>
    <td><small class="text-muted">{{ oc.fecha_estimada_pago.strftime('%d/%m/%Y') if oc.fecha_estimada_pago else '—' }}</small></td>
    <td>{% if oc.transportista %}<small><i class="bi bi-truck-front text-success me-1"></i>{{ oc.transportista.empresa or oc.transportista.nombre }}{% if oc.fecha_estimada_recogida %}<div class="text-muted" style="font-size:.72rem">Recogida: {{ oc.fecha_estimada_recogida.strftime('%d/%m/%Y') }}</div>{% endif %}</small>{% else %}<span class="text-muted">—</span>{% endif %}</td>
    <td class="fw-semibold">$ {{ '{:,.0f}'.format(oc.total).replace(',','.') }}</td>
    <td><div class="d-flex gap-1 flex-wrap">
      <a href="{{ url_for('orden_compra_editar', id=oc.id) }}" class="btn btn-sm btn-outline-secondary"><i class="bi bi-pencil"></i></a>
      {% if oc.estado == 'borrador' %}
      <form method="POST" action="{{ url_for('orden_compra_estado', id=oc.id) }}" class="d-inline">
        <input type="hidden" name="estado" value="enviada">
        <button class="btn btn-sm btn-outline-primary" title="Marcar como enviada"><i class="bi bi-send"></i></button>
      </form>
      {% elif oc.estado == 'enviada' %}
      <form method="POST" action="{{ url_for('orden_compra_estado', id=oc.id) }}" class="d-inline" onsubmit="return confirm('¿Marcar como recibida? Se agendará la entrega en el calendario.')">
        <input type="hidden" name="estado" value="recibida">
        <button class="btn btn-sm btn-outline-success" title="Marcar como recibida → agenda entrega en calendario"><i class="bi bi-box-arrow-in-down me-1"></i>Recibida</button>
      </form>
      {% endif %}
      <form method="POST" action="{{ url_for('orden_compra_eliminar', id=oc.id) }}" onsubmit="return confirm('¿Eliminar?')">
        <button class="btn btn-sm btn-outline-danger"><i class="bi bi-trash"></i></button></form>
    </div></td>
  </tr>{% endfor %}</tbody>
</table></div>
{% else %}<div class="text-center text-muted py-5"><i class="bi bi-cart-check" style="font-size:3rem"></i>
  <p class="mt-3">Sin órdenes de compra.</p><a href="{{ url_for('orden_compra_nueva') }}" class="btn btn-primary">Crear primera</a></div>
{% endif %}</div>{% endblock %}"""

T['ordenes_compra/form.html'] = """{% extends 'base.html' %}
{% block title %}{{ titulo }}{% endblock %}{% block page_title %}{{ titulo }}{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('ordenes_compra') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Volver</a>{% endblock %}
{% block content %}<div class="fc" style="max-width:1040px"><form method="POST">
<h6 class="text-muted text-uppercase mb-3" style="font-size:.75rem;letter-spacing:1px">Información general</h6>
<div class="row g-3 mb-3">
  <div class="col-md-4"><label class="form-label">Proveedor</label>
    <select name="proveedor_id" class="form-select" id="selectProv">
      <option value="">— Sin proveedor —</option>
      {% for pv in proveedores_list %}<option value="{{ pv.id }}" {% if obj and obj.proveedor_id==pv.id %}selected{% endif %}>{{ pv.empresa or pv.nombre }}</option>{% endfor %}
    </select></div>
  <div class="col-md-4"><label class="form-label">Cotización de referencia <small class="text-muted">(calcula fecha entrega)</small></label>
    <select name="cotizacion_id" class="form-select" id="selectCot" onchange="onCotChange()">
      <option value="">— Sin cotización —</option>
      {% for c in cotizaciones_list %}<option value="{{ c.id }}"
        data-plazo="{{ c.plazo_entrega_dias or 0 }}"
        data-prov="{{ c.proveedor_id or '' }}"
        data-nombre="{{ c.nombre_producto }}"
        data-precio="{{ c.precio_unitario }}"
        data-unidad="{{ c.unidad }}"
        data-cond="{{ c.condicion_pago_tipo }}"
        {% if obj and obj.cotizacion_id==c.id %}selected{% endif %}>
        {{ c.numero or '' }} — {{ c.nombre_producto }} ({{ c.proveedor.empresa if c.proveedor else '—' }})
      </option>{% endfor %}
    </select></div>
  <div class="col-md-2"><label class="form-label">Estado</label>
    <select name="estado" class="form-select">
      {% for est,lbl in [('borrador','Borrador'),('enviada','Enviada'),('recibida','Recibida'),('cancelada','Cancelada')] %}
      <option value="{{ est }}" {% if obj and obj.estado==est %}selected{% elif not obj and est=='borrador' %}selected{% endif %}>{{ lbl }}</option>{% endfor %}
    </select></div>
</div>
<div class="row g-3 mb-3">
  <div class="col-md-3"><label class="form-label">Fecha emisión</label>
    <input type="date" name="fecha_emision" id="fechaEmision" class="form-control" value="{{ obj.fecha_emision.strftime('%Y-%m-%d') if obj and obj.fecha_emision else today }}" oninput="calcFechaEntrega()"></div>
  <div class="col-md-3"><label class="form-label">Fecha entrega esperada <small class="text-muted" id="plazoHint"></small></label>
    <input type="date" name="fecha_esperada" id="fechaEsperada" class="form-control" value="{{ obj.fecha_esperada.strftime('%Y-%m-%d') if obj and obj.fecha_esperada else '' }}"></div>
  <div class="col-md-3"><label class="form-label">Fecha estimada pago</label>
    <input type="date" name="fecha_estimada_pago" class="form-control" value="{{ obj.fecha_estimada_pago.strftime('%Y-%m-%d') if obj and obj.fecha_estimada_pago else '' }}"></div>
</div>
<!-- Info de condición de pago desde cotización -->
<div id="condPagoInfo" class="alert alert-info py-2 mb-3" style="display:none;font-size:.85rem">
  <i class="bi bi-credit-card me-2"></i><span id="condPagoText"></span>
</div>
<hr class="my-3">
<h6 class="text-muted text-uppercase mb-3" style="font-size:.75rem;letter-spacing:1px">Transportista y logística</h6>
<div class="row g-3 mb-3">
  <div class="col-md-4"><label class="form-label">Transportista</label>
    <select name="transportista_id" class="form-select" id="selectTrans">
      <option value="">— Sin transportista —</option>
      {% for t in transportistas_list %}<option value="{{ t.id }}" {% if obj and obj.transportista_id==t.id %}selected{% endif %}>{{ t.empresa or t.nombre }}</option>{% endfor %}
    </select></div>
  <div class="col-md-3"><label class="form-label">Fecha estimada recogida</label>
    <input type="date" name="fecha_estimada_recogida" id="fechaRecogida" class="form-control" value="{{ obj.fecha_estimada_recogida.strftime('%Y-%m-%d') if obj and obj.fecha_estimada_recogida else '' }}"></div>
  <div class="col-md-5"><div class="alert alert-warning py-2 mb-0 mt-4" style="font-size:.82rem" id="transAlert" {% if not (obj and obj.transportista_id) %}style="display:none"{% endif %}>
    <i class="bi bi-exclamation-triangle me-1"></i>Al guardar se creará automáticamente una <strong>tarea</strong> para contactar al transportista 2 días antes de la recogida.</div></div>
</div>
<hr class="my-3">
<div class="d-flex justify-content-between align-items-center mb-2">
  <h6 class="text-muted mb-0 text-uppercase" style="letter-spacing:1px;font-size:.75rem">Ítems de la orden</h6>
  <div class="d-flex gap-2">
    <button type="button" class="btn btn-sm btn-outline-success" onclick="importarDesdeCot()" id="btnImportCot" {% if not cotizaciones_list %}disabled{% endif %} title="Importar ítem desde cotización seleccionada">
      <i class="bi bi-cloud-download me-1"></i>Importar de cotización</button>
    <button type="button" class="btn btn-sm btn-outline-primary" onclick="addItem()"><i class="bi bi-plus-lg me-1"></i>Agregar ítem manual</button>
  </div>
</div>
<div id="itemsContainer"></div>
<div class="mt-3 mb-4"><div class="row g-2 justify-content-end"><div class="col-md-4">
  <div class="d-flex justify-content-between py-1 border-bottom"><span class="text-muted">Subtotal:</span><strong id="lblSub">$ 0</strong></div>
  <div class="d-flex justify-content-between py-1 border-bottom"><span class="text-muted">IVA 19%:</span><strong id="lblIva">$ 0</strong></div>
  <div class="d-flex justify-content-between py-1"><span class="fw-bold">Total:</span><strong id="lblTot" style="color:#5e72e4;font-size:1.1rem">$ 0</strong></div>
</div></div></div>
<input type="hidden" name="subtotal_calc" id="subtotalCalc">
<input type="hidden" name="iva_calc" id="ivaCalc">
<input type="hidden" name="total_calc" id="totalCalc">
<div class="mb-3"><label class="form-label">Notas</label>
  <textarea name="notas" class="form-control" rows="2">{{ obj.notas if obj else '' }}</textarea></div>
<div class="d-flex gap-2 mt-4">
  <button type="submit" class="btn btn-primary"><i class="bi bi-check-lg me-1"></i>{{ 'Actualizar' if obj else 'Crear Orden' }}</button>
  <a href="{{ url_for('ordenes_compra') }}" class="btn btn-outline-secondary">Cancelar</a>
</div></form></div>
{% block scripts %}<script>
var ITEMS_EDIT = {{ items_json|tojson }};
var UNIDADES = ['unidades','kg','g','litros','ml','libras','galones','metros','cm','piezas','cajas','bolsas'];
function fmt(n){return'$ '+Math.round(n).toLocaleString('es-CO');}

function recalc(){
  var sub=0;
  document.querySelectorAll('.oc-row').forEach(function(row){
    var c=parseFloat(row.querySelector('[data-f="cant"]').value)||0;
    var p=parseFloat(row.querySelector('[data-f="precio"]').value)||0;
    var st=c*p; sub+=st;
    row.querySelector('[data-f="sub"]').textContent=fmt(st);
  });
  var iva=sub*0.19; var tot=sub+iva;
  document.getElementById('lblSub').textContent=fmt(sub);
  document.getElementById('lblIva').textContent=fmt(iva);
  document.getElementById('lblTot').textContent=fmt(tot);
  document.getElementById('subtotalCalc').value=sub.toFixed(2);
  document.getElementById('ivaCalc').value=iva.toFixed(2);
  document.getElementById('totalCalc').value=tot.toFixed(2);
}

function addItem(d){
  d=d||{};
  var unOpts=UNIDADES.map(function(u){return'<option value="'+u+'"'+(d.unidad===u?' selected':'')+'>'+u+'</option>';}).join('');
  var html='<div class="oc-row row g-2 mb-2 align-items-end" style="background:#f8f9fe;border-radius:8px;padding:.5rem">'
    +'<input type="hidden" name="item_cot_id[]" value="'+(d.cot_id||'')+'"/>'
    +'<div class="col-md-3"><label class="form-label mb-1" style="font-size:.8rem">Ítem *</label>'
    +'<input type="text" name="item_nombre[]" class="form-control form-control-sm" value="'+(d.nombre||'')+'" required></div>'
    +'<div class="col-md-2"><label class="form-label mb-1" style="font-size:.8rem">Descripción</label>'
    +'<input type="text" name="item_desc[]" class="form-control form-control-sm" value="'+(d.desc||'')+'"></div>'
    +'<div class="col-md-1"><label class="form-label mb-1" style="font-size:.8rem">Cant.</label>'
    +'<input type="number" name="item_cant[]" data-f="cant" class="form-control form-control-sm" step="0.001" min="0.001" value="'+(d.cant||1)+'" oninput="recalc()"></div>'
    +'<div class="col-md-2"><label class="form-label mb-1" style="font-size:.8rem">Unidad</label>'
    +'<select name="item_unidad[]" class="form-select form-select-sm">'+unOpts+'</select></div>'
    +'<div class="col-md-2"><label class="form-label mb-1" style="font-size:.8rem">Precio unit. $</label>'
    +'<input type="number" name="item_precio[]" data-f="precio" class="form-control form-control-sm" step="1" min="0" value="'+(d.precio||0)+'" oninput="recalc()"></div>'
    +'<div class="col-md-1"><label class="form-label mb-1" style="font-size:.8rem">Subtotal</label>'
    +'<span class="form-control-plaintext form-control-sm fw-semibold" data-f="sub" style="font-size:.82rem">$ 0</span></div>'
    +'<div class="col-auto d-flex align-items-end"><button type="button" class="btn btn-sm btn-outline-danger" onclick="this.closest(\'.oc-row\').remove();recalc()"><i class="bi bi-trash"></i></button></div>'
    +(d.cot_id ? '<div class="col-12"><small class="text-success"><i class="bi bi-link-45deg me-1"></i>Vinculado a cotización</small></div>' : '')
    +'</div>';
  document.getElementById('itemsContainer').insertAdjacentHTML('beforeend',html);
  recalc();
}

function onCotChange(){
  var sel=document.getElementById('selectCot');
  var opt=sel.options[sel.selectedIndex];
  if(!opt||!opt.value){ document.getElementById('condPagoInfo').style.display='none'; return; }
  var plazo=parseInt(opt.dataset.plazo)||0;
  var prov=opt.dataset.prov;
  var cond=opt.dataset.cond;
  // Auto-fill proveedor if empty
  if(prov && !document.getElementById('selectProv').value){
    document.getElementById('selectProv').value=prov;
  }
  // Recalculate delivery date
  calcFechaEntrega();
  // Show payment condition info
  var condTexts={'contado':'Pago de contado al recibir','credito':'Pago a crédito','anticipo_saldo':'Anticipo + saldo contra entrega','consignacion':'Pago por consignación'};
  document.getElementById('condPagoText').textContent='Condición de pago: '+(condTexts[cond]||cond);
  document.getElementById('condPagoInfo').style.display='';
}

function calcFechaEntrega(){
  var sel=document.getElementById('selectCot');
  var opt=sel&&sel.options[sel.selectedIndex];
  var plazo=opt&&opt.dataset?parseInt(opt.dataset.plazo)||0:0;
  if(!plazo) return;
  var feVal=document.getElementById('fechaEmision').value;
  if(!feVal) return;
  var fe=new Date(feVal+'T12:00:00');
  fe.setDate(fe.getDate()+plazo);
  var yy=fe.getFullYear();
  var mm=String(fe.getMonth()+1).padStart(2,'0');
  var dd=String(fe.getDate()).padStart(2,'0');
  document.getElementById('fechaEsperada').value=yy+'-'+mm+'-'+dd;
  document.getElementById('plazoHint').textContent='(auto: '+plazo+'d desde emisión)';
}

function importarDesdeCot(){
  var sel=document.getElementById('selectCot');
  var opt=sel.options[sel.selectedIndex];
  if(!opt||!opt.value){alert('Selecciona una cotización primero.');return;}
  addItem({
    nombre: opt.dataset.nombre||'',
    desc: '',
    cant: 1,
    unidad: opt.dataset.unidad||'unidades',
    precio: parseFloat(opt.dataset.precio)||0,
    cot_id: opt.value
  });
}

// Transportista alert
document.getElementById('selectTrans').addEventListener('change',function(){
  document.getElementById('transAlert').style.display=this.value?'':'none';
});

// Init
ITEMS_EDIT.forEach(function(it){addItem(it);});
if(ITEMS_EDIT.length===0) addItem();
onCotChange();
</script>{% endblock %}{% endblock %}"""

T['proveedores/index.html'] = """{% extends 'base.html' %}
{% block title %}Proveedores{% endblock %}{% block page_title %}Proveedores y Transportistas{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('proveedor_nuevo') }}" class="btn btn-primary btn-sm"><i class="bi bi-plus-lg me-1"></i>Nuevo</a>{% endblock %}
{% block content %}
<div class="tc mb-3"><div class="p-3"><form method="GET" class="row g-2 align-items-end">
  <div class="col-sm-4"><input type="text" name="buscar" class="form-control form-control-sm" placeholder="Nombre, empresa, NIT..." value="{{ busqueda }}"></div>
  <div class="col-sm-3">
    <select name="tipo_f" class="form-select form-select-sm">
      <option value="">Todos los tipos</option>
      {% for v,l in [('proveedor','Proveedores'),('transportista','Transportistas'),('ambos','Proveedor + Transportista')] %}
      <option value="{{ v }}" {% if request.args.get('tipo_f')==v %}selected{% endif %}>{{ l }}</option>{% endfor %}
    </select></div>
  <div class="col-auto">
    <button type="submit" class="btn btn-primary btn-sm"><i class="bi bi-search"></i></button>
    <a href="{{ url_for('proveedores') }}" class="btn btn-outline-secondary btn-sm">Limpiar</a>
  </div>
</form></div></div>
<!-- Resumen por tipo -->
<div class="row g-2 mb-3">
  {% set n_prov = items|selectattr('tipo','equalto','proveedor')|list|length + items|selectattr('tipo','equalto','ambos')|list|length %}
  {% set n_trans = items|selectattr('tipo','equalto','transportista')|list|length + items|selectattr('tipo','equalto','ambos')|list|length %}
  <div class="col-auto"><div class="fc px-3 py-2 d-flex gap-2 align-items-center"><i class="bi bi-truck text-primary"></i><span class="fw-semibold">{{ n_prov }}</span><small class="text-muted">proveedores</small></div></div>
  <div class="col-auto"><div class="fc px-3 py-2 d-flex gap-2 align-items-center"><i class="bi bi-truck-front text-success"></i><span class="fw-semibold">{{ n_trans }}</span><small class="text-muted">transportistas</small></div></div>
</div>
<div class="tc"><div class="ch"><i class="bi bi-truck me-2"></i>{{ items|length }} registro(s)</div>
{% if items %}<div class="table-responsive"><table class="table">
  <thead><tr><th>Empresa / Nombre</th><th>Tipo</th><th>NIT</th><th>Categoría</th><th>Email</th><th>Teléfono</th><th></th></tr></thead>
  <tbody>{% for p in items %}<tr>
    <td><span class="fw-semibold" style="color:#1a1f36">{{ p.empresa or p.nombre }}</span>
      {% if p.empresa %}<div><small class="text-muted">{{ p.nombre }}</small></div>{% endif %}</td>
    <td>
      {% if p.tipo == 'transportista' %}<span class="badge bg-success">Transportista</span>
      {% elif p.tipo == 'ambos' %}<span class="badge bg-primary">Prov. + Transp.</span>
      {% else %}<span class="badge bg-secondary">Proveedor</span>{% endif %}
    </td>
    <td><small class="text-muted">{{ p.nit or '—' }}</small></td>
    <td><small>{{ p.categoria or '—' }}</small></td>
    <td><small>{{ p.email or '—' }}</small></td>
    <td><small>{{ p.telefono or '—' }}</small></td>
    <td><div class="d-flex gap-1">
      <a href="{{ url_for('proveedor_editar', id=p.id) }}" class="btn btn-sm btn-outline-secondary"><i class="bi bi-pencil"></i></a>
      <form method="POST" action="{{ url_for('proveedor_eliminar', id=p.id) }}" onsubmit="return confirm('¿Eliminar?')">
        <button class="btn btn-sm btn-outline-danger"><i class="bi bi-trash"></i></button></form>
    </div></td>
  </tr>{% endfor %}</tbody>
</table></div>
{% else %}<div class="text-center text-muted py-5"><i class="bi bi-truck" style="font-size:3rem"></i>
  <p class="mt-3">Sin registros.</p><a href="{{ url_for('proveedor_nuevo') }}" class="btn btn-primary">Agregar</a></div>
{% endif %}</div>{% endblock %}"""

T['proveedores/form.html'] = """{% extends 'base.html' %}
{% block title %}{{ titulo }}{% endblock %}{% block page_title %}{{ titulo }}{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('proveedores') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Volver</a>{% endblock %}
{% block content %}<div class="fc" style="max-width:900px"><form method="POST">
<h6 class="text-muted mb-3 text-uppercase" style="letter-spacing:1px;font-size:.75rem">Tipo de registro</h6>
<div class="row g-3 mb-3">
  <div class="col-12">
    <div class="d-flex gap-3 flex-wrap">
      {% for v,ic,col,lbl,desc in [('proveedor','bi-truck','secondary','Proveedor','Vende o suministra productos/materias primas'),('transportista','bi-truck-front','success','Transportista','Presta servicios de logística y transporte'),('ambos','bi-boxes','primary','Proveedor + Transportista','Cumple ambas funciones')] %}
      <label class="d-flex align-items-center gap-2 p-3 rounded border" style="cursor:pointer;flex:1;min-width:180px;max-width:260px">
        <input type="radio" name="tipo" value="{{ v }}" {% if (obj and obj.tipo==v) or (not obj and v=='proveedor') %}checked{% endif %}>
        <i class="bi {{ ic }} text-{{ col }}" style="font-size:1.3rem"></i>
        <div><div class="fw-semibold">{{ lbl }}</div><small class="text-muted">{{ desc }}</small></div>
      </label>{% endfor %}
    </div>
  </div>
</div>
<h6 class="text-muted mb-2 text-uppercase" style="letter-spacing:1px;font-size:.75rem">Información general</h6>
<div class="row g-3 mb-4">
  <div class="col-md-4"><label class="form-label">Nombre contacto</label>
    <input type="text" name="nombre" class="form-control" value="{{ obj.nombre if obj else '' }}"></div>
  <div class="col-md-5"><label class="form-label">Empresa *</label>
    <input type="text" name="empresa" class="form-control" value="{{ obj.empresa if obj else '' }}" required></div>
  <div class="col-md-3"><label class="form-label">NIT</label>
    <input type="text" name="nit" class="form-control" placeholder="900.123.456-7" value="{{ obj.nit if obj else '' }}"></div>
  <div class="col-md-4"><label class="form-label">Categoría</label>
    <input type="text" name="categoria" class="form-control" placeholder="Ej: insumos, logística..." value="{{ obj.categoria if obj else '' }}"></div>
  <div class="col-md-4"><label class="form-label">Email</label>
    <input type="email" name="email" class="form-control" value="{{ obj.email if obj else '' }}"></div>
  <div class="col-md-4"><label class="form-label">Teléfono</label>
    <input type="text" name="telefono" class="form-control" value="{{ obj.telefono if obj else '' }}"></div>
  <div class="col-12"><label class="form-label">Dirección</label>
    <input type="text" name="direccion" class="form-control" value="{{ obj.direccion if obj else '' }}"></div>
  <div class="col-12"><label class="form-label">Notas</label>
    <textarea name="notas" class="form-control" rows="2">{{ obj.notas if obj else '' }}</textarea></div>
</div>
<div class="d-flex gap-2 mt-4">
  <button type="submit" class="btn btn-primary"><i class="bi bi-check-lg me-1"></i>{{ 'Actualizar' if obj else 'Crear' }}</button>
  <a href="{{ url_for('proveedores') }}" class="btn btn-outline-secondary">Cancelar</a>
</div></form></div>{% endblock %}"""

T['ventas/index.html'] = """{% extends 'base.html' %}
{% block title %}Ventas{% endblock %}{% block page_title %}Ventas{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('venta_nueva') }}" class="btn btn-primary btn-sm"><i class="bi bi-plus-lg me-1"></i>Nueva</a>{% endblock %}
{% block content %}
<div class="mb-3 d-flex gap-2 flex-wrap">
  {% for est,lbl in [('','Todas'),('prospecto','Prospecto'),('negociacion','Negociación'),('anticipo_pagado','Anticipo'),('ganado','Ganado'),('perdido','Perdido')] %}
  <a href="{{ url_for('ventas', estado=est) }}" class="btn btn-sm {{ 'btn-primary' if estado_f==est else 'btn-outline-secondary' }}">{{ lbl }}</a>{% endfor %}
</div>
<div class="tc"><div class="ch"><i class="bi bi-graph-up-arrow me-2"></i>{{ items|length }} venta(s)</div>
{% if items %}<div class="table-responsive"><table class="table">
  <thead><tr><th>N°</th><th>Título</th><th>Cliente</th><th>Total</th><th>Anticipo</th><th>Saldo</th><th>Estado</th><th>Entrega est.</th><th></th></tr></thead>
  <tbody>{% for v in items %}<tr>
    <td><small class="text-muted fw-semibold">{{ v.numero or '—' }}</small></td>
    <td class="fw-semibold" style="color:#1a1f36">{{ v.titulo }}</td>
    <td>{% if v.cliente %}<a href="{{ url_for('cliente_ver', id=v.cliente.id) }}" class="text-decoration-none">{{ v.cliente.empresa or v.cliente.nombre }}</a>{% else %}—{% endif %}</td>
    <td class="fw-semibold">$ {{ '{:,.0f}'.format(v.total).replace(',','.') }}</td>
    <td>$ {{ '{:,.0f}'.format(v.monto_anticipo).replace(',','.') }}</td>
    <td class="{{ 'text-danger fw-semibold' if v.saldo > 0 else '' }}">$ {{ '{:,.0f}'.format(v.saldo).replace(',','.') }}</td>
    <td><span class="b b-{{ v.estado }}">{{ v.estado.replace('_',' ').title() }}</span></td>
    <td>{% if v.fecha_entrega_est %}<small>{{ v.fecha_entrega_est.strftime('%d/%m/%Y') }}</small>{% else %}—{% endif %}</td>
    <td><div class="d-flex gap-1 flex-wrap">
      <a href="{{ url_for('venta_factura', id=v.id) }}" class="btn btn-sm btn-outline-primary" target="_blank" title="Ver factura/cotización"><i class="bi bi-file-earmark-text"></i></a>
      <a href="{{ url_for('venta_editar', id=v.id) }}" class="btn btn-sm btn-outline-secondary"><i class="bi bi-pencil"></i></a>
      {# Botón Informar al cliente: visible si hay producción completada y aún no se informó #}
      {% set ops_completadas = v.ordenes_produccion|selectattr('estado','equalto','completado')|list %}
      {% if ops_completadas and not v.cliente_informado_en and not v.entregado_en %}
      <form method="POST" action="{{ url_for('venta_informar_cliente', id=v.id) }}" class="d-inline">
        <button class="btn btn-sm btn-warning" title="Producción lista — Informar al cliente">
          <i class="bi bi-envelope-check me-1"></i>Informar cliente
        </button>
      </form>
      {% endif %}
      {# Botón Entregado: visible si ya se informó al cliente y aún no está entregado #}
      {% if v.cliente_informado_en and not v.entregado_en %}
      <form method="POST" action="{{ url_for('venta_entregar', id=v.id) }}" class="d-inline"
            onsubmit="return confirm('¿Marcar como entregado? Esto descontará las unidades del inventario.')">
        <button class="btn btn-sm btn-success" title="Marcar como entregado y descontar stock">
          <i class="bi bi-box-arrow-up me-1"></i>Entregado
        </button>
      </form>
      {% endif %}
      {# Estado entregado #}
      {% if v.entregado_en %}
      <span class="badge bg-success align-self-center"><i class="bi bi-check2-all me-1"></i>Entregado</span>
      {% endif %}
      <form method="POST" action="{{ url_for('venta_eliminar', id=v.id) }}" onsubmit="return confirm('¿Eliminar?')">
        <button class="btn btn-sm btn-outline-danger"><i class="bi bi-trash"></i></button></form>
    </div></td>
  </tr>{% endfor %}</tbody>
</table></div>
{% else %}<div class="text-center text-muted py-5"><i class="bi bi-graph-up-arrow" style="font-size:3rem"></i>
  <p class="mt-3">Sin ventas.</p><a href="{{ url_for('venta_nueva') }}" class="btn btn-primary">Crear primera</a></div>
{% endif %}</div>{% endblock %}"""

T['ventas/form.html'] = """{% extends 'base.html' %}
{% block title %}{{ titulo }}{% endblock %}{% block page_title %}{{ titulo }}{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('ventas') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Volver</a>{% endblock %}
{% block content %}<div class="fc" style="max-width:960px"><form method="POST">
<div class="row g-3 mb-3">
  <div class="col-md-8"><label class="form-label">Título *</label>
    <input type="text" name="titulo" class="form-control" value="{{ obj.titulo if obj else '' }}" required></div>
  <div class="col-md-4"><label class="form-label">Cliente</label>
    <select name="cliente_id" class="form-select"><option value="">Sin cliente</option>
      {% for c in clientes_list %}<option value="{{ c.id }}" {% if obj and obj.cliente_id==c.id %}selected{% endif %}>{{ c.empresa or c.nombre }}</option>{% endfor %}
    </select></div>
</div><hr class="my-3">
<div class="d-flex justify-content-between align-items-center mb-2">
  <h6 class="text-muted mb-0 text-uppercase" style="letter-spacing:1px;font-size:.75rem">Productos</h6>
  <button type="button" class="btn btn-sm btn-outline-primary" onclick="addProd()"><i class="bi bi-plus-lg me-1"></i>Agregar</button>
</div>
<div id="prodsContainer"></div>
<div class="totales-box mt-3 mb-4"><div class="row g-2 justify-content-end"><div class="col-md-4">
  <div class="d-flex justify-content-between py-1 border-bottom"><span class="text-muted">Subtotal:</span><strong id="lblSub">$ 0</strong></div>
  <div class="d-flex justify-content-between py-1 border-bottom"><span class="text-muted">IVA 19%:</span><strong id="lblIva">$ 0</strong></div>
  <div class="d-flex justify-content-between py-1"><span class="fw-bold">Total:</span><strong id="lblTot" style="color:#5e72e4;font-size:1.1rem">$ 0</strong></div>
</div></div></div>
<hr class="my-3">
<div class="row g-3 mb-3">
  <div class="col-md-3"><label class="form-label">Estado</label>
    <select name="estado" class="form-select">
      {% for est,lbl in [('prospecto','Prospecto'),('negociacion','Negociación'),('anticipo_pagado','Anticipo Pagado'),('ganado','Ganado'),('perdido','Perdido')] %}
      <option value="{{ est }}" {% if obj and obj.estado==est %}selected{% elif not obj and est=='prospecto' %}selected{% endif %}>{{ lbl }}</option>{% endfor %}
    </select></div>
  <div class="col-md-3"><label class="form-label">% Anticipo</label>
    <div class="input-group"><input type="number" name="porcentaje_anticipo" id="pctA" class="form-control" min="0" max="100" step="1"
      value="{{ obj.porcentaje_anticipo|int if obj else '0' }}" oninput="calcA()"><span class="input-group-text">%</span></div></div>
  <div class="col-md-3"><label class="form-label">Monto anticipo COP</label>
    <input type="text" id="montoAVis" class="form-control" readonly style="background:#f8f9fe">
    <input type="hidden" name="monto_anticipo" id="montoAHid"></div>
  <div class="col-md-3"><label class="form-label">Saldo a pagar</label>
    <input type="text" id="saldoVis" class="form-control" readonly style="background:#f8f9fe">
    <input type="hidden" name="saldo" id="saldoHid"></div>
  <div class="col-md-4"><label class="form-label">Fecha pago anticipo</label>
    <input type="date" name="fecha_anticipo" id="fAnticipo" class="form-control"
      value="{{ obj.fecha_anticipo.strftime('%Y-%m-%d') if obj and obj.fecha_anticipo else '' }}" onchange="calcEnt()"></div>
  <div class="col-md-4"><label class="form-label">Días de producción</label>
    <select name="dias_entrega" id="diasEnt" class="form-select" onchange="calcEnt()">
      <option value="30" {% if not obj or obj.dias_entrega==30 %}selected{% endif %}>30 días</option>
      <option value="45" {% if obj and obj.dias_entrega==45 %}selected{% endif %}>45 días</option>
    </select></div>
  <div class="col-md-4"><label class="form-label">Entrega estimada</label>
    <input type="text" id="entVis" class="form-control" readonly style="background:#f8f9fe">
    <input type="hidden" name="fecha_entrega_est" id="entHid" value="{{ obj.fecha_entrega_est.strftime('%Y-%m-%d') if obj and obj.fecha_entrega_est else '' }}"></div>
  <div class="col-12"><label class="form-label">Notas</label>
    <textarea name="notas" class="form-control" rows="2">{{ obj.notas if obj else '' }}</textarea></div>
</div>
<input type="hidden" name="subtotal_calc" id="subHid">
<input type="hidden" name="iva_calc" id="ivaHid">
<input type="hidden" name="total_calc" id="totHid">
<div class="d-flex gap-2 mt-4">
  <button type="submit" class="btn btn-primary"><i class="bi bi-check-lg me-1"></i>{{ 'Actualizar' if obj else 'Crear Venta' }}</button>
  <a href="{{ url_for('ventas') }}" class="btn btn-outline-secondary">Cancelar</a>
</div></form></div>{% endblock %}
{% block scripts %}<script>
const PRODS={{ productos_json|tojson }};const ITEMS={{ items_json|tojson }};let totG=0;
function fCOP(n){return '$ '+Math.round(n).toLocaleString('es-CO');}
function stockInfo(p,cant){
  if(!p||p.stock===undefined) return '';
  var en_stock=Math.min(p.stock,cant);
  var producir=Math.max(0,cant-p.stock);
  var cls=p.stock>=cant?'text-success':'text-warning';
  return '<div class="mt-1" style="font-size:.76rem"><span class="'+cls+'"><i class="bi bi-box-seam me-1"></i>Stock: '+p.stock+'</span>'
    +(producir>0?'<span class="ms-2 text-warning"><i class="bi bi-gear me-1"></i>A producir: '+producir+'</span>':'<span class="ms-2 text-success"> ✓ Hay stock suficiente</span>')
    +'</div>';
}
function addProd(pid='',cant=1,precio=0){
  const opts=PRODS.map(p=>`<option value="${p.id}" data-p="${p.precio}" data-stock="${p.stock}" ${p.id==pid?'selected':''}>${p.nombre}${p.sku?' ('+p.sku+')':''}</option>`).join('');
  const idx=Date.now();
  const selProd=PRODS.find(p=>p.id==pid)||{};
  document.getElementById('prodsContainer').insertAdjacentHTML('beforeend',`<div class="prod-row mb-2" id="pr${idx}">
    <div class="row g-2 align-items-end">
      <div class="col-md-5"><label class="form-label mb-1" style="font-size:.8rem">Producto</label>
        <select name="prod_id[]" class="form-select form-select-sm" onchange="onPC(this,${idx})">${opts}</select>
        <div id="si${idx}">${stockInfo(selProd,cant)}</div></div>
      <div class="col-md-2"><label class="form-label mb-1" style="font-size:.8rem">Cantidad</label>
        <input type="number" name="prod_cant[]" class="form-control form-control-sm" min="0.01" step="0.01" value="${cant}" oninput="onCantChange(this,${idx})"></div>
      <div class="col-md-3"><label class="form-label mb-1" style="font-size:.8rem">Precio unit. COP sin IVA</label>
        <input type="number" name="prod_precio[]" id="pp${idx}" class="form-control form-control-sm" min="0" step="1" value="${precio}" oninput="calcRow(${idx})"></div>
      <div class="col-md-1"><label class="form-label mb-1" style="font-size:.8rem">Subtotal</label>
        <input type="text" id="ps${idx}" class="form-control form-control-sm" readonly style="background:#f8f9fe"></div>
      <div class="col-md-1"><label class="form-label mb-1" style="font-size:.8rem"> </label>
        <button type="button" class="btn btn-sm btn-outline-danger w-100" onclick="document.getElementById('pr${idx}').remove();calcTot()"><i class="bi bi-trash"></i></button></div>
    </div></div>`);
  calcRow(idx);
}
function onPC(sel,idx){
  const o=sel.options[sel.selectedIndex];
  document.getElementById('pp'+idx).value=Math.round(o.dataset.p||0);
  const stock=parseInt(o.dataset.stock||0);
  const pr=document.getElementById('pr'+idx);
  const cant=parseFloat(pr.querySelector('[name="prod_cant[]"]').value)||1;
  const p={stock:stock};
  document.getElementById('si'+idx).innerHTML=stockInfo(p,cant);
  calcRow(idx);
}
function onCantChange(el,idx){
  const pr=document.getElementById('pr'+idx);
  const sel=pr.querySelector('select');
  const o=sel.options[sel.selectedIndex];
  const stock=parseInt(o.dataset.stock||0);
  const cant=parseFloat(el.value)||0;
  document.getElementById('si'+idx).innerHTML=stockInfo({stock:stock},cant);
  calcRow(idx);
}
function calcRow(idx){
  const pr=document.getElementById('pr'+idx);if(!pr)return;
  const c=parseFloat(pr.querySelector('[name="prod_cant[]"]').value)||0;
  const p=parseFloat(document.getElementById('pp'+idx).value)||0;
  document.getElementById('ps'+idx).value=fCOP(c*p);calcTot();
}
function calcTot(){
  let sub=0;
  document.querySelectorAll('[name="prod_cant[]"]').forEach((el,i)=>{
    sub+=(parseFloat(el.value)||0)*(parseFloat(document.querySelectorAll('[name="prod_precio[]"]')[i].value)||0);
  });
  const iva=sub*.19,tot=sub+iva;totG=tot;
  document.getElementById('lblSub').textContent=fCOP(sub);
  document.getElementById('lblIva').textContent=fCOP(iva);
  document.getElementById('lblTot').textContent=fCOP(tot);
  document.getElementById('subHid').value=Math.round(sub);
  document.getElementById('ivaHid').value=Math.round(iva);
  document.getElementById('totHid').value=Math.round(tot);calcA();
}
function calcA(){
  const pct=parseFloat(document.getElementById('pctA').value)||0;
  const ant=totG*pct/100,sal=totG-ant;
  document.getElementById('montoAVis').value=fCOP(ant);document.getElementById('montoAHid').value=Math.round(ant);
  document.getElementById('saldoVis').value=fCOP(sal);document.getElementById('saldoHid').value=Math.round(sal);
}
function calcEnt(){
  const fa=document.getElementById('fAnticipo').value;const d=parseInt(document.getElementById('diasEnt').value)||30;
  if(fa){const dt=new Date(fa);dt.setDate(dt.getDate()+d);const iso=dt.toISOString().split('T')[0];
    const p=iso.split('-');document.getElementById('entVis').value=p[2]+'/'+p[1]+'/'+p[0];document.getElementById('entHid').value=iso;}
}
if(ITEMS.length>0){ITEMS.forEach(it=>addProd(it.pid,it.cant,it.precio));}else{addProd();}
calcEnt();
const feh=document.getElementById('entHid').value;if(feh){const p=feh.split('-');document.getElementById('entVis').value=p[2]+'/'+p[1]+'/'+p[0];}
// Auto-título: cliente + estado + fecha anticipo
setupAutoTitulo('[name="titulo"]',['[name="cliente_id"]','[name="estado"]','[name="fecha_anticipo"]']);
</script>{% endblock %}"""

T['tareas/index.html'] = """{% extends 'base.html' %}
{% block title %}Tareas{% endblock %}{% block page_title %}Tareas{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('tarea_nueva') }}" class="btn btn-primary btn-sm"><i class="bi bi-plus-lg me-1"></i>Nueva</a>{% endblock %}
{% block content %}
<div class="mb-3 d-flex gap-2 flex-wrap">
  {% for est,lbl in [('','Todas'),('pendiente','Pendiente'),('en_progreso','En progreso'),('completada','Completada')] %}
  <a href="{{ url_for('tareas', estado=est) }}" class="btn btn-sm {{ 'btn-primary' if estado_f==est else 'btn-outline-secondary' }}">{{ lbl }}</a>{% endfor %}
  <span class="text-muted mx-1">|</span>
  <a href="{{ url_for('tareas', prioridad='alta') }}" class="btn btn-sm {{ 'btn-danger' if prioridad_f=='alta' else 'btn-outline-secondary' }}">Alta prioridad</a>
</div>
<div class="tc"><div class="ch"><i class="bi bi-check2-square me-2"></i>{{ items|length }} tarea(s)</div>
{% if items %}<div class="table-responsive"><table class="table">
  <thead><tr><th>Tarea</th><th>Prioridad</th><th>Estado</th><th>Asignada a</th><th>Vence</th><th></th></tr></thead>
  <tbody>{% for t in items %}<tr {% if t.estado=='completada' %}style="opacity:.6"{% endif %}>
    <td><a href="{{ url_for('tarea_ver', id=t.id) }}" class="fw-semibold text-decoration-none {% if t.estado=='completada' %}text-decoration-line-through{% endif %}" style="color:#1a1f36">{{ t.titulo }}</a>
      {% if t.comentarios %}<span class="badge bg-light text-secondary ms-1" style="font-size:.65rem"><i class="bi bi-chat-dots"></i> {{ t.comentarios|length }}</span>{% endif %}</td>
    <td><span class="b b-{{ t.prioridad }}">{{ t.prioridad.title() }}</span></td>
    <td><span class="b b-{{ t.estado }}">{{ t.estado.replace('_',' ').title() }}</span></td>
    <td>{{ t.asignado_user.nombre if t.asignado_user else '—' }}</td>
    <td>{% if t.fecha_vencimiento %}
      <small class="{{ 'text-danger fw-bold' if t.estado!='completada' and t.fecha_vencimiento < now.date() else 'text-muted' }}">{{ t.fecha_vencimiento.strftime('%d/%m/%Y') }}</small>
      {% else %}—{% endif %}</td>
    <td><div class="d-flex gap-1">
      <a href="{{ url_for('tarea_ver', id=t.id) }}" class="btn btn-sm btn-outline-primary" title="Ver y chatear"><i class="bi bi-chat-dots"></i></a>
      {% if t.estado!='completada' %}<form method="POST" action="{{ url_for('tarea_completar', id=t.id) }}">
        <button class="btn btn-sm btn-outline-success"><i class="bi bi-check2"></i></button></form>{% endif %}
      <a href="{{ url_for('tarea_editar', id=t.id) }}" class="btn btn-sm btn-outline-secondary"><i class="bi bi-pencil"></i></a>
      <form method="POST" action="{{ url_for('tarea_eliminar', id=t.id) }}" onsubmit="return confirm('¿Eliminar?')">
        <button class="btn btn-sm btn-outline-danger"><i class="bi bi-trash"></i></button></form>
    </div></td>
  </tr>{% endfor %}</tbody>
</table></div>
{% else %}<div class="text-center text-muted py-5"><i class="bi bi-check2-all" style="font-size:3rem"></i>
  <p class="mt-3">Sin tareas.</p><a href="{{ url_for('tarea_nueva') }}" class="btn btn-primary">Crear primera</a></div>
{% endif %}</div>{% endblock %}"""

T['tareas/ver.html'] = """{% extends 'base.html' %}
{% block title %}{{ tarea.titulo }}{% endblock %}{% block page_title %}{{ tarea.titulo }}{% endblock %}
{% block topbar_actions %}
<a href="{{ url_for('tarea_editar', id=tarea.id) }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-pencil me-1"></i>Editar</a>
{% if tarea.estado != 'completada' %}
<form method="POST" action="{{ url_for('tarea_completar', id=tarea.id) }}" class="d-inline">
  <button class="btn btn-success btn-sm"><i class="bi bi-check2 me-1"></i>Completar</button></form>{% endif %}
<a href="{{ url_for('tareas') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Volver</a>
{% endblock %}
{% block content %}<div class="row g-4">
  <div class="col-lg-4">
    <div class="fc">
      <div class="mb-3">
        <span class="b b-{{ tarea.estado }} me-2">{{ tarea.estado.replace('_',' ').title() }}</span>
        <span class="b b-{{ tarea.prioridad }}">{{ tarea.prioridad.title() }} prioridad</span>
      </div>
      {% if tarea.descripcion %}<p style="font-size:.9rem">{{ tarea.descripcion }}</p><hr>{% endif %}
      <dl class="row mb-0" style="font-size:.88rem">
        <dt class="col-5 text-muted">Creada por</dt><dd class="col-7">{{ tarea.creador.nombre if tarea.creador else '—' }}</dd>
        <dt class="col-5 text-muted">Asignada a</dt><dd class="col-7">{{ tarea.asignado_user.nombre if tarea.asignado_user else '—' }}</dd>
        {% if tarea.asignados %}
        <dt class="col-5 text-muted">También</dt>
        <dd class="col-7">{% for a in tarea.asignados %}{{ a.user.nombre }}{% if not loop.last %}, {% endif %}{% endfor %}</dd>{% endif %}
        {% if tarea.fecha_vencimiento %}
        <dt class="col-5 text-muted">Vence</dt>
        <dd class="col-7 {{ 'text-danger fw-bold' if tarea.estado!='completada' and tarea.fecha_vencimiento < now.date() else '' }}">{{ tarea.fecha_vencimiento.strftime('%d/%m/%Y') }}</dd>{% endif %}
        <dt class="col-5 text-muted">Creada</dt><dd class="col-7">{{ tarea.creado_en.strftime('%d/%m/%Y') }}</dd>
      </dl>
    </div>
  </div>
  <div class="col-lg-8">
    <div class="tc">
      <div class="ch"><i class="bi bi-chat-dots me-2"></i>Conversación ({{ tarea.comentarios|length }} mensajes)</div>
      <div style="max-height:420px;overflow-y:auto;padding:1rem" id="chatBox">
        {% if tarea.comentarios %}
          {% for c in tarea.comentarios %}
          <div class="d-flex {% if c.autor_id == current_user.id %}justify-content-end{% endif %} mb-2">
            <div class="chat-bubble {{ 'mine' if c.autor_id == current_user.id else '' }}">
              <div>{{ c.mensaje }}</div>
              <div class="chat-meta">{{ c.autor.nombre }} · {{ c.creado_en.strftime('%d/%m %H:%M') }}</div>
            </div>
          </div>{% endfor %}
        {% else %}
        <div class="text-center text-muted py-4"><i class="bi bi-chat" style="font-size:2rem"></i>
          <p class="mt-2 mb-0">Sin mensajes aún. ¡Sé el primero!</p></div>{% endif %}
      </div>
      <div class="p-3 border-top">
        <form method="POST" action="{{ url_for('tarea_comentar', id=tarea.id) }}">
          <div class="d-flex gap-2">
            <input type="text" name="mensaje" class="form-control" placeholder="Escribe un mensaje..." required autofocus>
            <button type="submit" class="btn btn-primary"><i class="bi bi-send"></i></button>
          </div>
        </form>
      </div>
    </div>
  </div>
</div>{% endblock %}
{% block scripts %}<script>
const cb=document.getElementById('chatBox');if(cb)cb.scrollTop=cb.scrollHeight;
</script>{% endblock %}"""

T['tareas/form.html'] = """{% extends 'base.html' %}
{% block title %}{{ titulo }}{% endblock %}{% block page_title %}{{ titulo }}{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('tareas') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Volver</a>{% endblock %}
{% block content %}<div class="fc"><form method="POST"><div class="row g-3">
  <div class="col-12"><label class="form-label">Título *</label>
    <input type="text" name="titulo" class="form-control" value="{{ obj.titulo if obj else '' }}" required></div>
  <div class="col-md-4"><label class="form-label">Estado</label>
    <select name="estado" class="form-select">
      <option value="pendiente" {% if not obj or obj.estado=='pendiente' %}selected{% endif %}>Pendiente</option>
      <option value="en_progreso" {% if obj and obj.estado=='en_progreso' %}selected{% endif %}>En progreso</option>
      <option value="completada" {% if obj and obj.estado=='completada' %}selected{% endif %}>Completada</option>
    </select></div>
  <div class="col-md-4"><label class="form-label">Prioridad</label>
    <select name="prioridad" class="form-select">
      <option value="baja" {% if obj and obj.prioridad=='baja' %}selected{% endif %}>Baja</option>
      <option value="media" {% if not obj or obj.prioridad=='media' %}selected{% endif %}>Media</option>
      <option value="alta" {% if obj and obj.prioridad=='alta' %}selected{% endif %}>Alta</option>
    </select></div>
  <div class="col-md-4"><label class="form-label">Fecha vencimiento</label>
    <input type="date" name="fecha_vencimiento" class="form-control"
      value="{{ obj.fecha_vencimiento.strftime('%Y-%m-%d') if obj and obj.fecha_vencimiento else '' }}"></div>
  <div class="col-md-6"><label class="form-label">Asignado principal</label>
    <select name="asignado_a" class="form-select">
      {% for u in usuarios %}<option value="{{ u.id }}" {% if (obj and obj.asignado_a==u.id) or (not obj and u.id==current_user.id) %}selected{% endif %}>
        {{ u.nombre }}{% if u.id==current_user.id %} (yo){% endif %}</option>{% endfor %}
    </select></div>
  <div class="col-md-6"><label class="form-label">También asignar a</label>
    <select name="otros_asignados[]" class="form-select" multiple style="height:90px">
      {% for u in usuarios %}{% if u.id != current_user.id %}
      <option value="{{ u.id }}" {% if obj and u.id in asignados_ids %}selected{% endif %}>{{ u.nombre }}</option>
      {% endif %}{% endfor %}
    </select>
    <div class="form-text">Ctrl+clic para seleccionar varios</div></div>
  <div class="col-12"><label class="form-label">Descripción</label>
    <textarea name="descripcion" class="form-control" rows="4">{{ obj.descripcion if obj else '' }}</textarea></div>
</div>
<div class="d-flex gap-2 mt-4">
  <button type="submit" class="btn btn-primary"><i class="bi bi-check-lg me-1"></i>{{ 'Actualizar' if obj else 'Crear Tarea' }}</button>
  <a href="{{ url_for('tareas') }}" class="btn btn-outline-secondary">Cancelar</a>
</div></form></div>
{% block scripts %}<script>
setupAutoTitulo('[name="titulo"]',['[name="prioridad"]','[name="asignado_a"]','[name="fecha_vencimiento"]']);
</script>{% endblock %}{% endblock %}"""

T['inventario/index.html'] = """{% extends 'base.html' %}
{% block title %}Inventario{% endblock %}{% block page_title %}Inventario{% endblock %}
{% block topbar_actions %}
<a href="{{ url_for('lotes') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-layers me-1"></i>Lotes</a>
<a href="{{ url_for('inventario_ingresos') }}" class="btn btn-outline-primary btn-sm"><i class="bi bi-plus-square me-1"></i>Multi-Ingreso</a>
<a href="{{ url_for('producto_nuevo') }}" class="btn btn-primary btn-sm"><i class="bi bi-plus-lg me-1"></i>Nuevo Producto</a>
{% endblock %}
{% block content %}
<div class="tc mb-3"><div class="p-3"><form method="GET" class="row g-2 align-items-end">
  <div class="col-sm-5"><input type="text" name="buscar" class="form-control form-control-sm" placeholder="Nombre, SKU, NSO..." value="{{ busqueda }}"></div>
  <div class="col-sm-3"><select name="categoria" class="form-select form-select-sm">
    <option value="">Todas las categorías</option>
    {% for cat in categorias %}<option value="{{ cat }}" {% if categoria_f==cat %}selected{% endif %}>{{ cat }}</option>{% endfor %}
  </select></div>
  <div class="col-auto">
    <button type="submit" class="btn btn-primary btn-sm"><i class="bi bi-search"></i></button>
    <a href="{{ url_for('inventario') }}" class="btn btn-outline-secondary btn-sm">Limpiar</a>
  </div>
</form></div></div>
<div class="tc"><div class="ch"><i class="bi bi-box-seam-fill me-2"></i>{{ items|length }} producto(s)</div>
{% if items %}<div class="table-responsive"><table class="table">
  <thead><tr><th>Producto</th><th>SKU</th><th>NSO (INVIMA)</th><th>Caducidad</th><th>Precio venta</th><th>Stock</th><th></th></tr></thead>
  <tbody>{% for p in items %}<tr>
    <td><div class="fw-semibold" style="color:#1a1f36">{{ p.nombre }}</div>
      {% if p.categoria %}<small class="text-muted">{{ p.categoria }}</small>{% endif %}</td>
    <td><small class="text-muted">{{ p.sku or '—' }}</small></td>
    <td><small class="text-muted">{{ p.nso or '—' }}</small></td>
    <td>{% if p.fecha_caducidad %}
      {% set dias_cad = (p.fecha_caducidad - now.date()).days %}
      <span class="b {% if dias_cad < 30 %}b-alta{% elif dias_cad < 90 %}b-media{% else %}b-activo{% endif %}"
            title="{{ dias_cad }} días">{{ p.fecha_caducidad.strftime('%d/%m/%Y') }}</span>
      {% else %}<small class="text-muted">—</small>{% endif %}</td>
    <td class="fw-semibold">$ {{ '{:,.0f}'.format(p.precio).replace(',','.') }}</td>
    <td><span class="fw-semibold {{ 'text-danger' if p.stock <= p.stock_minimo else 'text-success' }}">{{ p.stock }}</span>
      <small class="text-muted"> / mín {{ p.stock_minimo }}</small>
      {% if p.stock <= p.stock_minimo %}<span class="badge bg-danger ms-1" style="font-size:.65rem">BAJO</span>{% endif %}</td>
    <td><div class="d-flex gap-1">
      <a href="{{ url_for('producto_editar', id=p.id) }}" class="btn btn-sm btn-outline-secondary"><i class="bi bi-pencil"></i></a>
      <form method="POST" action="{{ url_for('producto_eliminar', id=p.id) }}" onsubmit="return confirm('¿Eliminar?')">
        <button class="btn btn-sm btn-outline-danger"><i class="bi bi-trash"></i></button></form>
    </div></td>
  </tr>{% endfor %}</tbody>
</table></div>
{% else %}<div class="text-center text-muted py-5"><i class="bi bi-box-seam" style="font-size:3rem"></i>
  <p class="mt-3">Sin productos.</p><a href="{{ url_for('producto_nuevo') }}" class="btn btn-primary">Agregar</a></div>
{% endif %}</div>{% endblock %}"""

T['inventario/form.html'] = """{% extends 'base.html' %}
{% block title %}{{ titulo }}{% endblock %}{% block page_title %}{{ titulo }}{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('inventario') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Volver</a>{% endblock %}
{% block content %}<div class="fc"><form method="POST"><div class="row g-3">
  <div class="col-md-6"><label class="form-label">Nombre *</label>
    <input type="text" name="nombre" class="form-control" value="{{ obj.nombre if obj else '' }}" required></div>
  <div class="col-md-3"><label class="form-label">SKU</label>
    <input type="text" name="sku" class="form-control" value="{{ obj.sku if obj else '' }}"></div>
  <div class="col-md-3"><label class="form-label">NSO (INVIMA)</label>
    <input type="text" name="nso" class="form-control" placeholder="NSO-XXXX-XXXX" value="{{ obj.nso if obj else '' }}"></div>
  <div class="col-md-3"><label class="form-label">Fecha de caducidad</label>
    <input type="date" name="fecha_caducidad" class="form-control" value="{{ obj.fecha_caducidad.strftime('%Y-%m-%d') if obj and obj.fecha_caducidad else '' }}"></div>
  <div class="col-md-3"><label class="form-label">Precio venta COP</label>
    <div class="input-group"><span class="input-group-text">$</span>
      <input type="number" name="precio" class="form-control" step="1" min="0" value="{{ obj.precio|int if obj else '0' }}"></div></div>
  <div class="col-md-3"><label class="form-label">Costo COP <small class="text-muted">(auto desde Producción)</small></label>
    <div class="input-group"><span class="input-group-text">$</span>
      <input type="number" name="costo" class="form-control" step="1" min="0" value="{{ obj.costo|int if obj else '0' }}"></div></div>
  <div class="col-md-3"><label class="form-label">Stock actual</label>
    <input type="number" name="stock" class="form-control" min="0" value="{{ obj.stock if obj else '0' }}"></div>
  <div class="col-md-3"><label class="form-label">Stock mínimo</label>
    <input type="number" name="stock_minimo" class="form-control" min="0" value="{{ obj.stock_minimo if obj else '5' }}"></div>
  <div class="col-md-6"><label class="form-label">Categoría</label>
    <input type="text" name="categoria" class="form-control" value="{{ obj.categoria if obj else '' }}"></div>
  <div class="col-12"><label class="form-label">Descripción</label>
    <textarea name="descripcion" class="form-control" rows="3">{{ obj.descripcion if obj else '' }}</textarea></div>
</div>

<div class="d-flex gap-2 mt-4">
  <button type="submit" class="btn btn-primary"><i class="bi bi-check-lg me-1"></i>{{ 'Actualizar' if obj else 'Crear Producto' }}</button>
  <a href="{{ url_for('inventario') }}" class="btn btn-outline-secondary">Cancelar</a>
</div></form></div>
{% endblock %}"""

T['produccion/index.html'] = """{% extends 'base.html' %}
{% block title %}Producción{% endblock %}{% block page_title %}Producción{% endblock %}
{% block content %}
<div class="row g-3 mb-4">
  <div class="col-md-3"><div class="sc"><div class="sv valor-cop" data-cop="{{ total_compras }}">$ {{ '{:,.0f}'.format(total_compras).replace(',','.') }}</div>
    <div class="sl">Total en compras</div></div></div>
  <div class="col-md-3"><div class="sc"><div class="sv valor-cop" data-cop="{{ compras_mes }}">$ {{ '{:,.0f}'.format(compras_mes).replace(',','.') }}</div>
    <div class="sl">Compras este mes</div></div></div>
  <div class="col-md-3"><div class="sc"><div class="sv">{{ cotizaciones_vigentes }}</div>
    <div class="sl">Cotizaciones granel vigentes</div></div></div>
  <div class="col-md-3"><div class="sc"><div class="sv">{{ ordenes_activas }}</div>
    <div class="sl">Órdenes de compra activas</div></div></div>
</div>
<div class="row g-4">
  <div class="col-md-6"><div class="tc">
    <div class="ch d-flex justify-content-between align-items-center">
      <span><i class="bi bi-cart3 me-2 text-primary"></i>Compras recientes</span>
      <a href="{{ url_for('compras') }}" class="btn btn-sm btn-outline-primary">Ver todas</a></div>
    {% if compras_recientes %}<table class="table"><tbody>
    {% for c in compras_recientes %}<tr>
      <td><div class="fw-semibold" style="font-size:.88rem;color:#1a1f36">{{ c.nombre_item }}</div>
        <small class="text-muted">{{ c.proveedor or '—' }} · {{ c.fecha.strftime('%d/%m/%Y') }}</small></td>
      <td class="fw-semibold" style="white-space:nowrap">$ {{ '{:,.0f}'.format(c.costo_total).replace(',','.') }}</td>
    </tr>{% endfor %}</tbody></table>
    {% else %}<div class="text-center text-muted py-3"><p class="mb-2">Sin compras</p>
      <a href="{{ url_for('compra_nueva') }}" class="btn btn-sm btn-primary">Registrar</a></div>{% endif %}
  </div></div>
  <div class="col-md-6"><div class="tc">
    <div class="ch d-flex justify-content-between align-items-center">
      <span><i class="bi bi-building me-2 text-success"></i>Cotizaciones Granel</span>
      <a href="{{ url_for('granel') }}" class="btn btn-sm btn-outline-success">Ver todas</a></div>
    {% if granel_recientes %}<table class="table"><tbody>
    {% for g in granel_recientes %}<tr>
      <td><div class="fw-semibold" style="font-size:.88rem;color:#1a1f36">{{ g.nombre_producto }}</div>
        <small class="text-muted">SKU: {{ g.sku or '—' }} · NSO: {{ g.nso or '—' }}</small></td>
      <td><span class="b b-{{ g.estado }}">{{ g.estado.title() }}</span></td>
    </tr>{% endfor %}</tbody></table>
    {% else %}<div class="text-center text-muted py-3"><p class="mb-2">Sin cotizaciones</p>
      <a href="{{ url_for('granel_nuevo') }}" class="btn btn-sm btn-success">Agregar</a></div>{% endif %}
  </div></div>
</div>
<div class="row g-3 mt-2">
  <div class="col-md-4"><div class="fc d-flex align-items-center gap-3 py-3">
    <i class="bi bi-boxes fs-2 text-primary"></i>
    <div>
      <div class="fw-semibold">Materias Primas</div>
      <small class="text-muted">Stock, costos y proveedores</small>
    </div>
    <a href="{{ url_for('materias') }}" class="btn btn-sm btn-outline-primary ms-auto">Gestionar</a>
  </div></div>
  <div class="col-md-4"><div class="fc d-flex align-items-center gap-3 py-3">
    <i class="bi bi-diagram-3 fs-2 text-success"></i>
    <div>
      <div class="fw-semibold">Recetas / BOM</div>
      <small class="text-muted">Ingredientes por producto</small>
    </div>
    <a href="{{ url_for('recetas') }}" class="btn btn-sm btn-outline-success ms-auto">Ver recetas</a>
  </div></div>
  <div class="col-md-4"><div class="fc d-flex align-items-center gap-3 py-3">
    <i class="bi bi-bookmark-check fs-2 text-warning"></i>
    <div>
      <div class="fw-semibold">Reservas</div>
      <small class="text-muted">Materiales reservados para producción</small>
    </div>
    <a href="{{ url_for('reservas') }}" class="btn btn-sm btn-outline-warning ms-auto">Ver reservas</a>
  </div></div>
</div>
<div class="row g-3 mt-2">
  <div class="col-md-4"><div class="fc d-flex align-items-center gap-3 py-3">
    <i class="bi bi-list-check fs-2 text-info"></i>
    <div>
      <div class="fw-semibold">Órdenes de Producción</div>
      <small class="text-muted">Seguimiento de productos en proceso</small>
    </div>
    <a href="{{ url_for('ordenes_produccion') }}" class="btn btn-sm btn-outline-info ms-auto">Ver órdenes</a>
  </div></div>
  <div class="col-md-4"><div class="fc d-flex align-items-center gap-3 py-3">
    <i class="bi bi-bar-chart-steps fs-2 text-primary"></i>
    <div>
      <div class="fw-semibold">Diagrama Gantt</div>
      <small class="text-muted">Visualizar línea de tiempo de producción</small>
    </div>
    <a href="{{ url_for('gantt') }}" class="btn btn-sm btn-outline-primary ms-auto">Ver Gantt</a>
  </div></div>
</div>
{% endblock %}"""

T['produccion/compras.html'] = """{% extends 'base.html' %}
{% block title %}Compras{% endblock %}{% block page_title %}Compras de Materia Prima{% endblock %}
{% block topbar_actions %}
<a href="{{ url_for('compra_nueva') }}" class="btn btn-primary btn-sm"><i class="bi bi-plus-lg me-1"></i>Nueva compra</a>
<a href="{{ url_for('produccion_index') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Producción</a>
{% endblock %}
{% block content %}
<div class="row g-3 mb-3">
  <div class="col-md-3"><div class="sc"><div class="sv">$ {{ '{:,.0f}'.format(total_general).replace(',','.') }}</div><div class="sl">Total acumulado</div></div></div>
  <div class="col-md-3"><div class="sc"><div class="sv">$ {{ '{:,.0f}'.format(total_mes).replace(',','.') }}</div><div class="sl">Este mes</div></div></div>
</div>
<div class="tc"><div class="ch"><i class="bi bi-cart3 me-2"></i>{{ items|length }} compra(s)</div>
{% if items %}<div class="table-responsive"><table class="table">
  <thead><tr><th>Ítem</th><th>Proveedor</th><th>Fecha</th><th># Factura</th><th>Cto. Producto</th><th>Impuestos</th><th>Transporte</th><th>Total</th><th>Cant.</th><th>P. Unitario</th><th>Producto vinculado</th><th></th></tr></thead>
  <tbody>{% for c in items %}<tr>
    <td class="fw-semibold" style="color:#1a1f36">{{ c.nombre_item }}</td>
    <td>{{ c.proveedor or '—' }}</td>
    <td><small>{{ c.fecha.strftime('%d/%m/%Y') }}</small></td>
    <td><small class="text-muted">{{ c.nro_factura or '—' }}</small></td>
    <td>$ {{ '{:,.0f}'.format(c.costo_producto).replace(',','.') }}</td>
    <td>$ {{ '{:,.0f}'.format(c.impuestos).replace(',','.') }}</td>
    <td>$ {{ '{:,.0f}'.format(c.transporte).replace(',','.') }}</td>
    <td class="fw-semibold">$ {{ '{:,.0f}'.format(c.costo_total).replace(',','.') }}</td>
    <td>{{ c.cantidad }}</td>
    <td class="fw-semibold">$ {{ '{:,.0f}'.format(c.precio_unitario).replace(',','.') }}</td>
    <td>{% if c.producto %}<span class="badge bg-light text-dark">{{ c.producto.nombre }}</span>{% else %}—{% endif %}</td>
    <td><div class="d-flex gap-1">
      <a href="{{ url_for('compra_editar', id=c.id) }}" class="btn btn-sm btn-outline-secondary"><i class="bi bi-pencil"></i></a>
      <form method="POST" action="{{ url_for('compra_eliminar', id=c.id) }}" onsubmit="return confirm('¿Eliminar?')">
        <button class="btn btn-sm btn-outline-danger"><i class="bi bi-trash"></i></button></form>
    </div></td>
  </tr>{% endfor %}</tbody>
</table></div>
{% else %}<div class="text-center text-muted py-5"><i class="bi bi-cart3" style="font-size:3rem"></i>
  <p class="mt-3">Sin compras.</p><a href="{{ url_for('compra_nueva') }}" class="btn btn-primary">Registrar primera</a></div>
{% endif %}</div>{% endblock %}"""

T['produccion/compra_form.html'] = """{% extends 'base.html' %}
{% block title %}{{ titulo }}{% endblock %}{% block page_title %}{{ titulo }}{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('compras') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Volver</a>{% endblock %}
{% block content %}<div class="fc" style="max-width:940px"><form method="POST" id="frmC">
<div class="row g-3">
  <!-- Nombre auto-titulo -->
  <div class="col-12"><label class="form-label fw-semibold">Nombre del ítem <span class="text-muted fw-normal" style="font-size:.8rem">(se llena automáticamente)</span></label>
    <input type="text" name="nombre_item" id="nombreItem" class="form-control" value="{{ obj.nombre_item if obj else '' }}" required placeholder="Se genera automáticamente..."></div>

  <!-- Tipo de compra -->
  <div class="col-md-4"><label class="form-label">Tipo de compra *</label>
    <select name="tipo_compra" id="tipoCompra" class="form-select" onchange="onTipoChange()">
      <option value="materia_prima" {% if obj and obj.tipo_compra=='materia_prima' %}selected{% endif %}>🧪 Materia prima para producción</option>
      <option value="producto_terminado" {% if obj and obj.tipo_compra=='producto_terminado' %}selected{% endif %}>📦 Producto terminado</option>
      <option value="insumo" {% if (not obj) or (obj and obj.tipo_compra=='insumo') %}selected{% endif %}>🔧 Insumo / suministro</option>
      <option value="otro" {% if obj and obj.tipo_compra=='otro' %}selected{% endif %}>📋 Otro</option>
    </select></div>

  <!-- Vincular materia prima (solo si tipo = materia_prima) -->
  <div class="col-md-8" id="divMateria" style="display:none">
    <label class="form-label">Materia prima vinculada <small class="text-muted">(actualiza stock al guardar)</small></label>
    <select name="materia_id" id="materiaId" class="form-select" onchange="onMateriaChange()">
      <option value="">— Sin vincular —</option>
      {% for m in materias %}<option value="{{ m.id }}" data-unidad="{{ m.unidad }}" data-prodid="{{ m.producto_id or '' }}" {% if obj and obj.materia_id==m.id %}selected{% endif %}>{{ m.nombre }} ({{ m.unidad }}){% if m.producto %} → {{ m.producto.nombre }}{% endif %}</option>{% endfor %}
    </select></div>

  <!-- Vincular producto inventario -->
  <div class="col-md-8" id="divProducto" style="">
    <label class="form-label" id="lblProducto">Vincular a producto en inventario <small class="text-muted" id="lblProductoHint">(actualiza costo)</small></label>
    <select name="producto_id" id="productoId" class="form-select">
      <option value="">— No vincular —</option>
      {% for p in productos %}<option value="{{ p.id }}" {% if obj and obj.producto_id==p.id %}selected{% endif %}>{{ p.nombre }}{% if p.sku %} ({{ p.sku }}){% endif %}</option>{% endfor %}
    </select></div>

  <div class="col-md-6"><label class="form-label">Proveedor</label>
    <select name="proveedor_id" id="proveedorSel" class="form-select" onchange="autoNombreFromSel()">
      <option value="">— Sin proveedor —</option>
      {% for pv in proveedores_list %}<option value="{{ pv.id }}" {% if obj and obj.proveedor_id==pv.id %}selected{% endif %}>{{ pv.empresa or pv.nombre }}</option>{% endfor %}
    </select>
    <input type="hidden" name="proveedor" id="proveedor" value="{{ obj.proveedor if obj else '' }}"></div>
  <div class="col-md-3"><label class="form-label">Fecha *</label>
    <input type="date" name="fecha" id="fechaC" class="form-control" value="{{ obj.fecha.strftime('%Y-%m-%d') if obj else today }}" required oninput="autoNombre()"></div>
  <div class="col-md-3"><label class="form-label"># Factura</label>
    <input type="text" name="nro_factura" class="form-control" value="{{ obj.nro_factura if obj else '' }}"></div>

  <!-- Cantidad + Unidad -->
  <div class="col-md-4"><label class="form-label">Cantidad</label>
    <input type="number" name="cantidad" id="cant" class="form-control" step="0.001" min="0.001" value="{{ obj.cantidad if obj else '1' }}" oninput="calc()"></div>
  <div class="col-md-4"><label class="form-label">Unidad de medida</label>
    <select name="unidad" id="unidadSel" class="form-select">
      {% for u in ['unidades','kg','g','litros','ml','libras','galones','metros','cm','piezas','cajas','bolsas'] %}
      <option value="{{ u }}" {% if obj and obj.unidad==u %}selected{% elif not obj and u=='unidades' %}selected{% endif %}>{{ u }}</option>
      {% endfor %}
    </select></div>

  <!-- Caducidad -->
  <div class="col-md-4">
    <label class="form-label">¿Tiene caducidad?</label>
    <div class="form-check mt-2">
      <input class="form-check-input" type="checkbox" name="tiene_caducidad" id="chkCad" value="1"
        {% if obj and obj.tiene_caducidad %}checked{% endif %} onchange="toggleCad()">
      <label class="form-check-label" for="chkCad">Sí, tiene fecha de caducidad</label>
    </div>
  </div>
  <div class="col-md-4" id="divCad" style="display:{{ 'block' if obj and obj.tiene_caducidad else 'none' }}">
    <label class="form-label">Fecha de caducidad</label>
    <input type="date" name="fecha_caducidad" class="form-control" value="{{ obj.fecha_caducidad.strftime('%Y-%m-%d') if obj and obj.fecha_caducidad else '' }}"></div>

  <!-- Costos -->
  <div class="col-md-4"><label class="form-label">Costo del producto COP</label>
    <div class="input-group"><span class="input-group-text">$</span>
      <input type="number" name="costo_producto" id="cP" class="form-control" step="1" min="0" value="{{ obj.costo_producto|int if obj else '0' }}" oninput="calc()"></div></div>
  <div class="col-md-4"><label class="form-label">Impuestos COP</label>
    <div class="input-group"><span class="input-group-text">$</span>
      <input type="number" name="impuestos" id="imp" class="form-control" step="1" min="0" value="{{ obj.impuestos|int if obj else '0' }}" oninput="calc()"></div></div>
  <div class="col-md-4"><label class="form-label">Transporte COP</label>
    <div class="input-group"><span class="input-group-text">$</span>
      <input type="number" name="transporte" id="tsp" class="form-control" step="1" min="0" value="{{ obj.transporte|int if obj else '0' }}" oninput="calc()"></div></div>
  <div class="col-md-6"><label class="form-label fw-bold">Costo total COP</label>
    <input type="text" id="totVis" class="form-control fw-bold" readonly style="background:#f0f2ff;font-size:1.05rem">
    <input type="hidden" name="costo_total" id="totHid"></div>
  <div class="col-md-6"><label class="form-label fw-bold">Precio unitario COP</label>
    <input type="text" id="pUVis" class="form-control fw-bold" readonly style="background:#f0f2ff;font-size:1.05rem">
    <input type="hidden" name="precio_unitario" id="pUHid"></div>

  <div class="col-12"><label class="form-label">Notas</label>
    <textarea name="notas" class="form-control" rows="2">{{ obj.notas if obj else '' }}</textarea></div>
</div>
<div class="d-flex gap-2 mt-4">
  <button type="submit" class="btn btn-primary"><i class="bi bi-check-lg me-1"></i>{{ 'Actualizar' if obj else 'Guardar Compra' }}</button>
  <a href="{{ url_for('compras') }}" class="btn btn-outline-secondary">Cancelar</a>
</div></form></div>
{% block scripts %}<script>
var _tipoLabels = {materia_prima:'Materia prima',producto_terminado:'Producto terminado',insumo:'Insumo',otro:'Otro'};
function fCOP(n){return '$ '+Math.round(n).toLocaleString('es-CO');}
function calc(){
  var cp=parseFloat(document.getElementById('cP').value)||0;
  var imp=parseFloat(document.getElementById('imp').value)||0;
  var tsp=parseFloat(document.getElementById('tsp').value)||0;
  var cant=parseFloat(document.getElementById('cant').value)||1;
  var tot=cp+imp+tsp; var pu=cant>0?tot/cant:0;
  document.getElementById('totVis').value=fCOP(tot); document.getElementById('totHid').value=Math.round(tot);
  document.getElementById('pUVis').value=fCOP(pu); document.getElementById('pUHid').value=Math.round(pu);
}
function onTipoChange(){
  var t=document.getElementById('tipoCompra').value;
  var isMP=(t==='materia_prima');
  document.getElementById('divMateria').style.display=isMP?'':'none';
  // Producto siempre visible; ajustar etiqueta según contexto
  document.getElementById('divProducto').style.display='';
  document.getElementById('lblProductoHint').textContent=isMP?'(auto-detectado desde materia prima)':'(actualiza costo)';
  autoNombre();
}
function onMateriaChange(){
  var sel=document.getElementById('materiaId');
  var opt=sel.options[sel.selectedIndex];
  // Sync unidad
  var unidad=opt.getAttribute('data-unidad');
  if(unidad){
    var uSel=document.getElementById('unidadSel');
    for(var i=0;i<uSel.options.length;i++){if(uSel.options[i].value===unidad){uSel.selectedIndex=i;break;}}
  }
  // Auto-seleccionar el producto vinculado a la materia prima
  var prodId=opt.getAttribute('data-prodid');
  var pSel=document.getElementById('productoId');
  if(pSel){
    if(prodId){
      for(var j=0;j<pSel.options.length;j++){if(pSel.options[j].value===prodId){pSel.selectedIndex=j;break;}}
    } else {
      pSel.selectedIndex=0; // Sin vincular
    }
  }
  autoNombre();
}
function autoNombre(){
  var tipo=document.getElementById('tipoCompra');
  var tipoLbl=_tipoLabels[tipo.value]||tipo.value;
  var partes=[tipoLbl];
  if(tipo.value==='materia_prima'){
    var mSel=document.getElementById('materiaId');
    if(mSel.selectedIndex>0) partes.push(mSel.options[mSel.selectedIndex].text.split(' (')[0]);
  }
  var sel=document.getElementById('proveedorSel');
  var prov=sel.options[sel.selectedIndex]?sel.options[sel.selectedIndex].text.trim():'';
  if(prov) partes.push(prov);
  var fecha=document.getElementById('fechaC').value;
  if(fecha){var d=new Date(fecha+'T00:00:00');partes.push(d.toLocaleDateString('es-CO',{day:'2-digit',month:'2-digit',year:'numeric'}));}
  var nom=document.getElementById('nombreItem');
  if(!nom._editado) nom.value=partes.join(', ');
}
function autoNombreFromSel(){
  var sel=document.getElementById('proveedorSel');
  var txt=sel.options[sel.selectedIndex]?sel.options[sel.selectedIndex].text.trim():'';
  document.getElementById('proveedor').value=txt;
  autoNombre();
}
document.getElementById('nombreItem').addEventListener('input',function(){this._editado=true;});
function toggleCad(){
  document.getElementById('divCad').style.display=document.getElementById('chkCad').checked?'block':'none';
}
// Init
onTipoChange(); calc();
</script>{% endblock %}{% endblock %}"""

T['produccion/granel.html'] = """{% extends 'base.html' %}
{% block title %}Granel{% endblock %}{% block page_title %}Cotizaciones Granel{% endblock %}
{% block topbar_actions %}
<a href="{{ url_for('granel_nuevo') }}" class="btn btn-primary btn-sm"><i class="bi bi-plus-lg me-1"></i>Nueva cotización</a>
<a href="{{ url_for('produccion_index') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Producción</a>
{% endblock %}
{% block content %}
<div class="mb-3 d-flex gap-2">
  {% for est,lbl in [('','Todas'),('vigente','Vigentes'),('vencida','Vencidas'),('en_revision','En revisión')] %}
  <a href="{{ url_for('granel', estado=est) }}" class="btn btn-sm {{ 'btn-primary' if estado_f==est else 'btn-outline-secondary' }}">{{ lbl }}</a>{% endfor %}
</div>
<div class="tc"><div class="ch"><i class="bi bi-building me-2"></i>{{ items|length }} cotización(es)</div>
{% if items %}<div class="table-responsive"><table class="table">
  <thead><tr><th>Producto</th><th>SKU</th><th>NSO (INVIMA)</th><th>Proveedor</th><th>P. Unitario</th><th>Mín. unidades</th><th>Vigencia</th><th>Estado</th><th>Producto inv.</th><th></th></tr></thead>
  <tbody>{% for g in items %}<tr>
    <td class="fw-semibold" style="color:#1a1f36">{{ g.nombre_producto }}</td>
    <td><small class="text-muted">{{ g.sku or '—' }}</small></td>
    <td><small class="text-muted">{{ g.nso or '—' }}</small></td>
    <td>{{ g.proveedor or '—' }}</td>
    <td class="fw-semibold">$ {{ '{:,.0f}'.format(g.precio_unitario).replace(',','.') }}</td>
    <td>{{ g.unidades_minimas }}</td>
    <td>{% if g.vigencia %}<small class="{{ 'text-danger' if g.vigencia < now.date() else 'text-muted' }}">{{ g.vigencia.strftime('%d/%m/%Y') }}</small>{% else %}—{% endif %}</td>
    <td><span class="b b-{{ g.estado }}">{{ g.estado.replace('_',' ').title() }}</span></td>
    <td>{% if g.producto %}<small>{{ g.producto.nombre }}</small>{% else %}—{% endif %}</td>
    <td><div class="d-flex gap-1">
      <a href="{{ url_for('granel_editar', id=g.id) }}" class="btn btn-sm btn-outline-secondary"><i class="bi bi-pencil"></i></a>
      <form method="POST" action="{{ url_for('granel_eliminar', id=g.id) }}" onsubmit="return confirm('¿Eliminar?')">
        <button class="btn btn-sm btn-outline-danger"><i class="bi bi-trash"></i></button></form>
    </div></td>
  </tr>{% endfor %}</tbody>
</table></div>
{% else %}<div class="text-center text-muted py-5"><i class="bi bi-building" style="font-size:3rem"></i>
  <p class="mt-3">Sin cotizaciones.</p><a href="{{ url_for('granel_nuevo') }}" class="btn btn-primary">Agregar primera</a></div>
{% endif %}</div>{% endblock %}"""

T['produccion/granel_form.html'] = """{% extends 'base.html' %}
{% block title %}{{ titulo }}{% endblock %}{% block page_title %}{{ titulo }}{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('granel') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Volver</a>{% endblock %}
{% block content %}<div class="fc" style="max-width:900px"><form method="POST"><div class="row g-3">
  <div class="col-md-6"><label class="form-label">Nombre del producto *</label>
    <input type="text" name="nombre_producto" class="form-control" value="{{ obj.nombre_producto if obj else '' }}" required></div>
  <div class="col-md-3"><label class="form-label">SKU</label>
    <input type="text" name="sku" class="form-control" value="{{ obj.sku if obj else '' }}"></div>
  <div class="col-md-3"><label class="form-label">NSO (INVIMA)</label>
    <input type="text" name="nso" class="form-control" placeholder="NSO-XXXX" value="{{ obj.nso if obj else '' }}"></div>
  <div class="col-md-6"><label class="form-label">Proveedor / Fabricante</label>
    <input type="text" name="proveedor" class="form-control" value="{{ obj.proveedor if obj else '' }}"></div>
  <div class="col-md-3"><label class="form-label">Precio unitario COP</label>
    <div class="input-group"><span class="input-group-text">$</span>
      <input type="number" name="precio_unitario" class="form-control" step="1" min="0" value="{{ obj.precio_unitario|int if obj else '0' }}"></div></div>
  <div class="col-md-3"><label class="form-label">Mínimo de unidades</label>
    <input type="number" name="unidades_minimas" class="form-control" min="1" value="{{ obj.unidades_minimas if obj else '1' }}"></div>
  <div class="col-md-4"><label class="form-label">Fecha cotización</label>
    <input type="date" name="fecha_cotizacion" class="form-control"
      value="{{ obj.fecha_cotizacion.strftime('%Y-%m-%d') if obj and obj.fecha_cotizacion else today }}"></div>
  <div class="col-md-4"><label class="form-label">Vigencia hasta</label>
    <input type="date" name="vigencia" class="form-control"
      value="{{ obj.vigencia.strftime('%Y-%m-%d') if obj and obj.vigencia else '' }}"></div>
  <div class="col-md-4"><label class="form-label">Estado</label>
    <select name="estado" class="form-select">
      {% for est,lbl in [('vigente','Vigente'),('vencida','Vencida'),('en_revision','En revisión')] %}
      <option value="{{ est }}" {% if obj and obj.estado==est %}selected{% elif not obj and est=='vigente' %}selected{% endif %}>{{ lbl }}</option>{% endfor %}
    </select></div>
  <div class="col-12"><label class="form-label">Vincular a producto en inventario</label>
    <select name="producto_id" class="form-select">
      <option value="">— No vincular —</option>
      {% for p in productos %}<option value="{{ p.id }}" {% if obj and obj.producto_id==p.id %}selected{% endif %}>{{ p.nombre }}{% if p.sku %} ({{ p.sku }}){% endif %}</option>{% endfor %}
    </select>
    <div class="form-text">Si estado es "Vigente", actualizará el costo del producto en inventario.</div></div>
  <div class="col-12"><label class="form-label">Notas</label>
    <textarea name="notas" class="form-control" rows="2">{{ obj.notas if obj else '' }}</textarea></div>
</div>
<div class="d-flex gap-2 mt-4">
  <button type="submit" class="btn btn-primary"><i class="bi bi-check-lg me-1"></i>{{ 'Actualizar' if obj else 'Guardar Cotización' }}</button>
  <a href="{{ url_for('granel') }}" class="btn btn-outline-secondary">Cancelar</a>
</div></form></div>
{% block scripts %}<script>
setupAutoTitulo('[name="nombre_producto"]',['[name="proveedor"]','[name="estado"]','[name="fecha_cotizacion"]']);
</script>{% endblock %}{% endblock %}"""

T['finanzas/impuestos.html'] = """{% extends 'base.html' %}
{% block title %}Impuestos{% endblock %}{% block page_title %}Reglas Tributarias{% endblock %}
{% block topbar_actions %}
<a href="{{ url_for('impuesto_nuevo') }}" class="btn btn-primary btn-sm"><i class="bi bi-plus-lg me-1"></i>Nueva regla</a>
<a href="{{ url_for('contable_index') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Finanzas</a>
{% endblock %}
{% block content %}
<div class="alert alert-info" style="border-radius:10px;border:none">
  <i class="bi bi-info-circle me-2"></i>Configure aquí las reglas tributarias de la empresa. Estas reglas son usadas por el dashboard de contabilidad para calcular impuestos automáticamente sobre ingresos y utilidades.
</div>
<div class="tc"><div class="ch"><i class="bi bi-percent me-2"></i>{{ items|length }} regla(s)</div>
{% if items %}<div class="table-responsive"><table class="table">
  <thead><tr><th>Nombre</th><th>Descripción</th><th>%</th><th>Aplica a</th><th>Estado</th><th></th></tr></thead>
  <tbody>{% for r in items %}<tr>
    <td class="fw-semibold" style="color:#1a1f36">{{ r.nombre }}</td>
    <td><small class="text-muted">{{ r.descripcion or '—' }}</small></td>
    <td class="fw-semibold">{{ r.porcentaje }}%</td>
    <td><span class="badge bg-light text-dark border">
      {% if r.aplica_a=='ventas' %}Ventas (IVA)
      {% elif r.aplica_a=='ingresos' %}Ingresos generales
      {% elif r.aplica_a=='profit' %}Utilidad/Profit
      {% elif r.aplica_a=='proveedor_producto' %}Proveedor producto{% if r.proveedor_nombre %}: {{ r.proveedor_nombre }}{% endif %}
      {% elif r.aplica_a=='proveedor_granel' %}Proveedor granel{% if r.proveedor_nombre %}: {{ r.proveedor_nombre }}{% endif %}
      {% else %}{{ r.aplica_a or '—' }}{% endif %}
    </span></td>
    <td><span class="b b-{{ 'activo' if r.activo else 'inactivo' }}">{{ 'Activa' if r.activo else 'Inactiva' }}</span></td>
    <td><div class="d-flex gap-1">
      <a href="{{ url_for('impuesto_editar', id=r.id) }}" class="btn btn-sm btn-outline-secondary"><i class="bi bi-pencil"></i></a>
      <form method="POST" action="{{ url_for('impuesto_eliminar', id=r.id) }}" onsubmit="return confirm('¿Eliminar?')">
        <button class="btn btn-sm btn-outline-danger"><i class="bi bi-trash"></i></button></form>
    </div></td>
  </tr>{% endfor %}</tbody>
</table></div>
{% else %}<div class="text-center text-muted py-5"><i class="bi bi-percent" style="font-size:3rem"></i>
  <p class="mt-3">Sin reglas configuradas.</p><a href="{{ url_for('impuesto_nuevo') }}" class="btn btn-primary">Agregar primera</a></div>
{% endif %}</div>{% endblock %}"""

T['finanzas/impuesto_form.html'] = """{% extends 'base.html' %}
{% block title %}{{ titulo }}{% endblock %}{% block page_title %}{{ titulo }}{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('impuestos') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Volver</a>{% endblock %}
{% block content %}<div class="fc"><form method="POST"><div class="row g-3">
  <div class="col-md-6"><label class="form-label">Nombre *</label>
    <input type="text" name="nombre" class="form-control" placeholder="Ej: IVA compras insumos" value="{{ obj.nombre if obj else '' }}" required></div>
  <div class="col-md-3"><label class="form-label">Porcentaje %</label>
    <div class="input-group"><input type="number" name="porcentaje" class="form-control" step="0.01" min="0" max="100"
      value="{{ obj.porcentaje if obj else '0' }}"><span class="input-group-text">%</span></div></div>
  <div class="col-md-3"><label class="form-label">Estado</label>
    <select name="activo" class="form-select">
      <option value="1" {% if not obj or obj.activo %}selected{% endif %}>Activa</option>
      <option value="0" {% if obj and not obj.activo %}selected{% endif %}>Inactiva</option>
    </select></div>
  <div class="col-md-6"><label class="form-label">Aplica a *</label>
    <select name="aplica_a" id="aplicaA" class="form-select" onchange="toggleProveedor()" required>
      <option value="ventas" {% if not obj or obj.aplica_a=='ventas' %}selected{% endif %}>Ventas (IVA en ventas)</option>
      <option value="ingresos" {% if obj and obj.aplica_a=='ingresos' %}selected{% endif %}>Ingresos generales</option>
      <option value="profit" {% if obj and obj.aplica_a=='profit' %}selected{% endif %}>Sobre utilidad / profit</option>
      <option value="proveedor_producto" {% if obj and obj.aplica_a=='proveedor_producto' %}selected{% endif %}>Proveedor de productos</option>
      <option value="proveedor_granel" {% if obj and obj.aplica_a=='proveedor_granel' %}selected{% endif %}>Proveedor de granel</option>
    </select></div>
  <div class="col-md-6" id="divProveedor" style="display:{% if obj and 'proveedor' in (obj.aplica_a or '') %}block{% else %}none{% endif %}">
    <label class="form-label">Nombre del proveedor</label>
    <input type="text" name="proveedor_nombre" class="form-control" placeholder="Nombre del proveedor específico"
      value="{{ obj.proveedor_nombre if obj else '' }}"></div>
  <div class="col-12"><label class="form-label">Descripción / Instrucciones</label>
    <textarea name="descripcion" class="form-control" rows="3" placeholder="Explicar cuándo y cómo aplicar esta regla...">{{ obj.descripcion if obj else '' }}</textarea></div>
</div>
<div class="d-flex gap-2 mt-4">
  <button type="submit" class="btn btn-primary"><i class="bi bi-check-lg me-1"></i>{{ 'Actualizar' if obj else 'Guardar Regla' }}</button>
  <a href="{{ url_for('impuestos') }}" class="btn btn-outline-secondary">Cancelar</a>
</div></form></div>
{% block scripts %}<script>
function toggleProveedor(){
  var v=document.getElementById('aplicaA').value;
  document.getElementById('divProveedor').style.display=v.startsWith('proveedor')?'block':'none';
}
</script>{% endblock %}{% endblock %}"""

T['gastos/index.html'] = """{% extends 'base.html' %}
{% block title %}Gastos Operativos{% endblock %}{% block page_title %}Gastos Operativos{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('gasto_nuevo') }}" class="btn btn-primary btn-sm"><i class="bi bi-plus-lg me-1"></i>Agregar gasto</a>{% endblock %}
{% block content %}
<div class="row g-3 mb-4">
  <div class="col-md-3"><div class="sc"><div class="sv">$ {{ '{:,.0f}'.format(total_general).replace(',','.') }}</div><div class="sl">Total acumulado COP</div></div></div>
  <div class="col-md-3"><div class="sc"><div class="sv">$ {{ '{:,.0f}'.format(total_mes).replace(',','.') }}</div><div class="sl">Este mes</div></div></div>
  <div class="col-md-3"><div class="sc"><div class="sv">{{ total_registros }}</div><div class="sl">Registros totales</div></div></div>
</div>
{% if plantillas %}
<div class="tc mb-4">
  <div class="ch"><i class="bi bi-arrow-repeat me-2"></i>Gastos recurrentes mensuales — pendientes de registrar</div>
  <div class="p-3">
    <div class="row g-2">
      {% for p in plantillas %}
      <div class="col-md-4">
        <div class="border rounded p-3 d-flex align-items-center justify-content-between" style="background:#f8f9fe">
          <div>
            <div class="fw-semibold" style="color:#1a1f36">{{ p.tipo_custom if p.tipo=='Otro' and p.tipo_custom else p.tipo }}</div>
            <small class="text-muted">$ {{ '{:,.0f}'.format(p.monto).replace(',','.') }} · mensual</small>
          </div>
          <form method="POST" action="{{ url_for('gasto_plantilla_usar', id=p.id) }}">
            <button class="btn btn-sm btn-primary" title="Registrar gasto de este mes"><i class="bi bi-plus-lg"></i> Registrar</button>
          </form>
        </div>
      </div>{% endfor %}
    </div>
  </div>
</div>{% endif %}
<div class="tc mb-3"><div class="p-3"><form method="GET" class="row g-2 align-items-end">
  <div class="col-sm-3"><select name="tipo" class="form-select form-select-sm">
    <option value="">Todos los tipos</option>
    {% for t in tipos %}<option value="{{ t }}" {% if tipo_f==t %}selected{% endif %}>{{ t }}</option>{% endfor %}
  </select></div>
  <div class="col-sm-2"><input type="date" name="desde" class="form-control form-control-sm" value="{{ desde_f }}"></div>
  <div class="col-sm-2"><input type="date" name="hasta" class="form-control form-control-sm" value="{{ hasta_f }}"></div>
  <div class="col-auto">
    <button type="submit" class="btn btn-primary btn-sm"><i class="bi bi-search"></i></button>
    <a href="{{ url_for('gastos') }}" class="btn btn-outline-secondary btn-sm">Limpiar</a>
  </div>
</form></div></div>
<div class="tc"><div class="ch"><i class="bi bi-receipt me-2"></i>{{ items|length }} gasto(s)</div>
{% if items %}<div class="table-responsive"><table class="table">
  <thead><tr><th>Fecha</th><th>Tipo</th><th>Descripción</th><th>Monto COP</th><th>Recurrencia</th><th></th></tr></thead>
  <tbody>{% for g in items %}{% if not g.es_plantilla %}<tr>
    <td><small>{{ g.fecha.strftime('%d/%m/%Y') }}</small></td>
    <td>
      <span class="badge bg-secondary">{{ g.tipo_custom if g.tipo=='Otro' and g.tipo_custom else g.tipo }}</span>
    </td>
    <td>{{ g.descripcion or '—' }}</td>
    <td class="fw-semibold">$ {{ '{:,.0f}'.format(g.monto).replace(',','.') }}</td>
    <td><span class="b b-{{ g.recurrencia or 'unico' }}">{{ 'Mensual' if g.recurrencia=='mensual' else 'Único' }}</span></td>
    <td><div class="d-flex gap-1">
      <a href="{{ url_for('gasto_editar', id=g.id) }}" class="btn btn-sm btn-outline-secondary"><i class="bi bi-pencil"></i></a>
      <form method="POST" action="{{ url_for('gasto_eliminar', id=g.id) }}" onsubmit="return confirm('¿Eliminar?')">
        <button class="btn btn-sm btn-outline-danger"><i class="bi bi-trash"></i></button></form>
    </div></td>
  </tr>{% endif %}{% endfor %}</tbody>
</table></div>
{% else %}<div class="text-center text-muted py-5"><i class="bi bi-receipt" style="font-size:3rem"></i>
  <p class="mt-3">Sin gastos.</p><a href="{{ url_for('gasto_nuevo') }}" class="btn btn-primary">Agregar primero</a></div>
{% endif %}</div>{% endblock %}"""

T['gastos/form.html'] = """{% extends 'base.html' %}
{% block title %}{{ titulo }}{% endblock %}{% block page_title %}{{ titulo }}{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('gastos') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Volver</a>{% endblock %}
{% block content %}<div class="fc"><form method="POST"><div class="row g-3">
  <div class="col-md-4"><label class="form-label">Fecha *</label>
    <input type="date" name="fecha" class="form-control" value="{{ obj.fecha.strftime('%Y-%m-%d') if obj else today }}" required></div>
  <div class="col-md-4"><label class="form-label">Tipo *</label>
    <select name="tipo" id="tipoSel" class="form-select" onchange="chkTipo()" required>
      {% for t in ['Arriendo','Servicios públicos','Nómina','Transporte','Mercadeo','Materia prima','Maquinaria','Impuestos','Logística','Mantenimiento','Otro'] %}
      <option value="{{ t }}" {% if obj and obj.tipo==t %}selected{% endif %}>{{ t }}</option>{% endfor %}
    </select></div>
  <div class="col-md-4"><label class="form-label">Monto COP *</label>
    <div class="input-group"><span class="input-group-text">$</span>
      <input type="number" name="monto" class="form-control" step="1" min="0" value="{{ obj.monto|int if obj else '0' }}" required></div></div>
  <div class="col-md-6" id="divTipoCustom" style="display:{{ 'block' if obj and obj.tipo=='Otro' else 'none' }}">
    <label class="form-label">¿Cuál? (nombre del gasto)</label>
    <input type="text" name="tipo_custom" class="form-control" placeholder="Ej: Suscripción Canva, Gastos notariales..."
      value="{{ obj.tipo_custom if obj else '' }}"></div>
  <div class="col-md-6"><label class="form-label">Recurrencia</label>
    <div class="d-flex gap-3 mt-2">
      <div class="form-check"><input class="form-check-input" type="radio" name="recurrencia" value="unico" id="rUnico"
        {% if not obj or obj.recurrencia=='unico' %}checked{% endif %} onchange="chkRec()">
        <label class="form-check-label" for="rUnico">Único</label></div>
      <div class="form-check"><input class="form-check-input" type="radio" name="recurrencia" value="mensual" id="rMensual"
        {% if obj and obj.recurrencia=='mensual' %}checked{% endif %} onchange="chkRec()">
        <label class="form-check-label" for="rMensual">Mensual</label></div>
    </div></div>
  <div class="col-12" id="divPlantilla" style="display:{{ 'block' if obj and obj.recurrencia=='mensual' else 'none' }}">
    <div class="form-check">
      <input class="form-check-input" type="checkbox" name="es_plantilla" id="esPlantilla" value="1"
        {% if obj and obj.es_plantilla %}checked{% endif %}>
      <label class="form-check-label" for="esPlantilla">
        <strong>Guardar como plantilla mensual</strong> — aparecerá cada mes para registrar rápidamente
      </label>
    </div>
  </div>
  <div class="col-12"><label class="form-label">Descripción</label>
    <input type="text" name="descripcion" class="form-control" value="{{ obj.descripcion if obj else '' }}"></div>
  <div class="col-12"><label class="form-label">Notas</label>
    <textarea name="notas" class="form-control" rows="2">{{ obj.notas if obj else '' }}</textarea></div>
</div>
<div class="d-flex gap-2 mt-4">
  <button type="submit" class="btn btn-primary"><i class="bi bi-check-lg me-1"></i>{{ 'Actualizar' if obj else 'Guardar Gasto' }}</button>
  <a href="{{ url_for('gastos') }}" class="btn btn-outline-secondary">Cancelar</a>
</div></form></div>
{% block scripts %}<script>
function chkTipo(){document.getElementById('divTipoCustom').style.display=document.getElementById('tipoSel').value=='Otro'?'block':'none';}
function chkRec(){document.getElementById('divPlantilla').style.display=document.getElementById('rMensual').checked?'block':'none';}
</script>{% endblock %}{% endblock %}"""

T['admin/usuarios.html'] = """{% extends 'base.html' %}
{% block title %}Usuarios{% endblock %}{% block page_title %}Gestión de Usuarios{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('admin_usuario_nuevo') }}" class="btn btn-primary btn-sm"><i class="bi bi-plus-lg me-1"></i>Nuevo Usuario</a>{% endblock %}
{% block content %}<div class="tc"><div class="ch"><i class="bi bi-shield-person-fill me-2"></i>{{ items|length }} usuario(s)</div>
<table class="table"><thead><tr><th>Nombre</th><th>Email</th><th>Rol</th><th>Estado</th><th>Alta</th><th></th></tr></thead>
<tbody>{% for u in items %}<tr>
  <td><div class="d-flex align-items-center gap-2">
    <div class="rounded-circle d-flex align-items-center justify-content-center text-white fw-bold"
         style="width:30px;height:30px;background:#5e72e4;font-size:.8rem;flex-shrink:0">{{ u.nombre[0].upper() }}</div>
    <div class="fw-semibold" style="color:#1a1f36">{{ u.nombre }}</div></div></td>
  <td>{{ u.email }}</td>
  <td><span class="badge {{ 'bg-primary' if u.rol=='admin' else 'bg-secondary' }}">{{ u.rol.title() }}</span></td>
  <td><span class="b b-{{ 'activo' if u.activo else 'inactivo' }}">{{ 'Activo' if u.activo else 'Inactivo' }}</span></td>
  <td><small class="text-muted">{{ u.creado_en.strftime('%d/%m/%Y') }}</small></td>
  <td>
    <a href="{{ url_for('admin_usuario_editar', id=u.id) }}" class="btn btn-sm btn-outline-secondary me-1"><i class="bi bi-pencil"></i></a>
    {% if u.id != current_user.id %}
    <form method="POST" action="{{ url_for('admin_usuario_toggle', id=u.id) }}" class="d-inline">
      <button class="btn btn-sm {{ 'btn-outline-warning' if u.activo else 'btn-outline-success' }}">
        {{ 'Desactivar' if u.activo else 'Activar' }}</button></form>
    {% else %}<small class="text-muted">(tú)</small>{% endif %}</td>
</tr>{% endfor %}</tbody></table></div>{% endblock %}"""

T['admin/usuario_form.html'] = """{% extends 'base.html' %}
{% block title %}{{ titulo }}{% endblock %}{% block page_title %}{{ titulo }}{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('admin_usuarios') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Volver</a>{% endblock %}
{% block content %}<div class="fc"><form method="POST"><div class="row g-3">
  <div class="col-md-6"><label class="form-label">Nombre *</label>
    <input type="text" name="nombre" class="form-control" value="{{ obj.nombre if obj else '' }}" required></div>
  <div class="col-md-6"><label class="form-label">Email *</label>
    <input type="email" name="email" class="form-control" value="{{ obj.email if obj else '' }}" required></div>
  <div class="col-md-6"><label class="form-label">Contraseña {% if obj %}(dejar vacío para no cambiar){% else %}*{% endif %}</label>
    <input type="password" name="password" class="form-control" {% if not obj %}required{% endif %}></div>
  <div class="col-md-6"><label class="form-label">Rol base</label>
    <select name="rol" id="rolSel" class="form-select" onchange="applyRol()">
      <option value="usuario" {% if obj and obj.rol=='usuario' %}selected{% endif %}>Usuario básico</option>
      <option value="vendedor" {% if obj and obj.rol=='vendedor' %}selected{% endif %}>Vendedor</option>
      <option value="produccion" {% if obj and obj.rol=='produccion' %}selected{% endif %}>Producción</option>
      <option value="contador" {% if obj and obj.rol=='contador' %}selected{% endif %}>Contador</option>
      <option value="admin" {% if obj and obj.rol=='admin' %}selected{% endif %}>Administrador (acceso total)</option>
    </select></div>
  <div class="col-12" id="divModulos">
    <label class="form-label fw-semibold">Módulos permitidos <small class="text-muted fw-normal">(personalizar acceso)</small></label>
    <div class="row g-2 mt-1">
      {% set modulos_map = [('clientes','Clientes','people-fill'),('ventas','Ventas','graph-up-arrow'),('cotizaciones','Cotizaciones','file-earmark-text-fill'),('tareas','Tareas','check2-square'),('calendario','Calendario','calendar3'),('notas','Notas','sticky-fill'),('inventario','Inventario','box-seam-fill'),('produccion','Producción','gear-fill'),('gastos','Gastos','receipt'),('reportes','Reportes','bar-chart-fill')] %}
      {% for key,label,icon in modulos_map %}
      <div class="col-md-3 col-6">
        <div class="form-check p-2 border rounded" style="background:#f8f9fe">
          <input class="form-check-input mod-check" type="checkbox" name="modulos" value="{{ key }}" id="mod_{{ key }}"
            {% if obj and key in (obj.modulos_permitidos or '[]') %}checked
            {% elif not obj %}checked{% endif %}>
          <label class="form-check-label" for="mod_{{ key }}" style="font-size:.85rem">
            <i class="bi bi-{{ icon }} me-1 text-primary"></i>{{ label }}
          </label>
        </div>
      </div>{% endfor %}
    </div>
    <small class="text-muted">El rol "Administrador" ignora esta lista y tiene acceso a todo.</small>
  </div>
</div>
<div class="d-flex gap-2 mt-4">
  <button type="submit" class="btn btn-primary"><i class="bi bi-check-lg me-1"></i>{{ 'Actualizar' if obj else 'Crear Usuario' }}</button>
  <a href="{{ url_for('admin_usuarios') }}" class="btn btn-outline-secondary">Cancelar</a>
</div></form></div>
{% block scripts %}<script>
var ROL_MODULOS = {
  usuario:    ['tareas','notas','calendario'],
  vendedor:   ['clientes','ventas','cotizaciones','tareas','calendario','notas'],
  produccion: ['inventario','produccion','gastos','notas','calendario','tareas'],
  contador:   ['gastos','reportes','produccion','notas'],
  admin:      ['clientes','ventas','cotizaciones','tareas','calendario','notas','inventario','produccion','gastos','reportes']
};
function applyRol(){
  var rol=document.getElementById('rolSel').value;
  var mods=ROL_MODULOS[rol]||[];
  document.querySelectorAll('.mod-check').forEach(function(c){c.checked=mods.includes(c.value);});
  document.getElementById('divModulos').style.opacity=rol==='admin'?'0.5':'1';
}
</script>{% endblock %}{% endblock %}"""

T['perfil.html'] = """{% extends 'base.html' %}
{% block title %}Mi Perfil{% endblock %}{% block page_title %}Mi Perfil{% endblock %}
{% block content %}
<div class="row g-4">
  <div class="col-md-6">
    <div class="fc">
      <h5 class="mb-3"><i class="bi bi-person-circle me-2 text-primary"></i>Datos personales</h5>
      <form method="POST" action="{{ url_for('perfil') }}">
        <input type="hidden" name="accion" value="datos">
        <div class="row g-3">
          <div class="col-12"><label class="form-label">Nombre *</label>
            <input type="text" name="nombre" class="form-control" value="{{ current_user.nombre }}" required></div>
          <div class="col-12"><label class="form-label">Email *</label>
            <input type="email" name="email" class="form-control" value="{{ current_user.email }}" required></div>
        </div>
        <div class="mt-3">
          <button type="submit" class="btn btn-primary"><i class="bi bi-check-lg me-1"></i>Guardar cambios</button>
        </div>
      </form>
    </div>
  </div>
  <div class="col-md-6">
    <div class="fc">
      <h5 class="mb-3"><i class="bi bi-lock-fill me-2 text-warning"></i>Cambiar contraseña</h5>
      <form method="POST" action="{{ url_for('perfil') }}">
        <input type="hidden" name="accion" value="password">
        <div class="row g-3">
          <div class="col-12"><label class="form-label">Contraseña actual *</label>
            <input type="password" name="password_actual" class="form-control" required></div>
          <div class="col-12"><label class="form-label">Nueva contraseña *</label>
            <input type="password" name="password_nueva" class="form-control" minlength="6" required></div>
          <div class="col-12"><label class="form-label">Confirmar nueva contraseña *</label>
            <input type="password" name="password_confirmar" class="form-control" minlength="6" required></div>
        </div>
        <div class="mt-3">
          <button type="submit" class="btn btn-warning"><i class="bi bi-key me-1"></i>Cambiar contraseña</button>
        </div>
      </form>
    </div>
  </div>
</div>{% endblock %}"""

T['reportes.html'] = """{% extends 'base.html' %}
{% block title %}Reportes{% endblock %}{% block page_title %}Reportes{% endblock %}
{% block topbar_actions %}
<div class="dropdown">
  <button class="btn btn-success btn-sm dropdown-toggle" data-bs-toggle="dropdown"><i class="bi bi-file-earmark-excel me-1"></i>Exportar Excel</button>
  <ul class="dropdown-menu">
    <li><a class="dropdown-item" href="{{ url_for('exportar_ventas') }}"><i class="bi bi-graph-up-arrow me-2 text-primary"></i>Ventas</a></li>
    <li><a class="dropdown-item" href="{{ url_for('exportar_gastos') }}"><i class="bi bi-receipt me-2 text-danger"></i>Gastos Operativos</a></li>
    <li><a class="dropdown-item" href="{{ url_for('exportar_inventario') }}"><i class="bi bi-box-seam me-2 text-success"></i>Inventario</a></li>
    <li><a class="dropdown-item" href="{{ url_for('exportar_clientes') }}"><i class="bi bi-people me-2 text-info"></i>Clientes</a></li>
  </ul>
</div>{% endblock %}
{% block content %}
<div class="row g-3 mb-4">
  <div class="col-md-3"><div class="sc">
    <div class="sv" style="color:#5e72e4">{{ total_clientes }}</div>
    <div class="sl">Clientes activos</div></div></div>
  <div class="col-md-3"><div class="sc">
    <div class="sv valor-cop" data-cop="{{ ingresos_totales }}">$ {{ '{:,.0f}'.format(ingresos_totales).replace(',','.') }}</div>
    <div class="sl">Ingresos totales (ventas ganadas)</div></div></div>
  <div class="col-md-3"><div class="sc">
    <div class="sv valor-cop" data-cop="{{ gastos_totales }}">$ {{ '{:,.0f}'.format(gastos_totales).replace(',','.') }}</div>
    <div class="sl">Gastos operativos totales</div></div></div>
  <div class="col-md-3"><div class="sc">
    <div class="sv valor-cop" data-cop="{{ balance }}" style="color:{{ '#2dce89' if balance>=0 else '#f5365c' }}">$ {{ '{:,.0f}'.format(balance).replace(',','.') }}</div>
    <div class="sl">Balance</div></div></div>
</div>
<div class="row g-4 mb-4">
  <div class="col-lg-7">
    <div class="tc"><div class="ch"><i class="bi bi-bar-chart me-2"></i>Ventas por mes (últimos 6 meses)</div>
      <canvas id="chartVentas" height="100"></canvas>
    </div>
  </div>
  <div class="col-lg-5">
    <div class="tc"><div class="ch"><i class="bi bi-pie-chart me-2"></i>Gastos por tipo</div>
      <canvas id="chartGastos" height="160"></canvas>
    </div>
  </div>
</div>
<div class="row g-4">
  <div class="col-md-6">
    <div class="tc"><div class="ch d-flex justify-content-between align-items-center">
      <span><i class="bi bi-trophy me-2 text-warning"></i>Top 5 clientes por ventas</span></div>
      <table class="table"><tbody>
      {% for c in top_clientes %}
      <tr><td><a href="{{ url_for('cliente_ver', id=c.id) }}" class="fw-semibold text-decoration-none" style="color:#1a1f36">{{ c.empresa or c.nombre }}</a></td>
        <td class="text-end fw-semibold">$ {{ '{:,.0f}'.format(c.total_ventas or 0).replace(',','.') }}</td></tr>
      {% else %}<tr><td colspan="2" class="text-muted text-center">Sin datos</td></tr>{% endfor %}
      </tbody></table>
    </div>
  </div>
  <div class="col-md-6">
    <div class="tc"><div class="ch"><i class="bi bi-exclamation-triangle me-2 text-warning"></i>Stock bajo</div>
      {% if bajo_stock %}
      <table class="table"><thead><tr><th>Producto</th><th>Stock</th><th>Mínimo</th></tr></thead><tbody>
      {% for p in bajo_stock %}
      <tr><td>{{ p.nombre }}</td>
        <td class="fw-bold text-danger">{{ p.stock }}</td>
        <td class="text-muted">{{ p.stock_minimo }}</td></tr>
      {% endfor %}</tbody></table>
      {% else %}<div class="text-center text-muted py-3"><i class="bi bi-check-circle text-success" style="font-size:2rem"></i>
        <p class="mt-2">Todo el inventario está OK</p></div>{% endif %}
    </div>
  </div>
</div>{% endblock %}
{% block scripts %}
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.0/chart.umd.min.js"></script>
<script>
const meses = {{ meses_labels|tojson }};
const ventasMes = {{ ventas_por_mes|tojson }};
const gastosTipos = {{ gastos_tipos_labels|tojson }};
const gastosTotales = {{ gastos_tipos_values|tojson }};
new Chart(document.getElementById('chartVentas'),{type:'bar',data:{labels:meses,datasets:[{
  label:'Ventas COP',data:ventasMes,backgroundColor:'rgba(94,114,228,0.7)',borderRadius:6}]},
  options:{responsive:true,plugins:{legend:{display:false}},scales:{y:{ticks:{callback:v=>'$'+v.toLocaleString('es-CO')}}}}});
const colores=['#5e72e4','#2dce89','#fb6340','#11cdef','#f4f5f7','#adb5bd'];
new Chart(document.getElementById('chartGastos'),{type:'doughnut',data:{labels:gastosTipos,
  datasets:[{data:gastosTotales,backgroundColor:colores.slice(0,gastosTipos.length),borderWidth:2}]},
  options:{responsive:true,plugins:{legend:{position:'bottom'}}}});
</script>{% endblock %}"""

T['notas/index.html'] = """{% extends 'base.html' %}
{% block title %}Notas{% endblock %}{% block page_title %}Notas Rápidas{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('nota_nueva') }}" class="btn btn-primary btn-sm"><i class="bi bi-plus-lg me-1"></i>Nueva nota</a>{% endblock %}
{% block content %}
<div class="row g-3 mb-3">
  <div class="col-md-4">
    <form method="GET"><select name="cliente_id" class="form-select form-select-sm" onchange="this.form.submit()">
      <option value="">Todas las notas</option>
      {% for c in clientes_list %}<option value="{{ c.id }}" {% if c.id|string == cliente_f %}selected{% endif %}>{{ c.empresa or c.nombre }}</option>{% endfor %}
    </select></form>
  </div>
</div>
{% if items %}
<div class="row g-3">
{% for n in items %}
<div class="col-md-6 col-lg-4">
  <div class="tc h-100" style="position:relative">
    <div class="d-flex justify-content-between align-items-start mb-2">
      <div class="fw-semibold" style="color:#1a1f36;font-size:.95rem">{{ n.titulo or '(Sin título)' }}</div>
      <div class="d-flex gap-1">
        <a href="{{ url_for('nota_editar', id=n.id) }}" class="btn btn-sm btn-outline-secondary py-0 px-1"><i class="bi bi-pencil" style="font-size:.75rem"></i></a>
        <form method="POST" action="{{ url_for('nota_eliminar', id=n.id) }}" onsubmit="return confirm('¿Eliminar nota?')">
          <button class="btn btn-sm btn-outline-danger py-0 px-1"><i class="bi bi-trash" style="font-size:.75rem"></i></button></form>
      </div>
    </div>
    <p style="font-size:.88rem;color:#525f7f;white-space:pre-line;margin-bottom:.75rem">{{ n.contenido[:300] }}{% if n.contenido|length > 300 %}…{% endif %}</p>
    <div style="font-size:.75rem;color:#adb5bd;border-top:1px solid #f0f2ff;padding-top:.5rem;margin-top:auto">
      {% if n.cliente %}<span class="badge bg-light text-dark me-1"><i class="bi bi-person me-1"></i>{{ n.cliente.empresa or n.cliente.nombre }}</span>{% endif %}
      {% if n.producto %}<span class="badge bg-light text-dark me-1"><i class="bi bi-box me-1"></i>{{ n.producto.nombre }}</span>{% endif %}
      {% if n.modulo %}<span class="badge" style="background:#e8eeff;color:#5e72e4">{{ n.modulo }}</span>{% endif %}
      {% if n.fecha_revision %}<span class="badge" style="background:#f3e8ff;color:#8965e0"><i class="bi bi-calendar-check me-1"></i>Rev. {{ n.fecha_revision.strftime('%d/%m/%Y') }}</span>{% endif %}
      <br>{{ n.autor.nombre if n.autor else '' }} · {{ n.creado_en.strftime('%d/%m/%Y') }}
    </div>
  </div>
</div>{% endfor %}
</div>
{% else %}<div class="text-center text-muted py-5"><i class="bi bi-sticky" style="font-size:3rem"></i>
  <p class="mt-3">Sin notas.</p><a href="{{ url_for('nota_nueva') }}" class="btn btn-primary">Crear primera</a></div>{% endif %}
{% endblock %}"""

T['notas/form.html'] = """{% extends 'base.html' %}
{% block title %}{{ titulo }}{% endblock %}{% block page_title %}{{ titulo }}{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('notas') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Volver</a>{% endblock %}
{% block content %}<div class="fc"><form method="POST"><div class="row g-3">
  <div class="col-md-8"><label class="form-label">Título</label>
    <input type="text" name="titulo" class="form-control" placeholder="Título opcional" value="{{ obj.titulo if obj else '' }}"></div>
  <div class="col-md-4"><label class="form-label">Fecha de revisión</label>
    <input type="date" name="fecha_revision" class="form-control"
      value="{{ obj.fecha_revision.strftime('%Y-%m-%d') if obj and obj.fecha_revision else '' }}">
    <small class="text-muted">Aparecerá en el calendario</small></div>
  <div class="col-12"><label class="form-label fw-bold">Vincular a (opcional)</label></div>
  <div class="col-md-4"><label class="form-label">Cliente</label>
    <select name="cliente_id" class="form-select">
      <option value="">— Ninguno —</option>
      {% for c in clientes_list %}<option value="{{ c.id }}" {% if obj and obj.cliente_id==c.id %}selected{% endif %}>{{ c.empresa or c.nombre }}</option>{% endfor %}
    </select></div>
  <div class="col-md-4"><label class="form-label">Producto</label>
    <select name="producto_id" class="form-select">
      <option value="">— Ninguno —</option>
      {% for p in productos_list %}<option value="{{ p.id }}" {% if obj and obj.producto_id==p.id %}selected{% endif %}>{{ p.nombre }}</option>{% endfor %}
    </select></div>
  <div class="col-md-4"><label class="form-label">Módulo del sistema</label>
    <select name="modulo" class="form-select">
      <option value="">— Ninguno —</option>
      {% for m in ['Ventas','Cotizaciones','Producción','Inventario','Gastos','Tareas','Compras','Granel','Otro'] %}
      <option value="{{ m }}" {% if obj and obj.modulo==m %}selected{% endif %}>{{ m }}</option>{% endfor %}
    </select></div>
  <div class="col-12"><label class="form-label">Contenido *</label>
    <textarea name="contenido" class="form-control" rows="10" style="font-size:.9rem" required>{{ obj.contenido if obj else '' }}</textarea></div>
</div>
<div class="d-flex gap-2 mt-4">
  <button type="submit" class="btn btn-primary"><i class="bi bi-check-lg me-1"></i>{{ 'Actualizar' if obj else 'Guardar nota' }}</button>
  <a href="{{ url_for('notas') }}" class="btn btn-outline-secondary">Cancelar</a>
</div></form></div>
{% block scripts %}<script>
setupAutoTitulo('[name="titulo"]',['[name="modulo"]','[name="cliente_id"]','[name="producto_id"]','[name="fecha_revision"]']);
</script>{% endblock %}{% endblock %}"""

T['calendario.html'] = """{% extends 'base.html' %}
{% block title %}Calendario{% endblock %}{% block page_title %}Calendario{% endblock %}
{% block topbar_actions %}
<a href="{{ url_for('calendario') }}?mes={{ prev_mes }}&anio={{ prev_anio }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-chevron-left"></i></a>
<span class="fw-semibold mx-2" style="min-width:160px;display:inline-block;text-align:center">{{ mes_nombre }} {{ anio }}</span>
<a href="{{ url_for('calendario') }}?mes={{ next_mes }}&anio={{ next_anio }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-chevron-right"></i></a>
<a href="{{ url_for('calendario') }}" class="btn btn-outline-secondary btn-sm ms-1">Hoy</a>
<button class="btn btn-primary btn-sm ms-2" data-bs-toggle="modal" data-bs-target="#modalEvento"><i class="bi bi-plus-lg me-1"></i>Nuevo evento</button>
{% endblock %}
{% block content %}
<div class="row g-2 mb-3">
  <div class="col-auto"><span class="cal-ev ev-tarea px-2 py-1 rounded" style="font-size:.75rem;display:inline-block">● Tarea</span></div>
  <div class="col-auto"><span class="cal-ev ev-venta px-2 py-1 rounded" style="font-size:.75rem;display:inline-block">● Venta / Entrega</span></div>
  <div class="col-auto"><span class="cal-ev ev-evento px-2 py-1 rounded" style="font-size:.75rem;display:inline-block">● Evento / Cita</span></div>
  <div class="col-auto"><span class="cal-ev ev-nota px-2 py-1 rounded" style="font-size:.75rem;display:inline-block">● Revisión nota</span></div>
  <div class="col-auto"><span class="cal-ev ev-caducidad px-2 py-1 rounded" style="font-size:.75rem;display:inline-block">● Caducidad producto</span></div>
</div>
<div style="background:#fff;border-radius:16px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,.06)">
<table style="width:100%;border-collapse:collapse;table-layout:fixed">
<thead><tr>
  {% for dia in ['Lun','Mar','Mié','Jue','Vie','Sáb','Dom'] %}
  <th style="padding:.6rem;text-align:center;font-size:.8rem;color:#8898aa;background:#f8f9fe;font-weight:600;border:1px solid #f0f2ff">{{ dia }}</th>
  {% endfor %}
</tr></thead>
<tbody>
{% for week in month_days %}
<tr>
  {% for day in week %}
    {% if day == 0 %}
    <td class="cal-day other-month" style="border:1px solid #f0f2ff;height:80px;vertical-align:top;padding:4px"></td>
    {% else %}
      {% set dk = '%04d-%02d-%02d'|format(anio, mes, day) %}
      {% set evs = eventos.get(dk, []) %}
      {% set is_today = (dk == today_str) %}
      <td class="cal-day{% if is_today %} today{% endif %}"
          style="border:1px solid #f0f2ff;height:80px;vertical-align:top;padding:4px;cursor:pointer{% if is_today %};background:#f5f7ff{% endif %}"
          onclick="document.getElementById('evFecha').value='{{ dk }}';new bootstrap.Modal(document.getElementById('modalEvento')).show()">
        {% if is_today %}
        <div style="margin-bottom:3px">
          <span style="background:#5e72e4;color:#fff;border-radius:50%;width:22px;height:22px;display:inline-flex;align-items:center;justify-content:center;font-size:.78rem;font-weight:700">{{ day }}</span>
        </div>
        {% else %}
        <div style="font-size:.82rem;font-weight:500;margin-bottom:3px;color:#1a1f36">{{ day }}</div>
        {% endif %}
        {% for ev in evs[:3] %}
          {% set ev_cls = {'tarea':'ev-tarea','venta':'ev-venta','evento':'ev-evento','nota':'ev-nota','caducidad':'ev-caducidad'}.get(ev.t, 'ev-evento') %}
          <div class="cal-ev {{ ev_cls }}" title="{{ ev.n|e }}" style="font-size:.68rem;overflow:hidden;white-space:nowrap;text-overflow:ellipsis;max-width:100%">{{ ev.n|truncate(24,true,'…') }}</div>
        {% endfor %}
        {% if evs|length > 3 %}
          <div style="font-size:.65rem;color:#8898aa">+{{ evs|length - 3 }} más</div>
        {% endif %}
      </td>
    {% endif %}
  {% endfor %}
</tr>
{% endfor %}
</tbody>
</table>
</div>
<!-- Modal nuevo evento -->
<div class="modal fade" id="modalEvento" tabindex="-1">
  <div class="modal-dialog"><div class="modal-content">
    <div class="modal-header"><h5 class="modal-title"><i class="bi bi-calendar-plus me-2"></i>Nuevo Evento</h5>
      <button type="button" class="btn-close" data-bs-dismiss="modal"></button></div>
    <form method="POST" action="{{ url_for('evento_nuevo') }}">
    <div class="modal-body"><div class="row g-3">
      <div class="col-12"><label class="form-label">Título *</label>
        <input type="text" name="titulo" class="form-control" id="evTitulo" required></div>
      <div class="col-md-6"><label class="form-label">Tipo</label>
        <select name="tipo" class="form-select">
          <option value="recordatorio">Recordatorio</option>
          <option value="cita">Cita</option>
          <option value="reunion">Reunión</option>
        </select></div>
      <div class="col-md-6"><label class="form-label">Fecha *</label>
        <input type="date" name="fecha" class="form-control" id="evFecha" required></div>
      <div class="col-md-6"><label class="form-label">Hora inicio</label>
        <input type="time" name="hora_inicio" class="form-control"></div>
      <div class="col-md-6"><label class="form-label">Hora fin</label>
        <input type="time" name="hora_fin" class="form-control"></div>
      <div class="col-12"><label class="form-label">Descripción</label>
        <textarea name="descripcion" class="form-control" rows="3"></textarea></div>
    </div></div>
    <div class="modal-footer">
      <button type="button" class="btn btn-outline-secondary" data-bs-dismiss="modal">Cancelar</button>
      <button type="submit" class="btn btn-primary"><i class="bi bi-check-lg me-1"></i>Guardar evento</button>
    </div></form>
  </div></div>
</div>
{% endblock %}"""

T['ventas/factura.html'] = """<!DOCTYPE html>
<html lang="es"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{{ doc_tipo }} {{ doc_numero }} — Evore</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'Segoe UI',Arial,sans-serif;font-size:12px;color:#1d1d1b;background:#fff}
.page{max-width:800px;margin:0 auto;padding:30px 36px}
.no-print{position:fixed;top:16px;right:16px;display:flex;gap:8px;z-index:999}
.btn-print{background:#5e72e4;color:#fff;border:none;padding:8px 18px;border-radius:8px;cursor:pointer;font-size:13px;font-weight:600}
.btn-back{background:#f4f6fb;color:#1a1f36;border:1px solid #dee2e6;padding:8px 14px;border-radius:8px;cursor:pointer;font-size:13px}
.header{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:28px;padding-bottom:20px;border-bottom:2px solid #d3af37}
.logo-area{display:flex;align-items:center;gap:10px}
.brand-txt{font-size:22px;font-weight:700;letter-spacing:6px;font-family:Georgia,serif;color:#1d1d1b}
.company-info{font-size:11px;color:#525f7f;line-height:1.6;text-align:right}
.doc-box{background:#1a1f36;color:#fff;padding:16px 20px;border-radius:12px;text-align:center;margin-bottom:24px;min-width:160px}
.doc-tipo{font-size:16px;font-weight:700;letter-spacing:3px}
.doc-num{font-size:13px;opacity:.8;margin-top:2px}
.doc-date{font-size:11px;opacity:.65;margin-top:4px}
.parties{display:grid;grid-template-columns:1fr 1fr;gap:20px;margin-bottom:24px}
.party-box{border:1px solid #e9ecef;border-radius:8px;padding:12px 14px}
.party-label{font-size:10px;font-weight:700;letter-spacing:2px;color:#8898aa;margin-bottom:6px;text-transform:uppercase}
.party-name{font-size:14px;font-weight:700;color:#1a1f36;margin-bottom:3px}
.party-detail{font-size:11px;color:#525f7f;line-height:1.6}
table.items{width:100%;border-collapse:collapse;margin-bottom:16px}
table.items thead th{background:#1a1f36;color:#fff;padding:8px 10px;font-size:11px;font-weight:600;text-align:left}
table.items thead th:last-child,table.items thead th:nth-last-child(2),table.items thead th:nth-last-child(3){text-align:right}
table.items tbody td{padding:8px 10px;border-bottom:1px solid #f0f2ff;font-size:12px;vertical-align:top}
table.items tbody td:last-child,table.items tbody td:nth-last-child(2),table.items tbody td:nth-last-child(3){text-align:right}
table.items tbody tr:hover{background:#fafbff}
.totals{display:flex;justify-content:flex-end;margin-bottom:20px}
.totals-box{min-width:260px}
.totals-row{display:flex;justify-content:space-between;padding:4px 0;font-size:12px;border-bottom:1px solid #f0f2ff}
.totals-row.total{font-size:15px;font-weight:700;color:#1a1f36;border-bottom:2px solid #d3af37;padding:8px 0}
.totals-row.saldo{color:#f5365c;font-weight:600}
.notes{background:#f8f9fe;border-radius:8px;padding:12px 14px;font-size:11px;color:#525f7f;margin-bottom:20px}
.footer{text-align:center;font-size:10px;color:#adb5bd;border-top:1px solid #f0f2ff;padding-top:12px;margin-top:8px}
.gold{color:#d3af37;font-weight:600}
@media print{
  .no-print{display:none!important}
  body{background:#fff}
  .page{padding:10px 16px}
  @page{margin:1cm}
}
</style></head><body>
<div class="no-print">
  <button class="btn-back" onclick="window.close()"><i>←</i> Cerrar</button>
  <button class="btn-print" onclick="window.print()">🖨️ Imprimir / PDF</button>
</div>
<div class="page">
  <div class="header">
    <div class="logo-area">
      <svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 341.94 261.01" width="90" height="69" style="display:block"><path fill="#1d1d1b" d="M28.41,217.71V184.78C28.41,171.75,27,169,27,169H72.31V173s-4.09-2.22-23.45-2.22H32.41V195.7h30.9v4.07s-2.64-2.32-15.36-2.32H32.41v34.22H51.58c19.36,0,23.45-2.22,23.45-2.22v4.07H27S28.41,230.75,28.41,217.71Z"/><path fill="#1d1d1b" d="M79.67,169h7s-.28,2.78,4.72,15.81l17.91,46.33,18.45-46.33c5.18-13,3.82-15.81,3.82-15.81h5.81s-2.45,2.68-7.45,15.17L110.3,233.52H106l-19-49C82.12,171.65,79.67,169,79.67,169Z"/><path fill="#1d1d1b" d="M138.61,201.15c0-18.49,14.45-33.38,32.36-33.38s32.36,14.89,32.36,33.38S188.87,234.54,171,234.54,138.61,219.65,138.61,201.15Zm60.81,0c0-17.47-12.73-31.62-28.45-31.62s-28.45,14.15-28.45,31.62,12.72,31.63,28.45,31.63S199.42,218.63,199.42,201.15Z"/><path fill="#1d1d1b" d="M210.64,217.71V184.78c0-13-1.36-15.81-1.36-15.81h28.63c11,0,19.9,9.34,19.9,20.81s-8.72,20.62-19.54,20.81l7.27,7.39c12.55,12.86,16.64,15.54,16.64,15.54h-8.36s-1.28-2.68-13.55-15.17l-7.54-7.76H214.64v7.12c0,13,1.36,15.81,1.36,15.81h-6.72S210.64,230.75,210.64,217.71Zm43.27-27.93c0-10.36-7.18-19-16-19H214.64v37.92h23.27C246.73,208.74,253.91,200.14,253.91,189.78Z"/><path fill="#1d1d1b" d="M268.26,217.71V184.78c0-13-1.36-15.81-1.36-15.81h45.27V173s-4.09-2.22-23.45-2.22H272.26V195.7h30.91v4.07s-2.64-2.32-15.36-2.32H272.26v34.22h19.18c19.36,0,23.45-2.22,23.45-2.22v4.07h-48S268.26,230.75,268.26,217.71Z"/><path fill="#1d1d1b" d="M239.09,111.71C234.61,109,225,105.23,225,105.23s-1.42-.5-1.34-.9c.21-1.2,2.76-1.54,9.63-10.7,7.11-9.49,8.6-13.94,9.94-17.84,1.14-3.3,2.38-11,1.72-11.79s-9.12-1.35-16.48.3c-7.06,1.58-11.1,3.47-11.5,3-.64-.71.95-6.23,1.11-11.33.17-5.28.32-6.75-.41-11.49-.61-4-1.47-7.57-2.65-7.91s-5,.82-7,1.69A81.41,81.41,0,0,0,198.41,44c-2.91,2-6.52,5.48-7.55,5.69s-2.79-6-6.9-13.15a86,86,0,0,0-5.69-9.09c-1.74-2.45-4.91-7-7.28-7h0c-2.38,0-5.54,4.59-7.28,7A84.29,84.29,0,0,0,158,36.52c-4.12,7.13-5.76,13.38-6.9,13.15s-4.73-3.8-7.64-5.8a79.94,79.94,0,0,0-9.56-5.58c-2-.87-5.81-2-7-1.69s-2,3.94-2.65,7.91c-.73,4.74-.58,6.21-.41,11.49.16,5.1,1.74,10.62,1.11,11.33-.4.44-4.44-1.45-11.5-3C106.08,62.65,97.72,63.08,97,64s.57,8.49,1.71,11.79c1.35,3.9,2.83,8.35,10,17.84,6.87,9.16,9.41,9.5,9.63,10.7.07.4-1.35.9-1.35.9s-9.57,3.72-14.06,6.48C99.17,114,95,117.34,95,119.08c0,2.11,3.54,4.62,8.14,7.67s15.36,7.83,26.75,6.75c12.82-1.22,20.68-7.21,21.76-6.74.49.21.11,5.24,5.16,10.44,4.43,4.56,10.58,10,14.15,10s9.72-5.42,14.15-10c5-5.2,4.67-10.23,5.16-10.44,1.07-.47,8.93,5.52,21.76,6.74,11.39,1.08,22.07-3.65,26.74-6.75s8.17-5.56,8.15-7.67C246.93,117.34,242.76,114,239.09,111.71ZM223,68.71c8.36-2.84,18.31-3,19-2.1s-.68,15.69-17.47,32.53C213.08,110.56,203.15,112,194.13,113c-7.48.88-16.37-.72-16.37-.72a34.59,34.59,0,0,0,11.36-11.93c4.39-7.85,6.29-14.46,13.95-20.31S214.59,71.55,223,68.71ZM201.61,45.52c7.35-5,11.19-5.92,11.86-5.48s1.71,4.63,1.66,12.23a115.69,115.69,0,0,1-1.71,16.27s-2.31,1.26-8.63,5.68a50.76,50.76,0,0,0-9.48,8.13,74.94,74.94,0,0,0-.42-15.92,99,99,0,0,0-3-12.66S194.26,50.51,201.61,45.52Zm-43.8-1.86c3.92-7.94,10.94-19.47,13.17-19.47s9.25,11.53,13.17,19.47c4.74,9.58,8.13,21.9,8.14,33.28,0,9.88-4,19.13-9.93,26.62-5.15,6.47-10.79,8.16-11.38,8.16s-6.23-1.69-11.39-8.16c-6-7.49-9.94-16.74-9.92-26.62C149.68,65.56,153.07,53.24,157.81,43.66ZM128.49,40c.67-.44,4.51.49,11.86,5.48s9.77,8.25,9.77,8.25a99,99,0,0,0-3.05,12.66,74.94,74.94,0,0,0-.42,15.92,50.76,50.76,0,0,0-9.48-8.13c-6.32-4.42-8.63-5.68-8.63-5.68a115.69,115.69,0,0,1-1.71-16.27C126.78,44.67,127.8,40.5,128.49,40Zm-11,59.1C100.69,82.3,99.34,67.51,100,66.61s10.63-.74,19,2.1,12.22,5.51,19.88,11.37,9.56,12.46,14,20.31a34.59,34.59,0,0,0,11.36,11.93s-8.89,1.6-16.37.72C138.81,112,128.88,110.56,117.49,99.14Zm28.92,26.37c-5.29,2.62-12.37,5.73-23.29,5-6.53-.44-11.77-2.84-16.93-5.52-5.33-2.76-7.66-5.09-7.66-6s4.73-4.29,9.19-6.39a112.32,112.32,0,0,1,12-5.22c1.08-.37,1.26-.07,3.11.91a40.92,40.92,0,0,0,12.86,5.9c6.85,1.73,18.66,2.63,18.66,2.63S151.7,122.88,146.41,125.51Zm33.35,12.19c-5.8,5.61-8,5.86-8.78,5.86s-3-.25-8.78-5.86-8.28-10.87-6.47-16.09a7.73,7.73,0,0,1,2.13-3.58c3.63-3.47,13.1-3.09,13.1-3.09h0s9.47-.38,13.1,3.09a7.73,7.73,0,0,1,2.13,3.58C188,126.83,185.56,132.09,179.76,137.7Zm56-12.72c-5.16,2.68-10.4,5.08-16.93,5.52-10.92.74-18-2.37-23.29-5s-7.92-8.74-7.92-8.74,11.81-.9,18.66-2.63a40.92,40.92,0,0,0,12.86-5.9c1.84-1,2-1.28,3.11-.91a112.32,112.32,0,0,1,12,5.22c4.46,2.1,9.19,5.45,9.19,6.39S241.1,122.22,235.77,125Z"/></svg>
    </div>
    <div style="text-align:right">
      <div class="doc-box">
        <div class="doc-tipo">{{ doc_tipo }}</div>
        <div class="doc-num">N° {{ doc_numero }}</div>
        <div class="doc-date">{{ fecha_doc }}</div>
      </div>
    </div>
  </div>
  <div class="parties">
    <div class="party-box">
      <div class="party-label">Emisor</div>
      <div class="party-name">{{ empresa.nombre }}</div>
      <div class="party-detail">
        {% if empresa.nit %}NIT: {{ empresa.nit }}<br>{% endif %}
        {% if empresa.direccion %}{{ empresa.direccion }}<br>{% endif %}
        {% if empresa.ciudad %}{{ empresa.ciudad }}<br>{% endif %}
        {% if empresa.telefono %}Tel: {{ empresa.telefono }}<br>{% endif %}
        {% if empresa.email %}<span class="gold">{{ empresa.email }}</span>{% endif %}
      </div>
    </div>
    <div class="party-box">
      <div class="party-label">Cliente</div>
      {% if obj.cliente %}
      <div class="party-name">{{ obj.cliente.empresa or obj.cliente.nombre }}</div>
      <div class="party-detail">
        {% if obj.cliente.nit %}NIT: {{ obj.cliente.nit }}<br>{% endif %}
        {% if obj.cliente.dir_comercial %}{{ obj.cliente.dir_comercial }}<br>{% endif %}
        {% for c in obj.cliente.contactos[:1] %}
          {{ c.nombre }}{% if c.cargo %} — {{ c.cargo }}{% endif %}<br>
          {% if c.email %}<span class="gold">{{ c.email }}</span>{% if c.telefono %} · {% endif %}{% endif %}
          {% if c.telefono %}{{ c.telefono }}{% endif %}
        {% endfor %}
      </div>
      {% else %}
      <div class="party-name" style="color:#adb5bd">Sin cliente asignado</div>{% endif %}
    </div>
  </div>
  <div style="margin-bottom:16px">
    <div style="font-size:15px;font-weight:700;color:#1a1f36;margin-bottom:4px">{{ obj.titulo }}</div>
    {% if obj.dias_entrega %}<div style="font-size:11px;color:#525f7f">Entrega estimada: {{ obj.dias_entrega }} días{% if obj.fecha_entrega_est %} ({{ obj.fecha_entrega_est.strftime('%d/%m/%Y') }}){% endif %}</div>{% endif %}
  </div>
  <table class="items">
    <thead><tr><th>#</th><th>Producto / Descripción</th><th>Cant.</th><th>P. Unitario</th><th>Subtotal</th></tr></thead>
    <tbody>
    {% for it in obj.items %}
    <tr>
      <td style="color:#adb5bd">{{ loop.index }}</td>
      <td>{{ it.nombre_prod }}</td>
      <td>{{ it.cantidad|int if it.cantidad == it.cantidad|int else it.cantidad }}</td>
      <td>$ {{ '{:,.0f}'.format(it.precio_unit).replace(',','.') }}</td>
      <td>$ {{ '{:,.0f}'.format(it.subtotal).replace(',','.') }}</td>
    </tr>{% endfor %}
    </tbody>
  </table>
  <div class="totals">
    <div class="totals-box">
      <div class="totals-row"><span>Subtotal</span><span>$ {{ '{:,.0f}'.format(obj.subtotal).replace(',','.') }}</span></div>
      <div class="totals-row"><span>IVA 19%</span><span>$ {{ '{:,.0f}'.format(obj.iva).replace(',','.') }}</span></div>
      <div class="totals-row total"><span>TOTAL</span><span>$ {{ '{:,.0f}'.format(obj.total).replace(',','.') }}</span></div>
      {% if obj.monto_anticipo > 0 %}
      <div class="totals-row"><span>Anticipo ({{ obj.porcentaje_anticipo|int }}%)</span><span>$ {{ '{:,.0f}'.format(obj.monto_anticipo).replace(',','.') }}</span></div>
      <div class="totals-row saldo"><span>Saldo pendiente</span><span>$ {{ '{:,.0f}'.format(obj.saldo).replace(',','.') }}</span></div>{% endif %}
    </div>
  </div>
  {% if obj.notas %}<div class="notes"><strong>Notas:</strong> {{ obj.notas }}</div>{% endif %}
  <div class="footer">
    {{ empresa.nombre }}{% if empresa.sitio_web %} · {{ empresa.sitio_web }}{% endif %}{% if empresa.email %} · {{ empresa.email }}{% endif %}
    <br>Documento generado el {{ now.strftime('%d/%m/%Y %H:%M') }} · Evore CRM
  </div>
</div>
</body></html>"""

T['cotizaciones/index.html'] = """{% extends 'base.html' %}
{% block title %}Cotizaciones{% endblock %}{% block page_title %}Cotizaciones{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('cotizacion_nueva') }}" class="btn btn-primary btn-sm"><i class="bi bi-plus-lg me-1"></i>Nueva cotización</a>{% endblock %}
{% block content %}
<div class="row g-3 mb-4">
  <div class="col-md-3"><div class="sc"><div class="sv">{{ items|length }}</div><div class="sl">Total cotizaciones</div></div></div>
  <div class="col-md-3"><div class="sc"><div class="sv">{{ items|selectattr('estado','eq','borrador')|list|length }}</div><div class="sl">Borradores</div></div></div>
  <div class="col-md-3"><div class="sc"><div class="sv">{{ items|selectattr('estado','eq','enviada')|list|length }}</div><div class="sl">Enviadas</div></div></div>
  <div class="col-md-3"><div class="sc"><div class="sv">{{ items|selectattr('estado','eq','aprobada')|list|length }}</div><div class="sl">Aprobadas</div></div></div>
</div>
<div class="tc"><div class="ch"><i class="bi bi-file-earmark-text me-2"></i>Cotizaciones</div>
{% if items %}<div class="table-responsive"><table class="table">
  <thead><tr><th>Número</th><th>Título</th><th>Cliente</th><th>Total</th><th>Estado</th><th>Fecha</th><th></th></tr></thead>
  <tbody>{% for c in items %}<tr>
    <td><a href="{{ url_for('cotizacion_ver', id=c.id) }}" class="fw-semibold text-decoration-none" style="color:#5e72e4">{{ c.numero or '—' }}</a></td>
    <td style="color:#1a1f36">{{ c.titulo }}</td>
    <td><small class="text-muted">{{ c.cliente.empresa or c.cliente.nombre if c.cliente else '—' }}</small></td>
    <td class="fw-semibold">$ {{ '{:,.0f}'.format(c.total).replace(',','.') }}</td>
    <td><span class="b b-{% if c.estado=='borrador' %}media{% elif c.estado=='enviada' %}info{% elif c.estado=='aprobada' %}activo{% elif c.estado=='confirmacion_orden' %}alta{% else %}inactivo{% endif %}">
      {{ {'borrador':'Borrador','enviada':'Enviada','aprobada':'Aprobada','confirmacion_orden':'Orden confirmada'}.get(c.estado, c.estado) }}
    </span></td>
    <td><small class="text-muted">{{ c.fecha_emision.strftime('%d/%m/%Y') }}</small></td>
    <td><div class="d-flex gap-1">
      <a href="{{ url_for('cotizacion_ver', id=c.id) }}" class="btn btn-sm btn-outline-primary" title="Ver"><i class="bi bi-eye"></i></a>
      <a href="{{ url_for('cotizacion_editar', id=c.id) }}" class="btn btn-sm btn-outline-secondary" title="Editar"><i class="bi bi-pencil"></i></a>
      <a href="{{ url_for('cotizacion_pdf', id=c.id) }}" class="btn btn-sm btn-outline-dark" title="PDF" target="_blank"><i class="bi bi-file-pdf"></i></a>
      <form method="POST" action="{{ url_for('cotizacion_eliminar', id=c.id) }}" onsubmit="return confirm('¿Eliminar cotización?')">
        <button class="btn btn-sm btn-outline-danger"><i class="bi bi-trash"></i></button></form>
    </div></td>
  </tr>{% endfor %}</tbody>
</table></div>
{% else %}<div class="text-center text-muted py-5"><i class="bi bi-file-earmark-text" style="font-size:3rem"></i>
  <p class="mt-3">Sin cotizaciones.</p><a href="{{ url_for('cotizacion_nueva') }}" class="btn btn-primary">Crear primera</a></div>
{% endif %}</div>{% endblock %}"""

T['cotizaciones/form.html'] = """{% extends 'base.html' %}
{% block title %}{{ titulo }}{% endblock %}{% block page_title %}{{ titulo }}{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('cotizaciones') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Volver</a>{% endblock %}
{% block content %}<div class="fc" style="max-width:900px"><form method="POST"><div class="row g-3">
  <div class="col-md-8"><label class="form-label">Título de la cotización *</label>
    <input type="text" name="titulo" class="form-control" placeholder="Ej: Diseño y producción de catálogo Q2 2026" value="{{ obj.titulo if obj else '' }}" required></div>
  <div class="col-md-4"><label class="form-label">Cliente</label>
    <select name="cliente_id" class="form-select">
      <option value="">— Sin cliente —</option>
      {% for c in clientes_list %}<option value="{{ c.id }}" {% if obj and obj.cliente_id==c.id %}selected{% endif %}>{{ c.empresa or c.nombre }}</option>{% endfor %}
    </select></div>
  <div class="col-md-3"><label class="form-label">Fecha de emisión</label>
    <input type="date" name="fecha_emision" class="form-control" value="{{ obj.fecha_emision.strftime('%Y-%m-%d') if obj else today }}"></div>
  <div class="col-md-3"><label class="form-label">Válida hasta</label>
    <input type="date" name="fecha_validez" class="form-control" value="{{ obj.fecha_validez.strftime('%Y-%m-%d') if obj and obj.fecha_validez else '' }}"></div>
  <div class="col-md-3"><label class="form-label">Días de entrega estimados</label>
    <input type="number" name="dias_entrega" class="form-control" min="1" value="{{ obj.dias_entrega if obj else '30' }}"></div>
  <div class="col-md-3"><label class="form-label">% Anticipo</label>
    <div class="input-group"><input type="number" name="porcentaje_anticipo" id="pctAnticipo" class="form-control" min="0" max="100" step="1" value="{{ obj.porcentaje_anticipo|int if obj else '50' }}"><span class="input-group-text">%</span></div></div>
  <div class="col-12"><label class="form-label fw-semibold">Ítems / Servicios</label>
    <div id="itemsWrap">
      {% if obj and obj.items %}
        {% for it in obj.items %}
        <div class="row g-2 mb-2 item-row">
          <div class="col-md-5"><input type="text" name="item_nombre[]" class="form-control form-control-sm" placeholder="Nombre del ítem" value="{{ it.nombre_prod }}" required></div>
          <div class="col-md-2"><input type="number" name="item_cantidad[]" class="form-control form-control-sm item-cant" placeholder="Cant." step="0.01" min="0" value="{{ it.cantidad }}" oninput="calcRow(this)"></div>
          <div class="col-md-3"><div class="input-group input-group-sm"><span class="input-group-text">$</span><input type="number" name="item_precio[]" class="form-control item-precio" placeholder="Precio unit." step="1" min="0" value="{{ it.precio_unit|int }}" oninput="calcRow(this)"></div></div>
          <div class="col-md-1 d-flex align-items-center"><span class="item-sub fw-semibold" style="font-size:.85rem;color:#5e72e4">$ {{ '{:,.0f}'.format(it.subtotal).replace(',','.') }}</span></div>
          <div class="col-md-1 d-flex align-items-center"><button type="button" class="btn btn-sm btn-outline-danger" onclick="this.closest('.item-row').remove();calcTotal()"><i class="bi bi-trash"></i></button></div>
        </div>{% endfor %}
      {% else %}
        <div class="row g-2 mb-2 item-row">
          <div class="col-md-5"><input type="text" name="item_nombre[]" class="form-control form-control-sm" placeholder="Nombre del ítem o servicio" required></div>
          <div class="col-md-2"><input type="number" name="item_cantidad[]" class="form-control form-control-sm item-cant" placeholder="1" step="0.01" min="0" value="1" oninput="calcRow(this)"></div>
          <div class="col-md-3"><div class="input-group input-group-sm"><span class="input-group-text">$</span><input type="number" name="item_precio[]" class="form-control item-precio" placeholder="0" step="1" min="0" value="0" oninput="calcRow(this)"></div></div>
          <div class="col-md-1 d-flex align-items-center"><span class="item-sub fw-semibold" style="font-size:.85rem;color:#5e72e4">$ 0</span></div>
          <div class="col-md-1 d-flex align-items-center"><button type="button" class="btn btn-sm btn-outline-danger" onclick="this.closest('.item-row').remove();calcTotal()"><i class="bi bi-trash"></i></button></div>
        </div>
      {% endif %}
    </div>
    <button type="button" class="btn btn-sm btn-outline-primary mt-2" onclick="addRow()"><i class="bi bi-plus-lg me-1"></i>Agregar ítem</button>
  </div>
  <div class="col-12"><hr class="my-2"></div>
  <div class="col-md-6 offset-md-6">
    <div class="d-flex justify-content-between mb-1"><span class="text-muted">Subtotal:</span><span id="lblSub" class="fw-semibold">$ 0</span></div>
    <div class="d-flex justify-content-between mb-1 align-items-center">
      <span class="text-muted">IVA:</span>
      <div class="d-flex align-items-center gap-2">
        <input type="number" name="iva_pct" id="ivaPct" class="form-control form-control-sm" style="width:70px" step="0.01" min="0" max="100" value="{{ ((obj.iva / obj.subtotal * 100)|round(2)) if obj and obj.subtotal and obj.subtotal > 0 else iva_default }}" oninput="calcTotal()">
        <span class="text-muted">%</span>
        <span id="lblIva" class="fw-semibold">$ 0</span>
      </div>
    </div>
    <div class="d-flex justify-content-between mb-1"><span class="fw-bold">Total:</span><span id="lblTotal" class="fw-bold" style="color:#5e72e4;font-size:1.1rem">$ 0</span></div>
    <div class="d-flex justify-content-between text-muted"><span>Anticipo (<span id="lblPct">50</span>%):</span><span id="lblAnticipo">$ 0</span></div>
    <div class="d-flex justify-content-between text-muted"><span>Saldo:</span><span id="lblSaldo">$ 0</span></div>
  </div>
  <div class="col-12"><label class="form-label">Condiciones de pago</label>
    <textarea name="condiciones_pago" class="form-control" rows="2" placeholder="Ej: 50% anticipo, 50% contra entrega">{{ obj.condiciones_pago if obj else '' }}</textarea></div>
  <div class="col-12"><label class="form-label">Notas adicionales</label>
    <textarea name="notas" class="form-control" rows="2" placeholder="Observaciones, alcance, exclusiones...">{{ obj.notas if obj else '' }}</textarea></div>
</div>
<div class="d-flex gap-2 mt-4">
  <button type="submit" class="btn btn-primary"><i class="bi bi-check-lg me-1"></i>{{ 'Actualizar' if obj else 'Crear Cotización' }}</button>
  <a href="{{ url_for('cotizaciones') }}" class="btn btn-outline-secondary">Cancelar</a>
</div></form></div>
{% block scripts %}<script>
var IVA_DEFAULT = {{ iva_default|tojson }};
function fmt(n){return'$ '+Math.round(n).toLocaleString('es-CO');}
function calcRow(el){
  var row=el.closest('.item-row');
  var c=parseFloat(row.querySelector('.item-cant').value)||0;
  var p=parseFloat(row.querySelector('.item-precio').value)||0;
  row.querySelector('.item-sub').textContent=fmt(c*p);
  calcTotal();
}
function calcTotal(){
  var sub=0;
  document.querySelectorAll('.item-row').forEach(function(r){
    var c=parseFloat(r.querySelector('.item-cant').value)||0;
    var p=parseFloat(r.querySelector('.item-precio').value)||0;
    sub+=c*p;
  });
  var ivaPct=parseFloat(document.getElementById('ivaPct').value)||0;
  var iva=sub*ivaPct/100;
  var total=sub+iva;
  var pct=parseFloat(document.getElementById('pctAnticipo').value)||50;
  var anticipo=total*pct/100;
  document.getElementById('lblSub').textContent=fmt(sub);
  document.getElementById('lblIva').textContent=fmt(iva);
  document.getElementById('lblTotal').textContent=fmt(total);
  document.getElementById('lblPct').textContent=Math.round(pct);
  document.getElementById('lblAnticipo').textContent=fmt(anticipo);
  document.getElementById('lblSaldo').textContent=fmt(total-anticipo);
}
function addRow(){
  var wrap=document.getElementById('itemsWrap');
  var div=document.createElement('div');
  div.className='row g-2 mb-2 item-row';
  div.innerHTML='<div class="col-md-5"><input type="text" name="item_nombre[]" class="form-control form-control-sm" placeholder="Nombre del ítem o servicio" required></div>'
    +'<div class="col-md-2"><input type="number" name="item_cantidad[]" class="form-control form-control-sm item-cant" placeholder="1" step="0.01" min="0" value="1" oninput="calcRow(this)"></div>'
    +'<div class="col-md-3"><div class="input-group input-group-sm"><span class="input-group-text">$</span><input type="number" name="item_precio[]" class="form-control item-precio" placeholder="0" step="1" min="0" value="0" oninput="calcRow(this)"></div></div>'
    +'<div class="col-md-1 d-flex align-items-center"><span class="item-sub fw-semibold" style="font-size:.85rem;color:#5e72e4">$ 0</span></div>'
    +'<div class="col-md-1 d-flex align-items-center"><button type="button" class="btn btn-sm btn-outline-danger" onclick="this.closest(\'.item-row\').remove();calcTotal()"><i class="bi bi-trash"></i></button></div>';
  wrap.appendChild(div);
}
document.getElementById('pctAnticipo').addEventListener('input',calcTotal);
calcTotal();
// Auto-título desde cliente + fecha emisión
setupAutoTitulo('[name="titulo"]',['[name="cliente_id"]','[name="fecha_emision"]']);
</script>{% endblock %}{% endblock %}"""

T['cotizaciones/ver.html'] = """{% extends 'base.html' %}
{% block title %}{{ obj.numero or 'Cotización' }}{% endblock %}{% block page_title %}{{ obj.numero or 'Cotización' }} — {{ obj.titulo }}{% endblock %}
{% block topbar_actions %}
<a href="{{ url_for('cotizacion_editar', id=obj.id) }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-pencil me-1"></i>Editar</a>
<a href="{{ url_for('cotizacion_pdf', id=obj.id) }}" class="btn btn-outline-dark btn-sm" target="_blank"><i class="bi bi-file-pdf me-1"></i>PDF</a>
<a href="{{ url_for('cotizaciones') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Volver</a>
{% endblock %}
{% block content %}
<div class="row g-3 mb-3">
  <div class="col-md-8">
    <div class="tc h-100"><div class="ch">Datos generales</div><div class="p-3">
      <div class="row g-2">
        <div class="col-sm-6"><small class="text-muted d-block">Cliente</small><span class="fw-semibold">{{ obj.cliente.empresa or obj.cliente.nombre if obj.cliente else '—' }}</span></div>
        <div class="col-sm-3"><small class="text-muted d-block">Emisión</small><span>{{ obj.fecha_emision.strftime('%d/%m/%Y') }}</span></div>
        <div class="col-sm-3"><small class="text-muted d-block">Válida hasta</small><span>{{ obj.fecha_validez.strftime('%d/%m/%Y') if obj.fecha_validez else '—' }}</span></div>
        <div class="col-sm-3"><small class="text-muted d-block">Entrega est.</small><span>{{ obj.dias_entrega }} días</span></div>
        <div class="col-sm-3"><small class="text-muted d-block">Estado</small>
          <span class="b b-{% if obj.estado=='borrador' %}media{% elif obj.estado=='enviada' %}info{% elif obj.estado=='aprobada' %}activo{% elif obj.estado=='confirmacion_orden' %}alta{% else %}inactivo{% endif %}">
            {{ {'borrador':'Borrador','enviada':'Enviada','aprobada':'Aprobada','confirmacion_orden':'Orden confirmada'}.get(obj.estado,obj.estado) }}
          </span></div>
        {% if obj.condiciones_pago %}<div class="col-12"><small class="text-muted d-block">Condiciones de pago</small><span style="font-size:.9rem">{{ obj.condiciones_pago }}</span></div>{% endif %}
        {% if obj.notas %}<div class="col-12"><small class="text-muted d-block">Notas</small><span style="font-size:.9rem">{{ obj.notas }}</span></div>{% endif %}
      </div>
    </div></div>
  </div>
  <div class="col-md-4">
    <div class="tc h-100"><div class="ch">Resumen financiero</div><div class="p-3">
      <div class="d-flex justify-content-between mb-2"><span class="text-muted">Subtotal</span><span class="fw-semibold">$ {{ '{:,.0f}'.format(obj.subtotal).replace(',','.') }}</span></div>
      <div class="d-flex justify-content-between mb-2"><span class="text-muted">IVA</span><span class="fw-semibold">$ {{ '{:,.0f}'.format(obj.iva).replace(',','.') }}</span></div>
      <div class="d-flex justify-content-between mb-2 pb-2" style="border-bottom:2px solid #5e72e4"><span class="fw-bold">Total</span><span class="fw-bold" style="color:#5e72e4;font-size:1.15rem">$ {{ '{:,.0f}'.format(obj.total).replace(',','.') }}</span></div>
      <div class="d-flex justify-content-between mb-1"><span class="text-muted">Anticipo ({{ obj.porcentaje_anticipo|int }}%)</span><span>$ {{ '{:,.0f}'.format(obj.monto_anticipo).replace(',','.') }}</span></div>
      <div class="d-flex justify-content-between"><span class="text-muted">Saldo</span><span>$ {{ '{:,.0f}'.format(obj.saldo).replace(',','.') }}</span></div>
    </div></div>
  </div>
</div>
<div class="tc mb-3"><div class="ch">Ítems</div>
<div class="table-responsive"><table class="table">
  <thead><tr><th>#</th><th>Descripción</th><th class="text-end">Cant.</th><th class="text-end">Precio unit.</th><th class="text-end">Subtotal</th></tr></thead>
  <tbody>{% for it in obj.items %}
  <tr>
    <td><small class="text-muted">{{ loop.index }}</small></td>
    <td>{{ it.nombre_prod }}</td>
    <td class="text-end">{{ it.cantidad }}</td>
    <td class="text-end">$ {{ '{:,.0f}'.format(it.precio_unit).replace(',','.') }}</td>
    <td class="text-end fw-semibold">$ {{ '{:,.0f}'.format(it.subtotal).replace(',','.') }}</td>
  </tr>{% endfor %}</tbody>
</table></div></div>
<div class="tc"><div class="ch">Cambiar estado</div><div class="p-3">
  <div class="d-flex flex-wrap gap-2">
    {% for e,lbl in [('borrador','Borrador'),('enviada','Enviada al cliente'),('aprobada','Aprobada'),('confirmacion_orden','Orden confirmada')] %}
    <form method="POST" action="{{ url_for('cotizacion_cambiar_estado', id=obj.id) }}">
      <input type="hidden" name="estado" value="{{ e }}">
      <button class="btn btn-sm {% if obj.estado==e %}btn-primary{% else %}btn-outline-secondary{% endif %}">{{ lbl }}</button>
    </form>{% endfor %}
  </div>
</div></div>{% endblock %}"""

T['cotizaciones/pdf.html'] = """<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8">
<title>{{ obj.numero or 'Cotización' }} — {{ empresa.nombre }}</title>
<style>
{% raw %}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'Helvetica Neue',Arial,sans-serif;font-size:11pt;color:#1a1f36;background:#fff}
.page{width:210mm;min-height:297mm;margin:0 auto;padding:16mm 16mm 20mm}
.header{display:flex;justify-content:space-between;align-items:flex-start;border-bottom:3px solid #1a1f36;padding-bottom:12px;margin-bottom:18px}
.logo-wrap{display:flex;align-items:center}
.logo-wrap img{max-height:48px;max-width:180px}
.doc-title{text-align:right}
.doc-title h1{font-size:22pt;font-weight:800;color:#1a1f36;letter-spacing:-0.5px}
.doc-title .numero{font-size:10pt;color:#525f7f;margin-top:4px}
.info-grid{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:16px}
.info-box{background:#f8f9fe;border-radius:8px;padding:10px 14px}
.info-box h4{font-size:8pt;text-transform:uppercase;letter-spacing:.5px;color:#8898aa;margin-bottom:6px}
.info-box p{font-size:10pt;color:#1a1f36;line-height:1.5}
table{width:100%;border-collapse:collapse;margin:12px 0}
th{background:#1a1f36;color:#fff;font-size:9pt;padding:8px 10px;text-align:left}
th.r,td.r{text-align:right}
td{padding:7px 10px;font-size:10pt;border-bottom:1px solid #e8ecf5}
tr:nth-child(even) td{background:#f8f9fe}
.totals{margin-left:auto;width:260px;margin-top:12px}
.totals table td{font-size:10pt}
.totals .grand{background:#1a1f36;color:#fff;font-weight:700;font-size:12pt}
.anticipo-box{margin-top:10px;padding:10px 14px;background:#e8ecff;border-radius:8px;display:flex;justify-content:space-between}
.footer-notes{margin-top:18px;padding-top:12px;border-top:1px solid #e8ecf5;font-size:9pt;color:#525f7f}
.estado-badge{display:inline-block;padding:3px 12px;border-radius:20px;font-size:9pt;font-weight:600;background:#e8ecff;color:#5e72e4}
@media print{body{background:#fff}.page{padding:10mm 12mm}}
{% endraw %}
</style></head>
<body><div class="page">
<div class="header">
  <div class="logo-wrap">
    {% if empresa.nombre %}
    <svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 341.94 261.01" style="height:48px">
      <path d="M171.13,0C76.58,0,0,58.38,0,130.38s76.58,130.38,171.13,130.38,170.81-58.38,170.81-130.38S265.51,0,171.13,0Zm0,235.17c-59.28,0-107.34-46.99-107.34-104.93S111.85,25.32,171.13,25.32s107.02,46.99,107.02,104.93-47.9,104.93-107.02,104.93Z" fill="#1a1f36"/>
      <path d="M248.21,88.31c-3.52-5.7-9.76-9.16-16.49-9.16h-121.18c-6.73,0-12.97,3.46-16.49,9.16-3.52,5.7-3.76,12.79-.63,18.72l60.59,114.27c3.27,6.16,9.67,10.01,16.65,10.01s13.38-3.84,16.65-10.01l60.59-114.27c3.13-5.93,2.9-13.02-.63-18.72Zm-77.08,99.59l-43.65-82.33h87.3l-43.65,82.33Z" fill="#1a1f36"/>
    </svg>
    <span style="font-size:14pt;font-weight:700;color:#1a1f36;margin-left:10px">{{ empresa.nombre }}</span>
    {% endif %}
  </div>
  <div class="doc-title">
    <h1>COTIZACIÓN</h1>
    <div class="numero">{{ obj.numero or '—' }}</div>
    <div style="margin-top:6px"><span class="estado-badge">{{ {'borrador':'Borrador','enviada':'Enviada','aprobada':'Aprobada','confirmacion_orden':'Orden confirmada'}.get(obj.estado,obj.estado) }}</span></div>
  </div>
</div>
<div class="info-grid">
  <div class="info-box">
    <h4>Para</h4>
    <p><strong>{{ obj.cliente.empresa or obj.cliente.nombre if obj.cliente else '—' }}</strong><br>
    {% if obj.cliente and obj.cliente.email %}{{ obj.cliente.email }}<br>{% endif %}
    {% if obj.cliente and obj.cliente.telefono %}{{ obj.cliente.telefono }}{% endif %}</p>
  </div>
  <div class="info-box">
    <h4>De parte de</h4>
    <p><strong>{{ empresa.nombre }}</strong><br>
    {% if empresa.email %}{{ empresa.email }}<br>{% endif %}
    {% if empresa.telefono %}{{ empresa.telefono }}<br>{% endif %}
    {% if empresa.nit %}NIT: {{ empresa.nit }}{% endif %}</p>
  </div>
  <div class="info-box">
    <h4>Descripción del proyecto</h4>
    <p>{{ obj.titulo }}</p>
  </div>
  <div class="info-box">
    <h4>Fechas</h4>
    <p>Emisión: {{ obj.fecha_emision.strftime('%d/%m/%Y') }}<br>
    Validez: {{ obj.fecha_validez.strftime('%d/%m/%Y') if obj.fecha_validez else '—' }}<br>
    Entrega estimada: {{ obj.dias_entrega }} días hábiles</p>
  </div>
</div>
<table>
  <thead><tr><th>#</th><th>Descripción</th><th class="r">Cant.</th><th class="r">Precio unit.</th><th class="r">Subtotal</th></tr></thead>
  <tbody>{% for it in obj.items %}
  <tr><td>{{ loop.index }}</td><td>{{ it.nombre_prod }}</td><td class="r">{{ it.cantidad }}</td>
    <td class="r">$ {{ '{:,.0f}'.format(it.precio_unit).replace(',','.') }}</td>
    <td class="r">$ {{ '{:,.0f}'.format(it.subtotal).replace(',','.') }}</td></tr>
  {% endfor %}</tbody>
</table>
<div class="totals">
  <table>
    <tr><td>Subtotal</td><td class="r">$ {{ '{:,.0f}'.format(obj.subtotal).replace(',','.') }}</td></tr>
    <tr><td>IVA</td><td class="r">$ {{ '{:,.0f}'.format(obj.iva).replace(',','.') }}</td></tr>
    <tr class="grand"><td>TOTAL</td><td class="r">$ {{ '{:,.0f}'.format(obj.total).replace(',','.') }}</td></tr>
  </table>
  <div class="anticipo-box">
    <span><strong>Anticipo requerido ({{ obj.porcentaje_anticipo|int }}%)</strong></span>
    <span><strong>$ {{ '{:,.0f}'.format(obj.monto_anticipo).replace(',','.') }}</strong></span>
  </div>
  <div style="text-align:right;font-size:10pt;color:#525f7f;margin-top:6px">Saldo: $ {{ '{:,.0f}'.format(obj.saldo).replace(',','.') }}</div>
</div>
{% if obj.condiciones_pago or obj.notas %}
<div class="footer-notes">
  {% if obj.condiciones_pago %}<p><strong>Condiciones de pago:</strong> {{ obj.condiciones_pago }}</p>{% endif %}
  {% if obj.notas %}<p style="margin-top:6px"><strong>Notas:</strong> {{ obj.notas }}</p>{% endif %}
</div>{% endif %}
<div class="footer-notes" style="margin-top:12px;text-align:center;font-size:8pt">
  Generado por sistema Evore CRM · {{ empresa.nombre }} {% if empresa.sitio_web %}· {{ empresa.sitio_web }}{% endif %}
</div>
</div></body></html>"""

T['actividad.html'] = """{% extends 'base.html' %}
{% block title %}Actividad{% endblock %}{% block page_title %}Historial de Actividad{% endblock %}
{% block content %}
<div class="tc">
  <div class="ch"><i class="bi bi-clock-history me-2"></i>Últimas {{ items|length }} acciones</div>
  {% if items %}
  <div class="table-responsive"><table class="table">
    <thead><tr><th>Fecha</th><th>Usuario</th><th>Acción</th><th>Módulo</th><th>Descripción</th></tr></thead>
    <tbody>{% for a in items %}
    <tr>
      <td><small class="text-muted">{{ a.creado_en.strftime('%d/%m/%Y %H:%M') }}</small></td>
      <td><span class="fw-semibold" style="font-size:.88rem">{{ a.usuario.nombre if a.usuario else '—' }}</span></td>
      <td><span class="badge
        {% if a.tipo == 'crear' %}bg-success
        {% elif a.tipo == 'editar' %}bg-primary
        {% elif a.tipo == 'eliminar' %}bg-danger
        {% elif a.tipo == 'completar' %}bg-info
        {% else %}bg-secondary{% endif %}">{{ a.tipo.title() }}</span></td>
      <td><small class="text-muted">{{ a.entidad.title() }}</small></td>
      <td style="font-size:.88rem">{{ a.descripcion }}</td>
    </tr>{% endfor %}</tbody>
  </table></div>
  {% else %}<div class="text-center text-muted py-5"><i class="bi bi-clock-history" style="font-size:3rem"></i>
    <p class="mt-3">Sin actividad registrada aún.</p></div>{% endif %}
</div>{% endblock %}"""

T['admin/config.html'] = """{% extends 'base.html' %}
{% block title %}Configuración Empresa{% endblock %}{% block page_title %}Datos de la Empresa{% endblock %}
{% block content %}<div class="fc" style="max-width:700px">
<p class="text-muted mb-4" style="font-size:.9rem">Esta información aparece en las facturas y cotizaciones generadas por el sistema.</p>
<form method="POST"><div class="row g-3">
  <div class="col-md-8"><label class="form-label">Nombre de la empresa *</label>
    <input type="text" name="nombre" class="form-control" value="{{ obj.nombre or '' }}" required></div>
  <div class="col-md-4"><label class="form-label">NIT</label>
    <input type="text" name="nit" class="form-control" value="{{ obj.nit or '' }}"></div>
  <div class="col-md-6"><label class="form-label">Ciudad</label>
    <input type="text" name="ciudad" class="form-control" value="{{ obj.ciudad or '' }}"></div>
  <div class="col-md-6"><label class="form-label">Teléfono</label>
    <input type="text" name="telefono" class="form-control" value="{{ obj.telefono or '' }}"></div>
  <div class="col-md-6"><label class="form-label">Email</label>
    <input type="email" name="email" class="form-control" value="{{ obj.email or '' }}"></div>
  <div class="col-md-6"><label class="form-label">Sitio web</label>
    <input type="text" name="sitio_web" class="form-control" placeholder="evore.us" value="{{ obj.sitio_web or '' }}"></div>
  <div class="col-12"><label class="form-label">Dirección</label>
    <input type="text" name="direccion" class="form-control" value="{{ obj.direccion or '' }}"></div>
</div>
<div class="mt-4">
  <button type="submit" class="btn btn-primary"><i class="bi bi-check-lg me-1"></i>Guardar configuración</button>
</div></form></div>{% endblock %}"""

# =============================================================
# TEMPLATES V12 — Notificaciones
# =============================================================

T['notificaciones.html'] = """{% extends 'base.html' %}
{% block title %}Notificaciones{% endblock %}{% block page_title %}Notificaciones{% endblock %}
{% block topbar_actions %}
<form method="POST" action="{{ url_for('notificaciones_marcar_todas') }}">
  <button type="submit" class="btn btn-outline-secondary btn-sm"><i class="bi bi-check-all me-1"></i>Marcar todas como leídas</button>
</form>
{% endblock %}
{% block content %}
<div class="fc">
{% if items %}
  <div class="list-group list-group-flush">
  {% for n in items %}
    <div class="list-group-item list-group-item-action d-flex align-items-start gap-3 py-3 {% if not n.leida %}list-group-item-info{% endif %}">
      <div class="mt-1">
        {% if n.tipo == 'tarea_asignada' %}<i class="bi bi-check2-square text-primary fs-5"></i>
        {% elif n.tipo == 'alerta_stock' %}<i class="bi bi-exclamation-triangle-fill text-warning fs-5"></i>
        {% else %}<i class="bi bi-info-circle-fill text-secondary fs-5"></i>{% endif %}
      </div>
      <div class="flex-grow-1">
        <div class="fw-semibold">{{ n.titulo }}</div>
        <div class="text-muted small">{{ n.mensaje }}</div>
        <div class="text-muted" style="font-size:.78rem">{{ n.creado_en.strftime('%d/%m/%Y %H:%M') }}</div>
      </div>
      {% if n.url %}<a href="{{ n.url }}" class="btn btn-sm btn-outline-primary align-self-center">Ver</a>{% endif %}
    </div>
  {% endfor %}
  </div>
{% else %}
  <div class="text-center py-5 text-muted">
    <i class="bi bi-bell-slash fs-1 d-block mb-2"></i>No tienes notificaciones.
  </div>
{% endif %}
</div>{% endblock %}"""

# =============================================================
# TEMPLATES V12 — Inventario Lotes
# =============================================================

T['inventario/lotes.html'] = """{% extends 'base.html' %}
{% block title %}Lotes de Inventario{% endblock %}{% block page_title %}Lotes de Inventario{% endblock %}
{% block topbar_actions %}
<a href="{{ url_for('lote_nuevo') }}" class="btn btn-primary btn-sm"><i class="bi bi-plus-lg me-1"></i>Nuevo Lote</a>
{% endblock %}
{% block content %}
<div class="fc">
<table class="table table-hover align-middle">
  <thead><tr>
    <th>Producto</th><th>Lote</th><th>NSO</th>
    <th>Fecha Prod.</th><th>Vencimiento</th>
    <th class="text-end">Producidas</th><th class="text-end">Restantes</th>
    <th>Notas</th><th></th>
  </tr></thead>
  <tbody>
  {% for l in lotes %}
  <tr>
    <td>{{ l.producto.nombre }}</td>
    <td><span class="badge bg-secondary">{{ l.numero_lote }}</span></td>
    <td>{{ l.nso or '—' }}</td>
    <td>{{ l.fecha_produccion.strftime('%d/%m/%Y') if l.fecha_produccion else '—' }}</td>
    <td>
      {% if l.fecha_vencimiento %}
        {% set hoy = now.date() %}
        {% set dias = (l.fecha_vencimiento - hoy).days %}
        <span class="badge {% if dias < 0 %}bg-danger{% elif dias <= 30 %}bg-warning text-dark{% else %}bg-success{% endif %}">
          {{ l.fecha_vencimiento.strftime('%d/%m/%Y') }}
          {% if dias < 0 %} (vencido){% elif dias <= 30 %} ({{ dias }}d){% endif %}
        </span>
      {% else %}—{% endif %}
    </td>
    <td class="text-end">{{ l.unidades_producidas }}</td>
    <td class="text-end">{{ l.unidades_restantes }}</td>
    <td><small class="text-muted">{{ l.notas or '' }}</small></td>
    <td>
      <a href="{{ url_for('lote_editar', id=l.id) }}" class="btn btn-sm btn-outline-secondary"><i class="bi bi-pencil"></i></a>
      <form method="POST" action="{{ url_for('lote_eliminar', id=l.id) }}" class="d-inline"
            onsubmit="return confirm('¿Eliminar lote?')">
        <button class="btn btn-sm btn-outline-danger"><i class="bi bi-trash"></i></button>
      </form>
    </td>
  </tr>
  {% else %}
  <tr><td colspan="9" class="text-center text-muted py-4">No hay lotes registrados.</td></tr>
  {% endfor %}
  </tbody>
</table>
</div>{% endblock %}"""

T['inventario/lote_form.html'] = """{% extends 'base.html' %}
{% block title %}{{ titulo }}{% endblock %}{% block page_title %}{{ titulo }}{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('lotes') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Volver</a>{% endblock %}
{% block content %}<div class="fc"><form method="POST"><div class="row g-3">
  <div class="col-md-6"><label class="form-label">Producto *</label>
    <select name="producto_id" class="form-select" required>
      {% for p in productos %}
      <option value="{{ p.id }}" {% if obj and obj.producto_id==p.id %}selected{% endif %}>{{ p.nombre }}</option>
      {% endfor %}
    </select></div>
  <div class="col-md-6"><label class="form-label">Número de Lote *</label>
    <input type="text" name="numero_lote" class="form-control" value="{{ obj.numero_lote if obj else '' }}" required placeholder="LOT-2025-001"></div>
  <div class="col-md-6"><label class="form-label">NSO (Número de Serie / Orden)</label>
    <input type="text" name="nso" class="form-control" value="{{ obj.nso if obj else '' }}" placeholder="NSO-001"></div>
  <div class="col-md-6"><label class="form-label">Fecha de Producción</label>
    <input type="date" name="fecha_produccion" class="form-control" value="{{ obj.fecha_produccion.strftime('%Y-%m-%d') if obj and obj.fecha_produccion else '' }}"></div>
  <div class="col-md-6"><label class="form-label">Fecha de Vencimiento</label>
    <input type="date" name="fecha_vencimiento" class="form-control" value="{{ obj.fecha_vencimiento.strftime('%Y-%m-%d') if obj and obj.fecha_vencimiento else '' }}"></div>
  <div class="col-md-3"><label class="form-label">Unidades Producidas *</label>
    <input type="number" name="unidades_producidas" class="form-control" min="0" value="{{ obj.unidades_producidas if obj else 0 }}" required></div>
  <div class="col-md-3"><label class="form-label">Unidades Restantes *</label>
    <input type="number" name="unidades_restantes" class="form-control" min="0" value="{{ obj.unidades_restantes if obj else 0 }}" required></div>
  <div class="col-12"><label class="form-label">Notas</label>
    <textarea name="notas" class="form-control" rows="2">{{ obj.notas if obj else '' }}</textarea></div>
</div>
<div class="d-flex gap-2 mt-4">
  <button type="submit" class="btn btn-primary"><i class="bi bi-check-lg me-1"></i>{{ 'Actualizar' if obj else 'Crear Lote' }}</button>
  <a href="{{ url_for('lotes') }}" class="btn btn-outline-secondary">Cancelar</a>
</div></form></div>
{% block scripts %}<script>
setupAutoTitulo('[name="numero_lote"]',['[name="producto_id"]','[name="fecha_produccion"]']);
</script>{% endblock %}{% endblock %}"""

# =============================================================
# TEMPLATES V12 — Materias Primas
# =============================================================

T['produccion/materias.html'] = """{% extends 'base.html' %}
{% block title %}Materias Primas{% endblock %}{% block page_title %}Materias Primas{% endblock %}
{% block topbar_actions %}
<a href="{{ url_for('materia_nueva') }}" class="btn btn-primary btn-sm"><i class="bi bi-plus-lg me-1"></i>Nueva Materia Prima</a>
{% endblock %}
{% block content %}
<div class="fc">
<table class="table table-hover align-middle">
  <thead><tr>
    <th>Nombre</th><th>Categoría</th><th>Unidad</th>
    <th class="text-end">Stock Disponible</th><th class="text-end">Stock Reservado</th>
    <th class="text-end">Mínimo</th><th class="text-end">Costo Unit.</th>
    <th>Proveedor</th><th></th>
  </tr></thead>
  <tbody>
  {% for m in materias %}
  <tr>
    <td>
      <strong>{{ m.nombre }}</strong>
      {% if m.descripcion %}<br><small class="text-muted">{{ m.descripcion }}</small>{% endif %}
    </td>
    <td>{{ m.categoria or '—' }}</td>
    <td><span class="badge bg-light text-dark border">{{ m.unidad }}</span></td>
    <td class="text-end {% if m.stock_disponible <= m.stock_minimo %}text-danger fw-bold{% endif %}">
      {{ '%.2f'|format(m.stock_disponible) }}</td>
    <td class="text-end text-warning">{{ '%.2f'|format(m.stock_reservado) }}</td>
    <td class="text-end">{{ '%.2f'|format(m.stock_minimo) }}</td>
    <td class="text-end">${{ '%.2f'|format(m.costo_unitario) }}</td>
    <td>{{ m.proveedor or '—' }}</td>
    <td>
      <a href="{{ url_for('materia_editar', id=m.id) }}" class="btn btn-sm btn-outline-secondary"><i class="bi bi-pencil"></i></a>
      <form method="POST" action="{{ url_for('materia_eliminar', id=m.id) }}" class="d-inline"
            onsubmit="return confirm('¿Eliminar materia prima?')">
        <button class="btn btn-sm btn-outline-danger"><i class="bi bi-trash"></i></button>
      </form>
    </td>
  </tr>
  {% else %}
  <tr><td colspan="9" class="text-center text-muted py-4">No hay materias primas registradas.</td></tr>
  {% endfor %}
  </tbody>
</table>
</div>{% endblock %}"""

T['produccion/materia_form.html'] = """{% extends 'base.html' %}
{% block title %}{{ titulo }}{% endblock %}{% block page_title %}{{ titulo }}{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('materias') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Volver</a>{% endblock %}
{% block content %}<div class="fc"><form method="POST"><div class="row g-3">
  <div class="col-md-8"><label class="form-label">Nombre *</label>
    <input type="text" name="nombre" class="form-control" value="{{ obj.nombre if obj else '' }}" required></div>
  <div class="col-md-4"><label class="form-label">Categoría</label>
    <input type="text" name="categoria" class="form-control" value="{{ obj.categoria if obj else '' }}" placeholder="Ingredientes, Químicos..."></div>
  <div class="col-12"><label class="form-label">Descripción</label>
    <textarea name="descripcion" class="form-control" rows="2">{{ obj.descripcion if obj else '' }}</textarea></div>

  <!-- Producto al que pertenece esta materia prima -->
  <div class="col-12">
    <label class="form-label fw-semibold">Producto asociado <small class="text-muted fw-normal">(opcional — para qué producto se usa principalmente esta materia)</small></label>
    <select name="producto_id" class="form-select">
      <option value="">— Sin asignar —</option>
      {% for p in productos %}
      <option value="{{ p.id }}" {% if obj and obj.producto_id==p.id %}selected{% endif %}>{{ p.nombre }}{% if p.sku %} ({{ p.sku }}){% endif %}</option>
      {% endfor %}
    </select>
  </div>

  <div class="col-md-4"><label class="form-label">Unidad de medida *</label>
    <select name="unidad" class="form-select" required>
      {% for u in ['kg','g','litros','ml','unidades','piezas','metros','cm'] %}
      <option value="{{ u }}" {% if obj and obj.unidad==u %}selected{% endif %}>{{ u }}</option>
      {% endfor %}
    </select></div>
  <div class="col-md-4"><label class="form-label">Stock disponible</label>
    <input type="number" name="stock_disponible" step="0.001" class="form-control" value="{{ obj.stock_disponible if obj else 0 }}"></div>
  <div class="col-md-4"><label class="form-label">Stock mínimo</label>
    <input type="number" name="stock_minimo" step="0.001" class="form-control" value="{{ obj.stock_minimo if obj else 0 }}"></div>
  <div class="col-md-4"><label class="form-label">Costo unitario</label>
    <div class="input-group"><span class="input-group-text">$</span>
    <input type="number" name="costo_unitario" step="0.01" class="form-control" value="{{ obj.costo_unitario if obj else 0 }}"></div></div>
  <div class="col-md-8"><label class="form-label">Proveedor</label>
    <select name="proveedor_id" class="form-select">
      <option value="">— Sin proveedor —</option>
      {% for pv in proveedores_list %}<option value="{{ pv.id }}" {% if obj and obj.proveedor_id==pv.id %}selected{% endif %}>{{ pv.empresa or pv.nombre }}</option>{% endfor %}
    </select>
    <input type="hidden" name="proveedor" value="{{ obj.proveedor if obj else '' }}"></div>
</div>
<div class="d-flex gap-2 mt-4">
  <button type="submit" class="btn btn-primary"><i class="bi bi-check-lg me-1"></i>{{ 'Actualizar' if obj else 'Crear Materia Prima' }}</button>
  <a href="{{ url_for('materias') }}" class="btn btn-outline-secondary">Cancelar</a>
</div></form></div>{% endblock %}"""

# =============================================================
# TEMPLATES V12 — Recetas / BOM
# =============================================================

T['produccion/recetas.html'] = """{% extends 'base.html' %}
{% block title %}Recetas de Producción{% endblock %}{% block page_title %}Recetas de Producción (BOM){% endblock %}
{% block topbar_actions %}
<a href="{{ url_for('receta_nueva') }}" class="btn btn-primary btn-sm"><i class="bi bi-plus-lg me-1"></i>Nueva Receta</a>
{% endblock %}
{% block content %}
<div class="row g-3">
{% for r in recetas %}
<div class="col-md-6">
  <div class="fc h-100">
    <div class="d-flex justify-content-between align-items-start mb-3">
      <div>
        <h5 class="mb-0">{{ r.producto.nombre }}</h5>
        <small class="text-muted">Produce {{ r.unidades_produce }} unidad(es) por lote</small>
      </div>
      <div class="d-flex gap-2">
        <a href="{{ url_for('receta_editar', id=r.id) }}" class="btn btn-sm btn-outline-secondary"><i class="bi bi-pencil"></i></a>
        <form method="POST" action="{{ url_for('receta_eliminar', id=r.id) }}" class="d-inline"
              onsubmit="return confirm('¿Eliminar receta?')">
          <button class="btn btn-sm btn-outline-danger"><i class="bi bi-trash"></i></button>
        </form>
      </div>
    </div>
    {% if r.descripcion %}<p class="text-muted small mb-3">{{ r.descripcion }}</p>{% endif %}
    <table class="table table-sm mb-0">
      <thead><tr><th>Materia Prima</th><th class="text-end">Cantidad/unidad</th><th>Unidad</th>
        <th class="text-end">Disponible</th><th class="text-end">Reservado</th></tr></thead>
      <tbody>
      {% for item in r.items %}
      <tr>
        <td>{{ item.materia.nombre }}</td>
        <td class="text-end">{{ item.cantidad_por_unidad }}</td>
        <td>{{ item.materia.unidad }}</td>
        <td class="text-end {% if item.materia.stock_disponible < item.cantidad_por_unidad * r.unidades_produce %}text-danger{% else %}text-success{% endif %}">
          {{ '%.2f'|format(item.materia.stock_disponible) }}</td>
        <td class="text-end text-warning">{{ '%.2f'|format(item.materia.stock_reservado) }}</td>
      </tr>
      {% endfor %}
      </tbody>
    </table>
  </div>
</div>
{% else %}
<div class="col-12"><div class="fc text-center py-5 text-muted">
  <i class="bi bi-diagram-3 fs-1 d-block mb-2"></i>No hay recetas de producción.
  <a href="{{ url_for('receta_nueva') }}" class="btn btn-primary mt-3">Crear primera receta</a>
</div></div>
{% endfor %}
</div>{% endblock %}"""

T['produccion/receta_form.html'] = """{% extends 'base.html' %}
{% block title %}{{ titulo }}{% endblock %}{% block page_title %}{{ titulo }}{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('recetas') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Volver</a>{% endblock %}
{% block content %}<div class="fc">
<form method="POST" id="recetaForm">
<div class="row g-3 mb-4">
  <div class="col-md-6"><label class="form-label">Producto terminado *</label>
    <select name="producto_id" class="form-select" required>
      <option value="">Seleccionar producto...</option>
      {% for p in productos %}
      <option value="{{ p.id }}" {% if obj and obj.producto_id==p.id %}selected{% endif %}>{{ p.nombre }}</option>
      {% endfor %}
    </select></div>
  <div class="col-md-3"><label class="form-label">Unidades por lote *</label>
    <input type="number" name="unidades_produce" min="1" class="form-control" value="{{ obj.unidades_produce if obj else 1 }}" required></div>
  <div class="col-12"><label class="form-label">Descripción</label>
    <textarea name="descripcion" class="form-control" rows="2">{{ obj.descripcion if obj else '' }}</textarea></div>
</div>

<h5 class="mb-3"><i class="bi bi-list-ul me-2 text-primary"></i>Ingredientes / Materias Primas</h5>
<div id="itemsContainer">
{% if obj %}
  {% for item in obj.items %}
  <div class="row g-2 mb-2 item-row">
    <div class="col-md-6">
      <select name="materia_id[]" class="form-select form-select-sm" required>
        <option value="">Seleccionar materia prima...</option>
        {% for m in materias %}
        <option value="{{ m.id }}" {% if item.materia_prima_id==m.id %}selected{% endif %}>{{ m.nombre }} ({{ m.unidad }})</option>
        {% endfor %}
      </select>
    </div>
    <div class="col-md-4">
      <input type="number" name="cantidad[]" step="0.001" min="0.001" class="form-control form-control-sm"
             value="{{ item.cantidad_por_unidad }}" placeholder="Cantidad por unidad" required>
    </div>
    <div class="col-md-2">
      <button type="button" class="btn btn-sm btn-outline-danger w-100" onclick="this.closest('.item-row').remove()"><i class="bi bi-trash"></i></button>
    </div>
  </div>
  {% endfor %}
{% endif %}
</div>
<button type="button" class="btn btn-outline-secondary btn-sm mb-4" onclick="addItem()">
  <i class="bi bi-plus-lg me-1"></i>Agregar ingrediente
</button>

<div class="d-flex gap-2">
  <button type="submit" class="btn btn-primary"><i class="bi bi-check-lg me-1"></i>{{ 'Actualizar' if obj else 'Crear Receta' }}</button>
  <a href="{{ url_for('recetas') }}" class="btn btn-outline-secondary">Cancelar</a>
</div>
</form>
</div>
{% block scripts %}<script>
var MATERIAS = {{ materias_json|tojson }};
function addItem(){
  var cont=document.getElementById('itemsContainer');
  if(!cont){alert('Error: contenedor no encontrado');return;}
  // Crear fila
  var row=document.createElement('div');
  row.className='row g-2 mb-2 item-row align-items-center';
  // Columna select materia prima
  var colA=document.createElement('div'); colA.className='col-md-5';
  var sel=document.createElement('select');
  sel.name='materia_id[]'; sel.className='form-select form-select-sm'; sel.required=true;
  var optDef=document.createElement('option'); optDef.value=''; optDef.textContent='Seleccionar materia prima...'; sel.appendChild(optDef);
  MATERIAS.forEach(function(m){
    var opt=document.createElement('option'); opt.value=m.id;
    opt.textContent=m.nombre+' ('+m.unidad+')';
    opt.setAttribute('data-unidad',m.unidad); sel.appendChild(opt);
  });
  sel.addEventListener('change',function(){
    var opt=this.options[this.selectedIndex];
    var u=opt.getAttribute('data-unidad')||'';
    this.closest('.item-row').querySelector('.unidad-lbl').textContent=u;
  });
  colA.appendChild(sel);
  // Columna cantidad
  var colB=document.createElement('div'); colB.className='col-md-3';
  var inp=document.createElement('input');
  inp.type='number'; inp.name='cantidad[]'; inp.step='0.001'; inp.min='0.001';
  inp.className='form-control form-control-sm'; inp.placeholder='Cantidad por unidad'; inp.required=true;
  colB.appendChild(inp);
  // Columna unidad label
  var colC=document.createElement('div'); colC.className='col-md-2';
  var uLbl=document.createElement('span'); uLbl.className='form-control-plaintext form-control-sm text-muted unidad-lbl';
  colC.appendChild(uLbl);
  // Columna eliminar
  var colD=document.createElement('div'); colD.className='col-md-2';
  var btn=document.createElement('button'); btn.type='button'; btn.className='btn btn-sm btn-outline-danger w-100';
  btn.innerHTML='<i class="bi bi-trash"></i>';
  btn.addEventListener('click',function(){this.closest('.item-row').remove();});
  colD.appendChild(btn);
  row.appendChild(colA); row.appendChild(colB); row.appendChild(colC); row.appendChild(colD);
  cont.appendChild(row);
}
// Bind existing delete buttons (for edit mode)
document.querySelectorAll('.item-row .btn-outline-danger').forEach(function(b){
  b.addEventListener('click',function(){this.closest('.item-row').remove();});
});
</script>{% endblock %}{% endblock %}"""

# =============================================================
# TEMPLATES V12 — Reservas de Producción
# =============================================================

# ── COTIZACIONES PROVEEDOR TEMPLATES ──────────────────────────
T['proveedores/cotizaciones.html'] = """{% extends 'base.html' %}
{% block title %}Cotizaciones Proveedores{% endblock %}{% block page_title %}Cotizaciones de Proveedores{% endblock %}
{% block topbar_actions %}
<a href="{{ url_for('cotizacion_proveedor_nueva', tipo='granel') }}" class="btn btn-outline-primary btn-sm"><i class="bi bi-plus-lg me-1"></i>Nueva Granel</a>
<a href="{{ url_for('cotizacion_proveedor_nueva', tipo='general') }}" class="btn btn-primary btn-sm"><i class="bi bi-plus-lg me-1"></i>Nueva General</a>
{% endblock %}
{% block content %}
<!-- Tabs Granel / General -->
<div class="d-flex gap-2 mb-3 flex-wrap align-items-center">
  <a href="{{ url_for('cotizaciones_proveedor') }}" class="btn btn-sm {{ 'btn-dark' if not tipo_f else 'btn-outline-secondary' }}">
    <i class="bi bi-grid me-1"></i>Todas <span class="badge bg-secondary ms-1">{{ (totales.granel or 0) + (totales.general or 0) }}</span></a>
  <a href="{{ url_for('cotizaciones_proveedor', tipo='granel') }}" class="btn btn-sm {{ 'btn-warning' if tipo_f=='granel' else 'btn-outline-warning' }}">
    <i class="bi bi-building me-1"></i>Granel <span class="badge bg-secondary ms-1">{{ totales.granel or 0 }}</span></a>
  <a href="{{ url_for('cotizaciones_proveedor', tipo='general') }}" class="btn btn-sm {{ 'btn-info' if tipo_f=='general' else 'btn-outline-info' }}">
    <i class="bi bi-box me-1"></i>General <span class="badge bg-secondary ms-1">{{ totales.general or 0 }}</span></a>
  <span class="text-muted ms-2" style="font-size:.8rem">|</span>
  {% for est,lbl in [('','Todos estados'),('vigente','Vigentes'),('vencida','Vencidas'),('en_revision','En revisión')] %}
  <a href="{{ url_for('cotizaciones_proveedor', tipo=tipo_f, estado=est) }}" class="btn btn-sm {{ 'btn-primary' if estado_f==est else 'btn-outline-secondary' }}">{{ lbl }}</a>{% endfor %}
</div>
<div class="tc"><div class="ch d-flex align-items-center gap-2">
  <i class="bi bi-file-earmark-check me-1"></i>{{ items|length }} cotización(es)
  {% if tipo_f %}<span class="badge {{ 'bg-warning text-dark' if tipo_f=='granel' else 'bg-info text-dark' }} ms-2">{{ tipo_f.title() }}</span>{% endif %}
</div>
{% if items %}<div class="table-responsive"><table class="table">
  <thead><tr><th>N°</th><th>Tipo</th><th>Producto / Servicio</th><th>Proveedor</th><th>P. Unit.</th><th>Plazo</th><th>Condición Pago</th><th>Vigencia</th><th>Estado</th><th></th></tr></thead>
  <tbody>{% for c in items %}<tr>
    <td><small class="fw-semibold text-muted">{{ c.numero or '—' }}</small></td>
    <td><span class="badge {{ 'bg-warning text-dark' if c.tipo_cotizacion=='granel' else 'bg-info text-dark' }}">{{ c.tipo_cotizacion.title() }}</span>
      {% if c.tipo_producto_servicio %}<div><small class="text-muted">{{ c.tipo_producto_servicio }}</small></div>{% endif %}</td>
    <td><span class="fw-semibold">{{ c.nombre_producto }}</span>{% if c.sku %}<div><small class="text-muted">SKU: {{ c.sku }}</small></div>{% endif %}</td>
    <td>{{ c.proveedor.empresa or c.proveedor.nombre if c.proveedor else '—' }}</td>
    <td class="fw-semibold">$ {{ '{:,.0f}'.format(c.precio_unitario).replace(',','.') }}
      <div><small class="text-muted">{{ c.unidades_minimas }} {{ c.unidad }} mín.</small></div></td>
    <td>{% if c.plazo_entrega_dias %}<span class="badge bg-light text-dark border">{{ c.plazo_entrega_dias }} días</span>{% else %}<small class="text-muted">{{ c.plazo_entrega or '—' }}</small>{% endif %}</td>
    <td>
      {% if c.condicion_pago_tipo == 'contado' %}<span class="badge bg-success">Contado</span>
      {% elif c.condicion_pago_tipo == 'credito' %}<span class="badge bg-primary">Crédito {{ c.condicion_pago_dias }}d</span>
      {% elif c.condicion_pago_tipo == 'anticipo_saldo' %}<span class="badge bg-warning text-dark">{{ c.anticipo_porcentaje|int }}% anticipo</span>
      {% elif c.condicion_pago_tipo == 'consignacion' %}<span class="badge bg-info text-dark">Consignación</span>
      {% else %}<small class="text-muted">{{ c.condiciones_pago or '—' }}</small>{% endif %}
    </td>
    <td>{% if c.vigencia %}<small class="{{ 'text-danger fw-semibold' if c.vigencia < now.date() else 'text-muted' }}">{{ c.vigencia.strftime('%d/%m/%Y') }}</small>{% else %}—{% endif %}</td>
    <td><span class="b b-{{ c.estado }}">{{ c.estado.replace('_',' ').title() }}</span></td>
    <td><div class="d-flex gap-1">
      <a href="{{ url_for('cotizacion_proveedor_editar', id=c.id) }}" class="btn btn-sm btn-outline-secondary"><i class="bi bi-pencil"></i></a>
      <form method="POST" action="{{ url_for('cotizacion_proveedor_eliminar', id=c.id) }}" onsubmit="return confirm('¿Eliminar?')">
        <button class="btn btn-sm btn-outline-danger"><i class="bi bi-trash"></i></button></form>
    </div></td>
  </tr>{% endfor %}</tbody>
</table></div>
{% else %}<div class="text-center text-muted py-5"><i class="bi bi-file-earmark-check" style="font-size:3rem"></i>
  <p class="mt-3">Sin cotizaciones{% if tipo_f %} de tipo <strong>{{ tipo_f }}</strong>{% endif %}.</p>
  <a href="{{ url_for('cotizacion_proveedor_nueva', tipo=tipo_f or 'general') }}" class="btn btn-primary btn-sm">Crear primera cotización</a></div>
{% endif %}</div>{% endblock %}"""

T['proveedores/cotizacion_form.html'] = """{% extends 'base.html' %}
{% block title %}{{ titulo }}{% endblock %}{% block page_title %}{{ titulo }}{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('cotizaciones_proveedor') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Volver</a>{% endblock %}
{% block content %}<div class="fc" style="max-width:960px"><form method="POST">
<!-- Tipo de cotización -->
<div class="row g-3 mb-3">
  <div class="col-12">
    <label class="form-label fw-bold">Tipo de cotización *</label>
    <div class="d-flex gap-3 mt-1" id="tipoBtns">
      <label class="d-flex align-items-center gap-2 p-3 rounded border cursor-pointer tipo-opt {{ 'border-warning bg-warning bg-opacity-10' if (obj and obj.tipo_cotizacion=='granel') or (not obj and tipo_default=='granel') else '' }}" style="cursor:pointer;flex:1;max-width:240px" onclick="setTipo('granel',this)">
        <input type="radio" name="tipo_cotizacion" value="granel" class="d-none" {% if (obj and obj.tipo_cotizacion=='granel') or (not obj and tipo_default=='granel') %}checked{% endif %}>
        <i class="bi bi-building text-warning" style="font-size:1.4rem"></i>
        <div><div class="fw-semibold">Granel / Materia Prima</div><small class="text-muted">Compras de materias primas a granel o grandes volúmenes</small></div>
      </label>
      <label class="d-flex align-items-center gap-2 p-3 rounded border cursor-pointer tipo-opt {{ 'border-info bg-info bg-opacity-10' if (obj and obj.tipo_cotizacion=='general') or (not obj and tipo_default=='general') else '' }}" style="cursor:pointer;flex:1;max-width:240px" onclick="setTipo('general',this)">
        <input type="radio" name="tipo_cotizacion" value="general" class="d-none" {% if (obj and obj.tipo_cotizacion=='general') or (not obj and tipo_default=='general') %}checked{% endif %}>
        <i class="bi bi-box text-info" style="font-size:1.4rem"></i>
        <div><div class="fw-semibold">General / Servicio</div><small class="text-muted">Productos terminados, servicios, insumos, empaques, etc.</small></div>
      </label>
    </div>
  </div>
  <div class="col-md-4"><label class="form-label">Tipo de producto / servicio</label>
    <input type="text" name="tipo_producto_servicio" class="form-control" placeholder="Ej: materia prima, empaque, logística..." value="{{ obj.tipo_producto_servicio if obj else '' }}"></div>
  <div class="col-md-5"><label class="form-label">Nombre del producto / ítem *</label>
    <input type="text" name="nombre_producto" class="form-control" value="{{ obj.nombre_producto if obj else '' }}" required></div>
  <div class="col-md-3"><label class="form-label">SKU</label>
    <input type="text" name="sku" class="form-control" value="{{ obj.sku if obj else '' }}"></div>
  <div class="col-12"><label class="form-label">Descripción</label>
    <input type="text" name="descripcion" class="form-control" value="{{ obj.descripcion if obj else '' }}"></div>
</div>
<hr class="my-3">
<h6 class="text-muted text-uppercase mb-3" style="font-size:.75rem;letter-spacing:1px">Proveedor y precio</h6>
<div class="row g-3 mb-3">
  <div class="col-md-6"><label class="form-label">Proveedor</label>
    <select name="proveedor_id" class="form-select">
      <option value="">— Sin proveedor —</option>
      {% for pv in proveedores_list %}<option value="{{ pv.id }}" {% if obj and obj.proveedor_id==pv.id %}selected{% endif %}>{{ pv.empresa or pv.nombre }}</option>{% endfor %}
    </select></div>
  <div class="col-md-3"><label class="form-label">Precio unitario COP</label>
    <div class="input-group"><span class="input-group-text">$</span>
      <input type="number" name="precio_unitario" class="form-control" step="1" min="0" value="{{ obj.precio_unitario|int if obj else '0' }}"></div></div>
  <div class="col-md-3"><label class="form-label">Unidades mínimas</label>
    <div class="input-group">
      <input type="number" name="unidades_minimas" class="form-control" min="1" value="{{ obj.unidades_minimas if obj else '1' }}">
      <select name="unidad" class="form-select" style="max-width:110px">
        {% for u in ['unidades','kg','g','litros','ml','piezas','cajas','metros'] %}
        <option value="{{ u }}" {% if obj and obj.unidad==u %}selected{% endif %}>{{ u }}</option>{% endfor %}
      </select></div></div>
</div>
<hr class="my-3">
<h6 class="text-muted text-uppercase mb-3" style="font-size:.75rem;letter-spacing:1px">Plazos y condiciones de pago</h6>
<div class="row g-3 mb-3">
  <div class="col-md-3"><label class="form-label">Plazo de entrega (días) *</label>
    <div class="input-group">
      <input type="number" name="plazo_entrega_dias" class="form-control" min="0" placeholder="0" value="{{ obj.plazo_entrega_dias if obj else '0' }}" id="plazoDias" oninput="updatePlazoText()">
      <span class="input-group-text">días</span></div>
    <small class="text-muted" id="plazoDiasLabel">Se calculará desde la fecha de recepción de OC</small></div>
  <div class="col-md-3"><label class="form-label">Descripción del plazo</label>
    <input type="text" name="plazo_entrega" class="form-control" placeholder="Ej: 5 días hábiles" value="{{ obj.plazo_entrega if obj else '' }}"></div>
  <div class="col-md-3"><label class="form-label">Condición de pago</label>
    <select name="condicion_pago_tipo" class="form-select" id="condPagoTipo" onchange="togglePagoFields()">
      {% for v,l in [('contado','Contado'),('credito','Crédito'),('anticipo_saldo','Anticipo + Saldo'),('consignacion','Consignación')] %}
      <option value="{{ v }}" {% if obj and obj.condicion_pago_tipo==v %}selected{% endif %}>{{ l }}</option>{% endfor %}
    </select></div>
  <div class="col-md-3" id="creditoDiasField" style="{{ 'display:none' if not (obj and obj.condicion_pago_tipo=='credito') else '' }}">
    <label class="form-label">Días de crédito</label>
    <div class="input-group"><input type="number" name="condicion_pago_dias" class="form-control" min="0" placeholder="30" value="{{ obj.condicion_pago_dias if obj else '30' }}"><span class="input-group-text">días</span></div></div>
  <div class="col-md-3" id="anticipoField" style="{{ 'display:none' if not (obj and obj.condicion_pago_tipo=='anticipo_saldo') else '' }}">
    <label class="form-label">% Anticipo</label>
    <div class="input-group"><input type="number" name="anticipo_porcentaje" class="form-control" min="0" max="100" placeholder="50" value="{{ obj.anticipo_porcentaje|int if obj else '50' }}"><span class="input-group-text">%</span></div></div>
  <div class="col-md-6"><label class="form-label">Notas condición de pago</label>
    <input type="text" name="condiciones_pago" class="form-control" placeholder="Notas adicionales..." value="{{ obj.condiciones_pago if obj else '' }}"></div>
</div>
<hr class="my-3">
<h6 class="text-muted text-uppercase mb-3" style="font-size:.75rem;letter-spacing:1px">Vigencia y estado</h6>
<div class="row g-3 mb-3">
  <div class="col-md-4"><label class="form-label">Fecha cotización</label>
    <input type="date" name="fecha_cotizacion" class="form-control" value="{{ obj.fecha_cotizacion.strftime('%Y-%m-%d') if obj and obj.fecha_cotizacion else today }}"></div>
  <div class="col-md-4"><label class="form-label">Vigencia hasta</label>
    <input type="date" name="vigencia" class="form-control" value="{{ obj.vigencia.strftime('%Y-%m-%d') if obj and obj.vigencia else '' }}"></div>
  <div class="col-md-4"><label class="form-label">Estado</label>
    <select name="estado" class="form-select">
      {% for est,lbl in [('vigente','Vigente'),('vencida','Vencida'),('en_revision','En revisión')] %}
      <option value="{{ est }}" {% if obj and obj.estado==est %}selected{% elif not obj and est=='vigente' %}selected{% endif %}>{{ lbl }}</option>{% endfor %}
    </select></div>
  <div class="col-12"><label class="form-label">Notas</label>
    <textarea name="notas" class="form-control" rows="2">{{ obj.notas if obj else '' }}</textarea></div>
</div>
<div class="d-flex gap-2 mt-4">
  <button type="submit" class="btn btn-primary"><i class="bi bi-check-lg me-1"></i>{{ 'Actualizar' if obj else 'Guardar Cotización' }}</button>
  <a href="{{ url_for('cotizaciones_proveedor') }}" class="btn btn-outline-secondary">Cancelar</a>
</div></form></div>
{% block scripts %}<script>
function setTipo(val, el){
  document.querySelectorAll('.tipo-opt').forEach(function(e){
    e.classList.remove('border-warning','bg-warning','bg-opacity-10','border-info','bg-info');
  });
  if(val==='granel'){ el.classList.add('border-warning','bg-warning','bg-opacity-10'); }
  else { el.classList.add('border-info','bg-info','bg-opacity-10'); }
  el.querySelector('input[type=radio]').checked=true;
}
function togglePagoFields(){
  var t=document.getElementById('condPagoTipo').value;
  document.getElementById('creditoDiasField').style.display = (t==='credito')?'':'none';
  document.getElementById('anticipoField').style.display = (t==='anticipo_saldo')?'':'none';
}
function updatePlazoText(){
  var d=parseInt(document.getElementById('plazoDias').value)||0;
  var msg = d>0 ? 'Al recibir la OC, se agendará la entrega en '+d+' días en el calendario.' : 'Se calculará desde la fecha de recepción de OC';
  document.getElementById('plazoDiasLabel').textContent=msg;
}
togglePagoFields();
updatePlazoText();
</script>{% endblock %}{% endblock %}"""

# ── LEGAL TEMPLATES ────────────────────────────────────────────
T['legal/index.html'] = """{% extends 'base.html' %}
{% block title %}Legal{% endblock %}{% block page_title %}Módulo Legal{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('legal_nuevo') }}" class="btn btn-primary btn-sm"><i class="bi bi-plus-lg me-1"></i>Nuevo documento</a>{% endblock %}
{% block content %}
{% if alertas %}
<div class="alert alert-warning mb-3" style="border-radius:10px;border:none">
  <i class="bi bi-exclamation-triangle-fill me-2"></i><strong>{{ alertas|length }} documento(s) próximo(s) a vencer:</strong>
  {% for a in alertas %}<span class="badge bg-danger ms-2">{{ a.titulo }}</span>{% endfor %}
</div>
{% endif %}
<div class="mb-3 d-flex gap-2 flex-wrap">
  <a href="{{ url_for('legal_index') }}" class="btn btn-sm {{ 'btn-primary' if not tipo_f else 'btn-outline-secondary' }}">Todos</a>
  {% for t,l in [('permiso_sanitario','Permisos'),('registro_invima','INVIMA'),('nso','NSO'),('contrato','Contratos'),('licencia','Licencias'),('certificado','Certificados'),('otro','Otros')] %}
  <a href="{{ url_for('legal_index', tipo=t) }}" class="btn btn-sm {{ 'btn-primary' if tipo_f==t else 'btn-outline-secondary' }}">{{ l }}</a>{% endfor %}
</div>
<div class="tc"><div class="ch"><i class="bi bi-shield-check me-2"></i>{{ items|length }} documento(s)</div>
{% if items %}<div class="table-responsive"><table class="table">
  <thead><tr><th>Tipo</th><th>Título</th><th>N°</th><th>Entidad</th><th>Emisión</th><th>Vencimiento</th><th>Estado</th><th></th></tr></thead>
  <tbody>{% for d in items %}
  {% set dias_venc = (d.fecha_vencimiento - now.date()).days if d.fecha_vencimiento else 999 %}
  <tr class="{{ 'table-danger' if dias_venc < 0 and d.estado=='vigente' else ('table-warning' if dias_venc <= d.recordatorio_dias and d.estado=='vigente' else '') }}">
    <td><span class="badge bg-light text-dark border">{{ d.tipo.replace('_',' ').title() }}</span></td>
    <td class="fw-semibold">{{ d.titulo }}</td>
    <td><small class="text-muted">{{ d.numero or '—' }}</small></td>
    <td><small>{{ d.entidad or '—' }}</small></td>
    <td><small>{{ d.fecha_emision.strftime('%d/%m/%Y') if d.fecha_emision else '—' }}</small></td>
    <td>
      {% if d.fecha_vencimiento %}
        <small class="{{ 'text-danger fw-semibold' if dias_venc < 0 else ('text-warning fw-semibold' if dias_venc <= d.recordatorio_dias else 'text-muted') }}">
          {{ d.fecha_vencimiento.strftime('%d/%m/%Y') }}
          {% if dias_venc >= 0 and dias_venc <= d.recordatorio_dias %}
            <span class="badge bg-warning text-dark ms-1">{{ dias_venc }}d</span>
          {% elif dias_venc < 0 %}
            <span class="badge bg-danger ms-1">VENCIDO</span>
          {% endif %}
        </small>
      {% else %}—{% endif %}
    </td>
    <td>
      {% if d.estado=='vigente' %}<span class="badge bg-success">Vigente</span>
      {% elif d.estado=='vencido' %}<span class="badge bg-danger">Vencido</span>
      {% elif d.estado=='en_tramite' %}<span class="badge bg-warning text-dark">En trámite</span>
      {% else %}<span class="badge bg-secondary">{{ d.estado.title() }}</span>{% endif %}
    </td>
    <td><div class="d-flex gap-1">
      <a href="{{ url_for('legal_editar', id=d.id) }}" class="btn btn-sm btn-outline-secondary"><i class="bi bi-pencil"></i></a>
      {% if d.archivo_url %}<a href="{{ d.archivo_url }}" target="_blank" class="btn btn-sm btn-outline-info" title="Ver archivo"><i class="bi bi-file-earmark-arrow-down"></i></a>{% endif %}
      <form method="POST" action="{{ url_for('legal_eliminar', id=d.id) }}" onsubmit="return confirm('¿Eliminar?')">
        <button class="btn btn-sm btn-outline-danger"><i class="bi bi-trash"></i></button></form>
    </div></td>
  </tr>{% endfor %}</tbody>
</table></div>
{% else %}<div class="text-center text-muted py-5"><i class="bi bi-shield-check" style="font-size:3rem"></i>
  <p class="mt-3">Sin documentos legales.</p><a href="{{ url_for('legal_nuevo') }}" class="btn btn-primary">Agregar</a></div>
{% endif %}</div>{% endblock %}"""

T['legal/form.html'] = """{% extends 'base.html' %}
{% block title %}{{ titulo }}{% endblock %}{% block page_title %}{{ titulo }}{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('legal_index') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Volver</a>{% endblock %}
{% block content %}<div class="fc" style="max-width:900px"><form method="POST">
<div class="row g-3">
  <div class="col-md-4"><label class="form-label">Tipo de documento *</label>
    <select name="tipo" class="form-select" required>
      {% for t,l in [('permiso_sanitario','Permiso Sanitario'),('registro_invima','Registro INVIMA'),('nso','NSO'),('contrato','Contrato'),('licencia','Licencia'),('certificado','Certificado'),('otro','Otro')] %}
      <option value="{{ t }}" {% if obj and obj.tipo==t %}selected{% endif %}>{{ l }}</option>{% endfor %}
    </select></div>
  <div class="col-md-8"><label class="form-label">Título *</label>
    <input type="text" name="titulo" class="form-control" value="{{ obj.titulo if obj else '' }}" required></div>
  <div class="col-md-4"><label class="form-label">Número del documento</label>
    <input type="text" name="numero" class="form-control" placeholder="Número de resolución, contrato..." value="{{ obj.numero if obj else '' }}"></div>
  <div class="col-md-4"><label class="form-label">Entidad emisora</label>
    <input type="text" name="entidad" class="form-control" placeholder="INVIMA, Secretaría, etc." value="{{ obj.entidad if obj else '' }}"></div>
  <div class="col-md-4"><label class="form-label">Estado</label>
    <select name="estado" class="form-select">
      {% for est,lbl in [('vigente','Vigente'),('en_tramite','En trámite'),('vencido','Vencido'),('suspendido','Suspendido')] %}
      <option value="{{ est }}" {% if obj and obj.estado==est %}selected{% elif not obj and est=='vigente' %}selected{% endif %}>{{ lbl }}</option>{% endfor %}
    </select></div>
  <div class="col-md-4"><label class="form-label">Fecha de emisión</label>
    <input type="date" name="fecha_emision" class="form-control" value="{{ obj.fecha_emision.strftime('%Y-%m-%d') if obj and obj.fecha_emision else '' }}"></div>
  <div class="col-md-4"><label class="form-label">Fecha de vencimiento</label>
    <input type="date" name="fecha_vencimiento" class="form-control" value="{{ obj.fecha_vencimiento.strftime('%Y-%m-%d') if obj and obj.fecha_vencimiento else '' }}"></div>
  <div class="col-md-4"><label class="form-label">Recordatorio (días antes de vencer)</label>
    <input type="number" name="recordatorio_dias" class="form-control" min="1" max="365" value="{{ obj.recordatorio_dias if obj else '30' }}"></div>
  <div class="col-12"><label class="form-label">Descripción</label>
    <textarea name="descripcion" class="form-control" rows="2">{{ obj.descripcion if obj else '' }}</textarea></div>
  <div class="col-12"><label class="form-label">URL del archivo (Google Drive, Dropbox, etc.)</label>
    <input type="url" name="archivo_url" class="form-control" placeholder="https://..." value="{{ obj.archivo_url if obj else '' }}"></div>
  <div class="col-12"><label class="form-label">Notas</label>
    <textarea name="notas" class="form-control" rows="2">{{ obj.notas if obj else '' }}</textarea></div>
</div>
<div class="d-flex gap-2 mt-4">
  <button type="submit" class="btn btn-primary"><i class="bi bi-check-lg me-1"></i>{{ 'Actualizar' if obj else 'Crear Documento' }}</button>
  <a href="{{ url_for('legal_index') }}" class="btn btn-outline-secondary">Cancelar</a>
</div></form></div>{% endblock %}"""

# ── CONTABLE TEMPLATES ─────────────────────────────────────────
T['contable/index.html'] = """{% extends 'base.html' %}
{% block title %}Contabilidad{% endblock %}{% block page_title %}Módulo Contable{% endblock %}
{% block topbar_actions %}
<div class="d-flex gap-2 flex-wrap">
  {% for m in meses_nav %}<a href="{{ url_for('contable_index', mes=m.val) }}" class="btn btn-sm {{ 'btn-primary' if mes_str==m.val else 'btn-outline-secondary' }}">{{ m.lbl }}</a>{% endfor %}
  <a href="{{ url_for('impuestos') }}" class="btn btn-sm btn-outline-warning ms-2"><i class="bi bi-percent me-1"></i>Reglas</a>
</div>
{% endblock %}
{% block content %}
{# ── Fila 1: indicadores principales ─────────────────── #}
<div class="row g-3 mb-3">
  <div class="col-6 col-md-3"><div class="fc text-center py-3">
    <div style="font-size:.72rem;color:#8898aa;text-transform:uppercase;letter-spacing:1px">Ingresos del mes</div>
    <div class="fw-bold mt-1" style="font-size:1.35rem;color:#2dce89">$ {{ '{:,.0f}'.format(total_ingresos).replace(',','.') }}</div>
    <small class="text-muted">{{ ventas_mes|length }} venta(s)</small>
  </div></div>
  <div class="col-6 col-md-3"><div class="fc text-center py-3">
    <div style="font-size:.72rem;color:#8898aa;text-transform:uppercase;letter-spacing:1px">Egresos del mes</div>
    <div class="fw-bold mt-1" style="font-size:1.35rem;color:#f5365c">$ {{ '{:,.0f}'.format(total_egresos).replace(',','.') }}</div>
    <small class="text-muted">Gastos + compras</small>
  </div></div>
  <div class="col-6 col-md-3"><div class="fc text-center py-3">
    <div style="font-size:.72rem;color:#8898aa;text-transform:uppercase;letter-spacing:1px">Utilidad bruta</div>
    <div class="fw-bold mt-1" style="font-size:1.35rem;color:{{ '#2dce89' if utilidad >= 0 else '#f5365c' }}">$ {{ '{:,.0f}'.format(utilidad).replace(',','.') }}</div>
    <small class="text-muted">Ingresos − Egresos</small>
  </div></div>
  <div class="col-6 col-md-3"><div class="fc text-center py-3">
    <div style="font-size:.72rem;color:#8898aa;text-transform:uppercase;letter-spacing:1px">Impuestos estimados</div>
    <div class="fw-bold mt-1" style="font-size:1.35rem;color:#e83e8c">$ {{ '{:,.0f}'.format(total_impuestos).replace(',','.') }}</div>
    <small class="text-muted">{{ detalle_impuestos|length }} regla(s) activa(s)</small>
  </div></div>
</div>
{# ── Fila 2: utilidad neta + otros ─────────────────────── #}
<div class="row g-3 mb-3">
  <div class="col-6 col-md-3"><div class="fc text-center py-3" style="border:2px solid {{ '#2dce89' if utilidad_neta >= 0 else '#f5365c' }}">
    <div style="font-size:.72rem;color:#8898aa;text-transform:uppercase;letter-spacing:1px">Utilidad neta</div>
    <div class="fw-bold mt-1" style="font-size:1.45rem;color:{{ '#2dce89' if utilidad_neta >= 0 else '#f5365c' }}">$ {{ '{:,.0f}'.format(utilidad_neta).replace(',','.') }}</div>
    <small class="text-muted">Utilidad − Impuestos</small>
  </div></div>
  <div class="col-6 col-md-3"><div class="fc text-center py-3">
    <div style="font-size:.72rem;color:#8898aa;text-transform:uppercase;letter-spacing:1px">Cuentas por cobrar</div>
    <div class="fw-bold mt-1" style="font-size:1.2rem;color:#fb6340">$ {{ '{:,.0f}'.format(total_cxc).replace(',','.') }}</div>
    <small class="text-muted">{{ cxc|length }} venta(s) pendientes</small>
  </div></div>
  <div class="col-6 col-md-3"><div class="fc text-center py-3">
    <div style="font-size:.72rem;color:#8898aa;text-transform:uppercase;letter-spacing:1px">Inventario valorizado</div>
    <div class="fw-bold mt-1" style="font-size:1.2rem;color:#5e72e4">$ {{ '{:,.0f}'.format(inventario_valor).replace(',','.') }}</div>
  </div></div>
  <div class="col-6 col-md-3"><div class="fc text-center py-3">
    <div style="font-size:.72rem;color:#8898aa;text-transform:uppercase;letter-spacing:1px">Anticipo cobrado</div>
    <div class="fw-bold mt-1" style="font-size:1.2rem;color:#11cdef">$ {{ '{:,.0f}'.format(total_anticipo).replace(',','.') }}</div>
  </div></div>
</div>
{# ── Desglose de impuestos ──────────────────────────────── #}
{% if detalle_impuestos %}
<div class="tc mb-3">
  <div class="ch d-flex justify-content-between align-items-center">
    <span><i class="bi bi-percent me-2 text-danger"></i>Desglose de impuestos estimados — {{ anio }}/{{ '%02d'|format(mes) }}</span>
    <a href="{{ url_for('impuestos') }}" class="btn btn-sm btn-outline-secondary"><i class="bi bi-gear me-1"></i>Configurar reglas</a>
  </div>
  <div class="table-responsive">
  <table class="table table-sm mb-0">
    <thead><tr><th>Regla</th><th>Aplica a</th><th class="text-end">Base</th><th class="text-center">%</th><th class="text-end">Impuesto estimado</th></tr></thead>
    <tbody>
    {% for d in detalle_impuestos %}<tr>
      <td class="fw-semibold">{{ d.nombre }}</td>
      <td><small class="badge bg-light text-dark border">{{ d.base_label }}</small></td>
      <td class="text-end"><small>$ {{ '{:,.0f}'.format(d.base_monto).replace(',','.') }}</small></td>
      <td class="text-center"><small>{{ d.porcentaje }}%</small></td>
      <td class="text-end text-danger fw-semibold">$ {{ '{:,.0f}'.format(d.monto).replace(',','.') }}</td>
    </tr>{% endfor %}
    </tbody>
    <tfoot><tr class="table-light fw-bold">
      <td colspan="4" class="text-end">Total impuestos</td>
      <td class="text-end text-danger">$ {{ '{:,.0f}'.format(total_impuestos).replace(',','.') }}</td>
    </tr></tfoot>
  </table></div>
</div>
{% else %}
<div class="alert alert-warning mb-3" style="border-radius:10px;border:none">
  <i class="bi bi-exclamation-triangle me-2"></i>No hay reglas tributarias activas.
  <a href="{{ url_for('impuestos') }}" class="alert-link">Configurar reglas tributarias</a> para calcular impuestos automáticamente.
</div>
{% endif %}
{# ── Tablas de detalle ─────────────────────────────────── #}
<div class="row g-3">
  <div class="col-md-6"><div class="tc">
    <div class="ch"><i class="bi bi-graph-up me-2 text-success"></i>Ventas del mes</div>
    {% if ventas_mes %}<table class="table table-sm">
      <thead><tr><th>N°</th><th>Título</th><th>Estado</th><th class="text-end">Total</th></tr></thead>
      <tbody>{% for v in ventas_mes %}<tr>
        <td><small>{{ v.numero or '—' }}</small></td>
        <td><small class="fw-semibold">{{ v.titulo[:30] }}</small></td>
        <td><small><span class="b b-{{ v.estado }}" style="font-size:.7rem">{{ v.estado.replace('_',' ').title() }}</span></small></td>
        <td class="text-end"><small>$ {{ '{:,.0f}'.format(v.total).replace(',','.') }}</small></td>
      </tr>{% endfor %}</tbody>
    </table>
    {% else %}<div class="text-center text-muted py-3"><i class="bi bi-graph-up"></i> Sin ventas este mes</div>{% endif %}
  </div></div>
  <div class="col-md-6"><div class="tc">
    <div class="ch"><i class="bi bi-arrow-down-circle me-2 text-danger"></i>Cuentas por cobrar</div>
    {% if cxc %}<table class="table table-sm">
      <thead><tr><th>N°</th><th>Cliente</th><th class="text-end">Saldo</th></tr></thead>
      <tbody>{% for v in cxc|sort(attribute='saldo', reverse=True)|list %}{% if loop.index <= 8 %}<tr>
        <td><small>{{ v.numero or '—' }}</small></td>
        <td><small>{{ v.cliente.empresa or v.cliente.nombre if v.cliente else '—' }}</small></td>
        <td class="text-end text-danger fw-semibold"><small>$ {{ '{:,.0f}'.format(v.saldo).replace(',','.') }}</small></td>
      </tr>{% endif %}{% endfor %}</tbody>
    </table>
    {% else %}<div class="text-center text-muted py-3 text-success"><i class="bi bi-check-circle me-1"></i>Sin cuentas por cobrar</div>{% endif %}
  </div></div>
</div>
<div class="mt-3 d-flex gap-2 flex-wrap">
  <a href="{{ url_for('contable_ingresos') }}" class="btn btn-outline-success btn-sm"><i class="bi bi-graph-up me-1"></i>Ver todos los ingresos</a>
  <a href="{{ url_for('contable_egresos') }}" class="btn btn-outline-danger btn-sm"><i class="bi bi-graph-down me-1"></i>Ver todos los egresos</a>
  <a href="{{ url_for('contable_libro_diario') }}" class="btn btn-outline-primary btn-sm"><i class="bi bi-journal-text me-1"></i>Libro diario</a>
  <a href="{{ url_for('contable_exportar', tipo='ventas') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-download me-1"></i>Exportar ventas CSV</a>
  <a href="{{ url_for('contable_exportar', tipo='gastos') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-download me-1"></i>Exportar gastos CSV</a>
  <a href="{{ url_for('contable_exportar', tipo='compras') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-download me-1"></i>Exportar compras CSV</a>
</div>
{% endblock %}"""

T['contable/ingresos.html'] = """{% extends 'base.html' %}
{% block title %}Ingresos{% endblock %}{% block page_title %}Ingresos{% endblock %}
{% block topbar_actions %}
<a href="{{ url_for('contable_exportar', tipo='ventas', desde=desde_s, hasta=hasta_s) }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-download me-1"></i>CSV</a>
<a href="{{ url_for('contable_index') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Contable</a>
{% endblock %}
{% block content %}
<div class="tc mb-3"><div class="p-3"><form method="GET" class="row g-2 align-items-end">
  <div class="col-md-3"><label class="form-label mb-1" style="font-size:.8rem">Desde</label>
    <input type="date" name="desde" class="form-control form-control-sm" value="{{ desde_s }}"></div>
  <div class="col-md-3"><label class="form-label mb-1" style="font-size:.8rem">Hasta</label>
    <input type="date" name="hasta" class="form-control form-control-sm" value="{{ hasta_s }}"></div>
  <div class="col-auto"><button type="submit" class="btn btn-primary btn-sm">Filtrar</button>
    <a href="{{ url_for('contable_ingresos') }}" class="btn btn-outline-secondary btn-sm">Limpiar</a></div>
</form></div></div>
<div class="tc"><div class="ch d-flex justify-content-between align-items-center">
  <span><i class="bi bi-graph-up me-2 text-success"></i>{{ items|length }} venta(s)</span>
  <span class="fw-bold text-success">Total: $ {{ '{:,.0f}'.format(items|sum(attribute='total')).replace(',','.') }}</span>
</div>
{% if items %}<div class="table-responsive"><table class="table">
  <thead><tr><th>N°</th><th>Fecha</th><th>Cliente</th><th>Título</th><th class="text-end">Subtotal</th><th class="text-end">IVA</th><th class="text-end">Total</th><th class="text-end">Anticipo</th><th class="text-end">Saldo</th><th>Estado</th></tr></thead>
  <tbody>{% for v in items %}<tr>
    <td><small class="fw-semibold">{{ v.numero or '—' }}</small></td>
    <td><small>{{ v.creado_en.strftime('%d/%m/%Y') }}</small></td>
    <td><small>{{ v.cliente.empresa or v.cliente.nombre if v.cliente else '—' }}</small></td>
    <td>{{ v.titulo[:40] }}</td>
    <td class="text-end">$ {{ '{:,.0f}'.format(v.subtotal).replace(',','.') }}</td>
    <td class="text-end">$ {{ '{:,.0f}'.format(v.iva).replace(',','.') }}</td>
    <td class="text-end fw-semibold">$ {{ '{:,.0f}'.format(v.total).replace(',','.') }}</td>
    <td class="text-end text-success">$ {{ '{:,.0f}'.format(v.monto_anticipo).replace(',','.') }}</td>
    <td class="text-end {{ 'text-danger fw-semibold' if v.saldo > 0 else 'text-success' }}">$ {{ '{:,.0f}'.format(v.saldo).replace(',','.') }}</td>
    <td><span class="b b-{{ v.estado }}" style="font-size:.75rem">{{ v.estado.replace('_',' ').title() }}</span></td>
  </tr>{% endfor %}
  <tr class="table-light fw-bold">
    <td colspan="4">TOTAL</td>
    <td class="text-end">$ {{ '{:,.0f}'.format(items|sum(attribute='subtotal')).replace(',','.') }}</td>
    <td class="text-end">$ {{ '{:,.0f}'.format(items|sum(attribute='iva')).replace(',','.') }}</td>
    <td class="text-end text-success">$ {{ '{:,.0f}'.format(items|sum(attribute='total')).replace(',','.') }}</td>
    <td class="text-end">$ {{ '{:,.0f}'.format(items|sum(attribute='monto_anticipo')).replace(',','.') }}</td>
    <td class="text-end text-danger">$ {{ '{:,.0f}'.format(items|sum(attribute='saldo')).replace(',','.') }}</td>
    <td></td>
  </tr>
  </tbody>
</table></div>{% else %}<div class="text-center text-muted py-5">Sin ingresos en el período.</div>{% endif %}</div>{% endblock %}"""

T['contable/egresos.html'] = """{% extends 'base.html' %}
{% block title %}Egresos{% endblock %}{% block page_title %}Egresos{% endblock %}
{% block topbar_actions %}
<a href="{{ url_for('contable_index') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Contable</a>
{% endblock %}
{% block content %}
<div class="tc mb-3"><div class="p-3"><form method="GET" class="row g-2 align-items-end">
  <div class="col-md-3"><label class="form-label mb-1" style="font-size:.8rem">Desde</label>
    <input type="date" name="desde" class="form-control form-control-sm" value="{{ desde_s }}"></div>
  <div class="col-md-3"><label class="form-label mb-1" style="font-size:.8rem">Hasta</label>
    <input type="date" name="hasta" class="form-control form-control-sm" value="{{ hasta_s }}"></div>
  <div class="col-auto"><button type="submit" class="btn btn-primary btn-sm">Filtrar</button>
    <a href="{{ url_for('contable_egresos') }}" class="btn btn-outline-secondary btn-sm">Limpiar</a></div>
</form></div></div>

<div class="row g-3 mb-4">
  <div class="col-md-4"><div class="fc text-center py-2">
    <div style="font-size:.75rem;color:#8898aa;text-transform:uppercase">Total gastos</div>
    <div class="fw-bold text-danger" style="font-size:1.2rem">$ {{ '{:,.0f}'.format(gastos|sum(attribute='monto')).replace(',','.') }}</div>
  </div></div>
  <div class="col-md-4"><div class="fc text-center py-2">
    <div style="font-size:.75rem;color:#8898aa;text-transform:uppercase">Total compras</div>
    <div class="fw-bold text-danger" style="font-size:1.2rem">$ {{ '{:,.0f}'.format(compras|sum(attribute='costo_total')).replace(',','.') }}</div>
  </div></div>
  <div class="col-md-4"><div class="fc text-center py-2">
    <div style="font-size:.75rem;color:#8898aa;text-transform:uppercase">Total egresos</div>
    <div class="fw-bold text-danger" style="font-size:1.3rem">$ {{ '{:,.0f}'.format((gastos|sum(attribute='monto'))+(compras|sum(attribute='costo_total'))).replace(',','.') }}</div>
  </div></div>
</div>

<div class="tc mb-3">
  <div class="ch"><i class="bi bi-receipt me-2"></i>Gastos operativos ({{ gastos|length }})
    <a href="{{ url_for('contable_exportar', tipo='gastos', desde=desde_s, hasta=hasta_s) }}" class="btn btn-sm btn-outline-secondary ms-2"><i class="bi bi-download"></i> CSV</a>
  </div>
  {% if gastos %}<table class="table table-sm"><thead><tr><th>Fecha</th><th>Tipo</th><th>Descripción</th><th class="text-end">Monto</th></tr></thead>
  <tbody>{% for g in gastos %}<tr>
    <td><small>{{ g.fecha.strftime('%d/%m/%Y') }}</small></td>
    <td><small><span class="badge bg-light text-dark border">{{ g.tipo }}</span></small></td>
    <td><small>{{ g.descripcion or '—' }}</small></td>
    <td class="text-end text-danger">$ {{ '{:,.0f}'.format(g.monto).replace(',','.') }}</td>
  </tr>{% endfor %}</tbody></table>
  {% else %}<div class="text-muted text-center py-3">Sin gastos en el período.</div>{% endif %}
</div>

<div class="tc">
  <div class="ch"><i class="bi bi-cart me-2"></i>Compras de materia ({{ compras|length }})
    <a href="{{ url_for('contable_exportar', tipo='compras', desde=desde_s, hasta=hasta_s) }}" class="btn btn-sm btn-outline-secondary ms-2"><i class="bi bi-download"></i> CSV</a>
  </div>
  {% if compras %}<table class="table table-sm"><thead><tr><th>Fecha</th><th>Ítem</th><th>Proveedor</th><th>Cant.</th><th class="text-end">Total</th></tr></thead>
  <tbody>{% for c in compras %}<tr>
    <td><small>{{ c.fecha.strftime('%d/%m/%Y') }}</small></td>
    <td><small class="fw-semibold">{{ c.nombre_item }}</small></td>
    <td><small>{{ c.proveedor or '—' }}</small></td>
    <td><small>{{ c.cantidad }} {{ c.unidad }}</small></td>
    <td class="text-end text-danger">$ {{ '{:,.0f}'.format(c.costo_total).replace(',','.') }}</td>
  </tr>{% endfor %}</tbody></table>
  {% else %}<div class="text-muted text-center py-3">Sin compras en el período.</div>{% endif %}
</div>{% endblock %}"""

T['contable/libro_diario.html'] = """{% extends 'base.html' %}
{% block title %}Libro Diario{% endblock %}{% block page_title %}Libro Diario{% endblock %}
{% block topbar_actions %}
<a href="{{ url_for('contable_exportar', tipo='asientos') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-download me-1"></i>CSV</a>
<a href="{{ url_for('contable_index') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Contable</a>
{% endblock %}
{% block content %}
<div class="row g-4">
<div class="col-md-4">
  <div class="fc">
    <h6 class="text-muted text-uppercase mb-3" style="font-size:.72rem;letter-spacing:1px">Nuevo asiento manual</h6>
    <form method="POST">
      <div class="mb-2"><label class="form-label mb-1" style="font-size:.82rem">Fecha *</label>
        <input type="date" name="fecha" class="form-control form-control-sm" value="{{ today }}" required></div>
      <div class="mb-2"><label class="form-label mb-1" style="font-size:.82rem">Descripción *</label>
        <input type="text" name="descripcion" class="form-control form-control-sm" placeholder="Descripción del asiento" required></div>
      <div class="row g-2 mb-2">
        <div class="col-6"><label class="form-label mb-1" style="font-size:.82rem">Debe $</label>
          <input type="number" name="debe" class="form-control form-control-sm" step="1" min="0" value="0"></div>
        <div class="col-6"><label class="form-label mb-1" style="font-size:.82rem">Haber $</label>
          <input type="number" name="haber" class="form-control form-control-sm" step="1" min="0" value="0"></div>
      </div>
      <div class="mb-2"><label class="form-label mb-1" style="font-size:.82rem">Cuenta Debe</label>
        <input type="text" name="cuenta_debe" class="form-control form-control-sm" placeholder="Ej: Caja, Bancos..."></div>
      <div class="mb-2"><label class="form-label mb-1" style="font-size:.82rem">Cuenta Haber</label>
        <input type="text" name="cuenta_haber" class="form-control form-control-sm" placeholder="Ej: Ventas, IVA..."></div>
      <div class="mb-3"><label class="form-label mb-1" style="font-size:.82rem">Referencia</label>
        <input type="text" name="referencia" class="form-control form-control-sm" placeholder="Factura, N° venta..."></div>
      <button type="submit" class="btn btn-primary w-100"><i class="bi bi-check-lg me-1"></i>Registrar asiento</button>
    </form>
  </div>
</div>
<div class="col-md-8">
  <div class="tc">
    <div class="ch"><i class="bi bi-journal-text me-2"></i>{{ asientos|length }} asiento(s)</div>
    {% if asientos %}<table class="table table-sm">
      <thead><tr><th>N°</th><th>Fecha</th><th>Descripción</th><th>Cta. Debe</th><th>Cta. Haber</th><th class="text-end">Debe</th><th class="text-end">Haber</th></tr></thead>
      <tbody>{% for a in asientos %}<tr>
        <td><small class="fw-semibold text-muted">{{ a.numero or '—' }}</small></td>
        <td><small>{{ a.fecha.strftime('%d/%m/%Y') }}</small></td>
        <td><small>{{ a.descripcion[:40] }}</small></td>
        <td><small class="text-muted">{{ a.cuenta_debe or '—' }}</small></td>
        <td><small class="text-muted">{{ a.cuenta_haber or '—' }}</small></td>
        <td class="text-end"><small class="text-success">{% if a.debe > 0 %}$ {{ '{:,.0f}'.format(a.debe).replace(',','.') }}{% else %}—{% endif %}</small></td>
        <td class="text-end"><small class="text-danger">{% if a.haber > 0 %}$ {{ '{:,.0f}'.format(a.haber).replace(',','.') }}{% else %}—{% endif %}</small></td>
      </tr>{% endfor %}</tbody>
    </table>
    {% else %}<div class="text-center text-muted py-4">Sin asientos registrados.</div>{% endif %}
  </div>
</div>
</div>{% endblock %}"""

T['buscador.html'] = """{% extends 'base.html' %}
{% block title %}Buscador{% endblock %}{% block page_title %}Buscador Global{% endblock %}
{% block content %}
<div class="fc" style="max-width:900px">
<form method="GET" class="mb-4">
  <div class="input-group input-group-lg">
    <span class="input-group-text"><i class="bi bi-search"></i></span>
    <input type="text" name="q" class="form-control" placeholder="Buscar por producto, cliente, proveedor, NSO, número de venta, orden..." value="{{ q or '' }}" autofocus>
    <button type="submit" class="btn btn-primary">Buscar</button>
    {% if q %}<a href="{{ url_for('buscador') }}" class="btn btn-outline-secondary">Limpiar</a>{% endif %}
  </div>
</form>

{% if q %}
<div class="mb-2 text-muted" style="font-size:.85rem">Resultados para: <strong>{{ q }}</strong></div>

{# CLIENTES #}
{% if resultados.clientes %}
<div class="tc mb-3">
  <div class="ch"><i class="bi bi-people-fill me-2 text-primary"></i>Clientes ({{ resultados.clientes|length }})</div>
  <table class="table table-hover">
    <thead><tr><th>Empresa</th><th>NIT</th><th>Relación</th><th>Contacto</th><th></th></tr></thead>
    <tbody>{% for c in resultados.clientes %}<tr>
      <td><a href="{{ url_for('cliente_ver', id=c.id) }}" class="fw-semibold text-decoration-none">{{ c.empresa or c.nombre }}</a></td>
      <td><small>{{ c.nit or '—' }}</small></td>
      <td><span class="b b-{{ c.estado_relacion }}">{{ c.estado_relacion.replace('_',' ').title() }}</span></td>
      <td><small>{% if c.contactos %}{{ c.contactos[0].nombre }}{% else %}—{% endif %}</small></td>
      <td><a href="{{ url_for('cliente_ver', id=c.id) }}" class="btn btn-sm btn-outline-primary"><i class="bi bi-eye"></i></a></td>
    </tr>{% endfor %}</tbody>
  </table>
</div>
{% endif %}

{# PROVEEDORES #}
{% if resultados.proveedores %}
<div class="tc mb-3">
  <div class="ch"><i class="bi bi-truck me-2 text-info"></i>Proveedores ({{ resultados.proveedores|length }})</div>
  <table class="table table-hover">
    <thead><tr><th>Empresa</th><th>NIT</th><th>Categoría</th><th>Email</th></tr></thead>
    <tbody>{% for p in resultados.proveedores %}<tr>
      <td class="fw-semibold">{{ p.empresa or p.nombre }}</td>
      <td><small>{{ p.nit or '—' }}</small></td>
      <td><small>{{ p.categoria or '—' }}</small></td>
      <td><small>{{ p.email or '—' }}</small></td>
    </tr>{% endfor %}</tbody>
  </table>
</div>
{% endif %}

{# PRODUCTOS #}
{% if resultados.productos %}
<div class="tc mb-3">
  <div class="ch"><i class="bi bi-box-seam me-2 text-success"></i>Productos ({{ resultados.productos|length }})</div>
  <table class="table table-hover">
    <thead><tr><th>Nombre</th><th>SKU</th><th>NSO</th><th>Stock</th><th>Precio</th></tr></thead>
    <tbody>{% for p in resultados.productos %}<tr>
      <td class="fw-semibold">{{ p.nombre }}</td>
      <td><small>{{ p.sku or '—' }}</small></td>
      <td><small>{{ p.nso or '—' }}</small></td>
      <td>{{ p.stock }}</td>
      <td>$ {{ '{:,.0f}'.format(p.precio).replace(',','.') }}</td>
    </tr>{% endfor %}</tbody>
  </table>
</div>
{% endif %}

{# VENTAS #}
{% if resultados.ventas %}
<div class="tc mb-3">
  <div class="ch"><i class="bi bi-graph-up-arrow me-2 text-warning"></i>Ventas ({{ resultados.ventas|length }})</div>
  <table class="table table-hover">
    <thead><tr><th>N°</th><th>Título</th><th>Cliente</th><th>Total</th><th>Estado</th><th>Producción</th></tr></thead>
    <tbody>{% for v in resultados.ventas %}<tr>
      <td class="fw-semibold">{{ v.numero or '—' }}</td>
      <td>{{ v.titulo }}</td>
      <td>{{ v.cliente.empresa or v.cliente.nombre if v.cliente else '—' }}</td>
      <td>$ {{ '{:,.0f}'.format(v.total).replace(',','.') }}</td>
      <td><span class="b b-{{ v.estado }}">{{ v.estado.replace('_',' ').title() }}</span></td>
      <td>{% set ops = v.ordenes_produccion %}
        {% if ops %}
          {% set comp = ops|selectattr('estado','equalto','completado')|list|length %}
          <small>{{ comp }}/{{ ops|length }} completadas</small>
        {% else %}—{% endif %}
      </td>
    </tr>{% endfor %}</tbody>
  </table>
</div>
{% endif %}

{# ÓRDENES DE PRODUCCIÓN #}
{% if resultados.ordenes_prod %}
<div class="tc mb-3">
  <div class="ch"><i class="bi bi-gear me-2 text-danger"></i>Órdenes de Producción ({{ resultados.ordenes_prod|length }})</div>
  <table class="table table-hover">
    <thead><tr><th>ID</th><th>Producto</th><th>Cant.</th><th>Estado</th><th>Venta</th><th>Creada</th></tr></thead>
    <tbody>{% for o in resultados.ordenes_prod %}<tr>
      <td>#{{ o.id }}</td>
      <td class="fw-semibold">{{ o.producto.nombre }}</td>
      <td>{{ o.cantidad_producir|int }}</td>
      <td>
        {% if o.estado=='en_produccion' %}<span class="badge bg-primary">En producción</span>
        {% elif o.estado=='completado' %}<span class="badge bg-success">Completado</span>
        {% else %}<span class="badge bg-warning text-dark">Pendiente material</span>{% endif %}
      </td>
      <td>{% if o.venta %}{{ o.venta.numero or ('VNT-'+o.venta_id|string) }}{% else %}—{% endif %}</td>
      <td><small>{{ o.creado_en.strftime('%d/%m/%Y') }}</small></td>
    </tr>{% endfor %}</tbody>
  </table>
</div>
{% endif %}

{# ÓRDENES DE COMPRA #}
{% if resultados.ordenes_compra %}
<div class="tc mb-3">
  <div class="ch"><i class="bi bi-cart-check me-2"></i>Órdenes de Compra ({{ resultados.ordenes_compra|length }})</div>
  <table class="table table-hover">
    <thead><tr><th>N°</th><th>Proveedor</th><th>Estado</th><th>Total</th><th>Emisión</th></tr></thead>
    <tbody>{% for oc in resultados.ordenes_compra %}<tr>
      <td class="fw-semibold">{{ oc.numero or '—' }}</td>
      <td>{{ oc.proveedor.empresa or oc.proveedor.nombre if oc.proveedor else '—' }}</td>
      <td>{% if oc.estado=='recibida' %}<span class="badge bg-success">Recibida</span>
          {% elif oc.estado=='enviada' %}<span class="badge bg-primary">Enviada</span>
          {% elif oc.estado=='cancelada' %}<span class="badge bg-danger">Cancelada</span>
          {% else %}<span class="badge bg-secondary">Borrador</span>{% endif %}</td>
      <td>$ {{ '{:,.0f}'.format(oc.total).replace(',','.') }}</td>
      <td><small>{{ oc.fecha_emision.strftime('%d/%m/%Y') if oc.fecha_emision else '—' }}</small></td>
    </tr>{% endfor %}</tbody>
  </table>
</div>
{% endif %}

{% if not (resultados.clientes or resultados.proveedores or resultados.productos or resultados.ventas or resultados.ordenes_prod or resultados.ordenes_compra) %}
<div class="text-center text-muted py-5">
  <i class="bi bi-search" style="font-size:3rem"></i>
  <p class="mt-3">No se encontraron resultados para <strong>{{ q }}</strong></p>
</div>
{% endif %}

{% else %}
<div class="text-center text-muted py-5">
  <i class="bi bi-search" style="font-size:4rem;opacity:.3"></i>
  <p class="mt-3" style="font-size:1.1rem">Busca en toda la base de datos</p>
  <p class="text-muted" style="font-size:.88rem">Clientes, proveedores, productos, ventas (por número), órdenes de producción, órdenes de compra, NSO...</p>
</div>
{% endif %}
</div>{% endblock %}"""

T['produccion/reservas.html'] = """{% extends 'base.html' %}
{% block title %}Reservas de Producción{% endblock %}{% block page_title %}Reservas de Producción{% endblock %}
{% block topbar_actions %}
<a href="{{ url_for('reserva_nueva') }}" class="btn btn-primary btn-sm"><i class="bi bi-plus-lg me-1"></i>Nueva Reserva</a>
{% endblock %}
{% block content %}
<div>

{# ── Agrupar por venta ── #}
{% set reservas_con_venta = reservas|selectattr('venta_id')|list %}
{% set reservas_sin_venta = reservas|rejectattr('venta_id')|list %}

{# ── Ventas únicas ordenadas ── #}
{% set ventas_vistas = [] %}
{% for r in reservas_con_venta %}
  {% if r.venta_id not in ventas_vistas %}{% set _ = ventas_vistas.append(r.venta_id) %}{% endif %}
{% endfor %}

{% for vid in ventas_vistas %}
  {% set grupo = reservas_con_venta|selectattr('venta_id','equalto',vid)|list %}
  {% set venta_obj = grupo[0].venta %}
  {# Determinar si TODAS las materias están cubiertas #}
  {% set hay_faltante = namespace(v=false) %}
  {% for r in grupo %}{% if r.estado=='reservado' and r.materia.stock_disponible < r.cantidad %}{% set hay_faltante.v = true %}{% endif %}{% endfor %}
  {# Determinar si hay anticipo pagado #}
  {% set anticipo_ok = venta_obj and venta_obj.estado in ['anticipo_pagado','ganado'] %}

<div class="mb-4 border rounded overflow-hidden">
  <div class="d-flex align-items-center justify-content-between px-3 py-2" style="background:#f0f3ff;border-bottom:1px solid #dee2f0">
    <div>
      <span class="fw-bold" style="color:#1a1f36">
        {% if venta_obj %}
          <i class="bi bi-tag-fill me-1 text-primary"></i>Venta {{ venta_obj.numero or ('VNT-'+venta_obj.id|string) }} — {{ venta_obj.titulo }}
        {% else %}
          Venta #{{ vid }}
        {% endif %}
      </span>
      {% if anticipo_ok %}
        <span class="badge bg-success ms-2"><i class="bi bi-check-circle me-1"></i>Anticipo recibido</span>
      {% else %}
        <span class="badge bg-warning text-dark ms-2"><i class="bi bi-hourglass me-1"></i>Sin anticipo</span>
      {% endif %}
    </div>
    <div class="d-flex gap-2">
      {% if not hay_faltante.v %}
        {# Sin faltante: botón funcional para iniciar producción #}
        <form method="POST" action="{{ url_for('reserva_iniciar_produccion', venta_id=vid) }}" class="d-inline">
          <button type="submit" class="btn btn-sm btn-success"
                  onclick="return confirm('¿Iniciar producción para esta venta? Las órdenes cambiarán a En Producción.')">
            <i class="bi bi-play-circle me-1"></i>Iniciar producción
          </button>
        </form>
      {% endif %}
    </div>
  </div>
  <table class="table table-hover align-middle mb-0">
    <thead><tr>
      <th>Materia Prima</th>
      <th class="text-end">Reservado</th>
      <th class="text-end">Stock disp.</th>
      <th>Producto</th><th>Estado</th><th>Notas</th><th>Fecha</th><th></th>
    </tr></thead>
    <tbody>
    {% for r in grupo %}
    <tr>
      <td><strong>{{ r.materia.nombre }}</strong><br><small class="text-muted">{{ r.materia.unidad }}</small></td>
      <td class="text-end fw-semibold">{{ '%.2f'|format(r.cantidad) }}</td>
      <td class="text-end">
        {% set disp = r.materia.stock_disponible %}
        {% if disp < r.cantidad %}
          <span class="text-danger fw-semibold">{{ '%.2f'|format(disp) }}
            <i class="bi bi-exclamation-circle" title="Insuficiente"></i></span>
        {% else %}
          <span class="text-success">{{ '%.2f'|format(disp) }}</span>
        {% endif %}
      </td>
      <td>{{ r.producto.nombre if r.producto else '—' }}</td>
      <td>
        {% if r.estado == 'reservado' %}<span class="badge bg-primary">Reservado</span>
        {% elif r.estado == 'usado' %}<span class="badge bg-success">Usado</span>
        {% else %}<span class="badge bg-secondary">Cancelado</span>{% endif %}
      </td>
      <td><small class="text-muted">{{ r.notas or '' }}</small></td>
      <td><small>{{ r.creado_en.strftime('%d/%m/%Y') }}</small></td>
      <td>
        {% if r.estado == 'reservado' %}
        <div class="d-flex gap-1 flex-wrap">
          <form method="POST" action="{{ url_for('reserva_cancelar', id=r.id) }}" class="d-inline"
                onsubmit="return confirm('¿Cancelar esta reserva?')">
            <button class="btn btn-sm btn-outline-warning">Cancelar</button>
          </form>
          {% set faltante_real = r.cantidad - r.materia.stock_disponible %}
          {% if faltante_real > 0 %}
          <button type="button" class="btn btn-sm btn-outline-danger"
            data-bs-toggle="modal" data-bs-target="#modalFalta"
            data-reserva="{{ r.id }}"
            data-materia="{{ r.materia.nombre }}"
            data-unidad="{{ r.materia.unidad }}"
            data-reservado="{{ '%.2f'|format(r.cantidad) }}"
            data-disponible="{{ '%.2f'|format(r.materia.stock_disponible) }}"
            data-faltante="{{ '%.2f'|format(faltante_real) }}"
            onclick="prepFalta(this)">
            <i class="bi bi-exclamation-triangle me-1"></i>Falta {{ '%.2f'|format(faltante_real) }} {{ r.materia.unidad }}
          </button>
          {% endif %}
        </div>
        {% endif %}
      </td>
    </tr>
    {% endfor %}
    </tbody>
  </table>
</div>
{% endfor %}

{# Reservas sin venta asociada #}
{% if reservas_sin_venta %}
<div class="mb-4 border rounded overflow-hidden">
  <div class="px-3 py-2" style="background:#f8f9fe;border-bottom:1px solid #dee2e6">
    <span class="fw-semibold text-muted"><i class="bi bi-question-circle me-1"></i>Sin venta asociada</span>
  </div>
  <table class="table table-hover align-middle mb-0">
    <thead><tr>
      <th>Materia Prima</th><th class="text-end">Reservado</th><th class="text-end">Stock disp.</th>
      <th>Producto</th><th>Estado</th><th>Notas</th><th>Fecha</th><th></th>
    </tr></thead>
    <tbody>
    {% for r in reservas_sin_venta %}
    <tr>
      <td><strong>{{ r.materia.nombre }}</strong><br><small class="text-muted">{{ r.materia.unidad }}</small></td>
      <td class="text-end fw-semibold">{{ '%.2f'|format(r.cantidad) }}</td>
      <td class="text-end">
        {% set disp = r.materia.stock_disponible %}
        {% if disp < r.cantidad %}
          <span class="text-danger fw-semibold">{{ '%.2f'|format(disp) }} <i class="bi bi-exclamation-circle"></i></span>
        {% else %}
          <span class="text-success">{{ '%.2f'|format(disp) }}</span>
        {% endif %}
      </td>
      <td>{{ r.producto.nombre if r.producto else '—' }}</td>
      <td>
        {% if r.estado == 'reservado' %}<span class="badge bg-primary">Reservado</span>
        {% elif r.estado == 'usado' %}<span class="badge bg-success">Usado</span>
        {% else %}<span class="badge bg-secondary">Cancelado</span>{% endif %}
      </td>
      <td><small class="text-muted">{{ r.notas or '' }}</small></td>
      <td><small>{{ r.creado_en.strftime('%d/%m/%Y') }}</small></td>
      <td>
        {% if r.estado == 'reservado' %}
        <div class="d-flex gap-1 flex-wrap">
          <form method="POST" action="{{ url_for('reserva_cancelar', id=r.id) }}" class="d-inline"
                onsubmit="return confirm('¿Cancelar esta reserva?')">
            <button class="btn btn-sm btn-outline-warning">Cancelar</button>
          </form>
          {% set faltante_real = r.cantidad - r.materia.stock_disponible %}
          {% if faltante_real > 0 %}
          <button type="button" class="btn btn-sm btn-outline-danger"
            data-bs-toggle="modal" data-bs-target="#modalFalta"
            data-reserva="{{ r.id }}"
            data-materia="{{ r.materia.nombre }}"
            data-unidad="{{ r.materia.unidad }}"
            data-reservado="{{ '%.2f'|format(r.cantidad) }}"
            data-disponible="{{ '%.2f'|format(r.materia.stock_disponible) }}"
            data-faltante="{{ '%.2f'|format(faltante_real) }}"
            onclick="prepFalta(this)">
            <i class="bi bi-exclamation-triangle me-1"></i>Falta {{ '%.2f'|format(faltante_real) }} {{ r.materia.unidad }}
          </button>
          {% endif %}
        </div>
        {% endif %}
      </td>
    </tr>
    {% endfor %}
    </tbody>
  </table>
</div>
{% endif %}

{% if not reservas %}
<div class="text-center text-muted py-5"><i class="bi bi-layers" style="font-size:3rem"></i>
  <p class="mt-3">No hay reservas registradas.</p></div>
{% endif %}

</div>

<!-- Modal: Falta Material -->
<div class="modal fade" id="modalFalta" tabindex="-1">
  <div class="modal-dialog">
    <div class="modal-content">
      <form method="POST" action="{{ url_for('reserva_solicitar_compra') }}">
        <input type="hidden" name="reserva_id" id="faltaReservaId">
        <div class="modal-header">
          <h5 class="modal-title"><i class="bi bi-exclamation-triangle text-danger me-2"></i>Solicitar compra de material</h5>
          <button type="button" class="btn-close" data-bs-dismiss="modal"></button>
        </div>
        <div class="modal-body">
          <div class="p-3 rounded mb-3" style="background:#fff4f4;border:1px solid #f5c2c7">
            <div class="fw-semibold mb-1" id="faltaMateriaNombre"></div>
            <div class="row g-2 text-center mt-1">
              <div class="col-4">
                <div style="font-size:.72rem;color:#6c757d">Reservado</div>
                <div class="fw-bold" id="faltaReservado"></div>
              </div>
              <div class="col-4">
                <div style="font-size:.72rem;color:#6c757d">Disponible</div>
                <div class="fw-bold text-warning" id="faltaDisponible"></div>
              </div>
              <div class="col-4">
                <div style="font-size:.72rem;color:#6c757d">Faltante</div>
                <div class="fw-bold text-danger" id="faltaCantMostrar"></div>
              </div>
            </div>
          </div>
          <div class="mb-3">
            <label class="form-label fw-semibold">Cantidad a comprar *</label>
            <div class="input-group">
              <input type="number" name="cantidad_faltante" id="faltaCantInput"
                class="form-control" step="0.01" min="0.01" required>
              <span class="input-group-text" id="faltaUnidad"></span>
            </div>
          </div>
          <div class="mb-3">
            <label class="form-label fw-semibold">Asignar tareas a:</label>
            <select name="usuario_id" class="form-select" required>
              <option value="">Seleccionar usuario...</option>
              {% for u in usuarios %}
              <option value="{{ u.id }}">{{ u.nombre }} ({{ u.rol }})</option>
              {% endfor %}
            </select>
          </div>
          <div class="mb-3">
            <label class="form-label">Notas adicionales</label>
            <textarea name="descripcion" class="form-control" rows="2"
              placeholder="Proveedor sugerido, urgencia, precio estimado..."></textarea>
          </div>
          <div class="alert alert-info py-2 mb-0" style="font-size:.82rem">
            <i class="bi bi-info-circle me-1"></i>Se crearán dos tareas pareadas:
            <strong>Comprar materias</strong> + <strong>Verificar abono</strong>.
          </div>
        </div>
        <div class="modal-footer">
          <button type="button" class="btn btn-outline-secondary" data-bs-dismiss="modal">Cancelar</button>
          <button type="submit" class="btn btn-danger"><i class="bi bi-check-lg me-1"></i>Crear tareas</button>
        </div>
      </form>
    </div>
  </div>
</div>
{% block scripts %}<script>
function prepFalta(btn){
  document.getElementById('faltaReservaId').value   = btn.dataset.reserva;
  document.getElementById('faltaMateriaNombre').textContent = btn.dataset.materia;
  document.getElementById('faltaReservado').textContent    = btn.dataset.reservado + ' ' + btn.dataset.unidad;
  document.getElementById('faltaDisponible').textContent   = btn.dataset.disponible + ' ' + btn.dataset.unidad;
  document.getElementById('faltaCantMostrar').textContent  = btn.dataset.faltante + ' ' + btn.dataset.unidad;
  document.getElementById('faltaCantInput').value          = btn.dataset.faltante;
  document.getElementById('faltaUnidad').textContent       = btn.dataset.unidad;
}
</script>{% endblock %}
{% endblock %}"""

T['produccion/reserva_form.html'] = """{% extends 'base.html' %}
{% block title %}Nueva Reserva{% endblock %}{% block page_title %}Nueva Reserva de Producción{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('reservas') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Volver</a>{% endblock %}
{% block content %}<div class="fc"><form method="POST"><div class="row g-3">
  <div class="col-md-6"><label class="form-label">Materia Prima *</label>
    <select name="materia_prima_id" class="form-select" required>
      <option value="">Seleccionar...</option>
      {% for m in materias %}
      <option value="{{ m.id }}">{{ m.nombre }} — disponible: {{ '%.2f'|format(m.stock_disponible) }} {{ m.unidad }}</option>
      {% endfor %}
    </select></div>
  <div class="col-md-3"><label class="form-label">Cantidad a reservar *</label>
    <input type="number" name="cantidad" step="0.001" min="0.001" class="form-control" required></div>
  <div class="col-md-3"><label class="form-label">Producto destino</label>
    <select name="producto_id" class="form-select">
      <option value="">Sin especificar</option>
      {% for p in productos %}
      <option value="{{ p.id }}">{{ p.nombre }}</option>
      {% endfor %}
    </select></div>
  <div class="col-12"><label class="form-label">Notas</label>
    <textarea name="notas" class="form-control" rows="2" placeholder="Descripción de la producción planificada..."></textarea></div>
</div>
<div class="d-flex gap-2 mt-4">
  <button type="submit" class="btn btn-primary"><i class="bi bi-check-lg me-1"></i>Crear Reserva</button>
  <a href="{{ url_for('reservas') }}" class="btn btn-outline-secondary">Cancelar</a>
</div></form></div>{% endblock %}"""

T['inventario/ingresos.html'] = """{% extends 'base.html' %}
{% block title %}Multi-Ingreso de Productos{% endblock %}{% block page_title %}Multi-Ingreso al Inventario{% endblock %}
{% block topbar_actions %}<a href="{{ url_for('inventario') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Volver</a>{% endblock %}
{% block content %}
<div class="fc" style="max-width:900px">
<p class="text-muted mb-4">Ingresa múltiples unidades del mismo producto al inventario, identificando el número de lote para trazabilidad.</p>
<form method="POST">
<div class="row g-3 mb-4 p-3 border rounded" style="background:#f8f9fe">
  <div class="col-md-6">
    <label class="form-label fw-semibold">Producto *</label>
    <select name="producto_id" class="form-select" required>
      <option value="">Seleccionar producto...</option>
      {% for p in productos %}
      <option value="{{ p.id }}">{{ p.nombre }}{% if p.sku %} ({{ p.sku }}){% endif %} — stock: {{ p.stock }}</option>
      {% endfor %}
    </select>
  </div>
  <div class="col-md-3">
    <label class="form-label fw-semibold">Número de Lote *</label>
    <input type="text" name="numero_lote" class="form-control" placeholder="Ej: LOTE-2026-001" required>
  </div>
  <div class="col-md-3">
    <label class="form-label fw-semibold">NSO (INVIMA)</label>
    <input type="text" name="nso" class="form-control" placeholder="NSO-XXXX-XXXX">
  </div>
  <div class="col-md-3">
    <label class="form-label fw-semibold">Fecha de producción</label>
    <input type="date" name="fecha_produccion" class="form-control">
  </div>
  <div class="col-md-3">
    <label class="form-label fw-semibold">Fecha de vencimiento</label>
    <input type="date" name="fecha_vencimiento" class="form-control">
  </div>
</div>

<div class="d-flex justify-content-between align-items-center mb-3">
  <h6 class="fw-semibold text-uppercase text-muted" style="letter-spacing:1px;font-size:.75rem">Ingresos del lote</h6>
  <button type="button" class="btn btn-sm btn-outline-primary" onclick="addRow()"><i class="bi bi-plus-lg me-1"></i>Agregar fila</button>
</div>
<div id="rowsContainer">
  <!-- filas JS -->
</div>
<div class="d-flex gap-2 mt-4">
  <button type="submit" class="btn btn-primary"><i class="bi bi-check-lg me-1"></i>Registrar ingresos</button>
  <a href="{{ url_for('inventario') }}" class="btn btn-outline-secondary">Cancelar</a>
</div>
</form>
</div>
{% block scripts %}<script>
var rowIdx = 0;
function addRow(){
  rowIdx++;
  var cont = document.getElementById('rowsContainer');
  var div = document.createElement('div');
  div.className = 'row g-2 mb-2 ingreso-row';
  div.id = 'ir'+rowIdx;
  div.innerHTML =
    '<div class="col-md-3"><label class="form-label mb-1" style="font-size:.8rem">Cantidad *</label>' +
    '<input type="number" name="cantidad[]" step="0.01" min="0.01" class="form-control form-control-sm" placeholder="Unidades" required></div>' +
    '<div class="col-md-3"><label class="form-label mb-1" style="font-size:.8rem">Costo unitario COP</label>' +
    '<input type="number" name="costo_unit[]" step="1" min="0" class="form-control form-control-sm" placeholder="0"></div>' +
    '<div class="col-md-5"><label class="form-label mb-1" style="font-size:.8rem">Nota (opcional)</label>' +
    '<input type="text" name="nota_fila[]" class="form-control form-control-sm" placeholder="Ej: bodega norte, proveedor X..."></div>' +
    '<div class="col-md-1"><label class="form-label mb-1" style="font-size:.8rem"> </label>' +
    '<button type="button" class="btn btn-sm btn-outline-danger w-100" onclick="this.closest(&#39;.ingreso-row&#39;).remove()"><i class="bi bi-x"></i></button></div>';
  cont.appendChild(div);
}
addRow();
</script>{% endblock %}
{% endblock %}"""

T['produccion/ordenes.html'] = """{% extends 'base.html' %}
{% block title %}Órdenes de Producción{% endblock %}{% block page_title %}Órdenes de Producción{% endblock %}
{% block topbar_actions %}
<a href="{{ url_for('produccion_index') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Producción</a>
<a href="{{ url_for('gantt') }}" class="btn btn-outline-primary btn-sm"><i class="bi bi-bar-chart-steps me-1"></i>Diagrama Gantt</a>
{% endblock %}
{% block content %}
<div class="fc" style="max-width:100%">
{% if pendientes %}
<h6 class="fw-semibold text-uppercase text-muted mb-3" style="letter-spacing:1px;font-size:.75rem">En producción / Pendientes</h6>
<div class="table-responsive mb-4">
<table class="table table-hover align-middle">
  <thead><tr>
    <th>Producto</th><th class="text-end">A producir</th><th class="text-end">Stock previo</th>
    <th>Cotización</th><th>Lote</th><th>Estado</th><th>Fecha</th><th></th>
  </tr></thead>
  <tbody>
  {% for o in pendientes %}
  <tr>
    <td><strong>{{ o.producto.nombre }}</strong></td>
    <td class="text-end">{{ '%.2f'|format(o.cantidad_producir) }}</td>
    <td class="text-end text-muted">{{ '%.2f'|format(o.cantidad_stock) }}</td>
    <td>{% if o.cotizacion %}<small>#{{ o.cotizacion.numero or o.cotizacion_id }} {{ o.cotizacion.titulo[:30] }}...</small>{% else %}—{% endif %}</td>
    <td>{{ o.numero_lote or '—' }}</td>
    <td>
      {% if o.estado == 'pendiente_materiales' %}<span class="badge bg-warning">Sin materiales</span>
      {% elif o.estado == 'en_produccion' %}<span class="badge bg-primary">En producción</span>
      {% else %}<span class="badge bg-success">Completado</span>{% endif %}
    </td>
    <td><small>{{ o.creado_en.strftime('%d/%m/%Y') }}</small></td>
    <td>
      <button type="button" class="btn btn-sm btn-success"
        data-bs-toggle="modal" data-bs-target="#modalCompletar"
        data-orden="{{ o.id }}" data-prod="{{ o.producto.nombre }}"
        data-cant="{{ o.cantidad_producir }}"
        onclick="prepCompletar(this)">
        <i class="bi bi-check2-circle me-1"></i>Completar
      </button>
    </td>
  </tr>
  {% endfor %}
  </tbody>
</table>
</div>
{% else %}
<div class="text-center text-muted py-5">
  <i class="bi bi-check2-all fs-1 d-block mb-2 text-success"></i>
  No hay órdenes de producción en curso.
</div>
{% endif %}

{% if completados %}
<h6 class="fw-semibold text-uppercase text-muted mb-3 mt-4" style="letter-spacing:1px;font-size:.75rem">Completadas recientemente</h6>
<div class="table-responsive">
<table class="table table-sm align-middle">
  <thead><tr><th>Producto</th><th class="text-end">Producido</th><th>Lote asignado</th><th>Completado</th></tr></thead>
  <tbody>
  {% for o in completados %}
  <tr class="text-muted">
    <td>{{ o.producto.nombre }}</td>
    <td class="text-end">{{ '%.2f'|format(o.cantidad_producir) }}</td>
    <td>{{ o.numero_lote or '—' }}</td>
    <td><small>{{ o.completado_en.strftime('%d/%m/%Y %H:%M') if o.completado_en else '—' }}</small></td>
  </tr>
  {% endfor %}
  </tbody>
</table>
</div>
{% endif %}
</div>

<!-- Modal Completar Orden -->
<div class="modal fade" id="modalCompletar" tabindex="-1">
  <div class="modal-dialog">
    <div class="modal-content">
      <form method="POST" action="{{ url_for('orden_completar') }}">
        <input type="hidden" name="orden_id" id="compOrdenId">
        <div class="modal-header">
          <h5 class="modal-title"><i class="bi bi-check2-circle text-success me-2"></i>Completar producción</h5>
          <button type="button" class="btn-close" data-bs-dismiss="modal"></button>
        </div>
        <div class="modal-body">
          <p>Producto: <strong id="compProdNombre"></strong></p>
          <p class="text-muted">Cantidad producida: <span id="compCant"></span> unidades</p>
          <div class="mb-3">
            <label class="form-label fw-semibold">Número de lote (para inventario)</label>
            <input type="text" name="numero_lote" class="form-control" id="compLote"
              placeholder="Ej: LOTE-2026-001" required>
          </div>
          <div class="mb-3">
            <label class="form-label">Fecha de vencimiento (opcional)</label>
            <input type="date" name="fecha_vencimiento" class="form-control">
          </div>
          <div class="mb-0">
            <label class="form-label">Notas</label>
            <textarea name="notas" class="form-control" rows="2"
              placeholder="Observaciones de producción..."></textarea>
          </div>
        </div>
        <div class="modal-footer">
          <button type="button" class="btn btn-outline-secondary" data-bs-dismiss="modal">Cancelar</button>
          <button type="submit" class="btn btn-success"><i class="bi bi-check-lg me-1"></i>Mover al inventario</button>
        </div>
      </form>
    </div>
  </div>
</div>
{% block scripts %}<script>
function prepCompletar(btn){
  document.getElementById('compOrdenId').value = btn.dataset.orden;
  document.getElementById('compProdNombre').textContent = btn.dataset.prod;
  document.getElementById('compCant').textContent = btn.dataset.cant;
  document.getElementById('compLote').value = '';
}
</script>{% endblock %}
{% endblock %}"""

T['produccion/gantt.html'] = """{% extends 'base.html' %}
{% block title %}Gantt por Pedido{% endblock %}{% block page_title %}Gantt por Pedido{% endblock %}
{% block topbar_actions %}
<a href="{{ url_for('ordenes_produccion') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left me-1"></i>Órdenes</a>
<a href="{{ url_for('produccion_index') }}" class="btn btn-outline-secondary btn-sm"><i class="bi bi-gear me-1"></i>Producción</a>
{% endblock %}
{% block content %}
<style>
.body-content>.fc,.body-content .fc{max-width:100%!important}
.gantt-wrap{overflow-x:auto;padding-bottom:1rem}
.gantt-grid{display:grid;font-size:.82rem;min-width:1000px}
.gantt-section-hdr{grid-column:1/-1;background:#f0f2ff;padding:.4rem .8rem;font-weight:700;font-size:.78rem;text-transform:uppercase;letter-spacing:.5px;color:#1a1f36;border-left:4px solid #5e72e4;display:flex;align-items:center;gap:.6rem}
.gantt-section-hdr.no-venta{border-left-color:#8898aa}
.gantt-header{display:contents}
.gantt-header .gh-label{background:#f8f9fe;font-weight:700;color:#525f7f;padding:.55rem .8rem;border-bottom:2px solid #e8ecf0;font-size:.72rem;text-transform:uppercase;letter-spacing:.5px;position:sticky;top:0;z-index:2}
.gantt-row{display:contents}
.gantt-row:hover .gc{background:#f4f6fb}
.gc{padding:.45rem .8rem;border-bottom:1px solid #f0f2ff;display:flex;align-items:center;background:#fff;transition:background .12s}
.gc.timeline{padding:.3rem .3rem;position:relative;overflow:visible}
.bar-wrap{position:relative;height:44px;background:#f8f9fe;border-radius:8px;overflow:visible}
.bar-track{position:absolute;top:0;left:0;right:0;bottom:0;border-radius:8px;overflow:hidden}
.gantt-bar{position:absolute;height:18px;border-radius:5px;display:flex;align-items:center;padding:0 6px;font-size:.67rem;font-weight:700;color:#fff;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.gantt-bar.materiales{top:3px;background:linear-gradient(90deg,#fb6340,#f5365c)}
.gantt-bar.produccion{top:23px;background:linear-gradient(90deg,#5e72e4,#825ee4)}
.gantt-bar.completado{top:23px;background:linear-gradient(90deg,#2dce89,#26b07b)}
.gantt-bar.pendiente_materiales{top:23px;background:linear-gradient(90deg,#fb6340,#f5365c)}
.today-line{position:absolute;top:-4px;bottom:-4px;width:2px;background:#f5365c;z-index:5;border-radius:2px;pointer-events:none}
.today-line::before{content:'Hoy';position:absolute;top:-18px;left:50%;transform:translateX(-50%);background:#f5365c;color:#fff;font-size:.6rem;font-weight:700;padding:1px 5px;border-radius:4px;white-space:nowrap}
.legend-dot{width:12px;height:12px;border-radius:3px;display:inline-block;margin-right:5px;vertical-align:middle}
.empty-gantt{text-align:center;padding:4rem 2rem;color:#8898aa}
</style>

{% if not pedidos_json %}
<div class="empty-gantt">
  <i class="bi bi-bar-chart-steps" style="font-size:3rem;display:block;margin-bottom:1rem;opacity:.3"></i>
  <div class="fw-semibold mb-1">No hay pedidos con órdenes de producción</div>
  <small>Los pedidos aparecerán aquí cuando se creen órdenes de producción.</small>
</div>
{% else %}
<!-- Leyenda -->
<div class="d-flex gap-4 mb-3 flex-wrap align-items-center">
  <span style="font-size:.82rem"><span class="legend-dot" style="background:linear-gradient(90deg,#fb6340,#f5365c)"></span>Entrega materiales (OC)</span>
  <span style="font-size:.82rem"><span class="legend-dot" style="background:linear-gradient(90deg,#5e72e4,#825ee4)"></span>Producción en curso</span>
  <span style="font-size:.82rem"><span class="legend-dot" style="background:linear-gradient(90deg,#2dce89,#26b07b)"></span>Producción completada</span>
  <span style="font-size:.82rem"><span class="legend-dot" style="background:#f5365c"></span>Línea de hoy</span>
</div>
<!-- Resumen -->
<div class="row g-2 mb-4">
  <div class="col-auto"><div class="fc px-3 py-2 d-flex gap-2 align-items-center">
    <i class="bi bi-bag text-primary"></i>
    <span class="fw-semibold" id="nPedidos">0</span><small class="text-muted">pedidos</small>
  </div></div>
  <div class="col-auto"><div class="fc px-3 py-2 d-flex gap-2 align-items-center">
    <i class="bi bi-gear-fill text-warning"></i>
    <span class="fw-semibold" id="nProd">0</span><small class="text-muted">en producción</small>
  </div></div>
  <div class="col-auto"><div class="fc px-3 py-2 d-flex gap-2 align-items-center">
    <i class="bi bi-check-circle text-success"></i>
    <span class="fw-semibold" id="nComp">0</span><small class="text-muted">completadas</small>
  </div></div>
</div>

<div style="background:#fff;border-radius:12px;box-shadow:0 2px 8px rgba(0,0,0,.06);overflow:hidden;padding:1rem">
<div class="gantt-wrap">
  <div class="gantt-grid" id="ganttGrid" style="grid-template-columns:200px 90px 110px 1fr">
    <div class="gantt-header">
      <div class="gh-label">Producto</div>
      <div class="gh-label">Cantidad</div>
      <div class="gh-label">Estado</div>
      <div class="gh-label">Línea de tiempo (<span id="spanLabel"></span>)</div>
    </div>
    <div id="ganttRows"></div>
  </div>
</div>
</div>
{% endif %}

{% block scripts %}<script>
var PEDIDOS = {{ pedidos_json|tojson }};

if(PEDIDOS.length > 0){
  var allDates=[];
  PEDIDOS.forEach(function(p){
    p.ordenes.forEach(function(o){
      allDates.push(new Date(o.inicio));
      allDates.push(new Date(o.fin));
      if(o.mat_inicio) allDates.push(new Date(o.mat_inicio));
      if(o.mat_fin) allDates.push(new Date(o.mat_fin));
    });
  });
  var minD=new Date(Math.min.apply(null,allDates));
  var maxD=new Date(Math.max.apply(null,allDates));
  minD.setDate(minD.getDate()-3); maxD.setDate(maxD.getDate()+3);
  var totalMs=maxD-minD;
  var totalDays=Math.ceil(totalMs/86400000);
  var today=new Date(); today.setHours(0,0,0,0);
  document.getElementById('spanLabel').textContent=totalDays+' días';

  function pct(d){ return Math.max(0,Math.min(100,(new Date(d)-minD)/totalMs*100)); }
  function fmtD(s){ var d=new Date(s); return d.getDate()+'/'+(d.getMonth()+1)+'/'+d.getFullYear(); }

  function estadoBadge(s){
    if(s==='en_produccion') return '<span class="badge" style="background:#5e72e4;font-size:.7rem">En producción</span>';
    if(s==='pendiente_materiales') return '<span class="badge" style="background:#fb6340;font-size:.7rem">Sin materiales</span>';
    if(s==='completado') return '<span class="badge" style="background:#2dce89;font-size:.7rem">Completado</span>';
    return '<span class="badge bg-secondary" style="font-size:.7rem">'+s+'</span>';
  }

  var step=totalDays<=30?7:(totalDays<=90?14:30);
  var markerHtml='';
  for(var di=0;di<=totalDays;di+=step){
    var mp=di/totalDays*100;
    var md=new Date(minD); md.setDate(md.getDate()+di);
    markerHtml+='<div style="position:absolute;left:'+mp+'%;top:0;bottom:0;border-left:1px dashed #e0e4f0;">'
      +'<span style="position:absolute;top:2px;left:3px;font-size:.58rem;color:#c0c8d8;white-space:nowrap">'+md.getDate()+'/'+(md.getMonth()+1)+'</span></div>';
  }
  var todayPct=pct(today);
  var todayLine=(todayPct>=0&&todayPct<=100)?'<div class="today-line" style="left:'+todayPct+'%"></div>':'';

  var rowsHtml='';
  var nProd=0,nComp=0;
  PEDIDOS.forEach(function(p){
    // Section header (pedido)
    rowsHtml+='<div class="gantt-section-hdr'+(p.venta_id?'':' no-venta')+'">'
      +'<i class="bi bi-bag'+(p.venta_id?'-fill':'')+'"></i>'
      +(p.venta_numero?'<span style="color:#5e72e4">'+p.venta_numero+'</span>':'')
      +'<span>'+(p.titulo||'Sin pedido asociado')+'</span>'
      +(p.cliente?'<small class="text-muted ms-auto">'+p.cliente+'</small>':'')
      +(p.fecha_entrega?'<small class="badge bg-warning text-dark ms-1">Entrega: '+p.fecha_entrega+'</small>':'')
      +'</div>';

    p.ordenes.forEach(function(o){
      if(o.estado==='en_produccion') nProd++;
      if(o.estado==='completado') nComp++;

      var lProd=pct(o.inicio), rProd=pct(o.fin);
      var wProd=Math.max(0.5,rProd-lProd);
      var durDias=Math.round((new Date(o.fin)-new Date(o.inicio))/86400000);
      var hasMat=o.mat_inicio&&o.mat_fin;
      var lMat=hasMat?pct(o.mat_inicio):0, rMat=hasMat?pct(o.mat_fin):0;
      var wMat=hasMat?Math.max(0.5,rMat-lMat):0;

      rowsHtml+='<div class="gantt-row">'
        +'<div class="gc"><div>'
          +'<div class="fw-semibold" style="color:#1a1f36;font-size:.83rem">'+o.producto+'</div>'
          +(o.lote?'<small class="text-muted">Lote: '+o.lote+'</small>':'')
        +'</div></div>'
        +'<div class="gc" style="font-weight:600;color:#525f7f">'+o.cantidad+' uds</div>'
        +'<div class="gc">'+estadoBadge(o.estado)+'</div>'
        +'<div class="gc timeline"><div class="bar-wrap">'
          +'<div class="bar-track">'+markerHtml+'</div>'
          +todayLine
          // Material delivery bar (top track)
          +(hasMat?'<div class="gantt-bar materiales" style="left:'+lMat+'%;width:'+wMat+'%" title="Materiales: '+fmtD(o.mat_inicio)+' → '+fmtD(o.mat_fin)+'"><i class=\"bi bi-box me-1\"></i>Mat.</div>':'')
          // Production bar (bottom track)
          +'<div class="gantt-bar '+o.estado+'" style="left:'+lProd+'%;width:'+wProd+'%" title="Producción: '+fmtD(o.inicio)+' → '+fmtD(o.fin)+' ('+durDias+'d)">'
            +(wProd>6?durDias+'d':'')
          +'</div>'
        +'</div></div>'
      +'</div>';
    });
  });

  document.getElementById('ganttRows').innerHTML=rowsHtml;
  document.getElementById('nPedidos').textContent=PEDIDOS.length;
  document.getElementById('nProd').textContent=nProd;
  document.getElementById('nComp').textContent=nComp;
}
</script>{% endblock %}
{% endblock %}"""

app.jinja_loader = DictLoader(T)

# =============================================================
# RUTAS — AUTENTICACIÓN
# =============================================================

@app.route('/login', methods=['GET','POST'])
def login():
    if current_user.is_authenticated: return redirect(url_for('dashboard'))
    if request.method == 'POST':
        user = User.query.filter_by(email=request.form.get('email','').strip()).first()
        if user and user.check_password(request.form.get('password','')) and user.activo:
            login_user(user, remember=bool(request.form.get('remember')))
            flash(f'¡Bienvenido, {user.nombre}!', 'success')
            return redirect(request.args.get('next') or url_for('dashboard'))
        flash('Email o contraseña incorrectos.', 'danger')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user(); flash('Sesión cerrada.', 'info'); return redirect(url_for('login'))

# =============================================================
# DASHBOARD
# =============================================================

@app.route('/')
@login_required
def dashboard():
    from datetime import date
    hoy = date.today()
    mes_inicio = hoy.replace(day=1)
    ingresos = db.session.query(db.func.sum(Venta.total)).filter(Venta.estado.in_(['ganado','anticipo_pagado'])).scalar() or 0
    gastos_tot = db.session.query(db.func.sum(GastoOperativo.monto)).scalar() or 0
    gastos_mes = db.session.query(db.func.sum(GastoOperativo.monto)).filter(GastoOperativo.fecha >= mes_inicio).scalar() or 0
    compras_tot = db.session.query(db.func.sum(CompraMateria.costo_total)).scalar() or 0
    saldo_pend = db.session.query(db.func.sum(Venta.saldo)).filter(Venta.estado.in_(['anticipo_pagado','negociacion'])).scalar() or 0
    # Impuestos estimados globales (acumulado total, no solo mes)
    total_egresos_global = gastos_tot + compras_tot
    utilidad_global = ingresos - total_egresos_global
    impuestos_estimados, detalle_imp_dash = _calcular_impuestos(ingresos, utilidad_global)
    saldo_neto = ingresos - total_egresos_global - impuestos_estimados
    return render_template('dashboard.html',
        total_clientes       = Cliente.query.filter_by(estado='activo').count(),
        ventas_ganadas       = Venta.query.filter_by(estado='ganado').count(),
        tareas_pendientes    = Tarea.query.filter(Tarea.estado != 'completada').count(),
        ingresos_totales     = ingresos,
        gastos_totales       = gastos_tot,
        compras_totales      = compras_tot,
        balance              = ingresos - total_egresos_global,
        impuestos_estimados  = impuestos_estimados,
        saldo_neto           = saldo_neto,
        saldo_pendiente      = saldo_pend,
        productos_bajo_stock = Producto.query.filter(Producto.activo==True, Producto.stock<=Producto.stock_minimo).count(),
        tareas_recientes     = Tarea.query.filter(Tarea.estado!='completada').order_by(Tarea.creado_en.desc()).limit(5).all(),
        ventas_recientes     = Venta.query.order_by(Venta.creado_en.desc()).limit(6).all(),
        actividades_recientes= Actividad.query.order_by(Actividad.creado_en.desc()).limit(8).all(),
    )

# =============================================================
# CLIENTES
# =============================================================

@app.route('/clientes')
@login_required
def clientes():
    busqueda = request.args.get('buscar','')
    estado_rel_f = request.args.get('estado_rel','')
    q = Cliente.query
    if busqueda:
        q = q.filter(db.or_(Cliente.nombre.ilike(f'%{busqueda}%'),
                             Cliente.empresa.ilike(f'%{busqueda}%'),
                             Cliente.nit.ilike(f'%{busqueda}%')))
    if estado_rel_f: q = q.filter_by(estado_relacion=estado_rel_f)
    return render_template('clientes/index.html', items=q.order_by(Cliente.empresa, Cliente.nombre).all(),
                           busqueda=busqueda, estado_rel_f=estado_rel_f)

def _save_contactos(cliente_obj):
    ContactoCliente.query.filter_by(cliente_id=cliente_obj.id).delete()
    nombres   = request.form.getlist('c_nombre[]')
    cargos    = request.form.getlist('c_cargo[]')
    emails    = request.form.getlist('c_email[]')
    telefonos = request.form.getlist('c_telefono[]')
    for i, n in enumerate(nombres):
        if n.strip():
            db.session.add(ContactoCliente(
                cliente_id=cliente_obj.id,
                nombre=n.strip(),
                cargo=cargos[i] if i < len(cargos) else '',
                email=emails[i] if i < len(emails) else '',
                telefono=telefonos[i] if i < len(telefonos) else ''))

@app.route('/clientes/nuevo', methods=['GET','POST'])
@login_required
def cliente_nuevo():
    if request.method == 'POST':
        c = Cliente(nombre=request.form.get('empresa','') or request.form.get('nombre',''),
            empresa=request.form.get('empresa',''), nit=request.form.get('nit',''),
            estado_relacion=request.form.get('estado_relacion','prospecto'),
            dir_comercial=request.form.get('dir_comercial',''),
            dir_entrega=request.form.get('dir_entrega',''),
            notas=request.form.get('notas',''), estado='activo')
        db.session.add(c); db.session.flush()
        _save_contactos(c); db.session.commit()
        _log('crear','cliente',c.id,f'Cliente creado: {c.empresa or c.nombre}'); db.session.commit()
        flash('Cliente creado.','success'); return redirect(url_for('clientes'))
    return render_template('clientes/form.html', obj=None, titulo='Nuevo Cliente')

@app.route('/clientes/<int:id>')
@login_required
def cliente_ver(id):
    return render_template('clientes/ver.html', obj=Cliente.query.get_or_404(id))

@app.route('/clientes/<int:id>/editar', methods=['GET','POST'])
@login_required
def cliente_editar(id):
    obj = Cliente.query.get_or_404(id)
    if request.method == 'POST':
        obj.empresa=request.form.get('empresa',''); obj.nit=request.form.get('nit','')
        obj.nombre=request.form.get('empresa','') or obj.nombre
        obj.estado_relacion=request.form.get('estado_relacion','prospecto')
        obj.dir_comercial=request.form.get('dir_comercial','')
        obj.dir_entrega=request.form.get('dir_entrega','')
        obj.notas=request.form.get('notas',''); obj.actualizado_en=datetime.utcnow()
        db.session.flush(); _save_contactos(obj); db.session.commit()
        _log('editar','cliente',obj.id,f'Cliente editado: {obj.empresa or obj.nombre}'); db.session.commit()
        flash('Cliente actualizado.','success'); return redirect(url_for('cliente_ver', id=obj.id))
    return render_template('clientes/form.html', obj=obj, titulo='Editar Cliente')

@app.route('/clientes/<int:id>/eliminar', methods=['POST'])
@login_required
def cliente_eliminar(id):
    obj=Cliente.query.get_or_404(id); db.session.delete(obj); db.session.commit()
    flash('Cliente eliminado.','info'); return redirect(url_for('clientes'))

# =============================================================
# PROVEEDORES
# =============================================================

@app.route('/proveedores')
@login_required
def proveedores():
    busqueda = request.args.get('buscar','')
    tipo_f   = request.args.get('tipo_f','')
    q = Proveedor.query.filter_by(activo=True)
    if busqueda:
        q = q.filter(db.or_(Proveedor.nombre.ilike(f'%{busqueda}%'),
                             Proveedor.empresa.ilike(f'%{busqueda}%'),
                             Proveedor.nit.ilike(f'%{busqueda}%')))
    if tipo_f:
        q = q.filter_by(tipo=tipo_f)
    return render_template('proveedores/index.html', items=q.order_by(Proveedor.empresa, Proveedor.nombre).all(),
                           busqueda=busqueda)

@app.route('/proveedores/nuevo', methods=['GET','POST'])
@login_required
def proveedor_nuevo():
    if request.method == 'POST':
        p = Proveedor(
            nombre=request.form.get('nombre','') or request.form.get('empresa',''),
            empresa=request.form.get('empresa',''),
            nit=request.form.get('nit',''),
            email=request.form.get('email',''),
            telefono=request.form.get('telefono',''),
            direccion=request.form.get('direccion',''),
            categoria=request.form.get('categoria',''),
            tipo=request.form.get('tipo','proveedor'),
            notas=request.form.get('notas',''), activo=True)
        db.session.add(p); db.session.commit()
        flash('Registro creado.','success'); return redirect(url_for('proveedores'))
    return render_template('proveedores/form.html', obj=None, titulo='Nuevo Proveedor / Transportista')

@app.route('/proveedores/<int:id>/editar', methods=['GET','POST'])
@login_required
def proveedor_editar(id):
    obj = Proveedor.query.get_or_404(id)
    if request.method == 'POST':
        obj.nombre=request.form.get('nombre','') or request.form.get('empresa','') or obj.nombre
        obj.empresa=request.form.get('empresa','')
        obj.nit=request.form.get('nit','')
        obj.email=request.form.get('email','')
        obj.telefono=request.form.get('telefono','')
        obj.direccion=request.form.get('direccion','')
        obj.categoria=request.form.get('categoria','')
        obj.tipo=request.form.get('tipo','proveedor')
        obj.notas=request.form.get('notas','')
        db.session.commit()
        flash('Registro actualizado.','success'); return redirect(url_for('proveedores'))
    return render_template('proveedores/form.html', obj=obj, titulo='Editar Proveedor / Transportista')

@app.route('/proveedores/<int:id>/eliminar', methods=['POST'])
@login_required
def proveedor_eliminar(id):
    obj = Proveedor.query.get_or_404(id)
    obj.activo = False
    db.session.commit()
    flash('Registro eliminado.','info'); return redirect(url_for('proveedores'))

# =============================================================
# =============================================================
# COTIZACIONES PROVEEDOR
# =============================================================

@app.route('/cotizaciones-proveedor')
@login_required
def cotizaciones_proveedor():
    estado_f = request.args.get('estado','')
    tipo_f   = request.args.get('tipo','')   # granel / general
    q = CotizacionProveedor.query
    if estado_f: q = q.filter_by(estado=estado_f)
    if tipo_f:   q = q.filter_by(tipo_cotizacion=tipo_f)
    items = q.order_by(CotizacionProveedor.creado_en.desc()).all()
    totales = {
        'granel':  CotizacionProveedor.query.filter_by(tipo_cotizacion='granel').count(),
        'general': CotizacionProveedor.query.filter_by(tipo_cotizacion='general').count(),
    }
    return render_template('proveedores/cotizaciones.html',
                           items=items, estado_f=estado_f, tipo_f=tipo_f, totales=totales)

@app.route('/cotizaciones-proveedor/<int:id>/json')
@login_required
def cotizacion_proveedor_json(id):
    """API para traer datos de cotización al formulario de OC"""
    cp = CotizacionProveedor.query.get_or_404(id)
    return jsonify({
        'id': cp.id,
        'numero': cp.numero,
        'nombre_producto': cp.nombre_producto,
        'descripcion': cp.descripcion or '',
        'sku': cp.sku or '',
        'precio_unitario': cp.precio_unitario,
        'unidades_minimas': cp.unidades_minimas,
        'unidad': cp.unidad,
        'plazo_entrega_dias': cp.plazo_entrega_dias or 0,
        'condicion_pago_tipo': cp.condicion_pago_tipo or 'contado',
        'condicion_pago_dias': cp.condicion_pago_dias or 0,
        'anticipo_porcentaje': cp.anticipo_porcentaje or 0,
        'proveedor_id': cp.proveedor_id,
    })

@app.route('/cotizaciones-proveedor/nueva', methods=['GET','POST'])
@login_required
def cotizacion_proveedor_nueva():
    provs = Proveedor.query.filter(Proveedor.activo==True,
        Proveedor.tipo.in_(['proveedor','ambos'])).order_by(Proveedor.empresa).all()
    tipo_default = request.args.get('tipo','general')
    if request.method == 'POST':
        fc = request.form.get('fecha_cotizacion')
        fv = request.form.get('vigencia')
        cp = CotizacionProveedor(
            proveedor_id=int(request.form.get('proveedor_id')) if request.form.get('proveedor_id') else None,
            tipo_cotizacion=request.form.get('tipo_cotizacion','general'),
            tipo_producto_servicio=request.form.get('tipo_producto_servicio',''),
            nombre_producto=request.form['nombre_producto'],
            descripcion=request.form.get('descripcion',''),
            sku=request.form.get('sku',''),
            precio_unitario=float(request.form.get('precio_unitario') or 0),
            unidades_minimas=int(request.form.get('unidades_minimas') or 1),
            unidad=request.form.get('unidad','unidades'),
            plazo_entrega=request.form.get('plazo_entrega',''),
            plazo_entrega_dias=int(request.form.get('plazo_entrega_dias') or 0),
            condiciones_pago=request.form.get('condiciones_pago',''),
            condicion_pago_tipo=request.form.get('condicion_pago_tipo','contado'),
            condicion_pago_dias=int(request.form.get('condicion_pago_dias') or 0),
            anticipo_porcentaje=float(request.form.get('anticipo_porcentaje') or 0),
            fecha_cotizacion=datetime.strptime(fc,'%Y-%m-%d').date() if fc else datetime.utcnow().date(),
            vigencia=datetime.strptime(fv,'%Y-%m-%d').date() if fv else None,
            estado=request.form.get('estado','vigente'),
            notas=request.form.get('notas',''),
            creado_por=current_user.id
        )
        db.session.add(cp); db.session.flush()
        hoy = datetime.utcnow().date()
        ultimo = CotizacionProveedor.query.filter(
            CotizacionProveedor.numero.like(f'CP-{hoy.year}-%')
        ).order_by(CotizacionProveedor.id.desc()).first()
        if ultimo and ultimo.numero:
            try: seq = int(ultimo.numero.split('-')[-1]) + 1
            except: seq = 1
        else: seq = 1
        cp.numero = f'CP-{hoy.year}-{seq:03d}'
        db.session.commit()
        flash(f'Cotización {cp.numero} creada.','success')
        return redirect(url_for('cotizaciones_proveedor', tipo=cp.tipo_cotizacion))
    return render_template('proveedores/cotizacion_form.html', obj=None,
                           proveedores_list=provs, titulo='Nueva Cotización Proveedor',
                           tipo_default=tipo_default)

@app.route('/cotizaciones-proveedor/<int:id>/editar', methods=['GET','POST'])
@login_required
def cotizacion_proveedor_editar(id):
    obj = CotizacionProveedor.query.get_or_404(id)
    provs = Proveedor.query.filter(Proveedor.activo==True,
        Proveedor.tipo.in_(['proveedor','ambos'])).order_by(Proveedor.empresa).all()
    if request.method == 'POST':
        fc = request.form.get('fecha_cotizacion')
        fv = request.form.get('vigencia')
        obj.proveedor_id=int(request.form.get('proveedor_id')) if request.form.get('proveedor_id') else None
        obj.tipo_cotizacion=request.form.get('tipo_cotizacion','general')
        obj.tipo_producto_servicio=request.form.get('tipo_producto_servicio','')
        obj.nombre_producto=request.form.get('nombre_producto','')
        obj.descripcion=request.form.get('descripcion','')
        obj.sku=request.form.get('sku','')
        obj.precio_unitario=float(request.form.get('precio_unitario') or 0)
        obj.unidades_minimas=int(request.form.get('unidades_minimas') or 1)
        obj.unidad=request.form.get('unidad','unidades')
        obj.plazo_entrega=request.form.get('plazo_entrega','')
        obj.plazo_entrega_dias=int(request.form.get('plazo_entrega_dias') or 0)
        obj.condiciones_pago=request.form.get('condiciones_pago','')
        obj.condicion_pago_tipo=request.form.get('condicion_pago_tipo','contado')
        obj.condicion_pago_dias=int(request.form.get('condicion_pago_dias') or 0)
        obj.anticipo_porcentaje=float(request.form.get('anticipo_porcentaje') or 0)
        obj.fecha_cotizacion=datetime.strptime(fc,'%Y-%m-%d').date() if fc else obj.fecha_cotizacion
        obj.vigencia=datetime.strptime(fv,'%Y-%m-%d').date() if fv else None
        obj.estado=request.form.get('estado','vigente')
        obj.notas=request.form.get('notas','')
        db.session.commit()
        flash('Cotización actualizada.','success')
        return redirect(url_for('cotizaciones_proveedor', tipo=obj.tipo_cotizacion))
    return render_template('proveedores/cotizacion_form.html', obj=obj,
                           proveedores_list=provs, titulo='Editar Cotización Proveedor',
                           tipo_default=obj.tipo_cotizacion)

@app.route('/cotizaciones-proveedor/<int:id>/eliminar', methods=['POST'])
@login_required
def cotizacion_proveedor_eliminar(id):
    obj = CotizacionProveedor.query.get_or_404(id)
    tipo = obj.tipo_cotizacion
    db.session.delete(obj); db.session.commit()
    flash('Cotización eliminada.','info')
    return redirect(url_for('cotizaciones_proveedor', tipo=tipo))

# =============================================================
# ÓRDENES DE COMPRA
# =============================================================

@app.route('/ordenes-compra')
@login_required
def ordenes_compra():
    estado_f = request.args.get('estado','')
    q = OrdenCompra.query
    if estado_f: q = q.filter_by(estado=estado_f)
    return render_template('ordenes_compra/index.html',
                           items=q.order_by(OrdenCompra.creado_en.desc()).all(),
                           estado_f=estado_f)

def _oc_save_items(oc_id):
    """Guarda ítems del formulario de OC y devuelve la lista."""
    nombres  = request.form.getlist('item_nombre[]')
    descs    = request.form.getlist('item_desc[]')
    cants    = request.form.getlist('item_cant[]')
    units    = request.form.getlist('item_unidad[]')
    precios  = request.form.getlist('item_precio[]')
    cot_ids  = request.form.getlist('item_cot_id[]')
    items = []
    for i, nom in enumerate(nombres):
        if not nom.strip(): continue
        cant   = float(cants[i])  if i < len(cants)   else 1
        precio = float(precios[i]) if i < len(precios) else 0
        cot_id = int(cot_ids[i]) if i < len(cot_ids) and cot_ids[i].strip() else None
        items.append(OrdenCompraItem(
            orden_id=oc_id,
            nombre_item=nom.strip(),
            descripcion=descs[i] if i < len(descs) else '',
            cantidad=cant,
            unidad=units[i] if i < len(units) else 'unidades',
            precio_unit=precio,
            subtotal=cant*precio,
            cotizacion_id=cot_id
        ))
    return items

@app.route('/ordenes-compra/nueva', methods=['GET','POST'])
@login_required
def orden_compra_nueva():
    provs       = Proveedor.query.filter(Proveedor.activo==True, Proveedor.tipo.in_(['proveedor','ambos'])).order_by(Proveedor.empresa).all()
    transportistas = Proveedor.query.filter(Proveedor.activo==True, Proveedor.tipo.in_(['transportista','ambos'])).order_by(Proveedor.nombre).all()
    cotizaciones_disponibles = CotizacionProveedor.query.filter_by(estado='vigente').order_by(CotizacionProveedor.nombre_producto).all()
    if request.method == 'POST':
        fe  = request.form.get('fecha_emision')
        fes = request.form.get('fecha_esperada')
        fep = request.form.get('fecha_estimada_pago')
        fer = request.form.get('fecha_estimada_recogida')
        cot_id = int(request.form.get('cotizacion_id')) if request.form.get('cotizacion_id') else None
        tra_id = int(request.form.get('transportista_id')) if request.form.get('transportista_id') else None
        fecha_emision = datetime.strptime(fe,'%Y-%m-%d').date() if fe else datetime.utcnow().date()
        # Calcular fecha_esperada desde cotización si no se ingresó manualmente
        fecha_esp = None
        if fes:
            fecha_esp = datetime.strptime(fes,'%Y-%m-%d').date()
        elif cot_id:
            cot_obj = CotizacionProveedor.query.get(cot_id)
            if cot_obj and cot_obj.plazo_entrega_dias:
                fecha_esp = fecha_emision + timedelta(days=cot_obj.plazo_entrega_dias)
        oc = OrdenCompra(
            proveedor_id=int(request.form.get('proveedor_id')) if request.form.get('proveedor_id') else None,
            cotizacion_id=cot_id,
            transportista_id=tra_id,
            estado=request.form.get('estado','borrador'),
            fecha_emision=fecha_emision,
            fecha_esperada=fecha_esp,
            fecha_estimada_pago=datetime.strptime(fep,'%Y-%m-%d').date() if fep else None,
            fecha_estimada_recogida=datetime.strptime(fer,'%Y-%m-%d').date() if fer else None,
            subtotal=float(request.form.get('subtotal_calc') or 0),
            iva=float(request.form.get('iva_calc') or 0),
            total=float(request.form.get('total_calc') or 0),
            notas=request.form.get('notas',''),
            creado_por=current_user.id
        )
        db.session.add(oc); db.session.flush()
        hoy = datetime.utcnow().date()
        ultimo_oc = OrdenCompra.query.filter(OrdenCompra.numero.like(f'OC-{hoy.year}-%')).order_by(OrdenCompra.id.desc()).first()
        if ultimo_oc and ultimo_oc.numero:
            try: seq = int(ultimo_oc.numero.split('-')[-1]) + 1
            except: seq = 1
        else: seq = 1
        oc.numero = f'OC-{hoy.year}-{seq:03d}'
        for it in _oc_save_items(oc.id): db.session.add(it)
        # Auto-tarea para transportista
        if tra_id and fer:
            tra = Proveedor.query.get(tra_id)
            fecha_rec = datetime.strptime(fer,'%Y-%m-%d').date()
            t = Tarea(titulo=f'Contratar transporte para OC {oc.numero}',
                      descripcion=f'Contactar a {tra.nombre or tra.empresa} para coordinar recogida el {fecha_rec.strftime("%d/%m/%Y")}. OC: {oc.numero}',
                      estado='pendiente', prioridad='alta',
                      fecha_vencimiento=fecha_rec - timedelta(days=2),
                      creado_por=current_user.id, tarea_tipo='contratar_transporte')
            db.session.add(t)
        db.session.commit()
        flash(f'Orden de compra {oc.numero} creada.','success')
        return redirect(url_for('ordenes_compra'))
    return render_template('ordenes_compra/form.html', obj=None,
                           proveedores_list=provs, transportistas_list=transportistas,
                           cotizaciones_list=cotizaciones_disponibles,
                           titulo='Nueva Orden de Compra', items_json=[])

@app.route('/ordenes-compra/<int:id>/editar', methods=['GET','POST'])
@login_required
def orden_compra_editar(id):
    obj = OrdenCompra.query.get_or_404(id)
    provs       = Proveedor.query.filter(Proveedor.activo==True, Proveedor.tipo.in_(['proveedor','ambos'])).order_by(Proveedor.empresa).all()
    transportistas = Proveedor.query.filter(Proveedor.activo==True, Proveedor.tipo.in_(['transportista','ambos'])).order_by(Proveedor.nombre).all()
    cotizaciones_disponibles = CotizacionProveedor.query.filter_by(estado='vigente').order_by(CotizacionProveedor.nombre_producto).all()
    if request.method == 'POST':
        fe  = request.form.get('fecha_emision')
        fes = request.form.get('fecha_esperada')
        fep = request.form.get('fecha_estimada_pago')
        fer = request.form.get('fecha_estimada_recogida')
        cot_id = int(request.form.get('cotizacion_id')) if request.form.get('cotizacion_id') else None
        tra_id = int(request.form.get('transportista_id')) if request.form.get('transportista_id') else None
        fecha_emision = datetime.strptime(fe,'%Y-%m-%d').date() if fe else obj.fecha_emision
        fecha_esp = None
        if fes:
            fecha_esp = datetime.strptime(fes,'%Y-%m-%d').date()
        elif cot_id:
            cot_obj = CotizacionProveedor.query.get(cot_id)
            if cot_obj and cot_obj.plazo_entrega_dias:
                fecha_esp = fecha_emision + timedelta(days=cot_obj.plazo_entrega_dias)
        obj.proveedor_id        = int(request.form.get('proveedor_id')) if request.form.get('proveedor_id') else None
        obj.cotizacion_id       = cot_id
        obj.transportista_id    = tra_id
        obj.estado              = request.form.get('estado', obj.estado)
        obj.fecha_emision       = fecha_emision
        obj.fecha_esperada      = fecha_esp
        obj.fecha_estimada_pago = datetime.strptime(fep,'%Y-%m-%d').date() if fep else None
        obj.fecha_estimada_recogida = datetime.strptime(fer,'%Y-%m-%d').date() if fer else None
        obj.subtotal = float(request.form.get('subtotal_calc') or 0)
        obj.iva      = float(request.form.get('iva_calc') or 0)
        obj.total    = float(request.form.get('total_calc') or 0)
        obj.notas    = request.form.get('notas','')
        OrdenCompraItem.query.filter_by(orden_id=obj.id).delete()
        for it in _oc_save_items(obj.id): db.session.add(it)
        db.session.commit()
        flash('Orden de compra actualizada.','success')
        return redirect(url_for('ordenes_compra'))
    items_json = [{'nombre':it.nombre_item,'desc':it.descripcion or '','cant':it.cantidad,
                   'unidad':it.unidad,'precio':it.precio_unit,'cot_id':it.cotizacion_id or ''} for it in obj.items]
    return render_template('ordenes_compra/form.html', obj=obj,
                           proveedores_list=provs, transportistas_list=transportistas,
                           cotizaciones_list=cotizaciones_disponibles,
                           titulo='Editar Orden de Compra', items_json=items_json)

@app.route('/ordenes-compra/<int:id>/estado', methods=['POST'])
@login_required
def orden_compra_estado(id):
    obj = OrdenCompra.query.get_or_404(id)
    estado_anterior = obj.estado
    obj.estado = request.form.get('estado', obj.estado)
    # Al marcar como "recibida": calcular fecha de entrega con plazo de cotización y agendar en calendario
    if obj.estado == 'recibida' and estado_anterior != 'recibida':
        hoy_recv = datetime.utcnow().date()
        cot = obj.cotizacion_ref
        if cot and cot.plazo_entrega_dias and not cot.calendario_integrado:
            fecha_entrega = hoy_recv + timedelta(days=cot.plazo_entrega_dias)
            ev = Evento(
                titulo=f'Entrega esperada: {cot.nombre_producto} ({obj.numero})',
                tipo='recordatorio',
                fecha=fecha_entrega,
                descripcion=f'OC {obj.numero} recibida el {hoy_recv.strftime("%d/%m/%Y")}. Entrega esperada en {cot.plazo_entrega_dias} días desde recepción. Proveedor: {obj.proveedor.nombre if obj.proveedor else "—"}',
                usuario_id=current_user.id
            )
            db.session.add(ev)
            cot.calendario_integrado = True
        elif obj.fecha_esperada:
            # Si no hay cotizacion pero hay fecha_esperada, agendar igual
            ev = Evento(
                titulo=f'Entrega esperada OC {obj.numero}',
                tipo='recordatorio',
                fecha=obj.fecha_esperada,
                descripcion=f'Orden de compra {obj.numero} marcada como recibida. Entrega esperada: {obj.fecha_esperada.strftime("%d/%m/%Y")}.',
                usuario_id=current_user.id
            )
            db.session.add(ev)
    db.session.commit()
    flash(f'Estado actualizado a "{obj.estado}".','success')
    return redirect(url_for('ordenes_compra'))

@app.route('/ordenes-compra/<int:id>/eliminar', methods=['POST'])
@login_required
def orden_compra_eliminar(id):
    obj = OrdenCompra.query.get_or_404(id)
    db.session.delete(obj); db.session.commit()
    flash('Orden de compra eliminada.','info')
    return redirect(url_for('ordenes_compra'))

# =============================================================
# VENTAS
# =============================================================

@app.route('/ventas')
@login_required
def ventas():
    estado_f=request.args.get('estado','')
    q=Venta.query
    if estado_f: q=q.filter_by(estado=estado_f)
    return render_template('ventas/index.html', items=q.order_by(Venta.creado_en.desc()).all(), estado_f=estado_f)

def _prods_json():
    prods = Producto.query.filter_by(activo=True).order_by(Producto.nombre).all()
    return [{'id':p.id,'nombre':p.nombre,'sku':p.sku or '','precio':p.precio,
             'stock':p.stock,'stock_minimo':p.stock_minimo} for p in prods]

def _save_items(venta_obj):
    VentaProducto.query.filter_by(venta_id=venta_obj.id).delete()
    pids    = request.form.getlist('prod_id[]')
    cants   = request.form.getlist('prod_cant[]')
    precios = request.form.getlist('prod_precio[]')
    for i, pid in enumerate(pids):
        cant  = float(cants[i]) if i < len(cants) else 1
        precio= float(precios[i]) if i < len(precios) else 0
        prod  = Producto.query.get(int(pid)) if pid else None
        db.session.add(VentaProducto(
            venta_id=venta_obj.id,
            producto_id=int(pid) if pid else None,
            nombre_prod=prod.nombre if prod else '',
            cantidad=cant, precio_unit=precio, subtotal=cant*precio))

def _descontar_stock_venta(venta):
    """
    Descuenta del inventario (Producto.stock) las cantidades de los items
    de la venta. Se llama exactamente una vez, cuando la venta pasa a
    anticipo_pagado o ganado por primera vez.
    """
    try:
        for item in venta.items:
            if not item.producto_id:
                continue
            prod = Producto.query.get(item.producto_id)
            if not prod:
                continue
            cant = item.cantidad
            prod.stock = max(0, (prod.stock or 0) - cant)
    except Exception as ex:
        print(f'_descontar_stock_venta error: {ex}')

@app.route('/ventas/nueva', methods=['GET','POST'])
@login_required
def venta_nueva():
    cl = Cliente.query.order_by(Cliente.empresa, Cliente.nombre).all()
    if request.method == 'POST':
        fa = request.form.get('fecha_anticipo')
        fe = request.form.get('fecha_entrega_est')
        v = Venta(titulo=request.form['titulo'],
            cliente_id=request.form.get('cliente_id') or None,
            subtotal=float(request.form.get('subtotal_calc') or 0),
            iva=float(request.form.get('iva_calc') or 0),
            total=float(request.form.get('total_calc') or 0),
            porcentaje_anticipo=float(request.form.get('porcentaje_anticipo') or 0),
            monto_anticipo=float(request.form.get('monto_anticipo') or 0),
            saldo=float(request.form.get('saldo') or 0),
            estado=request.form.get('estado','prospecto'),
            fecha_anticipo=datetime.strptime(fa,'%Y-%m-%d').date() if fa else None,
            dias_entrega=int(request.form.get('dias_entrega') or 30),
            fecha_entrega_est=datetime.strptime(fe,'%Y-%m-%d').date() if fe else None,
            notas=request.form.get('notas',''), creado_por=current_user.id)
        db.session.add(v); db.session.flush()
        # Generar numero único VNT-YYYY-NNN
        hoy = datetime.utcnow().date()
        ultimo_vnt = Venta.query.filter(
            Venta.numero.like(f'VNT-{hoy.year}-%')
        ).order_by(Venta.id.desc()).first()
        if ultimo_vnt and ultimo_vnt.numero:
            try: seq = int(ultimo_vnt.numero.split('-')[-1]) + 1
            except: seq = 1
        else: seq = 1
        v.numero = f'VNT-{hoy.year}-{seq:03d}'
        _save_items(v); db.session.flush()
        _procesar_venta_produccion(v)
        db.session.commit()
        _log('crear','venta',v.id,f'Venta creada: {v.titulo} ({v.numero})'); db.session.commit()
        flash('Venta creada.','success'); return redirect(url_for('ventas'))
    return render_template('ventas/form.html', obj=None, clientes_list=cl,
                           titulo='Nueva Venta', productos_json=_prods_json(), items_json=[])

@app.route('/ventas/<int:id>/editar', methods=['GET','POST'])
@login_required
def venta_editar(id):
    obj = Venta.query.get_or_404(id)
    cl  = Cliente.query.order_by(Cliente.empresa, Cliente.nombre).all()
    if request.method == 'POST':
        fa = request.form.get('fecha_anticipo')
        fe = request.form.get('fecha_entrega_est')
        obj.titulo=request.form['titulo']; obj.cliente_id=request.form.get('cliente_id') or None
        obj.subtotal=float(request.form.get('subtotal_calc') or 0)
        obj.iva=float(request.form.get('iva_calc') or 0)
        obj.total=float(request.form.get('total_calc') or 0)
        obj.porcentaje_anticipo=float(request.form.get('porcentaje_anticipo') or 0)
        obj.monto_anticipo=float(request.form.get('monto_anticipo') or 0)
        obj.saldo=float(request.form.get('saldo') or 0)
        obj.estado=request.form.get('estado','prospecto')
        obj.fecha_anticipo=datetime.strptime(fa,'%Y-%m-%d').date() if fa else None
        obj.dias_entrega=int(request.form.get('dias_entrega') or 30)
        obj.fecha_entrega_est=datetime.strptime(fe,'%Y-%m-%d').date() if fe else None
        obj.notas=request.form.get('notas','')
        db.session.flush(); _save_items(obj); db.session.flush()
        _procesar_venta_produccion(obj)
        db.session.commit()
        _log('editar','venta',obj.id,f'Venta editada: {obj.titulo}'); db.session.commit()
        flash('Venta actualizada.','success'); return redirect(url_for('ventas'))
    items_j = [{'pid':it.producto_id or '','nombre':it.nombre_prod,
                'cant':it.cantidad,'precio':it.precio_unit} for it in obj.items]
    return render_template('ventas/form.html', obj=obj, clientes_list=cl,
                           titulo='Editar Venta', productos_json=_prods_json(), items_json=items_j)

@app.route('/ventas/<int:id>/eliminar', methods=['POST'])
@login_required
def venta_eliminar(id):
    obj=Venta.query.get_or_404(id); db.session.delete(obj); db.session.commit()
    flash('Venta eliminada.','info'); return redirect(url_for('ventas'))

@app.route('/ventas/<int:id>/informar_cliente', methods=['POST'])
@login_required
def venta_informar_cliente(id):
    venta = Venta.query.get_or_404(id)
    venta.cliente_informado_en = datetime.utcnow()
    db.session.commit()
    # Enviar email al cliente si tiene email
    email = venta.cliente.contactos[0].email if venta.cliente and venta.cliente.contactos else None
    if email:
        prod_nombres = ', '.join(
            it.nombre_prod for it in venta.items if it.nombre_prod
        ) or 'los productos'
        _send_email(
            email,
            f'Tu pedido está listo — {venta.titulo}',
            f'Hola {venta.cliente.nombre},\n\n'
            f'Nos complace informarte que {prod_nombres} de tu pedido "{venta.titulo}" '
            f'están listos para entrega.\n\n'
            f'Coordinaremos contigo la fecha y horario de entrega.\n\n'
            f'¡Gracias por tu compra!'
        )
        flash(f'Cliente {venta.cliente.nombre} notificado por email. Pedido listo para entrega.','success')
    else:
        flash('Venta marcada como "cliente informado". (Sin email del cliente para enviar notificación.)','info')
    return redirect(url_for('ventas'))

@app.route('/ventas/<int:id>/entregar', methods=['POST'])
@login_required
def venta_entregar(id):
    venta = Venta.query.get_or_404(id)
    venta.entregado_en = datetime.utcnow()
    # Descontar stock de productos entregados
    _descontar_stock_venta(venta)
    # Marcar órdenes de producción vinculadas como listas
    db.session.commit()
    flash(f'Venta marcada como entregada. Stock descontado del inventario.','success')
    return redirect(url_for('ventas'))

# =============================================================
# TAREAS (v8: múltiples asignados + chat)
# =============================================================

def _save_asignados(tarea_obj):
    TareaAsignado.query.filter_by(tarea_id=tarea_obj.id).delete()
    # asignado principal
    principal = request.form.get('asignado_a')
    # otros asignados (multi-select)
    otros = request.form.getlist('otros_asignados[]')
    seen = set()
    for uid_str in ([principal] if principal else []) + otros:
        try:
            uid = int(uid_str)
            if uid not in seen:
                seen.add(uid)
                db.session.add(TareaAsignado(tarea_id=tarea_obj.id, user_id=uid))
        except (ValueError, TypeError):
            pass
    if not seen:
        db.session.add(TareaAsignado(tarea_id=tarea_obj.id, user_id=current_user.id))

@app.route('/tareas')
@login_required
def tareas():
    estado_f=request.args.get('estado',''); prioridad_f=request.args.get('prioridad','')
    q=Tarea.query
    if estado_f: q=q.filter_by(estado=estado_f)
    if prioridad_f: q=q.filter_by(prioridad=prioridad_f)
    return render_template('tareas/index.html', items=q.order_by(Tarea.creado_en.desc()).all(),
        estado_f=estado_f, prioridad_f=prioridad_f)

@app.route('/tareas/nueva', methods=['GET','POST'])
@login_required
def tarea_nueva():
    us=User.query.filter_by(activo=True).all()
    if request.method == 'POST':
        fs=request.form.get('fecha_vencimiento')
        asignado_id=int(request.form.get('asignado_a') or current_user.id)
        t=Tarea(titulo=request.form['titulo'], descripcion=request.form.get('descripcion',''),
            estado=request.form.get('estado','pendiente'), prioridad=request.form.get('prioridad','media'),
            fecha_vencimiento=datetime.strptime(fs,'%Y-%m-%d').date() if fs else None,
            asignado_a=asignado_id, creado_por=current_user.id)
        db.session.add(t); db.session.flush()
        _save_asignados(t); db.session.commit()
        _log('crear','tarea',t.id,f'Tarea creada: {t.titulo}'); db.session.commit()
        # Notificación al asignado (si no es quien la crea)
        if asignado_id != current_user.id:
            _crear_notificacion(asignado_id, 'tarea_asignada',
                f'Nueva tarea asignada: {t.titulo}',
                f'Te asignó una tarea: {current_user.nombre}',
                url_for('tarea_ver', id=t.id))
            asignado = User.query.get(asignado_id)
            if asignado and asignado.email:
                _send_email(asignado.email, f'Nueva tarea: {t.titulo}',
                    f'Hola {asignado.nombre},\n\n{current_user.nombre} te asignó la tarea "{t.titulo}".\n\nDescripción: {t.descripcion or "—"}')
        flash('Tarea creada.','success'); return redirect(url_for('tareas'))
    return render_template('tareas/form.html', obj=None, usuarios=us, titulo='Nueva Tarea', asignados_ids=[])

@app.route('/tareas/<int:id>')
@login_required
def tarea_ver(id):
    obj=Tarea.query.get_or_404(id)
    return render_template('tareas/ver.html', obj=obj, tarea=obj)

@app.route('/tareas/<int:id>/comentar', methods=['POST'])
@login_required
def tarea_comentar(id):
    obj=Tarea.query.get_or_404(id)
    msg=request.form.get('mensaje','').strip()
    if msg:
        db.session.add(TareaComentario(tarea_id=obj.id, autor_id=current_user.id, mensaje=msg))
        db.session.commit()
        flash('Comentario agregado.','success')
    return redirect(url_for('tarea_ver', id=id))

@app.route('/tareas/<int:id>/editar', methods=['GET','POST'])
@login_required
def tarea_editar(id):
    obj=Tarea.query.get_or_404(id); us=User.query.filter_by(activo=True).all()
    if request.method == 'POST':
        fs=request.form.get('fecha_vencimiento')
        prev_asignado = obj.asignado_a
        obj.titulo=request.form['titulo']; obj.descripcion=request.form.get('descripcion','')
        obj.estado=request.form.get('estado','pendiente'); obj.prioridad=request.form.get('prioridad','media')
        obj.fecha_vencimiento=datetime.strptime(fs,'%Y-%m-%d').date() if fs else None
        obj.asignado_a=int(request.form.get('asignado_a') or current_user.id)
        db.session.flush(); _save_asignados(obj); db.session.commit()
        _log('editar','tarea',obj.id,f'Tarea editada: {obj.titulo}'); db.session.commit()
        # Notificar si cambió el asignado
        if obj.asignado_a != prev_asignado and obj.asignado_a != current_user.id:
            _crear_notificacion(obj.asignado_a, 'tarea_asignada',
                f'Tarea reasignada: {obj.titulo}',
                f'{current_user.nombre} te reasignó esta tarea.',
                url_for('tarea_ver', id=obj.id))
            asignado = User.query.get(obj.asignado_a)
            if asignado and asignado.email:
                _send_email(asignado.email, f'Tarea reasignada: {obj.titulo}',
                    f'Hola {asignado.nombre},\n\n{current_user.nombre} te reasignó la tarea "{obj.titulo}".')
        flash('Tarea actualizada.','success'); return redirect(url_for('tarea_ver', id=obj.id))
    asignados_ids=[a.user_id for a in obj.asignados]
    return render_template('tareas/form.html', obj=obj, usuarios=us, titulo='Editar Tarea', asignados_ids=asignados_ids)

@app.route('/tareas/<int:id>/completar', methods=['POST'])
@login_required
def tarea_completar(id):
    obj = Tarea.query.get_or_404(id)
    obj.estado = 'completada'
    db.session.commit()
    _log('completar','tarea',obj.id,f'Tarea completada: {obj.titulo}')
    db.session.commit()

    # Lógica tareas pareadas: comprar_materias + verificar_abono
    if obj.tarea_tipo in ('comprar_materias','verificar_abono') and obj.tarea_pareja_id:
        pareja = Tarea.query.get(obj.tarea_pareja_id)
        if pareja and pareja.estado == 'completada' and obj.cotizacion_id:
            # Ambas tareas completadas → enviar email al cliente
            cot = Cotizacion.query.get(obj.cotizacion_id)
            if cot and cot.cliente:
                cliente = cot.cliente
                email_dest = None
                # Buscar email del cliente
                if cliente.contactos:
                    for c in cliente.contactos:
                        if c.email:
                            email_dest = c.email; break
                # Enviar email
                empresa = ConfigEmpresa.query.first()
                nombre_empresa = empresa.nombre if empresa else 'Evore'
                if email_dest:
                    _send_email(
                        email_dest,
                        f'Producción iniciada — {cot.titulo}',
                        f'Estimado/a {cliente.nombre},\n\n'
                        f'Nos complace informarle que todas las materias primas para su pedido '
                        f'(cotización #{cot.numero or cot.id}) han sido recibidas y el anticipo '
                        f'verificado. La producción ha comenzado.\n\n'
                        f'Producto(s): {", ".join(i.nombre_prod for i in cot.items if i.nombre_prod)}\n\n'
                        f'Saludos,\n{nombre_empresa}'
                    )
                # Notificar a admins
                admins = User.query.filter_by(rol='admin', activo=True).all()
                for adm in admins:
                    _crear_notificacion(
                        adm.id, 'info',
                        f'✅ Producción iniciada — {cot.titulo}',
                        f'Materias y abono confirmados. Email enviado a {cliente.nombre}.',
                        url_for('cotizacion_ver', id=cot.id)
                    )

    flash('¡Tarea completada!','success'); return redirect(url_for('tareas'))

@app.route('/tareas/<int:id>/eliminar', methods=['POST'])
@login_required
def tarea_eliminar(id):
    obj=Tarea.query.get_or_404(id); db.session.delete(obj); db.session.commit()
    flash('Tarea eliminada.','info'); return redirect(url_for('tareas'))

# =============================================================
# INVENTARIO (v8: NSO + costo)
# =============================================================

@app.route('/inventario')
@login_required
def inventario():
    busqueda=request.args.get('buscar',''); categoria_f=request.args.get('categoria','')
    q=Producto.query.filter_by(activo=True)
    if busqueda:
        q=q.filter(db.or_(Producto.nombre.ilike(f'%{busqueda}%'),
                           Producto.sku.ilike(f'%{busqueda}%'),
                           Producto.nso.ilike(f'%{busqueda}%')))
    if categoria_f: q=q.filter_by(categoria=categoria_f)
    cats=[c[0] for c in db.session.query(Producto.categoria).filter(
        Producto.activo==True,Producto.categoria!=None,Producto.categoria!='').distinct().all()]
    return render_template('inventario/index.html', items=q.order_by(Producto.nombre).all(),
                           busqueda=busqueda, categoria_f=categoria_f, categorias=cats,
                           now=datetime.utcnow())

def _inv_form_ctx():
    materias = MateriaPrima.query.filter_by(activo=True).order_by(MateriaPrima.nombre).all()
    lotes = LoteProducto.query.order_by(LoteProducto.creado_en.desc()).all()
    mj = [{'id': m.id, 'nombre': m.nombre, 'unidad': m.unidad,
            'stock': m.stock_disponible} for m in materias]
    return {'materias_json': mj, 'lotes': lotes}

def _descontar_materias(form):
    """Si el formulario tiene usar_materias=1, descuenta stock de cada materia prima."""
    if not form.get('usar_materias'): return
    mp_ids = form.getlist('mp_id[]')
    mp_cants = form.getlist('mp_cant[]')
    lote_id = form.get('lote_id') or None
    errores = []
    for mid, cant_str in zip(mp_ids, mp_cants):
        if not mid or not cant_str: continue
        try:
            cant = float(cant_str)
            m = MateriaPrima.query.get(int(mid))
            if not m:
                errores.append(f'Materia prima ID {mid} no encontrada')
                continue
            if m.stock_disponible < cant:
                errores.append(f'Stock insuficiente de "{m.nombre}": disponible {m.stock_disponible:.3f} {m.unidad}, solicitado {cant:.3f}')
                continue
            m.stock_disponible -= cant
            m.stock_reservado = max(0, m.stock_reservado - cant)
            if lote_id:
                db.session.add(ReservaProduccion(
                    materia_prima_id=m.id, cantidad=cant,
                    lote_id=int(lote_id), estado='usado',
                    notas='Descontado desde inventario',
                    creado_por=None))
        except Exception as e:
            errores.append(str(e))
    return errores

@app.route('/inventario/nuevo', methods=['GET','POST'])
@login_required
def producto_nuevo():
    if request.method == 'POST':
        fd_cad = request.form.get('fecha_caducidad')
        db.session.add(Producto(
            nombre=request.form['nombre'], descripcion=request.form.get('descripcion',''),
            sku=request.form.get('sku') or None, nso=request.form.get('nso') or None,
            precio=float(request.form.get('precio',0) or 0),
            costo=float(request.form.get('costo',0) or 0),
            stock=int(request.form.get('stock',0) or 0),
            stock_minimo=int(request.form.get('stock_minimo',5) or 5),
            categoria=request.form.get('categoria',''),
            fecha_caducidad=datetime.strptime(fd_cad,'%Y-%m-%d').date() if fd_cad else None))
        db.session.commit()
        flash('Producto creado.','success'); return redirect(url_for('inventario'))
    return render_template('inventario/form.html', obj=None, titulo='Nuevo Producto')

@app.route('/inventario/<int:id>/editar', methods=['GET','POST'])
@login_required
def producto_editar(id):
    obj=Producto.query.get_or_404(id)
    if request.method == 'POST':
        fd_cad = request.form.get('fecha_caducidad')
        obj.nombre=request.form['nombre']; obj.descripcion=request.form.get('descripcion','')
        obj.sku=request.form.get('sku') or None; obj.nso=request.form.get('nso') or None
        obj.precio=float(request.form.get('precio',0) or 0)
        obj.costo=float(request.form.get('costo',0) or 0)
        obj.stock=int(request.form.get('stock',0) or 0)
        obj.stock_minimo=int(request.form.get('stock_minimo',5) or 5)
        obj.categoria=request.form.get('categoria','')
        obj.fecha_caducidad=datetime.strptime(fd_cad,'%Y-%m-%d').date() if fd_cad else None
        db.session.commit()
        flash('Producto actualizado.','success'); return redirect(url_for('inventario'))
    return render_template('inventario/form.html', obj=obj, titulo='Editar Producto')

@app.route('/inventario/<int:id>/eliminar', methods=['POST'])
@login_required
def producto_eliminar(id):
    obj=Producto.query.get_or_404(id); obj.activo=False; db.session.commit()
    flash('Producto eliminado.','info'); return redirect(url_for('inventario'))

# =============================================================
# PRODUCCIÓN
# =============================================================

@app.route('/produccion')
@login_required
def produccion_index():
    from datetime import date
    mes_ini = date.today().replace(day=1)
    total_compras  = db.session.query(db.func.sum(CompraMateria.costo_total)).scalar() or 0
    compras_mes    = db.session.query(db.func.sum(CompraMateria.costo_total)).filter(CompraMateria.fecha >= mes_ini).scalar() or 0
    cotizaciones_vigentes = CotizacionGranel.query.filter_by(estado='vigente').count()
    ordenes_activas = OrdenCompra.query.filter(OrdenCompra.estado.in_(['borrador','enviada','en_transito'])).count()
    compras_recientes = CompraMateria.query.order_by(CompraMateria.fecha.desc()).limit(5).all()
    granel_recientes  = CotizacionGranel.query.order_by(CotizacionGranel.creado_en.desc()).limit(5).all()
    return render_template('produccion/index.html',
        total_compras=total_compras, compras_mes=compras_mes,
        cotizaciones_vigentes=cotizaciones_vigentes, ordenes_activas=ordenes_activas,
        compras_recientes=compras_recientes, granel_recientes=granel_recientes)

# --- Compras de Materia Prima ---

@app.route('/produccion/compras')
@login_required
def compras():
    busqueda=request.args.get('buscar','')
    q=CompraMateria.query
    if busqueda:
        q=q.filter(db.or_(CompraMateria.nombre_item.ilike(f'%{busqueda}%'),
                           CompraMateria.proveedor.ilike(f'%{busqueda}%'),
                           CompraMateria.nro_factura.ilike(f'%{busqueda}%')))
    from datetime import date
    mes_ini = date.today().replace(day=1)
    total_general = db.session.query(db.func.sum(CompraMateria.costo_total)).scalar() or 0
    total_mes = db.session.query(db.func.sum(CompraMateria.costo_total)).filter(CompraMateria.fecha >= mes_ini).scalar() or 0
    return render_template('produccion/compras.html', items=q.order_by(CompraMateria.fecha.desc()).all(),
                           busqueda=busqueda, total_general=total_general, total_mes=total_mes)

def _save_compra(c, form):
    """Helper to parse and save compra fields from form."""
    fd = form.get('fecha')
    cant  = float(form.get('cantidad',1) or 1)
    costo_p = float(form.get('costo_producto',0) or 0)
    imp   = float(form.get('impuestos',0) or 0)
    trans = float(form.get('transporte',0) or 0)
    costo_total = costo_p + imp + trans
    precio_unit = (costo_total / cant) if cant > 0 else 0
    pid   = form.get('producto_id') or None
    mid   = form.get('materia_id') or None
    fvenc = form.get('fecha_caducidad') or None
    c.producto_id   = int(pid) if pid else None
    c.materia_id    = int(mid) if mid else None
    c.nombre_item   = form['nombre_item']
    c.tipo_compra   = form.get('tipo_compra','insumo')
    c.unidad        = form.get('unidad','unidades')
    c.proveedor     = form.get('proveedor','')
    c.proveedor_id  = int(form.get('proveedor_id')) if form.get('proveedor_id') else None
    c.fecha         = datetime.strptime(fd,'%Y-%m-%d').date() if fd else datetime.utcnow().date()
    c.nro_factura   = form.get('nro_factura','')
    c.cantidad      = cant
    c.costo_producto = costo_p; c.impuestos = imp; c.transporte = trans
    c.costo_total   = costo_total; c.precio_unitario = precio_unit
    c.tiene_caducidad = bool(form.get('tiene_caducidad'))
    c.fecha_caducidad = datetime.strptime(fvenc,'%Y-%m-%d').date() if fvenc else None
    c.notas = form.get('notas','')
    # Update linked product cost
    if pid:
        prod = Producto.query.get(int(pid))
        if prod: prod.costo = precio_unit
    # Update materia prima stock if tipo=materia_prima
    if mid and c.tipo_compra == 'materia_prima':
        m = MateriaPrima.query.get(int(mid))
        if m:
            m.stock_disponible = (m.stock_disponible or 0) + cant
    # Auto-register as GastoOperativo
    tipo_label = {'materia_prima': 'Materia prima', 'insumo': 'Insumo',
                  'producto': 'Producto', 'servicio': 'Servicio'}.get(c.tipo_compra, c.tipo_compra.capitalize())
    desc_gasto = (f'{c.nombre_item} — {tipo_label}')
    if c.proveedor: desc_gasto += f' | Proveedor: {c.proveedor}'
    if c.nro_factura: desc_gasto += f' | Factura: {c.nro_factura}'
    gasto = GastoOperativo(
        fecha=c.fecha,
        tipo='compra_produccion',
        tipo_custom=f'Compra / {tipo_label}',
        descripcion=desc_gasto,
        monto=costo_total,
        recurrencia='unica',
        es_plantilla=False,
        notas=f'Registrado automáticamente desde módulo Compras.',
        creado_por=getattr(c, 'creado_por', None)
    )
    db.session.add(gasto)
    return c

@app.route('/produccion/compras/nueva', methods=['GET','POST'])
@login_required
def compra_nueva():
    if request.method == 'POST':
        c = CompraMateria(creado_por=current_user.id)
        _save_compra(c, request.form)
        db.session.add(c); db.session.commit()
        flash('Compra registrada.','success')
        return redirect(url_for('compras'))
    return render_template('produccion/compra_form.html', obj=None, titulo='Nueva Compra',
                           productos=Producto.query.filter_by(activo=True).order_by(Producto.nombre).all(),
                           materias=MateriaPrima.query.filter_by(activo=True).order_by(MateriaPrima.nombre).all(),
                           proveedores_list=Proveedor.query.filter_by(activo=True).order_by(Proveedor.empresa).all(),
                           today=datetime.utcnow().strftime('%Y-%m-%d'))

@app.route('/produccion/compras/<int:id>/editar', methods=['GET','POST'])
@login_required
def compra_editar(id):
    obj=CompraMateria.query.get_or_404(id)
    if request.method == 'POST':
        _save_compra(obj, request.form)
        db.session.commit()
        flash('Compra actualizada.','success')
        return redirect(url_for('compras'))
    return render_template('produccion/compra_form.html', obj=obj, titulo='Editar Compra',
                           productos=Producto.query.filter_by(activo=True).order_by(Producto.nombre).all(),
                           materias=MateriaPrima.query.filter_by(activo=True).order_by(MateriaPrima.nombre).all(),
                           proveedores_list=Proveedor.query.filter_by(activo=True).order_by(Proveedor.empresa).all(),
                           today=datetime.utcnow().strftime('%Y-%m-%d'))

@app.route('/produccion/compras/<int:id>/eliminar', methods=['POST'])
@login_required
def compra_eliminar(id):
    obj=CompraMateria.query.get_or_404(id); db.session.delete(obj); db.session.commit()
    flash('Compra eliminada.','info'); return redirect(url_for('compras'))

# --- Cotizaciones Granel ---

@app.route('/produccion/granel')
@login_required
def granel():
    estado_f=request.args.get('estado','')
    q=CotizacionGranel.query
    if estado_f: q=q.filter_by(estado=estado_f)
    return render_template('produccion/granel.html', items=q.order_by(CotizacionGranel.creado_en.desc()).all(),
                           estado_f=estado_f)

@app.route('/produccion/granel/nueva', methods=['GET','POST'])
@login_required
def granel_nuevo():
    if request.method == 'POST':
        fc = request.form.get('fecha_cotizacion')
        fv = request.form.get('vigencia')
        pid = request.form.get('producto_id') or None
        estado = request.form.get('estado','vigente')
        precio_u = float(request.form.get('precio_unitario',0) or 0)
        g = CotizacionGranel(
            producto_id=int(pid) if pid else None,
            nombre_producto=request.form['nombre_producto'],
            sku=request.form.get('sku',''), nso=request.form.get('nso',''),
            proveedor=request.form.get('proveedor',''),
            precio_unitario=precio_u,
            unidades_minimas=int(request.form.get('unidades_minimas',1) or 1),
            fecha_cotizacion=datetime.strptime(fc,'%Y-%m-%d').date() if fc else None,
            vigencia=datetime.strptime(fv,'%Y-%m-%d').date() if fv else None,
            estado=estado, notas=request.form.get('notas',''), creado_por=current_user.id)
        db.session.add(g)
        if pid and estado == 'vigente':
            prod = Producto.query.get(int(pid))
            if prod: prod.costo = precio_u
        db.session.commit()
        flash('Cotización guardada.','success')
        return redirect(url_for('granel'))
    return render_template('produccion/granel_form.html', obj=None, titulo='Nueva Cotización Granel',
                           productos=Producto.query.filter_by(activo=True).order_by(Producto.nombre).all(),
                           today=datetime.utcnow().strftime('%Y-%m-%d'))

@app.route('/produccion/granel/<int:id>/editar', methods=['GET','POST'])
@login_required
def granel_editar(id):
    obj=CotizacionGranel.query.get_or_404(id)
    if request.method == 'POST':
        fc = request.form.get('fecha_cotizacion')
        fv = request.form.get('vigencia')
        pid = request.form.get('producto_id') or None
        estado = request.form.get('estado','vigente')
        precio_u = float(request.form.get('precio_unitario',0) or 0)
        obj.producto_id=int(pid) if pid else None
        obj.nombre_producto=request.form['nombre_producto']
        obj.sku=request.form.get('sku',''); obj.nso=request.form.get('nso','')
        obj.proveedor=request.form.get('proveedor',''); obj.precio_unitario=precio_u
        obj.unidades_minimas=int(request.form.get('unidades_minimas',1) or 1)
        obj.fecha_cotizacion=datetime.strptime(fc,'%Y-%m-%d').date() if fc else None
        obj.vigencia=datetime.strptime(fv,'%Y-%m-%d').date() if fv else None
        obj.estado=estado; obj.notas=request.form.get('notas','')
        if pid and estado == 'vigente':
            prod = Producto.query.get(int(pid))
            if prod: prod.costo = precio_u
        db.session.commit()
        flash('Cotización actualizada.','success')
        return redirect(url_for('granel'))
    return render_template('produccion/granel_form.html', obj=obj, titulo='Editar Cotización Granel',
                           productos=Producto.query.filter_by(activo=True).order_by(Producto.nombre).all(),
                           today=datetime.utcnow().strftime('%Y-%m-%d'))

@app.route('/produccion/granel/<int:id>/eliminar', methods=['POST'])
@login_required
def granel_eliminar(id):
    obj=CotizacionGranel.query.get_or_404(id); db.session.delete(obj); db.session.commit()
    flash('Cotización eliminada.','info'); return redirect(url_for('granel'))

# --- Helper: cálculo de impuestos estimados basado en reglas activas ---
def _calcular_impuestos(ingresos, utilidad):
    """Retorna (total_impuestos, lista_detalle) según reglas tributarias activas.
    Cada item del detalle: {nombre, aplica_a, porcentaje, base_label, base_monto, monto}
    No incluye reglas de proveedor (se aplican por compra individual)."""
    reglas = ReglaTributaria.query.filter_by(activo=True).all()
    total = 0.0
    detalle = []
    for r in reglas:
        if r.aplica_a in ('proveedor_producto', 'proveedor_granel'):
            continue  # estas aplican por compra, no de forma global
        base_label = ''
        base_monto = 0.0
        monto = 0.0
        if r.aplica_a == 'ventas':
            base_monto = ingresos
            base_label = 'Ingresos brutos'
        elif r.aplica_a == 'ingresos':
            base_monto = ingresos
            base_label = 'Ingresos brutos'
        elif r.aplica_a == 'profit':
            if utilidad > 0:
                base_monto = utilidad
                base_label = 'Utilidad'
            else:
                continue
        monto = base_monto * (r.porcentaje / 100.0)
        total += monto
        detalle.append({
            'nombre': r.nombre,
            'aplica_a': r.aplica_a,
            'porcentaje': r.porcentaje,
            'base_label': base_label,
            'base_monto': base_monto,
            'monto': monto,
        })
    return total, detalle

# --- Reglas Tributarias ---

@app.route('/finanzas/impuestos')
@login_required
def impuestos():
    return render_template('finanzas/impuestos.html',
                           items=ReglaTributaria.query.order_by(ReglaTributaria.nombre).all())

@app.route('/finanzas/impuestos/nuevo', methods=['GET','POST'])
@login_required
def impuesto_nuevo():
    if request.method == 'POST':
        db.session.add(ReglaTributaria(
            nombre=request.form['nombre'],
            descripcion=request.form.get('descripcion',''),
            porcentaje=float(request.form.get('porcentaje',0) or 0),
            aplica_a=request.form.get('aplica_a','ventas'),
            proveedor_nombre=request.form.get('proveedor_nombre','') or None,
            activo=True))
        db.session.commit(); flash('Regla tributaria creada.','success')
        return redirect(url_for('impuestos'))
    return render_template('finanzas/impuesto_form.html', obj=None, titulo='Nueva Regla Tributaria')

@app.route('/finanzas/impuestos/<int:id>/editar', methods=['GET','POST'])
@login_required
def impuesto_editar(id):
    obj=ReglaTributaria.query.get_or_404(id)
    if request.method == 'POST':
        obj.nombre=request.form['nombre']
        obj.descripcion=request.form.get('descripcion','')
        obj.porcentaje=float(request.form.get('porcentaje',0) or 0)
        obj.aplica_a=request.form.get('aplica_a','ventas')
        obj.proveedor_nombre=request.form.get('proveedor_nombre','') or None
        obj.activo = request.form.get('activo') == '1'
        db.session.commit(); flash('Regla actualizada.','success')
        return redirect(url_for('impuestos'))
    return render_template('finanzas/impuesto_form.html', obj=obj, titulo='Editar Regla Tributaria')

@app.route('/finanzas/impuestos/<int:id>/eliminar', methods=['POST'])
@login_required
def impuesto_eliminar(id):
    obj=ReglaTributaria.query.get_or_404(id); db.session.delete(obj); db.session.commit()
    flash('Regla eliminada.','info'); return redirect(url_for('impuestos'))

# =============================================================
# GASTOS OPERATIVOS
# =============================================================

@app.route('/gastos')
@login_required
def gastos():
    from datetime import date as date_t
    tipo_f  = request.args.get('tipo','')
    desde_f = request.args.get('desde','')
    hasta_f = request.args.get('hasta','')
    try:
        q = GastoOperativo.query.filter_by(es_plantilla=False)
        if tipo_f:  q = q.filter_by(tipo=tipo_f)
        if desde_f: q = q.filter(GastoOperativo.fecha >= datetime.strptime(desde_f,'%Y-%m-%d').date())
        if hasta_f: q = q.filter(GastoOperativo.fecha <= datetime.strptime(hasta_f,'%Y-%m-%d').date())
        items = q.order_by(GastoOperativo.fecha.desc()).all()
        total_g   = db.session.query(db.func.sum(GastoOperativo.monto)).filter_by(es_plantilla=False).scalar() or 0
        mes_ini   = date_t.today().replace(day=1)
        total_mes = db.session.query(db.func.sum(GastoOperativo.monto)).filter(
            GastoOperativo.es_plantilla==False, GastoOperativo.fecha>=mes_ini).scalar() or 0
        tipos     = [t[0] for t in db.session.query(GastoOperativo.tipo).filter_by(es_plantilla=False).distinct().order_by(GastoOperativo.tipo).all()]
        plantillas = GastoOperativo.query.filter_by(es_plantilla=True).order_by(GastoOperativo.tipo).all()
        total_reg = GastoOperativo.query.filter_by(es_plantilla=False).count()
    except Exception:
        db.session.rollback()
        q2 = GastoOperativo.query
        if tipo_f: q2 = q2.filter_by(tipo=tipo_f)
        items = q2.order_by(GastoOperativo.fecha.desc()).all()
        total_g = db.session.query(db.func.sum(GastoOperativo.monto)).scalar() or 0
        mes_ini = date_t.today().replace(day=1)
        total_mes = db.session.query(db.func.sum(GastoOperativo.monto)).filter(GastoOperativo.fecha>=mes_ini).scalar() or 0
        tipos = [t[0] for t in db.session.query(GastoOperativo.tipo).distinct().order_by(GastoOperativo.tipo).all()]
        plantillas = []; total_reg = len(items)
    return render_template('gastos/index.html', items=items, tipo_f=tipo_f,
        desde_f=desde_f, hasta_f=hasta_f, total_general=total_g,
        total_mes=total_mes, total_registros=total_reg,
        tipos=tipos, plantillas=plantillas)

@app.route('/gastos/nuevo', methods=['GET','POST'])
@login_required
def gasto_nuevo():
    if request.method == 'POST':
        fd = request.form.get('fecha')
        rec = request.form.get('recurrencia','unico')
        es_pl = request.form.get('es_plantilla') == '1' and rec == 'mensual'
        db.session.add(GastoOperativo(
            fecha=datetime.strptime(fd,'%Y-%m-%d').date() if fd else datetime.utcnow().date(),
            tipo=request.form['tipo'],
            tipo_custom=request.form.get('tipo_custom','') or None,
            descripcion=request.form.get('descripcion',''),
            monto=float(request.form.get('monto',0) or 0),
            recurrencia=rec,
            es_plantilla=es_pl,
            notas=request.form.get('notas',''), creado_por=current_user.id))
        db.session.commit(); flash('Gasto registrado.','success'); return redirect(url_for('gastos'))
    return render_template('gastos/form.html', obj=None, titulo='Nuevo Gasto',
                           today=datetime.utcnow().strftime('%Y-%m-%d'))

@app.route('/gastos/<int:id>/editar', methods=['GET','POST'])
@login_required
def gasto_editar(id):
    obj = GastoOperativo.query.get_or_404(id)
    if request.method == 'POST':
        fd = request.form.get('fecha')
        rec = request.form.get('recurrencia','unico')
        obj.fecha=datetime.strptime(fd,'%Y-%m-%d').date() if fd else obj.fecha
        obj.tipo=request.form['tipo']
        obj.tipo_custom=request.form.get('tipo_custom','') or None
        obj.descripcion=request.form.get('descripcion','')
        obj.monto=float(request.form.get('monto',0) or 0)
        obj.recurrencia=rec
        obj.es_plantilla=request.form.get('es_plantilla') == '1' and rec == 'mensual'
        obj.notas=request.form.get('notas','')
        db.session.commit(); flash('Gasto actualizado.','success'); return redirect(url_for('gastos'))
    return render_template('gastos/form.html', obj=obj, titulo='Editar Gasto',
                           today=datetime.utcnow().strftime('%Y-%m-%d'))

@app.route('/gastos/<int:id>/eliminar', methods=['POST'])
@login_required
def gasto_eliminar(id):
    obj=GastoOperativo.query.get_or_404(id); db.session.delete(obj); db.session.commit()
    flash('Gasto eliminado.','info'); return redirect(url_for('gastos'))

@app.route('/gastos/plantilla/<int:id>/usar', methods=['POST'])
@login_required
def gasto_plantilla_usar(id):
    from datetime import date as date_t
    plantilla = GastoOperativo.query.get_or_404(id)
    nuevo = GastoOperativo(
        fecha=date_t.today(),
        tipo=plantilla.tipo,
        tipo_custom=plantilla.tipo_custom,
        descripcion=plantilla.descripcion,
        monto=plantilla.monto,
        recurrencia='mensual',
        es_plantilla=False,
        notas=f'Registrado desde plantilla mensual',
        creado_por=current_user.id)
    db.session.add(nuevo); db.session.commit()
    flash(f'Gasto "{plantilla.tipo_custom or plantilla.tipo}" registrado para este mes.','success')
    return redirect(url_for('gastos'))

# =============================================================
# ADMIN
# =============================================================

@app.route('/admin/usuarios')
@login_required
def admin_usuarios():
    if current_user.rol != 'admin':
        flash('Sin permisos.','danger'); return redirect(url_for('dashboard'))
    return render_template('admin/usuarios.html', items=User.query.order_by(User.nombre).all())

@app.route('/admin/usuarios/nuevo', methods=['GET','POST'])
@login_required
def admin_usuario_nuevo():
    if current_user.rol != 'admin':
        flash('Sin permisos.','danger'); return redirect(url_for('dashboard'))
    if request.method == 'POST':
        _pwd = request.form.get('password','')
        if len(_pwd) < 8:
            flash('La contraseña debe tener al menos 8 caracteres.','danger')
        elif User.query.filter_by(email=request.form['email']).first():
            flash('Ya existe ese email.','danger')
        else:
            modulos_sel = request.form.getlist('modulos')
            u=User(nombre=request.form['nombre'],email=request.form['email'],
                   rol=request.form.get('rol','usuario'),
                   modulos_permitidos=json.dumps(modulos_sel) if modulos_sel else '[]')
            u.set_password(_pwd); db.session.add(u); db.session.commit()
            flash('Usuario creado.','success'); return redirect(url_for('admin_usuarios'))
    return render_template('admin/usuario_form.html', obj=None, titulo='Nuevo Usuario')

@app.route('/admin/usuarios/<int:id>/toggle', methods=['POST'])
@login_required
def admin_usuario_toggle(id):
    if current_user.rol != 'admin':
        flash('Sin permisos.','danger'); return redirect(url_for('dashboard'))
    u=User.query.get_or_404(id)
    if u.id != current_user.id:
        u.activo=not u.activo; db.session.commit()
        flash(f'Usuario {"activado" if u.activo else "desactivado"}.','info')
    return redirect(url_for('admin_usuarios'))

# =============================================================
# PERFIL DE USUARIO
# =============================================================

@app.route('/perfil', methods=['GET','POST'])
@login_required
def perfil():
    if request.method == 'POST':
        accion = request.form.get('accion')
        if accion == 'datos':
            nuevo_email = request.form.get('email','').strip()
            if nuevo_email != current_user.email and User.query.filter_by(email=nuevo_email).first():
                flash('Ese email ya está en uso.','danger')
            else:
                current_user.nombre = request.form.get('nombre','').strip() or current_user.nombre
                current_user.email  = nuevo_email or current_user.email
                db.session.commit()
                flash('Datos actualizados.','success')
        elif accion == 'password':
            pw_actual    = request.form.get('password_actual','')
            pw_nueva     = request.form.get('password_nueva','')
            pw_confirmar = request.form.get('password_confirmar','')
            if not current_user.check_password(pw_actual):
                flash('La contraseña actual es incorrecta.','danger')
            elif len(pw_nueva) < 8:
                flash('La nueva contraseña debe tener al menos 8 caracteres.','danger')
            elif pw_nueva != pw_confirmar:
                flash('Las contraseñas nuevas no coinciden.','danger')
            else:
                current_user.set_password(pw_nueva)
                db.session.commit()
                flash('Contraseña cambiada exitosamente.','success')
    return render_template('perfil.html')

# =============================================================
# REPORTES
# =============================================================

@app.route('/reportes')
@login_required
def reportes():
    from datetime import date
    from calendar import month_abbr
    # Estadísticas generales
    ingresos_totales = db.session.query(db.func.sum(Venta.total)).filter(Venta.estado.in_(['ganado','anticipo_pagado'])).scalar() or 0
    gastos_totales   = db.session.query(db.func.sum(GastoOperativo.monto)).scalar() or 0
    balance          = ingresos_totales - gastos_totales
    total_clientes   = Cliente.query.filter_by(estado='activo').count()
    # Ventas por mes (últimos 6 meses)
    hoy = date.today()
    meses_labels, ventas_por_mes = [], []
    for i in range(5, -1, -1):
        mes = (hoy.month - i - 1) % 12 + 1
        anio = hoy.year - ((hoy.month - i - 1) // 12 + (1 if (hoy.month - i - 1) < 0 else 0))
        total_mes = db.session.query(db.func.sum(Venta.total)).filter(
            db.extract('month', Venta.creado_en) == mes,
            db.extract('year', Venta.creado_en) == anio).scalar() or 0
        meses_labels.append(f'{month_abbr[mes]} {str(anio)[2:]}')
        ventas_por_mes.append(round(total_mes))
    # Gastos por tipo
    gastos_por_tipo = db.session.query(GastoOperativo.tipo, db.func.sum(GastoOperativo.monto))\
        .group_by(GastoOperativo.tipo).order_by(db.func.sum(GastoOperativo.monto).desc()).all()
    gastos_tipos_labels = [g[0] for g in gastos_por_tipo]
    gastos_tipos_values = [round(g[1]) for g in gastos_por_tipo]
    # Top 5 clientes por ventas totales
    from sqlalchemy import func as sqlfunc
    top_q = db.session.query(
        Cliente.id, Cliente.nombre, Cliente.empresa,
        sqlfunc.sum(Venta.total).label('total_ventas')
    ).join(Venta, Venta.cliente_id == Cliente.id)\
     .group_by(Cliente.id, Cliente.nombre, Cliente.empresa)\
     .order_by(sqlfunc.sum(Venta.total).desc()).limit(5).all()
    class _C:
        def __init__(self, r): self.id=r[0]; self.nombre=r[1]; self.empresa=r[2]; self.total_ventas=r[3]
    top_clientes = [_C(r) for r in top_q]
    # Stock bajo
    bajo_stock = Producto.query.filter(Producto.activo==True, Producto.stock<=Producto.stock_minimo).all()
    return render_template('reportes.html',
        total_clientes=total_clientes, ingresos_totales=ingresos_totales,
        gastos_totales=gastos_totales, balance=balance,
        meses_labels=meses_labels, ventas_por_mes=ventas_por_mes,
        gastos_tipos_labels=gastos_tipos_labels, gastos_tipos_values=gastos_tipos_values,
        top_clientes=top_clientes, bajo_stock=bajo_stock)

# --- Exportar Excel ---

def _make_xlsx(titulo, headers, rows):
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo
    # Cabecera empresa
    ws.merge_cells(f'A1:{chr(64+len(headers))}1')
    ws['A1'] = 'Evore CRM — ' + titulo
    ws['A1'].font = Font(bold=True, size=13, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor='5E72E4')
    ws['A1'].alignment = Alignment(horizontal='center')
    # Headers
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1A1F36')
        cell.alignment = Alignment(horizontal='center')
    # Datos
    for r_idx, row in enumerate(rows, 3):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    # Autofit columns
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

@app.route('/reportes/exportar/ventas.xlsx')
@login_required
def exportar_ventas():
    from flask import send_file
    ventas = Venta.query.order_by(Venta.creado_en.desc()).all()
    headers = ['Título','Cliente','Subtotal COP','IVA COP','Total COP','% Anticipo','Anticipo COP','Saldo COP','Estado','Fecha anticipo','Días entrega','Creada']
    rows = []
    for v in ventas:
        rows.append([
            v.titulo,
            v.cliente.empresa or v.cliente.nombre if v.cliente else '',
            round(v.subtotal), round(v.iva), round(v.total),
            v.porcentaje_anticipo, round(v.monto_anticipo), round(v.saldo),
            v.estado,
            v.fecha_anticipo.strftime('%d/%m/%Y') if v.fecha_anticipo else '',
            v.dias_entrega,
            v.creado_en.strftime('%d/%m/%Y')
        ])
    buf = _make_xlsx('Ventas', headers, rows)
    return send_file(buf, download_name='evore_ventas.xlsx',
                     as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/reportes/exportar/gastos.xlsx')
@login_required
def exportar_gastos():
    from flask import send_file
    items = GastoOperativo.query.order_by(GastoOperativo.fecha.desc()).all()
    headers = ['Fecha','Tipo','Descripción','Monto COP','Notas']
    rows = [[g.fecha.strftime('%d/%m/%Y'), g.tipo, g.descripcion or '', round(g.monto), g.notas or ''] for g in items]
    buf = _make_xlsx('Gastos Operativos', headers, rows)
    return send_file(buf, download_name='evore_gastos.xlsx',
                     as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/reportes/exportar/inventario.xlsx')
@login_required
def exportar_inventario():
    from flask import send_file
    items = Producto.query.filter_by(activo=True).order_by(Producto.nombre).all()
    headers = ['Nombre','SKU','NSO (INVIMA)','Precio COP','Costo COP','Stock','Stock Mínimo','Categoría']
    rows = [[p.nombre, p.sku or '', p.nso or '', round(p.precio), round(p.costo),
             p.stock, p.stock_minimo, p.categoria or ''] for p in items]
    buf = _make_xlsx('Inventario', headers, rows)
    return send_file(buf, download_name='evore_inventario.xlsx',
                     as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/reportes/exportar/clientes.xlsx')
@login_required
def exportar_clientes():
    from flask import send_file
    items = Cliente.query.order_by(Cliente.empresa, Cliente.nombre).all()
    headers = ['Empresa','NIT','Relación','Dirección comercial','Dirección entrega','Estado','Creado']
    rows = [[c.empresa or '', c.nit or '', c.estado_relacion or '', c.dir_comercial or '',
             c.dir_entrega or '', c.estado, c.creado_en.strftime('%d/%m/%Y')] for c in items]
    buf = _make_xlsx('Clientes', headers, rows)
    return send_file(buf, download_name='evore_clientes.xlsx',
                     as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# =============================================================
# HELPER: LOG DE ACTIVIDAD
# =============================================================

def _log(tipo, entidad, entidad_id, descripcion):
    try:
        db.session.add(Actividad(
            tipo=tipo, entidad=entidad, entidad_id=entidad_id,
            descripcion=descripcion, usuario_id=current_user.id))
    except Exception:
        pass

# =============================================================
# NOTAS RÁPIDAS
# =============================================================

@app.route('/notas')
@login_required
def notas():
    cliente_f = request.args.get('cliente_id','')
    q = Nota.query
    if cliente_f: q = q.filter_by(cliente_id=int(cliente_f))
    return render_template('notas/index.html',
        items=q.order_by(Nota.actualizado_en.desc()).all(),
        clientes_list=Cliente.query.order_by(Cliente.empresa, Cliente.nombre).all(),
        productos_list=Producto.query.filter_by(activo=True).order_by(Producto.nombre).all(),
        cliente_f=cliente_f)

@app.route('/notas/nueva', methods=['GET','POST'])
@login_required
def nota_nueva():
    cl = Cliente.query.order_by(Cliente.empresa, Cliente.nombre).all()
    pl = Producto.query.filter_by(activo=True).order_by(Producto.nombre).all()
    if request.method == 'POST':
        fd_rev = request.form.get('fecha_revision')
        n = Nota(titulo=request.form.get('titulo','').strip() or None,
            contenido=request.form['contenido'],
            cliente_id=request.form.get('cliente_id') or None,
            producto_id=request.form.get('producto_id') or None,
            modulo=request.form.get('modulo','') or None,
            fecha_revision=datetime.strptime(fd_rev,'%Y-%m-%d').date() if fd_rev else None,
            creado_por=current_user.id)
        db.session.add(n); db.session.commit()
        _log('crear','nota',n.id,f'Nota creada: {n.titulo or "(sin título)"}'); db.session.commit()
        flash('Nota guardada.','success'); return redirect(url_for('notas'))
    return render_template('notas/form.html', obj=None, titulo='Nueva Nota',
        clientes_list=cl, productos_list=pl)

@app.route('/notas/<int:id>/editar', methods=['GET','POST'])
@login_required
def nota_editar(id):
    obj = Nota.query.get_or_404(id)
    cl  = Cliente.query.order_by(Cliente.empresa, Cliente.nombre).all()
    pl  = Producto.query.filter_by(activo=True).order_by(Producto.nombre).all()
    if request.method == 'POST':
        fd_rev = request.form.get('fecha_revision')
        obj.titulo=request.form.get('titulo','').strip() or None
        obj.contenido=request.form['contenido']
        obj.cliente_id=request.form.get('cliente_id') or None
        obj.producto_id=request.form.get('producto_id') or None
        obj.modulo=request.form.get('modulo','') or None
        obj.fecha_revision=datetime.strptime(fd_rev,'%Y-%m-%d').date() if fd_rev else None
        obj.actualizado_en=datetime.utcnow()
        db.session.commit()
        _log('editar','nota',obj.id,f'Nota editada: {obj.titulo or "(sin título)"}'); db.session.commit()
        flash('Nota actualizada.','success'); return redirect(url_for('notas'))
    return render_template('notas/form.html', obj=obj, titulo='Editar Nota',
        clientes_list=cl, productos_list=pl)

@app.route('/notas/<int:id>/eliminar', methods=['POST'])
@login_required
def nota_eliminar(id):
    obj=Nota.query.get_or_404(id); db.session.delete(obj); db.session.commit()
    flash('Nota eliminada.','info'); return redirect(url_for('notas'))

# =============================================================
# CALENDARIO
# =============================================================

@app.route('/calendario')
@login_required
def calendario():
    from datetime import date as date_t
    hoy = date_t.today()
    anio = int(request.args.get('anio', hoy.year))
    mes  = int(request.args.get('mes',  hoy.month))
    eventos = {}
    # 1. Tareas con fecha de vencimiento
    try:
        for t in Tarea.query.filter(Tarea.fecha_vencimiento != None).all():
            k = t.fecha_vencimiento.strftime('%Y-%m-%d')
            eventos.setdefault(k, []).append({'t':'tarea','n':t.titulo,'s':t.estado})
    except Exception: db.session.rollback()
    # 2. Ventas con fecha entrega estimada
    try:
        for v in Venta.query.filter(Venta.fecha_entrega_est != None).all():
            k = v.fecha_entrega_est.strftime('%Y-%m-%d')
            eventos.setdefault(k, []).append({'t':'venta','n':v.titulo,'s':v.estado})
    except Exception: db.session.rollback()
    # 3. Eventos manuales
    try:
        for e in Evento.query.all():
            k = e.fecha.strftime('%Y-%m-%d')
            eventos.setdefault(k, []).append({'t':'evento','n':e.titulo,'s':e.tipo})
    except Exception: db.session.rollback()
    # 4. Notas con fecha de revisión (columna nueva en v11)
    try:
        for n in Nota.query.filter(Nota.fecha_revision != None).all():
            k = n.fecha_revision.strftime('%Y-%m-%d')
            eventos.setdefault(k, []).append({'t':'nota','n':n.titulo or '(nota sin título)','s':'revision'})
    except Exception: db.session.rollback()
    # 5. Productos con fecha de caducidad (columna nueva en v11)
    try:
        for p in Producto.query.filter(Producto.fecha_caducidad != None, Producto.activo == True).all():
            k = p.fecha_caducidad.strftime('%Y-%m-%d')
            eventos.setdefault(k, []).append({'t':'caducidad','n':p.nombre,'s':'caducidad'})
    except Exception: db.session.rollback()
    import calendar as cal_mod
    from datetime import date as _date
    # Build week matrix: list of 7-element lists; 0 = filler day from other month
    cal_obj = cal_mod.Calendar(firstweekday=0)   # 0 = Monday first
    month_days = cal_obj.monthdayscalendar(anio, mes)
    # Compute prev / next month for nav links
    prev_mes  = mes - 1 if mes > 1 else 12
    prev_anio = anio if mes > 1 else anio - 1
    next_mes  = mes + 1 if mes < 12 else 1
    next_anio = anio if mes < 12 else anio + 1
    today_str = _date.today().strftime('%Y-%m-%d')
    mes_nombres = ['','Enero','Febrero','Marzo','Abril','Mayo','Junio',
                   'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
    return render_template('calendario.html',
                           eventos=eventos, anio=anio, mes=mes,
                           month_days=month_days,
                           prev_mes=prev_mes, prev_anio=prev_anio,
                           next_mes=next_mes, next_anio=next_anio,
                           today_str=today_str,
                           mes_nombre=mes_nombres[mes])

# =============================================================
# EVENTOS
# =============================================================

@app.route('/eventos/nuevo', methods=['GET','POST'])
@login_required
def evento_nuevo():
    if request.method == 'POST':
        fd = request.form.get('fecha')
        ev = Evento(
            titulo=request.form['titulo'],
            tipo=request.form.get('tipo','recordatorio'),
            fecha=datetime.strptime(fd,'%Y-%m-%d').date() if fd else datetime.utcnow().date(),
            hora_inicio=request.form.get('hora_inicio','') or None,
            hora_fin=request.form.get('hora_fin','') or None,
            descripcion=request.form.get('descripcion',''),
            usuario_id=current_user.id)
        db.session.add(ev); db.session.commit()
        flash('Evento creado.','success')
        return redirect(url_for('calendario'))
    return redirect(url_for('calendario'))

@app.route('/eventos/<int:id>/editar', methods=['POST'])
@login_required
def evento_editar(id):
    obj = Evento.query.get_or_404(id)
    fd = request.form.get('fecha')
    obj.titulo = request.form.get('titulo', obj.titulo)
    obj.tipo = request.form.get('tipo', obj.tipo)
    if fd: obj.fecha = datetime.strptime(fd,'%Y-%m-%d').date()
    obj.hora_inicio = request.form.get('hora_inicio','') or None
    obj.hora_fin    = request.form.get('hora_fin','') or None
    obj.descripcion = request.form.get('descripcion','')
    db.session.commit(); flash('Evento actualizado.','success')
    return redirect(url_for('calendario'))

@app.route('/eventos/<int:id>/eliminar', methods=['POST'])
@login_required
def evento_eliminar(id):
    obj = Evento.query.get_or_404(id)
    db.session.delete(obj); db.session.commit()
    flash('Evento eliminado.','info')
    return redirect(url_for('calendario'))

# =============================================================
# FACTURA / COTIZACIÓN
# =============================================================

@app.route('/ventas/<int:id>/factura')
@login_required
def venta_factura(id):
    obj = Venta.query.get_or_404(id)
    empresa = ConfigEmpresa.query.first()
    if not empresa:
        empresa = ConfigEmpresa(nombre='Evore')
    doc_tipo = 'COTIZACIÓN' if obj.estado in ('prospecto','negociacion') else 'FACTURA'
    doc_numero = f'EV-{obj.creado_en.year}-{obj.id:04d}'
    fecha_doc = obj.creado_en.strftime('%d/%m/%Y')
    return render_template('ventas/factura.html',
        obj=obj, empresa=empresa, doc_tipo=doc_tipo,
        doc_numero=doc_numero, fecha_doc=fecha_doc)

# =============================================================
# COTIZACIONES
# =============================================================

@app.route('/cotizaciones')
@login_required
def cotizaciones():
    items = Cotizacion.query.order_by(Cotizacion.fecha_emision.desc()).all()
    return render_template('cotizaciones/index.html', items=items)

@app.route('/cotizaciones/nueva', methods=['GET','POST'])
@login_required
def cotizacion_nueva():
    from datetime import date as date_t
    clientes_list = Cliente.query.order_by(Cliente.empresa, Cliente.nombre).all()
    regla_iva = ReglaTributaria.query.filter_by(aplica_a='ventas', activo=True).first()
    iva_default = regla_iva.porcentaje if regla_iva else 19.0
    if request.method == 'POST':
        hoy = date_t.today()
        # Generar número secuencial
        ultimo = Cotizacion.query.filter(
            Cotizacion.numero.like(f'COT-{hoy.year}-%')
        ).order_by(Cotizacion.id.desc()).first()
        if ultimo and ultimo.numero:
            try: seq = int(ultimo.numero.split('-')[-1]) + 1
            except: seq = 1
        else: seq = 1
        numero = f'COT-{hoy.year}-{seq:03d}'
        fd_em = request.form.get('fecha_emision')
        fd_val = request.form.get('fecha_validez')
        iva_pct = float(request.form.get('iva_pct', iva_default) or iva_default)
        nombres = request.form.getlist('item_nombre[]')
        cantidades = request.form.getlist('item_cantidad[]')
        precios = request.form.getlist('item_precio[]')
        items_data = []
        subtotal = 0.0
        for i in range(len(nombres)):
            nm = nombres[i].strip() if i < len(nombres) else ''
            if not nm: continue
            cant = float(cantidades[i]) if i < len(cantidades) else 1.0
            precio = float(precios[i]) if i < len(precios) else 0.0
            sub = cant * precio
            subtotal += sub
            items_data.append({'nombre': nm, 'cantidad': cant, 'precio': precio, 'subtotal': sub})
        iva_monto = subtotal * iva_pct / 100.0
        total = subtotal + iva_monto
        pct_anticipo = float(request.form.get('porcentaje_anticipo', 50) or 50)
        monto_anticipo = total * pct_anticipo / 100.0
        saldo = total - monto_anticipo
        cot = Cotizacion(
            numero=numero,
            titulo=request.form['titulo'],
            cliente_id=request.form.get('cliente_id') or None,
            subtotal=subtotal, iva=iva_monto, total=total,
            porcentaje_anticipo=pct_anticipo,
            monto_anticipo=monto_anticipo, saldo=saldo,
            fecha_emision=datetime.strptime(fd_em,'%Y-%m-%d').date() if fd_em else date_t.today(),
            fecha_validez=datetime.strptime(fd_val,'%Y-%m-%d').date() if fd_val else None,
            dias_entrega=int(request.form.get('dias_entrega',30) or 30),
            condiciones_pago=request.form.get('condiciones_pago',''),
            notas=request.form.get('notas',''),
            estado='borrador', creado_por=current_user.id)
        db.session.add(cot); db.session.flush()
        for it in items_data:
            db.session.add(CotizacionItem(
                cotizacion_id=cot.id, nombre_prod=it['nombre'],
                cantidad=it['cantidad'], precio_unit=it['precio'], subtotal=it['subtotal']))
        db.session.commit()
        _log('crear','cotizacion',cot.id,f'Cotización {numero}: {cot.titulo}'); db.session.commit()
        flash(f'Cotización {numero} creada.','success')
        return redirect(url_for('cotizacion_ver', id=cot.id))
    return render_template('cotizaciones/form.html', obj=None, titulo='Nueva Cotización',
        clientes_list=clientes_list, today=datetime.utcnow().strftime('%Y-%m-%d'),
        iva_default=iva_default)

@app.route('/cotizaciones/<int:id>')
@login_required
def cotizacion_ver(id):
    obj = Cotizacion.query.get_or_404(id)
    empresa = ConfigEmpresa.query.first() or ConfigEmpresa(nombre='Evore')
    return render_template('cotizaciones/ver.html', obj=obj, empresa=empresa)

@app.route('/cotizaciones/<int:id>/editar', methods=['GET','POST'])
@login_required
def cotizacion_editar(id):
    from datetime import date as date_t
    obj = Cotizacion.query.get_or_404(id)
    clientes_list = Cliente.query.order_by(Cliente.empresa, Cliente.nombre).all()
    regla_iva = ReglaTributaria.query.filter_by(aplica_a='ventas', activo=True).first()
    iva_default = regla_iva.porcentaje if regla_iva else 19.0
    if request.method == 'POST':
        fd_em = request.form.get('fecha_emision')
        fd_val = request.form.get('fecha_validez')
        iva_pct = float(request.form.get('iva_pct', iva_default) or iva_default)
        nombres = request.form.getlist('item_nombre[]')
        cantidades = request.form.getlist('item_cantidad[]')
        precios = request.form.getlist('item_precio[]')
        # Borrar items existentes
        for it in obj.items: db.session.delete(it)
        db.session.flush()
        subtotal = 0.0
        for i in range(len(nombres)):
            nm = nombres[i].strip() if i < len(nombres) else ''
            if not nm: continue
            cant = float(cantidades[i]) if i < len(cantidades) else 1.0
            precio = float(precios[i]) if i < len(precios) else 0.0
            sub = cant * precio
            subtotal += sub
            db.session.add(CotizacionItem(
                cotizacion_id=obj.id, nombre_prod=nm,
                cantidad=cant, precio_unit=precio, subtotal=sub))
        iva_monto = subtotal * iva_pct / 100.0
        total = subtotal + iva_monto
        pct_anticipo = float(request.form.get('porcentaje_anticipo', 50) or 50)
        obj.titulo = request.form['titulo']
        obj.cliente_id = request.form.get('cliente_id') or None
        obj.subtotal = subtotal; obj.iva = iva_monto; obj.total = total
        obj.porcentaje_anticipo = pct_anticipo
        obj.monto_anticipo = total * pct_anticipo / 100.0
        obj.saldo = total - obj.monto_anticipo
        if fd_em: obj.fecha_emision = datetime.strptime(fd_em,'%Y-%m-%d').date()
        obj.fecha_validez = datetime.strptime(fd_val,'%Y-%m-%d').date() if fd_val else None
        obj.dias_entrega = int(request.form.get('dias_entrega',30) or 30)
        obj.condiciones_pago = request.form.get('condiciones_pago','')
        obj.notas = request.form.get('notas','')
        db.session.commit()
        flash('Cotización actualizada.','success')
        return redirect(url_for('cotizacion_ver', id=obj.id))
    return render_template('cotizaciones/form.html', obj=obj, titulo='Editar Cotización',
        clientes_list=clientes_list, today=datetime.utcnow().strftime('%Y-%m-%d'),
        iva_default=iva_default)

def _procesar_orden_produccion(cot):
    """
    Al confirmar una cotización, por cada item:
    1. Verifica stock actual vs requerido y crea OrdenProduccion
    2. Reserva materias primas disponibles según BOM
    3. Si faltan materias, crea tareas comprar_materias + verificar_abono y notifica admins
    """
    admins = User.query.filter_by(rol='admin', activo=True).all()
    primer_admin_id = admins[0].id if admins else current_user.id

    for item in cot.items:
        if not item.producto_id: continue
        prod = Producto.query.get(item.producto_id)
        if not prod: continue
        cant_requerida = item.cantidad
        cant_en_stock  = min(prod.stock, cant_requerida)
        cant_producir  = max(0, cant_requerida - cant_en_stock)

        # Crear orden de producción
        orden = OrdenProduccion(
            cotizacion_id=cot.id,
            producto_id=prod.id,
            cantidad_total=cant_requerida,
            cantidad_stock=cant_en_stock,
            cantidad_producir=cant_producir,
            estado='en_produccion' if cant_producir > 0 else 'completado',
            notas=f'Cotización #{cot.numero or cot.id} — {cot.titulo}',
            creado_por=current_user.id
        )
        db.session.add(orden)

        if cant_producir <= 0:
            continue  # ya hay suficiente stock, sin acción adicional

        # Verificar BOM
        receta = RecetaProducto.query.filter_by(producto_id=prod.id, activo=True).first()
        materias_faltantes = []
        if receta and receta.unidades_produce > 0:
            factor = cant_producir / receta.unidades_produce
            for ri in receta.items:
                mp = MateriaPrima.query.get(ri.materia_prima_id)
                if not mp: continue
                necesaria = ri.cantidad_por_unidad * factor
                disponible = mp.stock_disponible
                if disponible >= necesaria:
                    # Reservar automáticamente
                    mp.stock_disponible -= necesaria
                    mp.stock_reservado  += necesaria
                    db.session.add(ReservaProduccion(
                        materia_prima_id=mp.id,
                        cantidad=necesaria,
                        producto_id=prod.id,
                        estado='reservado',
                        notas=f'Auto-reserva cot #{cot.numero or cot.id}',
                        creado_por=current_user.id
                    ))
                else:
                    faltante = necesaria - disponible
                    materias_faltantes.append(
                        f'{mp.nombre}: falta {faltante:.3f} {mp.unidad} (disponible {disponible:.3f})'
                    )
                    # Reservar lo que haya
                    if disponible > 0:
                        mp.stock_disponible = 0
                        mp.stock_reservado += disponible
                        db.session.add(ReservaProduccion(
                            materia_prima_id=mp.id,
                            cantidad=disponible,
                            producto_id=prod.id,
                            estado='reservado',
                            notas=f'Parcial cot #{cot.numero or cot.id}',
                            creado_por=current_user.id
                        ))

        if materias_faltantes:
            descripcion_falta = '\n'.join(materias_faltantes)
            desc_tareas = (f'Cotización: #{cot.numero or cot.id} — {cot.titulo}\n'
                           f'Producto: {prod.nombre} (x{cant_producir})\n\n'
                           f'Materiales faltantes:\n{descripcion_falta}')
            from datetime import timedelta
            venc = (datetime.utcnow() + timedelta(days=3)).date()

            t_compra = Tarea(
                titulo=f'Comprar materias — {prod.nombre} (cot #{cot.numero or cot.id})',
                descripcion=desc_tareas,
                estado='pendiente', prioridad='alta',
                asignado_a=primer_admin_id,
                creado_por=current_user.id,
                fecha_vencimiento=venc,
                cotizacion_id=cot.id,
                tarea_tipo='comprar_materias'
            )
            db.session.add(t_compra); db.session.flush()

            t_abono = Tarea(
                titulo=f'Verificar abono — {cot.titulo}',
                descripcion=(f'Confirmar recepción del anticipo antes de comprar materias.\n'
                             f'Cotización: #{cot.numero or cot.id}'),
                estado='pendiente', prioridad='alta',
                asignado_a=primer_admin_id,
                creado_por=current_user.id,
                fecha_vencimiento=venc,
                cotizacion_id=cot.id,
                tarea_tipo='verificar_abono',
                tarea_pareja_id=t_compra.id
            )
            db.session.add(t_abono); db.session.flush()
            # link inverso
            t_compra.tarea_pareja_id = t_abono.id

            # Notificar a todos los admins
            for adm in admins:
                _crear_notificacion(
                    adm.id, 'alerta_stock',
                    f'⚠️ Materiales insuficientes — {prod.nombre}',
                    f'Faltan materias para cotización #{cot.numero or cot.id}. '
                    f'Se crearon tareas de compra y abono.',
                    url_for('tareas')
                )


def _procesar_venta_produccion(venta):
    """
    Al guardar una venta, por cada item con producto:
    1. Compara stock actual vs cantidad requerida
    2. Si falta stock → crea OrdenProduccion (vinculada a esta venta)
    3. Reserva/descuenta materias primas disponibles según BOM inmediatamente
    4. Si faltan materias, crea tareas comprar_materias + verificar_abono
    El check de duplicado es POR VENTA+PRODUCTO (no global), así cada venta
    genera sus propias órdenes y las ventas posteriores no se bloquean.
    """
    try:
        admins = User.query.filter_by(rol='admin', activo=True).all()
        primer_admin_id = admins[0].id if admins else current_user.id
        from datetime import timedelta

        for item in venta.items:
            if not item.producto_id:
                continue
            prod = Producto.query.get(item.producto_id)
            if not prod:
                continue
            cant_requerida = item.cantidad
            cant_en_stock  = min(prod.stock or 0, cant_requerida)
            cant_producir  = max(0.0, cant_requerida - cant_en_stock)

            if cant_producir <= 0:
                continue  # stock suficiente, sin producción necesaria

            # Check duplicado POR VENTA+PRODUCTO (no bloquea otras ventas del mismo producto)
            existente = OrdenProduccion.query.filter(
                OrdenProduccion.venta_id == venta.id,
                OrdenProduccion.producto_id == prod.id,
                OrdenProduccion.estado != 'completado'
            ).first()
            if existente:
                continue

            orden = OrdenProduccion(
                venta_id=venta.id,
                producto_id=prod.id,
                cantidad_total=cant_requerida,
                cantidad_stock=cant_en_stock,
                cantidad_producir=cant_producir,
                estado='en_produccion',
                notas=f'Venta: {venta.titulo} — {prod.nombre} x{cant_requerida}',
                creado_por=current_user.id
            )
            db.session.add(orden)

            # Descontar/reservar materias primas del BOM inmediatamente
            receta = RecetaProducto.query.filter_by(producto_id=prod.id, activo=True).first()
            materias_faltantes = []
            if receta and receta.unidades_produce > 0:
                factor = cant_producir / receta.unidades_produce
                for ri in receta.items:
                    mp = MateriaPrima.query.get(ri.materia_prima_id)
                    if not mp: continue
                    necesaria  = ri.cantidad_por_unidad * factor
                    disponible = mp.stock_disponible or 0
                    if disponible >= necesaria:
                        # Disponible: reservar completamente
                        mp.stock_disponible -= necesaria
                        mp.stock_reservado   = (mp.stock_reservado or 0) + necesaria
                        db.session.add(ReservaProduccion(
                            materia_prima_id=mp.id, cantidad=necesaria,
                            producto_id=prod.id, venta_id=venta.id, estado='reservado',
                            notas=f'Auto-reserva venta: {venta.titulo}',
                            creado_por=current_user.id
                        ))
                    else:
                        # Insuficiente: reservar lo que haya + anotar faltante
                        faltante = necesaria - disponible
                        materias_faltantes.append(
                            f'{mp.nombre}: falta {faltante:.3f} {mp.unidad} (disp. {disponible:.3f})'
                        )
                        if disponible > 0:
                            mp.stock_disponible = 0
                            mp.stock_reservado   = (mp.stock_reservado or 0) + disponible
                            db.session.add(ReservaProduccion(
                                materia_prima_id=mp.id, cantidad=disponible,
                                producto_id=prod.id, venta_id=venta.id, estado='reservado',
                                notas=f'Parcial venta: {venta.titulo}',
                                creado_por=current_user.id
                            ))

            if materias_faltantes:
                desc_falta = '\n'.join(materias_faltantes)
                desc_t = (f'Venta: {venta.titulo}\n'
                          f'Producto: {prod.nombre} (x{cant_producir:.2f})\n\n'
                          f'Materiales faltantes:\n{desc_falta}')
                venc = (datetime.utcnow() + timedelta(days=3)).date()
                t_compra = Tarea(
                    titulo=f'Comprar materiales — {prod.nombre} (venta)',
                    descripcion=desc_t, estado='pendiente', prioridad='alta',
                    asignado_a=primer_admin_id, creado_por=current_user.id,
                    fecha_vencimiento=venc, tarea_tipo='comprar_materias'
                )
                db.session.add(t_compra); db.session.flush()
                t_abono = Tarea(
                    titulo=f'Verificar abono — compra {prod.nombre}',
                    descripcion=f'Confirmar anticipo antes de comprar materiales para {prod.nombre}.\n\n{desc_falta}',
                    estado='pendiente', prioridad='alta',
                    asignado_a=primer_admin_id, creado_por=current_user.id,
                    fecha_vencimiento=venc, tarea_tipo='verificar_abono',
                    tarea_pareja_id=t_compra.id
                )
                db.session.add(t_abono); db.session.flush()
                t_compra.tarea_pareja_id = t_abono.id
                orden.estado = 'pendiente_materiales'
                for adm in admins:
                    _crear_notificacion(
                        adm.id, 'alerta_stock',
                        f'⚠️ Materiales insuficientes — {prod.nombre}',
                        f'Venta "{venta.titulo}" requiere producción. Faltan materias.',
                        url_for('tareas')
                    )
    except Exception as ex:
        db.session.rollback()
        print(f'_procesar_venta_produccion error: {ex}')


@app.route('/cotizaciones/<int:id>/estado', methods=['POST'])
@login_required
def cotizacion_cambiar_estado(id):
    obj = Cotizacion.query.get_or_404(id)
    nuevo = request.form.get('estado','borrador')
    if nuevo in ('borrador','enviada','aprobada','confirmacion_orden'):
        obj.estado = nuevo
        db.session.commit()
        if nuevo == 'confirmacion_orden':
            try:
                _procesar_orden_produccion(obj)
                db.session.commit()
            except Exception as ep:
                db.session.rollback()
                print(f'_procesar_orden_produccion error: {ep}')
        flash(f'Estado actualizado a: {nuevo}.','success')
    return redirect(url_for('cotizacion_ver', id=id))

@app.route('/cotizaciones/<int:id>/eliminar', methods=['POST'])
@login_required
def cotizacion_eliminar(id):
    obj = Cotizacion.query.get_or_404(id)
    db.session.delete(obj); db.session.commit()
    flash('Cotización eliminada.','info')
    return redirect(url_for('cotizaciones'))

@app.route('/cotizaciones/<int:id>/pdf')
@login_required
def cotizacion_pdf(id):
    obj = Cotizacion.query.get_or_404(id)
    empresa = ConfigEmpresa.query.first() or ConfigEmpresa(nombre='Evore')
    return render_template('cotizaciones/pdf.html', obj=obj, empresa=empresa)

# =============================================================
# HISTORIAL DE ACTIVIDAD
# =============================================================

@app.route('/actividad')
@login_required
def actividad():
    if current_user.rol != 'admin':
        flash('Sin permisos.','danger'); return redirect(url_for('dashboard'))
    items = Actividad.query.order_by(Actividad.creado_en.desc()).limit(300).all()
    return render_template('actividad.html', items=items)

# =============================================================
# ADMIN — CONFIG EMPRESA
# =============================================================

@app.route('/admin/empresa', methods=['GET','POST'])
@login_required
def admin_config():
    if current_user.rol != 'admin':
        flash('Sin permisos.','danger'); return redirect(url_for('dashboard'))
    obj = ConfigEmpresa.query.first()
    if not obj:
        obj = ConfigEmpresa(nombre='Evore'); db.session.add(obj); db.session.commit()
    if request.method == 'POST':
        obj.nombre   = request.form.get('nombre','Evore')
        obj.nit      = request.form.get('nit','')
        obj.ciudad   = request.form.get('ciudad','')
        obj.telefono = request.form.get('telefono','')
        obj.email    = request.form.get('email','')
        obj.sitio_web= request.form.get('sitio_web','')
        obj.direccion= request.form.get('direccion','')
        db.session.commit(); flash('Configuración guardada.','success')
    return render_template('admin/config.html', obj=obj)

# =============================================================
# NOTIFICACIONES
# =============================================================

@app.route('/notificaciones')
@login_required
def notificaciones():
    items = Notificacion.query.filter_by(usuario_id=current_user.id)\
                .order_by(Notificacion.creado_en.desc()).limit(100).all()
    # marcar todas como leídas al abrir
    Notificacion.query.filter_by(usuario_id=current_user.id, leida=False).update({'leida': True})
    db.session.commit()
    return render_template('notificaciones.html', items=items)

@app.route('/notificaciones/recientes')
@login_required
def notificaciones_recientes():
    items = Notificacion.query.filter_by(usuario_id=current_user.id, leida=False)\
                .order_by(Notificacion.creado_en.desc()).limit(10).all()
    return jsonify([{
        'id': n.id, 'tipo': n.tipo, 'titulo': n.titulo,
        'mensaje': n.mensaje, 'url': n.url or '',
        'creado_en': n.creado_en.strftime('%d/%m %H:%M')
    } for n in items])

@app.route('/notificaciones/marcar_todas', methods=['POST'])
@login_required
def notificaciones_marcar_todas():
    Notificacion.query.filter_by(usuario_id=current_user.id, leida=False).update({'leida': True})
    db.session.commit()
    flash('Todas las notificaciones marcadas como leídas.', 'success')
    return redirect(url_for('notificaciones'))

def _crear_notificacion(usuario_id, tipo, titulo, mensaje, url=None):
    try:
        n = Notificacion(usuario_id=usuario_id, tipo=tipo, titulo=titulo,
                         mensaje=mensaje, url=url)
        db.session.add(n)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f'Notificacion error: {e}')

# =============================================================
# INVENTARIO — LOTES
# =============================================================

@app.route('/inventario/lotes')
@login_required
@requiere_modulo('inventario')
def lotes():
    items = LoteProducto.query.order_by(LoteProducto.creado_en.desc()).all()
    return render_template('inventario/lotes.html', lotes=items)

@app.route('/inventario/lotes/nuevo', methods=['GET','POST'])
@login_required
@requiere_modulo('inventario')
def lote_nuevo():
    productos = Producto.query.filter_by(activo=True).order_by(Producto.nombre).all()
    if request.method == 'POST':
        fp = request.form.get('fecha_produccion')
        fv = request.form.get('fecha_vencimiento')
        l = LoteProducto(
            producto_id=int(request.form['producto_id']),
            numero_lote=request.form['numero_lote'],
            nso=request.form.get('nso','') or None,
            fecha_produccion=datetime.strptime(fp,'%Y-%m-%d').date() if fp else None,
            fecha_vencimiento=datetime.strptime(fv,'%Y-%m-%d').date() if fv else None,
            unidades_producidas=int(request.form.get('unidades_producidas',0)),
            unidades_restantes=int(request.form.get('unidades_restantes',0)),
            notas=request.form.get('notas','') or None,
            creado_por=current_user.id
        )
        db.session.add(l); db.session.commit()
        flash('Lote creado.','success'); return redirect(url_for('lotes'))
    return render_template('inventario/lote_form.html', obj=None, productos=productos, titulo='Nuevo Lote')

@app.route('/inventario/lotes/<int:id>/editar', methods=['GET','POST'])
@login_required
@requiere_modulo('inventario')
def lote_editar(id):
    obj = LoteProducto.query.get_or_404(id)
    productos = Producto.query.filter_by(activo=True).order_by(Producto.nombre).all()
    if request.method == 'POST':
        fp = request.form.get('fecha_produccion')
        fv = request.form.get('fecha_vencimiento')
        obj.producto_id=int(request.form['producto_id'])
        obj.numero_lote=request.form['numero_lote']
        obj.nso=request.form.get('nso','') or None
        obj.fecha_produccion=datetime.strptime(fp,'%Y-%m-%d').date() if fp else None
        obj.fecha_vencimiento=datetime.strptime(fv,'%Y-%m-%d').date() if fv else None
        obj.unidades_producidas=int(request.form.get('unidades_producidas',0))
        obj.unidades_restantes=int(request.form.get('unidades_restantes',0))
        obj.notas=request.form.get('notas','') or None
        db.session.commit()
        flash('Lote actualizado.','success'); return redirect(url_for('lotes'))
    return render_template('inventario/lote_form.html', obj=obj, productos=productos, titulo='Editar Lote')

@app.route('/inventario/lotes/<int:id>/eliminar', methods=['POST'])
@login_required
@requiere_modulo('inventario')
def lote_eliminar(id):
    obj = LoteProducto.query.get_or_404(id); db.session.delete(obj); db.session.commit()
    flash('Lote eliminado.','info'); return redirect(url_for('lotes'))

# =============================================================
# PRODUCCIÓN — MATERIAS PRIMAS
# =============================================================

@app.route('/produccion/materias')
@login_required
@requiere_modulo('produccion')
def materias():
    items = MateriaPrima.query.filter_by(activo=True).order_by(MateriaPrima.nombre).all()
    return render_template('produccion/materias.html', materias=items)

@app.route('/produccion/materias/nueva', methods=['GET','POST'])
@login_required
@requiere_modulo('produccion')
def materia_nueva():
    productos = Producto.query.filter_by(activo=True).order_by(Producto.nombre).all()
    if request.method == 'POST':
        pid_raw = request.form.get('producto_id','')
        m = MateriaPrima(
            nombre=request.form['nombre'],
            descripcion=request.form.get('descripcion','') or None,
            unidad=request.form.get('unidad','unidades'),
            stock_disponible=float(request.form.get('stock_disponible',0)),
            stock_minimo=float(request.form.get('stock_minimo',0)),
            costo_unitario=float(request.form.get('costo_unitario',0)),
            categoria=request.form.get('categoria','') or None,
            proveedor=request.form.get('proveedor','') or None,
            proveedor_id=int(request.form.get('proveedor_id')) if request.form.get('proveedor_id') else None,
            producto_id=int(pid_raw) if pid_raw else None
        )
        db.session.add(m); db.session.commit()
        flash('Materia prima creada.','success'); return redirect(url_for('materias'))
    return render_template('produccion/materia_form.html', obj=None, titulo='Nueva Materia Prima',
                           productos=productos, proveedores_list=Proveedor.query.filter_by(activo=True).order_by(Proveedor.empresa).all())

@app.route('/produccion/materias/<int:id>/editar', methods=['GET','POST'])
@login_required
@requiere_modulo('produccion')
def materia_editar(id):
    obj = MateriaPrima.query.get_or_404(id)
    productos = Producto.query.filter_by(activo=True).order_by(Producto.nombre).all()
    if request.method == 'POST':
        pid_raw = request.form.get('producto_id','')
        obj.nombre=request.form['nombre']
        obj.descripcion=request.form.get('descripcion','') or None
        obj.unidad=request.form.get('unidad','unidades')
        obj.stock_disponible=float(request.form.get('stock_disponible',0))
        obj.stock_minimo=float(request.form.get('stock_minimo',0))
        obj.costo_unitario=float(request.form.get('costo_unitario',0))
        obj.categoria=request.form.get('categoria','') or None
        obj.proveedor=request.form.get('proveedor','') or None
        obj.proveedor_id=int(request.form.get('proveedor_id')) if request.form.get('proveedor_id') else None
        obj.producto_id=int(pid_raw) if pid_raw else None
        db.session.commit()
        flash('Materia prima actualizada.','success'); return redirect(url_for('materias'))
    return render_template('produccion/materia_form.html', obj=obj, titulo='Editar Materia Prima',
                           productos=productos, proveedores_list=Proveedor.query.filter_by(activo=True).order_by(Proveedor.empresa).all())

@app.route('/produccion/materias/<int:id>/eliminar', methods=['POST'])
@login_required
@requiere_modulo('produccion')
def materia_eliminar(id):
    obj = MateriaPrima.query.get_or_404(id); obj.activo=False; db.session.commit()
    flash('Materia prima desactivada.','info'); return redirect(url_for('materias'))

# =============================================================
# PRODUCCIÓN — RECETAS / BOM
# =============================================================

@app.route('/produccion/recetas')
@login_required
@requiere_modulo('produccion')
def recetas():
    items = RecetaProducto.query.filter_by(activo=True).all()
    return render_template('produccion/recetas.html', recetas=items)

@app.route('/produccion/recetas/nueva', methods=['GET','POST'])
@login_required
@requiere_modulo('produccion')
def receta_nueva():
    productos = Producto.query.filter_by(activo=True).order_by(Producto.nombre).all()
    materias = MateriaPrima.query.filter_by(activo=True).order_by(MateriaPrima.nombre).all()
    materias_json = [{'id': m.id, 'nombre': m.nombre, 'unidad': m.unidad} for m in materias]
    if request.method == 'POST':
        r = RecetaProducto(
            producto_id=int(request.form['producto_id']),
            unidades_produce=int(request.form.get('unidades_produce',1)),
            descripcion=request.form.get('descripcion','') or None
        )
        db.session.add(r); db.session.flush()
        ids = request.form.getlist('materia_id[]')
        cants = request.form.getlist('cantidad[]')
        for mid, cant in zip(ids, cants):
            if mid and cant:
                db.session.add(RecetaItem(
                    receta_id=r.id,
                    materia_prima_id=int(mid),
                    cantidad_por_unidad=float(cant)
                ))
        db.session.commit()
        flash('Receta creada.','success'); return redirect(url_for('recetas'))
    return render_template('produccion/receta_form.html', obj=None, productos=productos,
                           materias=materias, materias_json=materias_json, titulo='Nueva Receta')

@app.route('/produccion/recetas/<int:id>/editar', methods=['GET','POST'])
@login_required
@requiere_modulo('produccion')
def receta_editar(id):
    obj = RecetaProducto.query.get_or_404(id)
    productos = Producto.query.filter_by(activo=True).order_by(Producto.nombre).all()
    materias = MateriaPrima.query.filter_by(activo=True).order_by(MateriaPrima.nombre).all()
    materias_json = [{'id': m.id, 'nombre': m.nombre, 'unidad': m.unidad} for m in materias]
    if request.method == 'POST':
        obj.producto_id=int(request.form['producto_id'])
        obj.unidades_produce=int(request.form.get('unidades_produce',1))
        obj.descripcion=request.form.get('descripcion','') or None
        # Rebuild items
        for item in obj.items: db.session.delete(item)
        db.session.flush()
        ids = request.form.getlist('materia_id[]')
        cants = request.form.getlist('cantidad[]')
        for mid, cant in zip(ids, cants):
            if mid and cant:
                db.session.add(RecetaItem(
                    receta_id=obj.id,
                    materia_prima_id=int(mid),
                    cantidad_por_unidad=float(cant)
                ))
        db.session.commit()
        flash('Receta actualizada.','success'); return redirect(url_for('recetas'))
    return render_template('produccion/receta_form.html', obj=obj, productos=productos,
                           materias=materias, materias_json=materias_json, titulo='Editar Receta')

@app.route('/produccion/recetas/<int:id>/eliminar', methods=['POST'])
@login_required
@requiere_modulo('produccion')
def receta_eliminar(id):
    obj = RecetaProducto.query.get_or_404(id); obj.activo=False; db.session.commit()
    flash('Receta eliminada.','info'); return redirect(url_for('recetas'))

# =============================================================
# PRODUCCIÓN — RESERVAS
# =============================================================

@app.route('/produccion/reservas')
@login_required
@requiere_modulo('produccion')
def reservas():
    items = ReservaProduccion.query.order_by(ReservaProduccion.creado_en.desc()).all()
    usuarios = User.query.filter_by(activo=True).order_by(User.nombre).all()
    return render_template('produccion/reservas.html', reservas=items, usuarios=usuarios)

@app.route('/produccion/reservas/solicitar_compra', methods=['POST'])
@login_required
@requiere_modulo('produccion')
def reserva_solicitar_compra():
    reserva_id = request.form.get('reserva_id')
    usuario_id = int(request.form.get('usuario_id'))
    descripcion = request.form.get('descripcion','')
    cantidad_faltante = request.form.get('cantidad_faltante','')
    r = ReservaProduccion.query.get_or_404(int(reserva_id))
    mp = r.materia
    from datetime import timedelta
    venc = (datetime.utcnow() + timedelta(days=3)).date()

    # Format the title with the missing quantity
    try:
        cant_fmt = f'{float(cantidad_faltante):.3f} {mp.unidad}' if cantidad_faltante else ''
    except (ValueError, TypeError):
        cant_fmt = ''
    titulo_compra = f'Comprar {cant_fmt} de {mp.nombre}' if cant_fmt else f'Comprar materia: {mp.nombre}'

    t_compra = Tarea(
        titulo=titulo_compra,
        descripcion=(f'Falta material en reserva de producción.\n'
                     f'Materia: {mp.nombre}\n'
                     f'Cantidad faltante: {cant_fmt or "ver descripción"}\n'
                     f'Producto: {r.producto.nombre if r.producto else "N/A"}\n\n{descripcion}'),
        estado='pendiente', prioridad='alta',
        asignado_a=usuario_id,
        creado_por=current_user.id,
        fecha_vencimiento=venc,
        tarea_tipo='comprar_materias'
    )
    db.session.add(t_compra); db.session.flush()

    t_abono = Tarea(
        titulo=f'Verificar abono — {titulo_compra}',
        descripcion=(f'Confirmar anticipo antes de comprar {mp.nombre}.\n'
                     f'Cantidad requerida: {cant_fmt or "ver descripción"}\n\n{descripcion}'),
        estado='pendiente', prioridad='alta',
        asignado_a=usuario_id,
        creado_por=current_user.id,
        fecha_vencimiento=venc,
        tarea_tipo='verificar_abono',
        tarea_pareja_id=t_compra.id
    )
    db.session.add(t_abono); db.session.flush()
    t_compra.tarea_pareja_id = t_abono.id
    db.session.commit()

    _crear_notificacion(
        usuario_id, 'tarea_asignada',
        f'Nueva tarea asignada: {t_compra.titulo}',
        f'Tienes 2 nuevas tareas de compra/abono para {mp.nombre}.',
        url_for('tareas')
    )
    flash(f'Tareas de compra y abono creadas y asignadas.', 'success')
    return redirect(url_for('reservas'))

@app.route('/produccion/reservas/nueva', methods=['GET','POST'])
@login_required
@requiere_modulo('produccion')
def reserva_nueva():
    materias = MateriaPrima.query.filter_by(activo=True).order_by(MateriaPrima.nombre).all()
    productos = Producto.query.filter_by(activo=True).order_by(Producto.nombre).all()
    if request.method == 'POST':
        mid = int(request.form['materia_prima_id'])
        cantidad = float(request.form['cantidad'])
        m = MateriaPrima.query.get_or_404(mid)
        if cantidad > m.stock_disponible:
            flash(f'Stock insuficiente. Disponible: {m.stock_disponible} {m.unidad}','danger')
        else:
            pid_raw = request.form.get('producto_id','')
            r = ReservaProduccion(
                materia_prima_id=mid, cantidad=cantidad,
                producto_id=int(pid_raw) if pid_raw else None,
                estado='reservado', notas=request.form.get('notas','') or None,
                creado_por=current_user.id
            )
            m.stock_disponible -= cantidad
            m.stock_reservado += cantidad
            db.session.add(r); db.session.commit()
            flash('Reserva creada. Stock actualizado.','success')
            return redirect(url_for('reservas'))
    return render_template('produccion/reserva_form.html', materias=materias, productos=productos)

@app.route('/produccion/reservas/<int:id>/cancelar', methods=['POST'])
@login_required
@requiere_modulo('produccion')
def reserva_cancelar(id):
    r = ReservaProduccion.query.get_or_404(id)
    if r.estado == 'reservado':
        m = MateriaPrima.query.get(r.materia_prima_id)
        if m:
            m.stock_disponible += r.cantidad
            m.stock_reservado = max(0, m.stock_reservado - r.cantidad)
        r.estado = 'cancelado'
        db.session.commit()
        flash('Reserva cancelada y stock devuelto.','info')
    return redirect(url_for('reservas'))

@app.route('/produccion/reservas/venta/<int:venta_id>/iniciar', methods=['POST'])
@login_required
@requiere_modulo('produccion')
def reserva_iniciar_produccion(venta_id):
    """Marca todas las órdenes de producción de una venta como en_produccion."""
    ordenes = OrdenProduccion.query.filter(
        OrdenProduccion.venta_id == venta_id,
        OrdenProduccion.estado.in_(['pendiente_materiales','en_produccion'])
    ).all()
    for o in ordenes:
        o.estado = 'en_produccion'
    db.session.commit()
    flash('Producción iniciada. Las órdenes están en progreso.','success')
    return redirect(url_for('reservas'))

# =============================================================
# INVENTARIO — MULTI-INGRESO
# =============================================================

@app.route('/inventario/ingresos', methods=['GET','POST'])
@login_required
@requiere_modulo('inventario')
def inventario_ingresos():
    productos = Producto.query.filter_by(activo=True).order_by(Producto.nombre).all()
    if request.method == 'POST':
        pid = request.form.get('producto_id')
        numero_lote = request.form.get('numero_lote','').strip()
        nso = request.form.get('nso','').strip() or None
        fp  = request.form.get('fecha_produccion')
        fv  = request.form.get('fecha_vencimiento')
        cantidades = request.form.getlist('cantidad[]')
        costos     = request.form.getlist('costo_unit[]')

        if not pid or not numero_lote:
            flash('Selecciona un producto e indica el número de lote.','danger')
            return render_template('inventario/ingresos.html', productos=productos)

        prod = Producto.query.get_or_404(int(pid))
        total_unidades = 0
        for cant_s, costo_s in zip(cantidades, costos):
            try:
                cant = float(cant_s)
                if cant > 0:
                    total_unidades += cant
                    if costo_s:
                        prod.costo = float(costo_s)   # actualiza último costo
            except: pass

        if total_unidades <= 0:
            flash('Ingresa al menos una cantidad válida.','danger')
            return render_template('inventario/ingresos.html', productos=productos)

        prod.stock += int(total_unidades)
        lote = LoteProducto(
            producto_id=prod.id,
            numero_lote=numero_lote,
            nso=nso,
            fecha_produccion=datetime.strptime(fp,'%Y-%m-%d').date() if fp else None,
            fecha_vencimiento=datetime.strptime(fv,'%Y-%m-%d').date() if fv else None,
            unidades_producidas=total_unidades,
            unidades_restantes=total_unidades,
            notas=f'Multi-ingreso: {total_unidades} unidades',
            creado_por=current_user.id
        )
        db.session.add(lote)
        db.session.commit()
        flash(f'{total_unidades:.0f} unidades de "{prod.nombre}" ingresadas al lote {numero_lote}.','success')
        return redirect(url_for('inventario'))
    return render_template('inventario/ingresos.html', productos=productos)

# =============================================================
# PRODUCCIÓN — ÓRDENES DE PRODUCCIÓN
# =============================================================

@app.route('/produccion/ordenes')
@login_required
@requiere_modulo('produccion')
def ordenes_produccion():
    pendientes = OrdenProduccion.query.filter(
        OrdenProduccion.estado.in_(['pendiente_materiales','en_produccion'])
    ).order_by(OrdenProduccion.creado_en.desc()).all()
    completados = OrdenProduccion.query.filter_by(estado='completado')\
        .order_by(OrdenProduccion.completado_en.desc()).limit(30).all()
    return render_template('produccion/ordenes.html',
                           pendientes=pendientes, completados=completados)

@app.route('/produccion/ordenes/completar', methods=['POST'])
@login_required
@requiere_modulo('produccion')
def orden_completar():
    orden_id  = int(request.form.get('orden_id'))
    numero_lote = request.form.get('numero_lote','').strip()
    notas       = request.form.get('notas','')
    fv          = request.form.get('fecha_vencimiento')

    orden = OrdenProduccion.query.get_or_404(orden_id)
    prod  = orden.producto

    # Añadir al stock
    prod.stock += int(orden.cantidad_producir)

    # Registrar lote
    lote = LoteProducto(
        producto_id=prod.id,
        numero_lote=numero_lote or f'OP-{orden.id}',
        fecha_vencimiento=datetime.strptime(fv,'%Y-%m-%d').date() if fv else None,
        unidades_producidas=orden.cantidad_producir,
        unidades_restantes=orden.cantidad_producir,
        notas=notas or f'Orden producción #{orden.id}',
        creado_por=current_user.id
    )
    db.session.add(lote)

    orden.estado = 'completado'
    orden.numero_lote = numero_lote or f'OP-{orden.id}'
    orden.completado_en = datetime.utcnow()
    db.session.commit()

    # Notificar admins
    admins = User.query.filter_by(rol='admin', activo=True).all()
    for adm in admins:
        _crear_notificacion(
            adm.id, 'info',
            f'✅ Producción completada: {prod.nombre}',
            f'{orden.cantidad_producir:.0f} unidades movidas al inventario. Lote: {orden.numero_lote}',
            url_for('inventario')
        )
    flash(f'Producción completada. {orden.cantidad_producir:.0f} unidades agregadas al inventario.','success')
    return redirect(url_for('ordenes_produccion'))

@app.route('/produccion/gantt')
@login_required
@requiere_modulo('produccion')
def gantt():
    # Agrupa órdenes de producción por pedido (venta)
    ordenes = OrdenProduccion.query.order_by(OrdenProduccion.venta_id, OrdenProduccion.creado_en).all()

    # Construir grupos por venta
    grupos = {}   # venta_id -> {'venta': obj|None, 'ordenes': [...]}
    sin_venta = {'venta_id': None, 'titulo': 'Sin pedido asociado', 'ordenes': [],
                 'cliente': None, 'venta_numero': None, 'fecha_entrega': None}

    for o in ordenes:
        inicio = (o.fecha_inicio_real or o.creado_en.date())
        if o.fecha_fin_estimada:
            fin = o.fecha_fin_estimada
        elif o.completado_en:
            fin = o.completado_en.date()
        elif o.venta and o.venta.fecha_entrega_est:
            fin = o.venta.fecha_entrega_est
        elif o.cotizacion and o.cotizacion.fecha_entrega_est:
            fin = o.cotizacion.fecha_entrega_est
        else:
            fin = inicio + timedelta(days=30)

        # Buscar OC de materiales vinculada (mejor estimación de entrega de materiales)
        mat_inicio = mat_fin = None
        if o.venta_id:
            oc_mat = OrdenCompra.query.filter(
                OrdenCompra.estado.in_(['borrador','enviada','recibida'])
            ).order_by(OrdenCompra.creado_en.desc()).first()
            if oc_mat:
                mat_inicio = oc_mat.fecha_emision
                mat_fin    = oc_mat.fecha_esperada or (oc_mat.fecha_emision + timedelta(days=15) if oc_mat.fecha_emision else None)

        ord_data = {
            'id':         o.id,
            'producto':   o.producto.nombre,
            'cantidad':   int(o.cantidad_producir),
            'estado':     o.estado,
            'lote':       o.numero_lote or '',
            'inicio':     inicio.strftime('%Y-%m-%d'),
            'fin':        fin.strftime('%Y-%m-%d'),
            'mat_inicio': mat_inicio.strftime('%Y-%m-%d') if mat_inicio else None,
            'mat_fin':    mat_fin.strftime('%Y-%m-%d')    if mat_fin    else None,
        }

        if o.venta_id:
            if o.venta_id not in grupos:
                v = o.venta
                grupos[o.venta_id] = {
                    'venta_id':     o.venta_id,
                    'titulo':       v.titulo if v else f'Venta #{o.venta_id}',
                    'venta_numero': v.numero if v else None,
                    'cliente':      (v.cliente.empresa or v.cliente.nombre) if v and v.cliente else None,
                    'fecha_entrega': v.fecha_entrega_est.strftime('%d/%m/%Y') if v and v.fecha_entrega_est else None,
                    'ordenes':      []
                }
            grupos[o.venta_id]['ordenes'].append(ord_data)
        else:
            sin_venta['ordenes'].append(ord_data)

    pedidos_json = list(grupos.values())
    if sin_venta['ordenes']:
        pedidos_json.append(sin_venta)

    return render_template('produccion/gantt.html', pedidos_json=pedidos_json)

# =============================================================
# ADMIN — EDITAR USUARIO
# =============================================================

@app.route('/admin/usuarios/<int:id>/editar', methods=['GET','POST'])
@login_required
def admin_usuario_editar(id):
    if current_user.rol != 'admin':
        flash('Sin permisos.','danger'); return redirect(url_for('dashboard'))
    u = User.query.get_or_404(id)
    if request.method == 'POST':
        _pwd = request.form.get('password','')
        u.nombre = request.form.get('nombre', u.nombre)
        u.email  = request.form.get('email', u.email)
        u.rol    = request.form.get('rol', u.rol)
        if _pwd and len(_pwd) >= 8:
            u.set_password(_pwd)
        elif _pwd and len(_pwd) < 8:
            flash('Contraseña muy corta (mín. 8 caracteres).','danger')
            return render_template('admin/usuario_form.html', obj=u, titulo='Editar Usuario')
        # Guardar módulos personalizados
        modulos_sel = request.form.getlist('modulos')
        u.modulos_permitidos = json.dumps(modulos_sel) if modulos_sel else '[]'
        db.session.commit()
        flash('Usuario actualizado.','success'); return redirect(url_for('admin_usuarios'))
    return render_template('admin/usuario_form.html', obj=u, titulo='Editar Usuario')

# =============================================================
# =============================================================
# MÓDULO LEGAL
# =============================================================

@app.route('/legal')
@login_required
def legal_index():
    tipo_f = request.args.get('tipo','')
    estado_f = request.args.get('estado','')
    q = DocumentoLegal.query.filter_by(activo=True)
    if tipo_f: q = q.filter_by(tipo=tipo_f)
    if estado_f: q = q.filter_by(estado=estado_f)
    items = q.order_by(DocumentoLegal.fecha_vencimiento).all()
    # Alertas: vencimientos próximos
    from datetime import timedelta
    hoy = datetime.utcnow().date()
    alertas = [d for d in items if d.fecha_vencimiento and
               (d.fecha_vencimiento - hoy).days <= d.recordatorio_dias
               and d.estado == 'vigente']
    return render_template('legal/index.html', items=items,
                           tipo_f=tipo_f, estado_f=estado_f, alertas=alertas)

@app.route('/legal/nuevo', methods=['GET','POST'])
@login_required
def legal_nuevo():
    if request.method == 'POST':
        fe = request.form.get('fecha_emision')
        fv = request.form.get('fecha_vencimiento')
        d = DocumentoLegal(
            tipo=request.form.get('tipo','otro'),
            titulo=request.form['titulo'],
            numero=request.form.get('numero',''),
            entidad=request.form.get('entidad',''),
            descripcion=request.form.get('descripcion',''),
            estado=request.form.get('estado','vigente'),
            fecha_emision=datetime.strptime(fe,'%Y-%m-%d').date() if fe else None,
            fecha_vencimiento=datetime.strptime(fv,'%Y-%m-%d').date() if fv else None,
            recordatorio_dias=int(request.form.get('recordatorio_dias') or 30),
            archivo_url=request.form.get('archivo_url',''),
            notas=request.form.get('notas',''),
            activo=True, creado_por=current_user.id
        )
        db.session.add(d); db.session.commit()
        flash('Documento legal creado.','success')
        return redirect(url_for('legal_index'))
    return render_template('legal/form.html', obj=None, titulo='Nuevo Documento Legal')

@app.route('/legal/<int:id>/editar', methods=['GET','POST'])
@login_required
def legal_editar(id):
    obj = DocumentoLegal.query.get_or_404(id)
    if request.method == 'POST':
        fe = request.form.get('fecha_emision')
        fv = request.form.get('fecha_vencimiento')
        obj.tipo=request.form.get('tipo','otro')
        obj.titulo=request.form.get('titulo','')
        obj.numero=request.form.get('numero','')
        obj.entidad=request.form.get('entidad','')
        obj.descripcion=request.form.get('descripcion','')
        obj.estado=request.form.get('estado','vigente')
        obj.fecha_emision=datetime.strptime(fe,'%Y-%m-%d').date() if fe else None
        obj.fecha_vencimiento=datetime.strptime(fv,'%Y-%m-%d').date() if fv else None
        obj.recordatorio_dias=int(request.form.get('recordatorio_dias') or 30)
        obj.archivo_url=request.form.get('archivo_url','')
        obj.notas=request.form.get('notas','')
        db.session.commit()
        flash('Documento actualizado.','success')
        return redirect(url_for('legal_index'))
    return render_template('legal/form.html', obj=obj, titulo='Editar Documento Legal')

@app.route('/legal/<int:id>/eliminar', methods=['POST'])
@login_required
def legal_eliminar(id):
    obj = DocumentoLegal.query.get_or_404(id)
    obj.activo = False; db.session.commit()
    flash('Documento eliminado.','info')
    return redirect(url_for('legal_index'))

# =============================================================
# MÓDULO CONTABLE
# =============================================================

@app.route('/contable')
@login_required
def contable_index():
    from datetime import timedelta
    hoy = datetime.utcnow().date()
    mes_inicio = hoy.replace(day=1)
    mes_str = request.args.get('mes', hoy.strftime('%Y-%m'))
    try:
        anio, mes = int(mes_str.split('-')[0]), int(mes_str.split('-')[1])
    except: anio, mes = hoy.year, hoy.month
    import calendar as cal_mod
    _, ultimo_dia = cal_mod.monthrange(anio, mes)
    desde = datetime(anio, mes, 1).date()
    hasta = datetime(anio, mes, ultimo_dia).date()
    # Ingresos del mes (ventas ganadas/anticipo)
    ventas_mes = Venta.query.filter(
        Venta.estado.in_(['anticipo_pagado','ganado']),
        db.func.date(Venta.creado_en) >= desde,
        db.func.date(Venta.creado_en) <= hasta
    ).all()
    total_ingresos = sum(v.total for v in ventas_mes)
    total_anticipo = sum(v.monto_anticipo for v in ventas_mes)
    # Egresos del mes
    gastos_mes = GastoOperativo.query.filter(
        GastoOperativo.fecha >= desde, GastoOperativo.fecha <= hasta
    ).all()
    compras_mes = CompraMateria.query.filter(
        CompraMateria.fecha >= desde, CompraMateria.fecha <= hasta
    ).all()
    total_gastos = sum(g.monto for g in gastos_mes)
    total_compras = sum(c.costo_total for c in compras_mes)
    total_egresos = total_gastos + total_compras
    utilidad = total_ingresos - total_egresos
    # Impuestos estimados según reglas tributarias activas
    total_impuestos, detalle_impuestos = _calcular_impuestos(total_ingresos, utilidad)
    utilidad_neta = utilidad - total_impuestos
    # Cuentas por cobrar (ventas con saldo > 0)
    cxc = Venta.query.filter(Venta.saldo > 0, Venta.estado.in_(['anticipo_pagado','ganado'])).all()
    total_cxc = sum(v.saldo for v in cxc)
    # Inventario valorizado
    inventario_valor = sum((p.stock or 0) * (p.costo or 0) for p in Producto.query.filter_by(activo=True).all())
    meses_nav = []
    for i in range(5, -1, -1):
        d = (hoy.replace(day=1) - timedelta(days=i*28)).replace(day=1)
        meses_nav.append({'val': d.strftime('%Y-%m'), 'lbl': d.strftime('%b %Y')})
    return render_template('contable/index.html',
        total_ingresos=total_ingresos, total_anticipo=total_anticipo,
        total_egresos=total_egresos, total_gastos=total_gastos,
        total_compras=total_compras, utilidad=utilidad,
        total_impuestos=total_impuestos, detalle_impuestos=detalle_impuestos,
        utilidad_neta=utilidad_neta,
        total_cxc=total_cxc, inventario_valor=inventario_valor,
        ventas_mes=ventas_mes, gastos_mes=gastos_mes, compras_mes=compras_mes,
        cxc=cxc, mes_str=mes_str, meses_nav=meses_nav,
        anio=anio, mes=mes)

@app.route('/contable/ingresos')
@login_required
def contable_ingresos():
    desde_s = request.args.get('desde','')
    hasta_s = request.args.get('hasta','')
    q = Venta.query.filter(Venta.estado.in_(['anticipo_pagado','ganado','prospecto','negociacion','perdido']))
    if desde_s:
        try: q = q.filter(db.func.date(Venta.creado_en) >= datetime.strptime(desde_s,'%Y-%m-%d').date())
        except: pass
    if hasta_s:
        try: q = q.filter(db.func.date(Venta.creado_en) <= datetime.strptime(hasta_s,'%Y-%m-%d').date())
        except: pass
    items = q.order_by(Venta.creado_en.desc()).all()
    return render_template('contable/ingresos.html', items=items, desde_s=desde_s, hasta_s=hasta_s)

@app.route('/contable/egresos')
@login_required
def contable_egresos():
    desde_s = request.args.get('desde','')
    hasta_s = request.args.get('hasta','')
    gastos = GastoOperativo.query
    compras = CompraMateria.query
    if desde_s:
        try:
            d = datetime.strptime(desde_s,'%Y-%m-%d').date()
            gastos = gastos.filter(GastoOperativo.fecha >= d)
            compras = compras.filter(CompraMateria.fecha >= d)
        except: pass
    if hasta_s:
        try:
            h = datetime.strptime(hasta_s,'%Y-%m-%d').date()
            gastos = gastos.filter(GastoOperativo.fecha <= h)
            compras = compras.filter(CompraMateria.fecha <= h)
        except: pass
    return render_template('contable/egresos.html',
        gastos=gastos.order_by(GastoOperativo.fecha.desc()).all(),
        compras=compras.order_by(CompraMateria.fecha.desc()).all(),
        desde_s=desde_s, hasta_s=hasta_s)

@app.route('/contable/libro-diario', methods=['GET','POST'])
@login_required
def contable_libro_diario():
    if request.method == 'POST':
        fd = request.form.get('fecha')
        a = AsientoContable(
            fecha=datetime.strptime(fd,'%Y-%m-%d').date() if fd else datetime.utcnow().date(),
            descripcion=request.form['descripcion'],
            tipo=request.form.get('tipo','manual'),
            referencia=request.form.get('referencia',''),
            debe=float(request.form.get('debe') or 0),
            haber=float(request.form.get('haber') or 0),
            cuenta_debe=request.form.get('cuenta_debe',''),
            cuenta_haber=request.form.get('cuenta_haber',''),
            notas=request.form.get('notas',''),
            creado_por=current_user.id
        )
        db.session.add(a); db.session.flush()
        hoy = datetime.utcnow().date()
        ultimo = AsientoContable.query.filter(
            AsientoContable.numero.like(f'AC-{hoy.year}-%')
        ).order_by(AsientoContable.id.desc()).first()
        if ultimo and ultimo.numero:
            try: seq = int(ultimo.numero.split('-')[-1]) + 1
            except: seq = 1
        else: seq = 1
        a.numero = f'AC-{hoy.year}-{seq:03d}'
        db.session.commit()
        flash(f'Asiento {a.numero} registrado.','success')
        return redirect(url_for('contable_libro_diario'))
    asientos = AsientoContable.query.order_by(AsientoContable.fecha.desc(), AsientoContable.id.desc()).limit(100).all()
    return render_template('contable/libro_diario.html', asientos=asientos)

@app.route('/contable/exportar')
@login_required
def contable_exportar():
    tipo = request.args.get('tipo','ventas')
    desde_s = request.args.get('desde','')
    hasta_s = request.args.get('hasta','')
    import csv, io
    si = io.StringIO()
    writer = csv.writer(si)
    if tipo == 'ventas':
        writer.writerow(['Numero','Titulo','Cliente','Fecha','Estado','Subtotal','IVA','Total','Anticipo','Saldo'])
        q = Venta.query
        if desde_s:
            try: q = q.filter(db.func.date(Venta.creado_en) >= datetime.strptime(desde_s,'%Y-%m-%d').date())
            except: pass
        if hasta_s:
            try: q = q.filter(db.func.date(Venta.creado_en) <= datetime.strptime(hasta_s,'%Y-%m-%d').date())
            except: pass
        for v in q.order_by(Venta.creado_en).all():
            writer.writerow([v.numero or '', v.titulo,
                v.cliente.empresa or v.cliente.nombre if v.cliente else '',
                v.creado_en.strftime('%Y-%m-%d'), v.estado,
                v.subtotal, v.iva, v.total, v.monto_anticipo, v.saldo])
    elif tipo == 'gastos':
        writer.writerow(['Fecha','Tipo','Descripcion','Monto','Recurrencia'])
        q = GastoOperativo.query
        if desde_s:
            try: q = q.filter(GastoOperativo.fecha >= datetime.strptime(desde_s,'%Y-%m-%d').date())
            except: pass
        if hasta_s:
            try: q = q.filter(GastoOperativo.fecha <= datetime.strptime(hasta_s,'%Y-%m-%d').date())
            except: pass
        for g in q.order_by(GastoOperativo.fecha).all():
            writer.writerow([g.fecha.strftime('%Y-%m-%d'), g.tipo, g.descripcion or '', g.monto, g.recurrencia])
    elif tipo == 'compras':
        writer.writerow(['Fecha','Item','Proveedor','Cantidad','Unidad','Costo Total','Factura'])
        q = CompraMateria.query
        if desde_s:
            try: q = q.filter(CompraMateria.fecha >= datetime.strptime(desde_s,'%Y-%m-%d').date())
            except: pass
        if hasta_s:
            try: q = q.filter(CompraMateria.fecha <= datetime.strptime(hasta_s,'%Y-%m-%d').date())
            except: pass
        for c in q.order_by(CompraMateria.fecha).all():
            writer.writerow([c.fecha.strftime('%Y-%m-%d'), c.nombre_item, c.proveedor or '',
                c.cantidad, c.unidad, c.costo_total, c.nro_factura or ''])
    elif tipo == 'asientos':
        writer.writerow(['Numero','Fecha','Descripcion','Tipo','Referencia','Debe','Haber','Cuenta Debe','Cuenta Haber'])
        for a in AsientoContable.query.order_by(AsientoContable.fecha).all():
            writer.writerow([a.numero or '', a.fecha.strftime('%Y-%m-%d'), a.descripcion,
                a.tipo, a.referencia or '', a.debe, a.haber, a.cuenta_debe or '', a.cuenta_haber or ''])
    output = si.getvalue()
    from flask import Response
    return Response(output, mimetype='text/csv',
        headers={'Content-Disposition': f'attachment;filename=contable_{tipo}_{desde_s or "todo"}.csv'})

# =============================================================
# BUSCADOR GLOBAL
# =============================================================

@app.route('/buscador')
@login_required
def buscador():
    q = request.args.get('q','').strip()
    resultados = {}
    if q:
        like = f'%{q}%'
        try:
            resultados['clientes'] = Cliente.query.filter(
                db.or_(Cliente.nombre.ilike(like), Cliente.empresa.ilike(like), Cliente.nit.ilike(like))
            ).limit(20).all()
        except: resultados['clientes'] = []
        try:
            resultados['proveedores'] = Proveedor.query.filter(
                db.or_(Proveedor.nombre.ilike(like), Proveedor.empresa.ilike(like), Proveedor.nit.ilike(like))
            ).limit(20).all()
        except: resultados['proveedores'] = []
        try:
            resultados['productos'] = Producto.query.filter(
                db.or_(Producto.nombre.ilike(like), Producto.sku.ilike(like), Producto.nso.ilike(like))
            ).filter_by(activo=True).limit(20).all()
        except: resultados['productos'] = []
        try:
            resultados['ventas'] = Venta.query.filter(
                db.or_(Venta.titulo.ilike(like), Venta.numero.ilike(like))
            ).limit(20).all()
        except: resultados['ventas'] = []
        try:
            resultados['ordenes_prod'] = OrdenProduccion.query.join(Producto).filter(
                db.or_(Producto.nombre.ilike(like),
                       db.cast(OrdenProduccion.id, db.String).ilike(like))
            ).limit(20).all()
        except: resultados['ordenes_prod'] = []
        try:
            resultados['ordenes_compra'] = OrdenCompra.query.filter(
                OrdenCompra.numero.ilike(like)
            ).limit(20).all()
        except: resultados['ordenes_compra'] = []
    return render_template('buscador.html', q=q, resultados=resultados)

# =============================================================
# DIAGNÓSTICO
# =============================================================

@app.route('/diagnostico')
@login_required
def diagnostico():
    if current_user.rol != 'admin':
        return jsonify({'error': 'Sin permisos'}), 403
    critico = []; atencion = []; ok = []
    try:
        # Verificar DB
        db.session.execute(db.text('SELECT 1'))
        ok.append({'msg': 'Base de datos conectada', 'detalle': ''})
    except Exception as e:
        critico.append({'msg': 'Error de base de datos', 'detalle': str(e)})
    try:
        total_users = User.query.count()
        admins = User.query.filter_by(rol='admin', activo=True).count()
        ok.append({'msg': f'{total_users} usuarios ({admins} admins activos)', 'detalle': ''})
    except Exception as e:
        atencion.append({'msg': 'Error consultando usuarios', 'detalle': str(e)})
    try:
        total_prod = Producto.query.filter_by(activo=True).count()
        stock_bajo = Producto.query.filter(
            Producto.activo==True, Producto.stock_cantidad < 10
        ).count()
        if stock_bajo > 0:
            atencion.append({'msg': f'{stock_bajo} productos con stock bajo (<10)', 'detalle': ''})
        else:
            ok.append({'msg': f'{total_prod} productos activos, stock normal', 'detalle': ''})
    except Exception as e:
        atencion.append({'msg': 'Error consultando inventario', 'detalle': str(e)})
    try:
        hoy = datetime.utcnow().date()
        prox = hoy + timedelta(days=30)
        venc = LoteProducto.query.filter(
            LoteProducto.fecha_vencimiento != None,
            LoteProducto.fecha_vencimiento <= prox
        ).count()
        if venc > 0:
            atencion.append({'msg': f'{venc} lote(s) vencen en 30 días', 'detalle': ''})
        else:
            ok.append({'msg': 'Sin lotes próximos a vencer', 'detalle': ''})
    except Exception as e:
        atencion.append({'msg': 'No se pudieron revisar lotes', 'detalle': str(e)})
    try:
        notif_pend = Notificacion.query.filter_by(leida=False).count()
        if notif_pend > 20:
            atencion.append({'msg': f'{notif_pend} notificaciones sin leer en el sistema', 'detalle': ''})
        else:
            ok.append({'msg': f'{notif_pend} notificaciones pendientes', 'detalle': ''})
    except Exception as e:
        atencion.append({'msg': 'Error en notificaciones', 'detalle': str(e)})
    try:
        tareas_vencidas = Tarea.query.filter(
            Tarea.fecha_vencimiento < datetime.utcnow().date(),
            Tarea.estado.notin_(['completada','cancelada'])
        ).count()
        if tareas_vencidas > 0:
            atencion.append({'msg': f'{tareas_vencidas} tarea(s) vencidas sin completar', 'detalle': ''})
        else:
            ok.append({'msg': 'Sin tareas vencidas', 'detalle': ''})
    except Exception as e:
        atencion.append({'msg': 'Error consultando tareas', 'detalle': str(e)})
    try:
        materias_bajo = MateriaPrima.query.filter(
            MateriaPrima.activo==True,
            MateriaPrima.stock_disponible < MateriaPrima.stock_minimo
        ).count()
        if materias_bajo > 0:
            critico.append({'msg': f'{materias_bajo} materia(s) prima(s) bajo stock mínimo', 'detalle': ''})
        else:
            ok.append({'msg': 'Materias primas con stock normal', 'detalle': ''})
    except Exception as e:
        ok.append({'msg': 'Materias primas (módulo nuevo)', 'detalle': ''})
    return jsonify({'critico': critico, 'atencion': atencion, 'ok': ok})

# =============================================================
# INICIALIZACIÓN
# =============================================================

def _migrate(conn):
    """Agrega columnas nuevas a tablas existentes sin romper datos actuales."""
    migrations = [
        # GastoOperativo — campos v11
        ("ALTER TABLE gastos_operativos ADD COLUMN IF NOT EXISTS tipo_custom VARCHAR(100)"),
        ("ALTER TABLE gastos_operativos ADD COLUMN IF NOT EXISTS recurrencia VARCHAR(20) DEFAULT 'unico'"),
        ("ALTER TABLE gastos_operativos ADD COLUMN IF NOT EXISTS es_plantilla BOOLEAN DEFAULT FALSE"),
        # Producto — fecha de caducidad
        ("ALTER TABLE productos ADD COLUMN IF NOT EXISTS fecha_caducidad DATE"),
        # Nota — campos v11
        ("ALTER TABLE notas ADD COLUMN IF NOT EXISTS producto_id INTEGER REFERENCES productos(id)"),
        ("ALTER TABLE notas ADD COLUMN IF NOT EXISTS modulo VARCHAR(50)"),
        ("ALTER TABLE notas ADD COLUMN IF NOT EXISTS fecha_revision DATE"),
        # ReglaTributaria — proveedor
        ("ALTER TABLE reglas_tributarias ADD COLUMN IF NOT EXISTS proveedor_nombre VARCHAR(200)"),
        # User — módulos personalizados v12
        ("ALTER TABLE users ADD COLUMN IF NOT EXISTS modulos_permitidos TEXT DEFAULT '[]'"),
        # CompraMateria — campos v12
        ("ALTER TABLE compras_materia ADD COLUMN IF NOT EXISTS materia_id INTEGER REFERENCES materias_primas(id)"),
        ("ALTER TABLE compras_materia ADD COLUMN IF NOT EXISTS tipo_compra VARCHAR(50) DEFAULT 'insumo'"),
        ("ALTER TABLE compras_materia ADD COLUMN IF NOT EXISTS unidad VARCHAR(30) DEFAULT 'unidades'"),
        ("ALTER TABLE compras_materia ADD COLUMN IF NOT EXISTS tiene_caducidad BOOLEAN DEFAULT FALSE"),
        ("ALTER TABLE compras_materia ADD COLUMN IF NOT EXISTS fecha_caducidad DATE"),
        # Tarea — campos v12.1 (parejas de tareas + cotizacion)
        ("ALTER TABLE tareas ADD COLUMN IF NOT EXISTS cotizacion_id INTEGER REFERENCES cotizaciones(id)"),
        ("ALTER TABLE tareas ADD COLUMN IF NOT EXISTS tarea_tipo VARCHAR(50)"),
        ("ALTER TABLE tareas ADD COLUMN IF NOT EXISTS tarea_pareja_id INTEGER REFERENCES tareas(id)"),
        # OrdenProduccion — v12.2 venta_id para vincular con ventas
        ("ALTER TABLE ordenes_produccion ADD COLUMN IF NOT EXISTS venta_id INTEGER REFERENCES ventas(id)"),
        # MateriaPrima — v12.2 producto principal al que pertenece
        ("ALTER TABLE materias_primas ADD COLUMN IF NOT EXISTS producto_id INTEGER REFERENCES productos(id)"),
        # Venta — v12.2 seguimiento de entrega
        ("ALTER TABLE ventas ADD COLUMN IF NOT EXISTS cliente_informado_en TIMESTAMP"),
        ("ALTER TABLE ventas ADD COLUMN IF NOT EXISTS entregado_en TIMESTAMP"),
        # v12.3 — numero único de venta
        ("ALTER TABLE ventas ADD COLUMN IF NOT EXISTS numero VARCHAR(20)"),
        # v12.3 — venta_id en reservas
        ("ALTER TABLE reservas_produccion ADD COLUMN IF NOT EXISTS venta_id INTEGER REFERENCES ventas(id)"),
        # v12.3 — proveedor_id en materias y compras
        ("ALTER TABLE materias_primas ADD COLUMN IF NOT EXISTS proveedor_id INTEGER REFERENCES proveedores(id)"),
        ("ALTER TABLE compras_materia ADD COLUMN IF NOT EXISTS proveedor_id INTEGER REFERENCES proveedores(id)"),
        # v12.4 — órdenes de compra
        ("CREATE TABLE IF NOT EXISTS ordenes_compra (id SERIAL PRIMARY KEY, numero VARCHAR(20), proveedor_id INTEGER REFERENCES proveedores(id), estado VARCHAR(30) DEFAULT 'borrador', fecha_emision DATE, fecha_esperada DATE, subtotal FLOAT DEFAULT 0, iva FLOAT DEFAULT 0, total FLOAT DEFAULT 0, notas TEXT, creado_por INTEGER REFERENCES users(id), creado_en TIMESTAMP DEFAULT NOW())"),
        ("CREATE TABLE IF NOT EXISTS ordenes_compra_items (id SERIAL PRIMARY KEY, orden_id INTEGER REFERENCES ordenes_compra(id) ON DELETE CASCADE, nombre_item VARCHAR(200) NOT NULL, descripcion TEXT, cantidad FLOAT DEFAULT 1, unidad VARCHAR(30) DEFAULT 'unidades', precio_unit FLOAT DEFAULT 0, subtotal FLOAT DEFAULT 0)"),
        # v13 — cotizaciones proveedor
        ("CREATE TABLE IF NOT EXISTS cotizaciones_proveedor (id SERIAL PRIMARY KEY, numero VARCHAR(20), proveedor_id INTEGER REFERENCES proveedores(id), nombre_producto VARCHAR(200) NOT NULL, descripcion TEXT, sku VARCHAR(50), precio_unitario FLOAT DEFAULT 0, unidades_minimas INTEGER DEFAULT 1, unidad VARCHAR(30) DEFAULT 'unidades', plazo_entrega VARCHAR(100), condiciones_pago VARCHAR(200), fecha_cotizacion DATE, vigencia DATE, estado VARCHAR(20) DEFAULT 'vigente', notas TEXT, creado_por INTEGER REFERENCES users(id), creado_en TIMESTAMP DEFAULT NOW())"),
        # v13 — documentos legales
        ("CREATE TABLE IF NOT EXISTS documentos_legales (id SERIAL PRIMARY KEY, tipo VARCHAR(50) NOT NULL, titulo VARCHAR(200) NOT NULL, numero VARCHAR(100), entidad VARCHAR(200), descripcion TEXT, estado VARCHAR(20) DEFAULT 'vigente', fecha_emision DATE, fecha_vencimiento DATE, recordatorio_dias INTEGER DEFAULT 30, archivo_url VARCHAR(500), notas TEXT, activo BOOLEAN DEFAULT TRUE, creado_por INTEGER REFERENCES users(id), creado_en TIMESTAMP DEFAULT NOW())"),
        # v13 — asientos contables
        ("CREATE TABLE IF NOT EXISTS asientos_contables (id SERIAL PRIMARY KEY, numero VARCHAR(20), fecha DATE NOT NULL, descripcion VARCHAR(300) NOT NULL, tipo VARCHAR(30) DEFAULT 'manual', referencia VARCHAR(100), debe FLOAT DEFAULT 0, haber FLOAT DEFAULT 0, cuenta_debe VARCHAR(100), cuenta_haber VARCHAR(100), notas TEXT, creado_por INTEGER REFERENCES users(id), creado_en TIMESTAMP DEFAULT NOW())"),
        # v14 — Proveedor tipo (proveedor/transportista/ambos)
        ("ALTER TABLE proveedores ADD COLUMN IF NOT EXISTS tipo VARCHAR(20) DEFAULT 'proveedor'"),
        # v14 — CotizacionProveedor campos estructurados
        ("ALTER TABLE cotizaciones_proveedor ADD COLUMN IF NOT EXISTS tipo_cotizacion VARCHAR(20) DEFAULT 'general'"),
        ("ALTER TABLE cotizaciones_proveedor ADD COLUMN IF NOT EXISTS tipo_producto_servicio VARCHAR(100)"),
        ("ALTER TABLE cotizaciones_proveedor ADD COLUMN IF NOT EXISTS plazo_entrega_dias INTEGER DEFAULT 0"),
        ("ALTER TABLE cotizaciones_proveedor ADD COLUMN IF NOT EXISTS condicion_pago_tipo VARCHAR(30) DEFAULT 'contado'"),
        ("ALTER TABLE cotizaciones_proveedor ADD COLUMN IF NOT EXISTS condicion_pago_dias INTEGER DEFAULT 0"),
        ("ALTER TABLE cotizaciones_proveedor ADD COLUMN IF NOT EXISTS anticipo_porcentaje FLOAT DEFAULT 0"),
        ("ALTER TABLE cotizaciones_proveedor ADD COLUMN IF NOT EXISTS calendario_integrado BOOLEAN DEFAULT FALSE"),
        # v14 — OrdenCompra campos nuevos
        ("ALTER TABLE ordenes_compra ADD COLUMN IF NOT EXISTS cotizacion_id INTEGER REFERENCES cotizaciones_proveedor(id)"),
        ("ALTER TABLE ordenes_compra ADD COLUMN IF NOT EXISTS fecha_estimada_pago DATE"),
        ("ALTER TABLE ordenes_compra ADD COLUMN IF NOT EXISTS transportista_id INTEGER REFERENCES proveedores(id)"),
        ("ALTER TABLE ordenes_compra ADD COLUMN IF NOT EXISTS fecha_estimada_recogida DATE"),
        # v14 — OrdenCompraItem link a cotizacion
        ("ALTER TABLE ordenes_compra_items ADD COLUMN IF NOT EXISTS cotizacion_id INTEGER REFERENCES cotizaciones_proveedor(id)"),
        # v14 — OrdenProduccion fechas para Gantt
        ("ALTER TABLE ordenes_produccion ADD COLUMN IF NOT EXISTS fecha_inicio_real DATE"),
        ("ALTER TABLE ordenes_produccion ADD COLUMN IF NOT EXISTS fecha_fin_estimada DATE"),
    ]
    for sql in migrations:
        try:
            conn.execute(db.text(sql))
            conn.commit()
        except Exception as em:
            try: conn.rollback()
            except: pass
            print(f'Migración omitida (puede ser SQLite o ya existe): {em}')

def init_db():
    with app.app_context():
        db.create_all()
        # Migraciones para columnas nuevas en tablas existentes
        try:
            with db.engine.connect() as conn:
                _migrate(conn)
        except Exception as em:
            print(f'Migrate error (no crítico): {em}')
        _admin_email = os.environ.get('ADMIN_EMAIL', 'admin@evore.us')
        if not User.query.filter_by(email=_admin_email).first():
            _admin_pass = os.environ.get('ADMIN_PASSWORD')
            if not _admin_pass:
                _admin_pass = secrets.token_urlsafe(14)
                print(f'ADMIN AUTO-GENERATED PASSWORD (save this!): {_admin_pass}')
            admin = User(nombre='Administrador', email=_admin_email, rol='admin')
            admin.set_password(_admin_pass)
            db.session.add(admin); db.session.commit()
            print(f'Admin creado: {_admin_email}')
        if not ConfigEmpresa.query.first():
            db.session.add(ConfigEmpresa(nombre='Evore', email='contacto@evore.us', sitio_web='evore.us'))
            db.session.commit()

try:
    init_db()
except Exception as _e:
    print(f'init_db() error (no crítico): {_e}')

if __name__ == '__main__':
    app.run(debug=os.environ.get('FLASK_DEBUG', '0') == '1')

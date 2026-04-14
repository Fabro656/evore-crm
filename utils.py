# utils.py — Helper functions and utilities
from flask import current_app, session, request, redirect, url_for, flash
from flask_login import current_user
from extensions import db
from datetime import datetime, timedelta, date as date_type
from functools import wraps
import json, io, re, logging

__all__ = [
    # Module/role constants
    '_MODULOS_TODOS', '_MODULOS_ROL',
    # Payroll constants
    'SMLMV_2025', 'AUXILIO_TRANSPORTE_2025',
    'TASA_SALUD_EMP', 'TASA_PENSION_EMP', 'TASA_SALUD_EMPR', 'TASA_PENSION_EMPR',
    'TASA_CAJA_COMP', 'TASA_SENA', 'TASA_ICBF', 'TASA_ARL',
    'TASA_CESANTIAS', 'TASA_INT_CESANTIAS', 'TASA_PRIMA', 'TASA_VACACIONES',
    # Helpers
    'cop', 'moneda', 'moneda0', 'requiere_modulo', 'inject_globals',
    '_send_email', '_log', '_crear_notificacion', '_crear_asiento_auto',
    '_calcular_nomina', '_calcular_liquidacion', '_calcular_impuestos',
    '_descontar_materias', '_descontar_stock_venta', '_save_contactos',
    '_oc_save_items', '_prods_json', '_save_items', '_save_asignados',
    '_inv_form_ctx', '_save_compra', '_make_xlsx',
    '_procesar_orden_produccion', '_procesar_venta_produccion',
    '_modulos_user', '_get_roles_usuario', '_get_rol_activo', '_ROL_LABELS',
    'register_app_hooks', '_registrar_movimiento', '_actualizar_score_proveedor',
    '_calcular_costo_receta', '_precio_minimo_venta',
    'ONBOARDING_STEPS',
]

# ── Module constants (originally in app.py global scope)
_MODULOS_TODOS = ['clientes','ventas','cotizaciones','tareas','calendario',
                  'notas','inventario','produccion','gastos','reportes','proveedores',
                  'ordenes_compra','legal','finanzas','cotizaciones_proveedor','comercial','config','nomina',
                  'empaques','servicios','aprobaciones','logistica']

# Agrupacion visual para el form de usuario
_MODULOS_GRUPOS = [
    ('Principal', [
        ('tareas','Tickets','check2-square'),
        ('calendario','Calendario','calendar3'),
        ('notas','Notas','sticky-fill'),
    ]),
    ('Comercial', [
        ('clientes','Clientes','people-fill'),
        ('ventas','Ventas','graph-up-arrow'),
        ('cotizaciones','Cotizaciones','file-earmark-text-fill'),
        ('servicios','Servicios','briefcase-fill'),
    ]),
    ('Proveedores y Compras', [
        ('proveedores','Proveedores','truck'),
        ('cotizaciones_proveedor','Cotizaciones proveedor','file-earmark-check'),
        ('ordenes_compra','Ordenes de Compra','cart-check'),
    ]),
    ('Operaciones', [
        ('inventario','Inventario','box-seam-fill'),
        ('produccion','Produccion','gear-fill'),
        ('empaques','Empaques y logistica','box2-heart-fill'),
        ('logistica','Transportistas','truck-front'),
    ]),
    ('Finanzas', [
        ('finanzas','Contabilidad y Asientos','calculator'),
        ('gastos','Gastos operativos','receipt'),
        ('reportes','Reportes','bar-chart-fill'),
        ('nomina','Nomina','people-fill'),
    ]),
    ('Administracion', [
        ('legal','Legal','shield-check'),
        ('aprobaciones','Aprobaciones','clipboard-check'),
        ('config','Configuracion','gear'),
    ]),
]

_MODULOS_ROL = {
    'admin':      _MODULOS_TODOS,
    'tester':     _MODULOS_TODOS,   # acceso a todo, sin crear usuarios ni reset
    'director_financiero': _MODULOS_TODOS,  # Dueño: ve todo, aprueba gastos/compras
    'director_operativo':  ['clientes','ventas','cotizaciones','tareas','calendario','notas',
                            'inventario','produccion','proveedores','ordenes_compra',
                            'cotizaciones_proveedor','empaques','servicios','nomina','reportes',
                            'logistica','aprobaciones','gastos','finanzas'],
    'vendedor':   ['clientes','ventas','cotizaciones','tareas','calendario','notas','nomina','servicios'],
    'produccion': ['inventario','produccion','gastos','notas','calendario','tareas','ordenes_compra','empaques','logistica'],
    'contador':   ['gastos','reportes','produccion','notas','nomina','finanzas','tareas'],
    'usuario':    ['tareas','notas','calendario'],
    'sales_manager': ['clientes','ventas','cotizaciones','tareas','calendario','notas','ordenes_compra','nomina'],
    'cliente':       ['portal_cliente'],
    'proveedor':     ['portal_proveedor'],
}

# ── Multi-rol: labels y helpers ────────────────────────────────────
_ROL_LABELS = {
    'admin': 'Administrador',
    'director_financiero': 'Director Financiero',
    'director_operativo': 'Director Operativo',
    'vendedor': 'Vendedor',
    'sales_manager': 'Sales Manager',
    'produccion': 'Produccion',
    'contador': 'Contador',
    'tester': 'Tester',
    'usuario': 'Usuario',
    'cliente': 'Portal Cliente',
    'proveedor': 'Portal Proveedor',
}
_ROL_ICONS = {
    'admin': 'bi-shield-lock-fill',
    'director_financiero': 'bi-currency-dollar',
    'director_operativo': 'bi-gear-wide-connected',
    'vendedor': 'bi-bag-fill',
    'sales_manager': 'bi-star-fill',
    'produccion': 'bi-gear-fill',
    'contador': 'bi-calculator',
    'tester': 'bi-bug-fill',
    'usuario': 'bi-person',
    'cliente': 'bi-building',
    'proveedor': 'bi-truck',
}

def _get_roles_usuario(user):
    """Retorna lista de roles disponibles para el usuario."""
    if not user or not user.is_authenticated:
        return []
    roles = set()
    roles.add(user.rol)  # rol principal siempre incluido
    try:
        asignados = json.loads(user.roles_asignados or '[]')
        if isinstance(asignados, list):
            roles.update(asignados)
    except Exception:
        pass
    # admin siempre tiene acceso a todo
    if 'admin' in roles:
        return list(_ROL_LABELS.keys())
    return sorted(roles, key=lambda r: list(_ROL_LABELS.keys()).index(r) if r in _ROL_LABELS else 99)

def _get_rol_activo(user):
    """Retorna el rol activo actual desde session, o el rol principal."""
    if not user or not user.is_authenticated:
        return 'usuario'
    rol_session = session.get('rol_activo')
    if rol_session and rol_session in _get_roles_usuario(user):
        return rol_session
    return user.rol

# ── Onboarding por rol ─────────────────────────────────────────────
ONBOARDING_STEPS = {
    'admin': [
        {'key':'empresa','titulo':'Configura tu empresa','desc':'Nombre, NIT, direccion, firma','url':'admin_config','icon':'bi-building'},
        {'key':'usuarios','titulo':'Invita a tu equipo','desc':'Crea usuarios para cada area','url':'admin_usuarios','icon':'bi-people'},
        {'key':'productos','titulo':'Agrega productos','desc':'Tu catalogo de productos terminados','url':'inventario','icon':'bi-box'},
        {'key':'receta','titulo':'Crea una receta','desc':'Formula/BOM de un producto','url':'produccion_index','icon':'bi-clipboard2-data'},
        {'key':'cliente','titulo':'Registra tu primer cliente','desc':'Empieza a vender','url':'clientes','icon':'bi-person-plus'},
        {'key':'ordenes_compra','titulo':'Ordenes de Compra','desc':'Compra materiales con cotizaciones multiples y seguimiento contable','url':'ordenes_compra','icon':'bi-cart-check'},
        {'key':'aprobaciones','titulo':'Sistema de aprobaciones','desc':'Revisa y aprueba solicitudes del equipo','url':'aprobaciones_pendientes','icon':'bi-clipboard-check'},
        {'key':'nomina','titulo':'Nomina y liquidaciones','desc':'Cierra nomina mensual con prorrateo por dias','url':'nomina_index','icon':'bi-people-fill'},
    ],
    'director_financiero': [
        {'key':'dashboard','titulo':'Tu panel financiero','desc':'Revisa ingresos, egresos y utilidad','url':'contable_index','icon':'bi-speedometer2'},
        {'key':'puc','titulo':'Revisa el Plan de Cuentas','desc':'PUC colombiano configurado','url':'contable_puc','icon':'bi-list-ol'},
        {'key':'aprobaciones','titulo':'Aprobaciones pendientes','desc':'Revisa y aprueba gastos del equipo','url':'aprobaciones_pendientes','icon':'bi-clipboard-check'},
        {'key':'balance','titulo':'Genera un Balance General','desc':'Activos = Pasivos + Patrimonio','url':'contable_balance_general','icon':'bi-bank2'},
        {'key':'asientos','titulo':'Asientos contables','desc':'Confirma pagos de OC e ingresos de ventas','url':'contable_asientos','icon':'bi-journal-bookmark-fill'},
    ],
    'director_operativo': [
        {'key':'produccion','titulo':'Ordenes de produccion','desc':'Revisa el estado de produccion','url':'ordenes_produccion','icon':'bi-gear'},
        {'key':'compras','titulo':'Ordenes de compra','desc':'Gestiona compras a proveedores','url':'ordenes_compra','icon':'bi-cart4'},
        {'key':'inventario','titulo':'Inventario de productos','desc':'Stock disponible y alertas','url':'inventario','icon':'bi-box-seam'},
        {'key':'proveedores','titulo':'Gestiona proveedores','desc':'Directorio de proveedores','url':'proveedores','icon':'bi-truck'},
    ],
    'vendedor': [
        {'key':'clientes','titulo':'Conoce tus clientes','desc':'Tu cartera de clientes','url':'clientes','icon':'bi-people'},
        {'key':'cotizacion','titulo':'Crea una cotizacion','desc':'Prepara una propuesta comercial','url':'cotizacion_nueva','icon':'bi-file-earmark-text'},
        {'key':'venta','titulo':'Registra una venta','desc':'O convierte una cotizacion aprobada','url':'venta_nueva','icon':'bi-cart-plus'},
        {'key':'tareas','titulo':'Gestiona tus tareas','desc':'Seguimiento a pendientes','url':'tareas','icon':'bi-check2-square'},
        {'key':'tickets','titulo':'Tickets de seguimiento','desc':'Crea y gestiona tickets para cada negocio','url':'tareas','icon':'bi-check2-square'},
    ],
    'produccion': [
        {'key':'inventario','titulo':'Revisa el inventario','desc':'Stock de productos y materias primas','url':'inventario','icon':'bi-box-seam'},
        {'key':'materias','titulo':'Materias primas','desc':'Stock disponible de insumos','url':'materias','icon':'bi-droplet-half'},
        {'key':'recetas','titulo':'Recetas / BOM','desc':'Formulas de cada producto','url':'recetas','icon':'bi-clipboard2-data'},
        {'key':'ordenes','titulo':'Ordenes de produccion','desc':'Produccion pendiente y en curso','url':'ordenes_produccion','icon':'bi-gear'},
        {'key':'recepcion','titulo':'Recepcion de material','desc':'Recibe material desde ordenes de compra','url':'compras','icon':'bi-box-arrow-in-down'},
    ],
    'contador': [
        {'key':'puc','titulo':'Plan de Cuentas (PUC)','desc':'Catalogo contable colombiano','url':'contable_puc','icon':'bi-list-ol'},
        {'key':'asiento','titulo':'Crea un asiento contable','desc':'Registra una operacion en el libro','url':'contable_asiento_nuevo','icon':'bi-journal-plus'},
        {'key':'impuestos','titulo':'Reglas tributarias','desc':'IVA, retencion, ICA','url':'impuestos','icon':'bi-percent'},
        {'key':'balance','titulo':'Balance de Prueba','desc':'Verifica que todo cuadre','url':'contable_balance_prueba','icon':'bi-table'},
        {'key':'resultados','titulo':'Estado de Resultados','desc':'Ingresos vs Gastos del periodo','url':'contable_estado_resultados','icon':'bi-graph-up-arrow'},
        {'key':'asientos_oc','titulo':'Asientos desde OC','desc':'Los pagos de ordenes de compra se confirman aqui','url':'contable_asientos','icon':'bi-journal-bookmark-fill'},
    ],
    'sales_manager': [
        {'key':'clientes','titulo':'Cartera de clientes','desc':'Clientes asignados a tu equipo','url':'clientes','icon':'bi-people'},
        {'key':'cotizaciones','titulo':'Cotizaciones','desc':'Pipeline comercial','url':'cotizaciones','icon':'bi-file-earmark-text'},
        {'key':'ventas','titulo':'Ventas del equipo','desc':'Estado de cada negocio','url':'ventas','icon':'bi-graph-up-arrow'},
        {'key':'portal','titulo':'Portal del cliente','desc':'Revisa pre-cotizaciones','url':'portal_cliente','icon':'bi-shop'},
    ],
    'cliente': [
        {'key':'portal','titulo':'Tu portal de compras','desc':'Revisa tus pedidos y cotizaciones','url':'portal_cliente','icon':'bi-shop'},
        {'key':'precotizacion','titulo':'Solicita una cotizacion','desc':'Envia tu pedido al equipo comercial','url':'portal_pre_cotizacion_nueva','icon':'bi-file-earmark-plus'},
    ],
    'proveedor': [
        {'key':'portal','titulo':'Tu portal de ordenes','desc':'Revisa ordenes de compra','url':'portal_proveedor','icon':'bi-truck'},
        {'key':'ticket','titulo':'Envia un mensaje','desc':'Contacta al equipo de compras','url':'portal_prov_ticket','icon':'bi-chat-left-text'},
    ],
}

# Acciones que requieren aprobación de director_financiero
REQUIERE_APROBACION = {
    'gasto_nuevo':       'Registrar gasto operativo',
    'compra_nueva':      'Registrar compra de materia prima',
    'nomina_cerrar':     'Cerrar nómina mensual',
    'orden_compra_nueva':'Crear orden de compra',
}

# ── Payroll constants from company config ──────────────────────────
from company_config import COMPANY, COMPANY_ID
_payroll = COMPANY['payroll']
SMLMV_2025              = _payroll.get('min_wage', 0)
AUXILIO_TRANSPORTE_2025 = _payroll.get('transport_subsidy', 0)
TASA_SALUD_EMP          = _payroll.get('health_employee', 0.04)
TASA_PENSION_EMP        = _payroll.get('pension_employee', 0.04)
TASA_SALUD_EMPR         = _payroll.get('health_employer', 0.085)
TASA_PENSION_EMPR       = _payroll.get('pension_employer', 0.12)
TASA_CAJA_COMP          = _payroll.get('caja_comp', 0.04)
TASA_SENA               = _payroll.get('sena', 0.02)
TASA_ICBF               = _payroll.get('icbf', 0.03)
TASA_ARL                = _payroll.get('arl', {1: 0.00522})
TASA_CESANTIAS          = _payroll.get('cesantias', 1/12)
TASA_INT_CESANTIAS      = _payroll.get('int_cesantias', 0.12)
TASA_PRIMA              = _payroll.get('prima', 1/12)
TASA_VACACIONES         = _payroll.get('vacaciones', 0.0417)

def _cargar_nomina_params():
    """Carga parametros de nomina desde DB (ConfigEmpresa.nomina_params) y sobreescribe globals."""
    global SMLMV_2025, AUXILIO_TRANSPORTE_2025, TASA_SALUD_EMP, TASA_PENSION_EMP
    global TASA_SALUD_EMPR, TASA_PENSION_EMPR, TASA_CAJA_COMP, TASA_SENA, TASA_ICBF
    global TASA_ARL, TASA_CESANTIAS, TASA_INT_CESANTIAS, TASA_PRIMA, TASA_VACACIONES
    try:
        from models import ConfigEmpresa
        import json
        cfg = ConfigEmpresa.query.first()
        if cfg and cfg.nomina_params:
            p = json.loads(cfg.nomina_params)
            if p.get('min_wage'):              SMLMV_2025 = float(p['min_wage'])
            if p.get('transport_subsidy'):     AUXILIO_TRANSPORTE_2025 = float(p['transport_subsidy'])
            if p.get('health_employee'):       TASA_SALUD_EMP = float(p['health_employee'])
            if p.get('pension_employee'):      TASA_PENSION_EMP = float(p['pension_employee'])
            if p.get('health_employer'):       TASA_SALUD_EMPR = float(p['health_employer'])
            if p.get('pension_employer'):      TASA_PENSION_EMPR = float(p['pension_employer'])
            if p.get('caja_comp'):             TASA_CAJA_COMP = float(p['caja_comp'])
            if p.get('sena'):                  TASA_SENA = float(p['sena'])
            if p.get('icbf'):                  TASA_ICBF = float(p['icbf'])
    except Exception:
        pass  # fallback to company_config defaults

# ── Mail setup (graceful degradation if Flask-Mail not installed/configured)
try:
    from flask_mail import Message as MailMessage
    _mail_ok = True   # confirmed against MAIL_SERVER config when first called
    _mail = None      # actual Mail instance lives in extensions.py
except ImportError:
    MailMessage = None
    _mail = None
    _mail_ok = False

# ── Model imports (no circular dependency: models.py only imports from extensions)
from models import (
    User, Cliente, ContactoCliente, Proveedor,
    Venta, VentaProducto, Producto, LoteProducto,
    MateriaPrima, RecetaProducto, OrdenProduccion, ReservaProduccion,
    OrdenCompraItem, Tarea, TareaAsignado,
    GastoOperativo, Actividad, Notificacion, ReglaTributaria,
    Aprobacion,
)

def _format_currency(value, decimals=None):
    """Formatea valor monetario segun la configuracion de la empresa activa."""
    try:
        v = float(value or 0)
        d = decimals if decimals is not None else COMPANY.get('currency_decimals', 0)
        sym = COMPANY.get('currency_symbol', '$')
        tsep = COMPANY.get('thousand_sep', '.')
        dsep = COMPANY.get('decimal_sep', ',')
        if d == 0:
            formatted = f'{v:,.0f}'
        else:
            formatted = f'{v:,.{d}f}'
        # Replace separators: Python uses , for thousands and . for decimal
        # Swap to temp, then to target
        formatted = formatted.replace(',', 'TEMP').replace('.', dsep).replace('TEMP', tsep)
        return f'{sym} {formatted}'
    except Exception:
        return f'{COMPANY.get("currency_symbol", "$")} 0'

def cop(value):
    return _format_currency(value, 0)

def moneda(value):
    return _format_currency(value)

def moneda0(value):
    return _format_currency(value, 0)

def requiere_modulo(modulo):
    def decorator(f):
        @wraps(f)
        def wrapped(*a, **kw):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))
            rol_activo = _get_rol_activo(current_user)
            if rol_activo == 'admin' or modulo in _modulos_user(current_user):
                return f(*a, **kw)
            flash('No tienes acceso a este módulo.', 'danger')
            return redirect(url_for('dashboard'))
        return wrapped
    return decorator

_inject_cache = {}
_INJECT_CACHE_TTL = 15  # seconds

def _cached_count(key, query_fn, user_id):
    """Cache DB count queries for 15 seconds per user."""
    import time
    cache_key = f'{key}_{user_id}'
    now = time.time()
    if cache_key in _inject_cache and now - _inject_cache[cache_key][1] < _INJECT_CACHE_TTL:
        return _inject_cache[cache_key][0]
    result = query_fn()
    _inject_cache[cache_key] = (result, now)
    return result

def inject_globals():
    modulos = _modulos_user(current_user) if current_user.is_authenticated else []
    notif_count = 0
    empresa_cliente_nombre = None
    empresa_proveedor_nombre = None
    if current_user.is_authenticated:
        # Validate session role: clear if user no longer has it
        if 'rol_activo' in session:
            _roles_validos = _get_roles_usuario(current_user)
            if session['rol_activo'] not in _roles_validos:
                session.pop('rol_activo', None)
        try:
            notif_count = _cached_count('notif',
                lambda: Notificacion.query.filter_by(usuario_id=current_user.id, leida=False).count(),
                current_user.id)
        except Exception: pass
        _rol_act = _get_rol_activo(current_user)
        if _rol_act == 'cliente':
            try:
                if current_user.cliente_id:
                    cli = db.session.get(Cliente, current_user.cliente_id)
                    if cli: empresa_cliente_nombre = cli.empresa or cli.nombre
            except Exception: pass
        if _rol_act == 'proveedor':
            try:
                prov_id = getattr(current_user, 'proveedor_id', None)
                if prov_id:
                    prov = db.session.get(Proveedor, prov_id)
                    if prov: empresa_proveedor_nombre = prov.nombre
            except Exception: pass
    # Onboarding por rol
    onboarding_data = None
    if current_user.is_authenticated and not getattr(current_user, 'onboarding_dismissed', False):
        rol = _get_rol_activo(current_user)
        steps = ONBOARDING_STEPS.get(rol, [])
        if steps:
            try:
                completed = json.loads(getattr(current_user, 'onboarding_role_config', '{}') or '{}')
            except Exception: completed = {}
            pending = [s for s in steps if s['key'] not in completed]
            if pending:
                onboarding_data = {
                    'steps': steps, 'completed': completed,
                    'pending': pending, 'total': len(steps),
                    'done': len(steps) - len(pending),
                    'current': pending[0] if pending else None
                }
    # Multi-rol
    rol_activo = _get_rol_activo(current_user) if current_user.is_authenticated else 'usuario'
    roles_disponibles = _get_roles_usuario(current_user) if current_user.is_authenticated else []
    # Quick counts for sidebar badges
    _tareas_pend = 0
    _aprob_pend = 0
    if current_user.is_authenticated:
        try:
            _tareas_pend = _cached_count('tareas',
                lambda: Tarea.query.filter(Tarea.estado != 'completada', Tarea.asignado_a == current_user.id).count(),
                current_user.id)
        except Exception: pass
        try:
            if _get_rol_activo(current_user) in ('admin','director_financiero','director_operativo'):
                _aprob_pend = _cached_count('aprob',
                    lambda: Aprobacion.query.filter_by(estado='pendiente').count(),
                    current_user.id)
        except Exception: pass
    return {'now': datetime.utcnow(), 'modulos_user': modulos, 'notif_count': notif_count,
            'empresa_cliente_nombre': empresa_cliente_nombre, 'empresa_proveedor_nombre': empresa_proveedor_nombre,
            'onboarding': onboarding_data,
            'rol_activo': rol_activo,
            'roles_disponibles': roles_disponibles,
            'rol_labels': _ROL_LABELS,
            'rol_icons': _ROL_ICONS,
            'tareas_pend': _tareas_pend, 'aprob_pend': _aprob_pend,
            'company_name': COMPANY['name'], 'company_config': COMPANY,
            'nit_label': COMPANY.get('nit_label', 'NIT'),
            'currency_code': COMPANY.get('currency_code', 'COP')}

def _send_email(to, subject, body):
    if not _mail_ok or not MailMessage: return
    try:
        from extensions import mail as _mail_ext, MAIL_AVAILABLE
        if not MAIL_AVAILABLE or not _mail_ext: return
        if not current_app.config.get('MAIL_SERVER'): return
        msg = MailMessage(subject, recipients=[to], body=body)
        _mail_ext.send(msg)
    except Exception as e:
        logging.warning(f'Email error: {e}')

def _log(tipo, entidad, entidad_id, descripcion):
    try:
        db.session.add(Actividad(
            tipo=tipo, entidad=entidad, entidad_id=entidad_id,
            descripcion=descripcion, usuario_id=current_user.id))
    except Exception:
        pass

# Mapeo de cuentas textuales a codigos PUC colombiano
_PUC_MAP = {
    'Gastos de nomina': '510506 Sueldos',
    'Gastos de nomina - Liquidaciones': '510530 Cesantias',
    'Bancos / Caja': '111005 Moneda nacional',
    'Cuentas por cobrar clientes': '130505 Nacionales',
    'Ingresos por ventas': '4135 Comercio al por mayor y menor',
    'Inventario materias primas': '1405 Materias primas',
    'Proveedores nacionales': '220505 Nacionales',
    'Gastos Nomina': '510506 Sueldos',
}

def _registrar_movimiento(producto_id=None, materia_prima_id=None, tipo='ajuste',
                          cantidad=0, stock_anterior=0, stock_posterior=0, referencia='', usuario_id=None):
    """Registra un movimiento de inventario en el audit trail."""
    try:
        from models import MovimientoInventario
        mov = MovimientoInventario(
            producto_id=producto_id, materia_prima_id=materia_prima_id,
            tipo=tipo, cantidad=cantidad,
            stock_anterior=stock_anterior, stock_posterior=stock_posterior,
            referencia=referencia, usuario_id=usuario_id
        )
        db.session.add(mov)
    except Exception as e:
        import logging
        logging.warning(f'_registrar_movimiento error: {e}')


def _actualizar_score_proveedor(proveedor_id):
    """Recalcula score del proveedor basado en historial de OC."""
    try:
        from models import Proveedor, OrdenCompra
        prov = db.session.get(Proveedor, proveedor_id)
        if not prov: return
        ocs = OrdenCompra.query.filter_by(proveedor_id=proveedor_id).all()
        prov.total_oc = len(ocs)
        # Score entrega: % de OC recibidas a tiempo
        recibidas = [oc for oc in ocs if oc.estado == 'recibida']
        if recibidas:
            a_tiempo = sum(1 for oc in recibidas if oc.fecha_esperada and oc.fecha_emision and
                          (oc.fecha_esperada >= oc.fecha_emision))
            prov.score_entrega = round(min(10, (a_tiempo / len(recibidas)) * 10), 1)
        # Score calidad: penalizar por rechazos
        rechazos = sum(1 for oc in ocs if oc.tiene_problema_calidad)
        prov.total_rechazos = rechazos
        if ocs:
            prov.score_calidad = round(max(1, 10 - (rechazos / len(ocs)) * 10), 1)
    except Exception:
        pass


def _resolver_puc(cuenta_texto):
    """Resuelve texto de cuenta a codigo PUC si existe mapeo."""
    return _PUC_MAP.get(cuenta_texto, cuenta_texto)

def _crear_asiento_auto(tipo, subtipo, descripcion, monto, cuenta_debe, cuenta_haber,
                        clasificacion='egreso', referencia=None, venta_id=None,
                        orden_compra_id=None, gasto_id=None, proveedor_id=None):
    """Crea un AsientoContable automático con auto-numeración."""
    try:
        asiento = AsientoContable(
            numero='AC-TEMP',
            fecha=datetime.utcnow().date(),
            descripcion=descripcion[:300],
            tipo=tipo, subtipo=subtipo,
            referencia=referencia,
            debe=float(monto), haber=float(monto),
            cuenta_debe=_resolver_puc(cuenta_debe), cuenta_haber=_resolver_puc(cuenta_haber),
            clasificacion=clasificacion,
            venta_id=venta_id, orden_compra_id=orden_compra_id,
            gasto_id=gasto_id, proveedor_id=proveedor_id,
            creado_por=current_user.id if current_user and current_user.is_authenticated else None
        )
        db.session.add(asiento)
        db.session.flush()
        asiento.numero = f'AC-{datetime.utcnow().year}-{asiento.id:04d}'

        # Crear LineaAsiento para que aparezca en balances PUC
        try:
            from models import LineaAsiento, CuentaPUC
            cuenta_debe_resolved = _resolver_puc(cuenta_debe)
            cuenta_haber_resolved = _resolver_puc(cuenta_haber)
            # Buscar CuentaPUC por codigo (primeros 6 digitos)
            codigo_debe = cuenta_debe_resolved.split(' ')[0] if cuenta_debe_resolved else ''
            codigo_haber = cuenta_haber_resolved.split(' ')[0] if cuenta_haber_resolved else ''
            puc_debe = CuentaPUC.query.filter(CuentaPUC.codigo == codigo_debe, CuentaPUC.activo == True).first()
            puc_haber = CuentaPUC.query.filter(CuentaPUC.codigo == codigo_haber, CuentaPUC.activo == True).first()
            if puc_debe:
                db.session.add(LineaAsiento(asiento_id=asiento.id, cuenta_puc_id=puc_debe.id,
                    descripcion=descripcion[:300], debe=float(monto), haber=0))
            if puc_haber:
                db.session.add(LineaAsiento(asiento_id=asiento.id, cuenta_puc_id=puc_haber.id,
                    descripcion=descripcion[:300], debe=0, haber=float(monto)))
        except Exception as le:
            logging.warning(f'_crear_asiento_auto LineaAsiento error: {le}')

        return asiento
    except Exception as e:
        logging.warning(f'_crear_asiento_auto error: {e}')
        return None

def _calcular_costo_receta(producto_id):
    """
    Calcula el costo de producción de un producto basado en su receta activa
    y los costos de materias primas (preferiblemente de cotización vigente).
    Retorna dict: {costo_total, costo_unitario, desglose: [{materia, cantidad, costo_unit, subtotal}], alertas: []}
    """
    from models import Producto, RecetaProducto, RecetaItem, MateriaPrima, CotizacionProveedor
    prod = db.session.get(Producto, producto_id)
    if not prod:
        return {'costo_total': 0, 'costo_unitario': 0, 'desglose': [], 'alertas': ['Producto no encontrado']}

    receta = RecetaProducto.query.filter_by(producto_id=producto_id, activo=True).first()
    if not receta:
        return {'costo_total': 0, 'costo_unitario': 0, 'desglose': [], 'alertas': ['Sin receta activa']}

    desglose = []
    alertas = []
    costo_total = 0

    for ri in receta.items:
        mp = db.session.get(MateriaPrima, ri.materia_prima_id)
        if not mp:
            alertas.append(f'Materia prima ID {ri.materia_prima_id} no encontrada')
            continue

        es_empaque = getattr(ri, 'es_empaque', False)
        rendimiento = float(getattr(ri, 'rendimiento', 1) or 1)

        # Buscar cotización vigente para esta materia prima
        from datetime import date
        cot_vigente = CotizacionProveedor.query.filter(
            CotizacionProveedor.materia_prima_id == mp.id,
            CotizacionProveedor.estado == 'vigente',
            CotizacionProveedor.vigencia >= date.today()
        ).order_by(CotizacionProveedor.precio_unitario.asc()).first()

        if cot_vigente:
            precio_compra = cot_vigente.precio_unitario  # precio antes de IVA
        elif mp.costo_unitario and mp.costo_unitario > 0:
            precio_compra = mp.costo_unitario
            alertas.append(f'{mp.nombre}: sin cotización vigente, usando costo registrado')
        else:
            precio_compra = 0
            alertas.append(f'{mp.nombre}: SIN COTIZACIÓN NI COSTO')

        if es_empaque and rendimiento > 1:
            # EMPAQUE: precio_compra es el precio de 1 unidad de empaque (ej: 1 caja = $8000)
            # rendimiento = cuantas piezas caben en 1 empaque (ej: 150)
            # Costo por pieza = precio / rendimiento = 8000 / 150 = $53.33
            costo_por_pieza = precio_compra / rendimiento
            # Para el lote completo: costo_por_pieza * unidades_produce
            cantidad_total = receta.unidades_produce  # se expresa en piezas
            subtotal = costo_por_pieza * receta.unidades_produce
            costo_unit_display = costo_por_pieza  # lo que se muestra como "costo unitario"
        else:
            # MATERIA PRIMA NORMAL: cantidad_por_unidad * lote * precio
            cantidad_total = ri.cantidad_por_unidad * receta.unidades_produce
            costo_unit_display = precio_compra
            subtotal = cantidad_total * precio_compra

        costo_total += subtotal

        desglose.append({
            'materia': mp.nombre,
            'materia_id': mp.id,
            'cantidad': cantidad_total,
            'unidad': 'piezas' if es_empaque and rendimiento > 1 else mp.unidad,
            'costo_unit': costo_unit_display,
            'precio_compra': precio_compra,
            'subtotal': subtotal,
            'tiene_cotizacion': cot_vigente is not None,
            'es_empaque': es_empaque,
            'rendimiento': rendimiento,
            'stock_disponible': mp.stock_disponible or 0,
            'stock_reservado': mp.stock_reservado or 0,
        })

    costo_unitario = (costo_total / receta.unidades_produce) if receta.unidades_produce > 0 else 0

    # Actualizar costo y precio de venta en el producto
    try:
        prod.costo_receta = round(costo_unitario, 2)
        prod.costo = round(costo_unitario, 2)
        # Recalcular precio de venta: costo + margen + IVA
        margen_pct = receta.margen_pct or 30
        try:
            from models import ReglaTributaria
            regla_iva = ReglaTributaria.query.filter_by(aplica_a='ventas', activo=True).first()
            iva_pct = float(regla_iva.porcentaje) if regla_iva else 19.0
        except Exception:
            iva_pct = 19.0
        precio_sin_iva = costo_unitario * (1 + margen_pct / 100)
        precio_venta = round(precio_sin_iva * (1 + iva_pct / 100), 2)
        receta.costo_calculado = round(costo_unitario, 2)
        receta.precio_venta_sugerido = precio_venta
        prod.precio = precio_venta
    except Exception:
        pass

    return {
        'costo_total': round(costo_total, 2),
        'costo_unitario': round(costo_unitario, 2),
        'unidades_produce': receta.unidades_produce,
        'desglose': desglose,
        'alertas': alertas
    }

def _precio_minimo_venta(producto_id, cantidad=1):
    """
    Calcula el precio mínimo de venta considerando costo receta + IVA.
    Retorna dict: {precio_minimo, costo_produccion, iva, margen_sugerido}
    """
    from models import ReglaTributaria
    costo = _calcular_costo_receta(producto_id)
    costo_unit = costo['costo_unitario']

    # IVA vigente
    try:
        regla = ReglaTributaria.query.filter_by(aplica_a='ventas', activo=True).first()
        iva_pct = float(regla.porcentaje) if regla else 19.0
    except Exception:
        iva_pct = 19.0

    costo_total = costo_unit * cantidad
    iva = round(costo_total * iva_pct / 100, 2)
    margen_30 = round(costo_total * 0.30, 2)  # 30% suggested margin
    precio_minimo = round(costo_total + iva, 2)
    precio_sugerido = round(costo_total + margen_30 + (costo_total + margen_30) * iva_pct / 100, 2)

    return {
        'costo_produccion': round(costo_total, 2),
        'costo_unitario': round(costo_unit, 2),
        'iva_pct': iva_pct,
        'iva': iva,
        'precio_minimo': precio_minimo,
        'precio_sugerido': precio_sugerido,
        'margen_sugerido_pct': 30,
        'alertas': costo['alertas'],
        'cantidad': cantidad
    }

def _crear_notificacion(usuario_id, tipo, titulo, mensaje, url=None):
    try:
        n = Notificacion(usuario_id=usuario_id, tipo=tipo, titulo=titulo,
                         mensaje=mensaje, url=url)
        db.session.add(n)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logging.warning(f'Notificacion error: {e}')


def _calcular_nomina_co(empleado):
    """Colombian payroll calculation. Returns a dict with all payroll items."""
    salario = empleado.salario_base
    aux_transporte = AUXILIO_TRANSPORTE_2025 if (empleado.auxilio_transporte and salario <= 2 * SMLMV_2025) else 0

    # Deducciones empleado
    deduccion_salud     = round(salario * TASA_SALUD_EMP)
    deduccion_pension   = round(salario * TASA_PENSION_EMP)
    # Fondo solidaridad: 1% si salario > 4 SMLMV, 1.2% si > 16 SMLMV
    fondo_solidaridad = 0
    if salario > 16 * SMLMV_2025:
        fondo_solidaridad = round(salario * 0.012)
    elif salario > 4 * SMLMV_2025:
        fondo_solidaridad = round(salario * 0.01)

    total_deducciones = deduccion_salud + deduccion_pension + fondo_solidaridad
    salario_neto = salario + aux_transporte - total_deducciones

    # Aportes empleador (no afectan neto del empleado)
    aporte_salud_empr    = round(salario * TASA_SALUD_EMPR)
    aporte_pension_empr  = round(salario * TASA_PENSION_EMPR)
    tasa_arl             = TASA_ARL.get(empleado.nivel_riesgo_arl, TASA_ARL[1])
    aporte_arl           = round(salario * tasa_arl)
    aporte_caja          = round(salario * TASA_CAJA_COMP)
    aporte_sena          = round(salario * TASA_SENA)
    aporte_icbf          = round(salario * TASA_ICBF)
    total_costo_empr     = salario + aux_transporte + aporte_salud_empr + aporte_pension_empr + aporte_arl + aporte_caja + aporte_sena + aporte_icbf

    # Prestaciones sociales (provisionadas mensualmente)
    provision_cesantias      = round((salario + aux_transporte) * TASA_CESANTIAS)
    provision_int_cesantias  = round(provision_cesantias * TASA_INT_CESANTIAS / 12)
    provision_prima          = round((salario + aux_transporte) * TASA_PRIMA)
    provision_vacaciones     = round(salario * TASA_VACACIONES)
    total_prestaciones       = provision_cesantias + provision_int_cesantias + provision_prima + provision_vacaciones

    return {
        'salario': salario,
        'aux_transporte': aux_transporte,
        'deduccion_salud': deduccion_salud,
        'deduccion_pension': deduccion_pension,
        'fondo_solidaridad': fondo_solidaridad,
        'total_deducciones': total_deducciones,
        'salario_neto': salario_neto,
        'aporte_salud_empr': aporte_salud_empr,
        'aporte_pension_empr': aporte_pension_empr,
        'aporte_arl': aporte_arl,
        'aporte_caja': aporte_caja,
        'aporte_sena': aporte_sena,
        'aporte_icbf': aporte_icbf,
        'total_costo_empr': total_costo_empr,
        'provision_cesantias': provision_cesantias,
        'provision_int_cesantias': provision_int_cesantias,
        'provision_prima': provision_prima,
        'provision_vacaciones': provision_vacaciones,
        'total_prestaciones': total_prestaciones,
        'costo_total_empresa': total_costo_empr + total_prestaciones,
    }


def _calcular_isr(ingreso_mensual):
    """Calculate Mexican ISR (income tax) using progressive brackets."""
    brackets = _payroll.get('isr_brackets', [])
    for lim_inf, lim_sup, cuota_fija, tasa in brackets:
        if lim_inf <= ingreso_mensual <= lim_sup:
            excedente = ingreso_mensual - lim_inf
            return round(cuota_fija + excedente * tasa)
    # If above all brackets, use the last one
    if brackets:
        lim_inf, _, cuota_fija, tasa = brackets[-1]
        return round(cuota_fija + (ingreso_mensual - lim_inf) * tasa)
    return 0


def _calcular_nomina_mx(empleado):
    """Mexican payroll calculation. Returns a dict with the same structure as Colombian version."""
    salario = empleado.salario_base
    aux_transporte = 0  # Mexico does not have transport subsidy

    # --- Employee deductions ---
    # ISR (income tax) — progressive brackets
    deduccion_salud = _calcular_isr(salario)  # reuse key name: maps to ISR
    # IMSS employee: enfermedad_maternidad + invalidez_vida + cesantia_vejez
    imss_enf_mat_emp  = round(salario * _payroll.get('imss_employee_enf_mat', 0.00625))
    imss_inv_vida_emp = round(salario * _payroll.get('imss_employee_invalidez', 0.00625))
    imss_ces_vej_emp  = round(salario * _payroll.get('imss_employee_cesantia', 0.01125))
    deduccion_pension = imss_enf_mat_emp + imss_inv_vida_emp + imss_ces_vej_emp  # reuse key name: maps to IMSS employee total
    fondo_solidaridad = 0  # Not applicable in Mexico

    total_deducciones = deduccion_salud + deduccion_pension + fondo_solidaridad
    salario_neto = salario + aux_transporte - total_deducciones

    # --- Employer contributions ---
    # IMSS employer
    imss_enf_mat_empr  = round(salario * _payroll.get('imss_employer_enf_mat', 0.105))
    imss_inv_vida_empr = round(salario * _payroll.get('imss_employer_invalidez', 0.0175))
    imss_ces_vej_empr  = round(salario * _payroll.get('imss_employer_cesantia', 0.0315))
    imss_riesgo_empr   = round(salario * _payroll.get('imss_employer_riesgo', 0.005))
    imss_guarderias    = round(salario * _payroll.get('imss_employer_guarderias', 0.01))
    imss_retiro        = round(salario * _payroll.get('imss_employer_retiro', 0.02))
    # INFONAVIT (employer only)
    infonavit          = round(salario * _payroll.get('infonavit', 0.05))

    # Map to same keys as Colombian version for template compatibility
    aporte_salud_empr   = imss_enf_mat_empr + imss_inv_vida_empr  # IMSS enfermedad + invalidez
    aporte_pension_empr = imss_ces_vej_empr + imss_retiro          # IMSS cesantia + retiro (SAR)
    aporte_arl          = imss_riesgo_empr                          # Riesgo de trabajo
    aporte_caja         = imss_guarderias                           # Guarderias
    aporte_sena         = infonavit                                 # INFONAVIT (reuse key)
    aporte_icbf         = 0                                         # No equivalent in Mexico

    total_costo_empr = salario + aux_transporte + aporte_salud_empr + aporte_pension_empr + aporte_arl + aporte_caja + aporte_sena + aporte_icbf

    # --- Prestaciones (provisioned monthly) ---
    # Aguinaldo: 15 days / 12 months
    aguinaldo_days = _payroll.get('aguinaldo_days', 15)
    salario_diario = salario / 30
    provision_cesantias     = round(salario_diario * aguinaldo_days / 12)  # monthly aguinaldo provision
    provision_int_cesantias = 0  # No interest on aguinaldo in Mexico

    # Prima vacacional: vacation days * 25% * daily salary / 12
    vac_days = _payroll.get('vacaciones_min_days', 12)
    prima_vac_pct = _payroll.get('prima_vacacional_pct', 0.25)
    provision_prima = round(salario_diario * vac_days * prima_vac_pct / 12)

    # Vacaciones: vacation days salary / 12
    provision_vacaciones = round(salario_diario * vac_days / 12)

    total_prestaciones = provision_cesantias + provision_int_cesantias + provision_prima + provision_vacaciones

    return {
        'salario': salario,
        'aux_transporte': aux_transporte,
        'deduccion_salud': deduccion_salud,
        'deduccion_pension': deduccion_pension,
        'fondo_solidaridad': fondo_solidaridad,
        'total_deducciones': total_deducciones,
        'salario_neto': salario_neto,
        'aporte_salud_empr': aporte_salud_empr,
        'aporte_pension_empr': aporte_pension_empr,
        'aporte_arl': aporte_arl,
        'aporte_caja': aporte_caja,
        'aporte_sena': aporte_sena,
        'aporte_icbf': aporte_icbf,
        'total_costo_empr': total_costo_empr,
        'provision_cesantias': provision_cesantias,
        'provision_int_cesantias': provision_int_cesantias,
        'provision_prima': provision_prima,
        'provision_vacaciones': provision_vacaciones,
        'total_prestaciones': total_prestaciones,
        'costo_total_empresa': total_costo_empr + total_prestaciones,
        # Mexican-specific detail (extra keys — templates can optionally show these)
        'isr': deduccion_salud,
        'imss_empleado': deduccion_pension,
        'imss_empleador_detalle': {
            'enfermedad_maternidad': imss_enf_mat_empr,
            'invalidez_vida': imss_inv_vida_empr,
            'cesantia_vejez': imss_ces_vej_empr,
            'riesgo_trabajo': imss_riesgo_empr,
            'guarderias': imss_guarderias,
            'retiro': imss_retiro,
        },
        'infonavit': infonavit,
        'aguinaldo_provision': provision_cesantias,
        'prima_vacacional_provision': provision_prima,
    }


def _calcular_nomina(empleado):
    """Dispatcher — calls Colombian or Mexican payroll based on company config."""
    system = _payroll.get('system', 'colombian')
    if system == 'mexican':
        return _calcular_nomina_mx(empleado)
    return _calcular_nomina_co(empleado)


def _calcular_liquidacion_co(empleado, motivo):
    """
    Colombian employee liquidation.
    motivo: 'renuncia' | 'despido_justa' | 'despido_sin_justa' | 'mutuo_acuerdo'
    """
    from datetime import date as date_t
    hoy = date_t.today()
    fecha_retiro = empleado.fecha_retiro or hoy
    fecha_ingreso = empleado.fecha_ingreso
    if not fecha_ingreso:
        return None

    # Días trabajados
    dias_trabajados = (fecha_retiro - fecha_ingreso).days
    anios = dias_trabajados / 365.25
    meses = dias_trabajados / 30.417

    salario = empleado.salario_base
    aux_transporte = AUXILIO_TRANSPORTE_2025 if (empleado.auxilio_transporte and salario <= 2 * SMLMV_2025) else 0
    salario_con_aux = salario + aux_transporte

    # Cesantías: 1 mes de salario con aux por año trabajado (proporcional)
    cesantias = round(salario_con_aux * dias_trabajados / 365.25)

    # Intereses sobre cesantías: 12% anual proporcional
    int_cesantias = round(cesantias * 0.12 * dias_trabajados / 365.25)

    # Prima de servicios: 15 días de salario con aux por semestre (proporcional al último semestre)
    # Calcular días del semestre en curso
    ultimo_1_julio = date_t(hoy.year, 7, 1) if hoy.month >= 7 else date_t(hoy.year - 1, 7, 1)
    ultimo_1_enero = date_t(hoy.year, 1, 1)
    inicio_semestre = max(ultimo_1_julio, ultimo_1_enero, fecha_ingreso)
    dias_semestre = max((fecha_retiro - inicio_semestre).days, 0)
    prima = round(salario_con_aux * dias_semestre / 360)

    # Vacaciones: 15 días por año (proporcional)
    vacaciones = round(salario * dias_trabajados / 730)  # 15/365 = 1/730 * salario * dias

    # Indemnización por despido sin justa causa (Art. 64 CST)
    indemnizacion = 0
    if motivo == 'despido_sin_justa':
        if empleado.tipo_contrato == 'indefinido':
            if anios < 1:
                indemnizacion = round(salario * 30 / 30)  # 30 días primer año
            elif salario <= 10 * SMLMV_2025:
                indemnizacion = round(salario * 30 / 30 + salario * 20 / 30 * (anios - 1))
            else:
                indemnizacion = round(salario * 20 / 30 * anios)
        elif empleado.tipo_contrato == 'fijo':
            # Suma de salarios al vencimiento del contrato
            dias_restantes = max(0, (empleado.fecha_fin_contrato - fecha_retiro).days) if hasattr(empleado, 'fecha_fin_contrato') else 90
            indemnizacion = round(salario * min(dias_restantes, 180) / 30)

    total = cesantias + int_cesantias + prima + vacaciones + indemnizacion

    return {
        'empleado': empleado,
        'motivo': motivo,
        'fecha_ingreso': fecha_ingreso,
        'fecha_retiro': fecha_retiro,
        'dias_trabajados': dias_trabajados,
        'anios': round(anios, 2),
        'salario': salario,
        'aux_transporte': aux_transporte,
        'cesantias': cesantias,
        'int_cesantias': int_cesantias,
        'prima': prima,
        'vacaciones': vacaciones,
        'indemnizacion': indemnizacion,
        'total': total,
    }


def _calcular_liquidacion_mx(empleado, motivo):
    """
    Mexican employee finiquito/liquidacion.
    motivo: 'renuncia' | 'despido_justa' | 'despido_sin_justa' | 'mutuo_acuerdo'
    """
    from datetime import date as date_t
    hoy = date_t.today()
    fecha_retiro = empleado.fecha_retiro or hoy
    fecha_ingreso = empleado.fecha_ingreso
    if not fecha_ingreso:
        return None

    # Dias trabajados
    dias_trabajados = (fecha_retiro - fecha_ingreso).days
    anios = dias_trabajados / 365.25

    salario = empleado.salario_base
    aux_transporte = 0  # No transport subsidy in Mexico
    salario_diario = salario / 30

    # --- Aguinaldo proporcional ---
    # 15 days minimum per year, proportional to days worked in the current year
    aguinaldo_days = _payroll.get('aguinaldo_days', 15)
    inicio_anio = date_t(fecha_retiro.year, 1, 1)
    fecha_inicio_periodo = max(fecha_ingreso, inicio_anio)
    dias_anio_trabajados = max((fecha_retiro - fecha_inicio_periodo).days, 0)
    cesantias = round(salario_diario * aguinaldo_days * dias_anio_trabajados / 365)  # aguinaldo proporcional

    # --- Prima vacacional proporcional ---
    vac_days = _payroll.get('vacaciones_min_days', 12)
    prima_vac_pct = _payroll.get('prima_vacacional_pct', 0.25)
    int_cesantias = round(salario_diario * vac_days * prima_vac_pct * dias_anio_trabajados / 365)  # prima vacacional

    # --- Vacaciones no gozadas (proporcional) ---
    # Proportional vacation days for the current year
    prima = 0  # Not applicable in Mexico (prima de servicios is Colombian)
    vacaciones = round(salario_diario * vac_days * dias_anio_trabajados / 365)

    # --- Indemnizacion (Art. 50 LFT — despido injustificado) ---
    indemnizacion = 0
    if motivo == 'despido_sin_justa':
        # 3 months salary (indemnizacion constitucional)
        indem_constitucional = round(salario_diario * 90)
        # 20 days per year worked (prima de antiguedad for >15 years, but applied in unjust dismissal)
        indem_20_dias = round(salario_diario * 20 * anios)
        # Prima de antiguedad: 12 days per year (capped at 2x min wage daily)
        salario_diario_tope = min(salario_diario, 2 * _payroll.get('min_wage', 7468) / 30)
        prima_antiguedad = round(salario_diario_tope * 12 * anios)
        indemnizacion = indem_constitucional + indem_20_dias + prima_antiguedad
    elif motivo == 'renuncia' and anios >= 15:
        # Prima de antiguedad solo aplica con 15+ years on voluntary resignation
        salario_diario_tope = min(salario_diario, 2 * _payroll.get('min_wage', 7468) / 30)
        indemnizacion = round(salario_diario_tope * 12 * anios)

    total = cesantias + int_cesantias + prima + vacaciones + indemnizacion

    return {
        'empleado': empleado,
        'motivo': motivo,
        'fecha_ingreso': fecha_ingreso,
        'fecha_retiro': fecha_retiro,
        'dias_trabajados': dias_trabajados,
        'anios': round(anios, 2),
        'salario': salario,
        'aux_transporte': aux_transporte,
        'cesantias': cesantias,            # aguinaldo proporcional
        'int_cesantias': int_cesantias,    # prima vacacional proporcional
        'prima': prima,                    # 0 (no prima de servicios in MX)
        'vacaciones': vacaciones,
        'indemnizacion': indemnizacion,
        'total': total,
        # Mexican-specific detail
        'aguinaldo_proporcional': cesantias,
        'prima_vacacional': int_cesantias,
        'vacaciones_no_gozadas': vacaciones,
    }


def _calcular_liquidacion(empleado, motivo):
    """Dispatcher — calls Colombian or Mexican liquidation based on company config."""
    system = _payroll.get('system', 'colombian')
    if system == 'mexican':
        return _calcular_liquidacion_mx(empleado, motivo)
    return _calcular_liquidacion_co(empleado, motivo)

def _calcular_impuestos(ingresos, utilidad):
    """Retorna (total_impuestos, lista_detalle) según reglas tributarias activas.
    Cada item del detalle: {nombre, aplica_a, porcentaje, base_label, base_monto, monto}
    No incluye reglas de proveedor (se aplican por compra individual)."""
    reglas = ReglaTributaria.query.filter_by(activo=True).all()
    total = 0.0
    detalle = []
    for r in reglas:
        if r.aplica_a in ('proveedor_producto', 'proveedor_granel'):
            continue  # estas aplican por compra, no de forma global
        base_label = ''
        base_monto = 0.0
        monto = 0.0
        if r.aplica_a == 'ventas':
            # IVA se calcula sobre la base SIN impuesto (total / (1 + pct/100))
            pct = r.porcentaje / 100.0
            base_monto = ingresos / (1.0 + pct) if pct > 0 else ingresos
            base_label = 'Base gravable (sin IVA)'
        elif r.aplica_a == 'ingresos':
            pct = r.porcentaje / 100.0
            base_monto = ingresos / (1.0 + pct) if pct > 0 else ingresos
            base_label = 'Base gravable (sin IVA)'
        elif r.aplica_a == 'profit':
            if utilidad > 0:
                base_monto = utilidad
                base_label = 'Utilidad'
            else:
                continue
        elif r.aplica_a == 'ica':
            # ICA municipal: se calcula sobre ingresos brutos
            base_monto = ingresos
            base_label = 'Ingresos brutos (base ICA)'
        elif r.aplica_a in ('retencion_servicios', 'retencion_honorarios'):
            # Retención en la fuente por servicios u honorarios sobre ingresos
            base_monto = ingresos
            base_label = 'Ingresos (base retención)'
        elif r.aplica_a == 'reteiva':
            # ReteIVA: porcentaje sobre el IVA estimado (19% de ingresos)
            iva_estimado = ingresos / 1.19 * 0.19 if ingresos > 0 else 0.0
            base_monto = iva_estimado
            base_label = 'IVA estimado (base ReteIVA)'
        monto = base_monto * (r.porcentaje / 100.0)
        total += monto
        detalle.append({
            'nombre': r.nombre,
            'aplica_a': r.aplica_a,
            'porcentaje': r.porcentaje,
            'base_label': base_label,
            'base_monto': base_monto,
            'monto': monto,
        })
    return total, detalle

def _descontar_materias(form):
    """Si el formulario tiene usar_materias=1, descuenta stock de cada materia prima."""
    if not form.get('usar_materias'): return
    mp_ids = form.getlist('mp_id[]')
    mp_cants = form.getlist('mp_cant[]')
    lote_id = form.get('lote_id') or None
    errores = []
    for mid, cant_str in zip(mp_ids, mp_cants):
        if not mid or not cant_str: continue
        try:
            cant = float(cant_str)
            m = db.session.get(MateriaPrima, int(mid))
            if not m:
                errores.append(f'Materia prima ID {mid} no encontrada')
                continue
            if m.stock_disponible < cant:
                errores.append(f'Stock insuficiente de "{m.nombre}": disponible {m.stock_disponible:.3f} {m.unidad}, solicitado {cant:.3f}')
                continue
            m.stock_disponible -= cant
            m.stock_reservado = max(0, m.stock_reservado - cant)
            if lote_id:
                db.session.add(ReservaProduccion(
                    materia_prima_id=m.id, cantidad=cant,
                    lote_id=int(lote_id), estado='usado',
                    notas='Descontado desde inventario',
                    creado_por=None))
            # Verificar stock mínimo
            from services.inventario import verificar_stock_minimo
            verificar_stock_minimo(m.id)
        except Exception as e:
            errores.append(str(e))
    return errores

def _descontar_stock_venta(venta):
    """
    Descuenta del inventario (Producto.stock) las cantidades de los items
    de la venta. Se llama exactamente una vez, cuando la venta pasa a
    anticipo_pagado o completado por primera vez.
    """
    try:
        for item in venta.items:
            if not item.producto_id:
                continue
            prod = db.session.get(Producto, item.producto_id)
            if not prod:
                continue
            cant = item.cantidad
            prod.stock = max(0, (prod.stock or 0) - cant)
    except Exception as ex:
        logging.warning(f'_descontar_stock_venta error: {ex}')

def _save_contactos(cliente_obj):
    ContactoCliente.query.filter_by(cliente_id=cliente_obj.id).delete()
    nombres   = request.form.getlist('c_nombre[]')
    cargos    = request.form.getlist('c_cargo[]')
    emails    = request.form.getlist('c_email[]')
    telefonos = request.form.getlist('c_telefono[]')
    for i, n in enumerate(nombres):
        if n.strip():
            db.session.add(ContactoCliente(
                cliente_id=cliente_obj.id,
                nombre=n.strip(),
                cargo=cargos[i] if i < len(cargos) else '',
                email=emails[i] if i < len(emails) else '',
                telefono=telefonos[i] if i < len(telefonos) else ''))

def _oc_save_items(oc_id):
    """Guarda ítems del formulario de OC y devuelve la lista."""
    nombres  = request.form.getlist('item_nombre[]')
    descs    = request.form.getlist('item_desc[]')
    cants    = request.form.getlist('item_cant[]')
    units    = request.form.getlist('item_unidad[]')
    precios  = request.form.getlist('item_precio[]')
    cot_ids  = request.form.getlist('item_cot_id[]')
    items = []
    for i, nom in enumerate(nombres):
        if not nom.strip(): continue
        cant   = float(cants[i])  if i < len(cants)   else 1
        precio = float(precios[i]) if i < len(precios) else 0
        cot_id = int(cot_ids[i]) if i < len(cot_ids) and cot_ids[i].strip() else None
        items.append(OrdenCompraItem(
            orden_id=oc_id,
            nombre_item=nom.strip(),
            descripcion=descs[i] if i < len(descs) else '',
            cantidad=cant,
            unidad=units[i] if i < len(units) else 'unidades',
            precio_unit=precio,
            subtotal=cant*precio,
            cotizacion_id=cot_id
        ))
    return items

def _prods_json():
    prods = Producto.query.filter_by(activo=True).order_by(Producto.nombre).all()
    return [{'id':p.id,'nombre':p.nombre,'sku':p.sku or '','precio':p.precio,
             'stock':p.stock,'stock_minimo':p.stock_minimo} for p in prods]

def _save_items(venta_obj):
    VentaProducto.query.filter_by(venta_id=venta_obj.id).delete()
    pids    = request.form.getlist('prod_id[]')
    cants   = request.form.getlist('prod_cant[]')
    precios = request.form.getlist('prod_precio[]')
    for i, pid in enumerate(pids):
        cant  = float(cants[i]) if i < len(cants) else 1
        precio= float(precios[i]) if i < len(precios) else 0
        prod  = db.session.get(Producto, int(pid)) if pid else None
        db.session.add(VentaProducto(
            venta_id=venta_obj.id,
            producto_id=int(pid) if pid else None,
            nombre_prod=prod.nombre if prod else '',
            cantidad=cant, precio_unit=precio, subtotal=cant*precio))

def _save_asignados(tarea_obj):
    TareaAsignado.query.filter_by(tarea_id=tarea_obj.id).delete()
    # asignado principal
    principal = request.form.get('asignado_a')
    # otros asignados (multi-select)
    otros = request.form.getlist('otros_asignados[]')
    seen = set()
    for uid_str in ([principal] if principal else []) + otros:
        try:
            uid = int(uid_str)
            if uid not in seen:
                seen.add(uid)
                db.session.add(TareaAsignado(tarea_id=tarea_obj.id, user_id=uid))
        except (ValueError, TypeError):
            pass
    if not seen:
        db.session.add(TareaAsignado(tarea_id=tarea_obj.id, user_id=current_user.id))

def _inv_form_ctx():
    materias = MateriaPrima.query.filter_by(activo=True).order_by(MateriaPrima.nombre).all()
    lotes = LoteProducto.query.order_by(LoteProducto.creado_en.desc()).all()
    mj = [{'id': m.id, 'nombre': m.nombre, 'unidad': m.unidad,
            'stock': m.stock_disponible} for m in materias]
    return {'materias_json': mj, 'lotes': lotes}

def _save_compra(c, form):
    """Helper to parse and save compra fields from form."""
    fd = form.get('fecha')
    cant  = float(form.get('cantidad',1) or 1)
    costo_p = float(form.get('costo_producto',0) or 0)
    imp   = float(form.get('impuestos',0) or 0)
    trans = float(form.get('transporte',0) or 0)
    costo_total = costo_p + imp + trans
    precio_unit = (costo_total / cant) if cant > 0 else 0
    pid   = form.get('producto_id') or None
    mid   = form.get('materia_id') or None
    fvenc = form.get('fecha_caducidad') or None
    c.producto_id   = int(pid) if pid else None
    c.materia_id    = int(mid) if mid else None
    c.nombre_item   = form['nombre_item']
    c.tipo_compra   = form.get('tipo_compra','insumo')
    c.unidad        = form.get('unidad','unidades')
    c.proveedor     = form.get('proveedor','')
    c.proveedor_id  = int(form.get('proveedor_id')) if form.get('proveedor_id') else None
    c.fecha         = datetime.strptime(fd,'%Y-%m-%d').date() if fd else datetime.utcnow().date()
    c.nro_factura   = form.get('nro_factura','')
    c.cantidad      = cant
    c.costo_producto = costo_p; c.impuestos = imp; c.transporte = trans
    c.costo_total   = costo_total; c.precio_unitario = precio_unit
    c.tiene_caducidad = bool(form.get('tiene_caducidad'))
    c.fecha_caducidad = datetime.strptime(fvenc,'%Y-%m-%d').date() if fvenc else None
    c.notas = form.get('notas','')
    # Update linked product cost
    if pid:
        prod = db.session.get(Producto, int(pid))
        if prod: prod.costo = precio_unit
    # Update materia prima stock if tipo=materia_prima — crea LoteMateriaPrima para trazabilidad FIFO
    if mid and c.tipo_compra == 'materia_prima':
        m = db.session.get(MateriaPrima, int(mid))
        if m:
            m.stock_disponible = (m.stock_disponible or 0) + cant
            try:
                from models import LoteMateriaPrima
                lote_mp = LoteMateriaPrima(
                    materia_prima_id=m.id,
                    numero_lote=form.get('nro_lote_materia','').strip() or None,
                    nro_factura=c.nro_factura or None,
                    proveedor=c.proveedor or None,
                    fecha_compra=c.fecha,
                    fecha_vencimiento=c.fecha_caducidad,
                    cantidad_inicial=cant,
                    cantidad_disponible=cant,
                    cantidad_reservada=0.0,
                    costo_unitario=precio_unit,
                    notas=form.get('notas','') or None,
                )
                db.session.add(lote_mp)
            except Exception as _le:
                import logging; logging.warning(f'LoteMateriaPrima create error: {_le}')
    # Auto-register/update GastoOperativo
    tipo_label = {'materia_prima': 'Materia prima', 'insumo': 'Insumo',
                  'producto': 'Producto', 'servicio': 'Servicio'}.get(c.tipo_compra, c.tipo_compra.capitalize())
    desc_gasto = (f'{c.nombre_item} — {tipo_label}')
    if c.proveedor: desc_gasto += f' | Proveedor: {c.proveedor}'
    if c.nro_factura: desc_gasto += f' | Factura: {c.nro_factura}'
    # Buscar gasto existente para esta compra (evitar duplicados en edición)
    gasto = None
    if c.id:
        gasto = GastoOperativo.query.filter_by(
            tipo='compra_produccion',
            descripcion=db.func.substr(GastoOperativo.descripcion, 1, 50).like(f'%{c.nombre_item[:30]}%')
        ).filter(
            GastoOperativo.fecha == c.fecha,
            GastoOperativo.creado_por == getattr(c, 'creado_por', None)
        ).first()
    if gasto:
        # Actualizar existente
        gasto.fecha = c.fecha
        gasto.tipo_custom = f'Compra / {tipo_label}'
        gasto.descripcion = desc_gasto
        gasto.monto = costo_total
        # Actualizar asiento vinculado
        asiento_link = AsientoContable.query.filter_by(gasto_id=gasto.id).first()
        if asiento_link:
            asiento_link.debe = float(costo_total)
            asiento_link.haber = float(costo_total)
            asiento_link.descripcion = f'Compra: {c.nombre_item} — {tipo_label}'[:300]
            asiento_link.fecha = c.fecha
            asiento_link.referencia = c.nro_factura or None
    else:
        # Crear nuevo
        gasto = GastoOperativo(
            fecha=c.fecha,
            tipo='compra_produccion',
            tipo_custom=f'Compra / {tipo_label}',
            descripcion=desc_gasto,
            monto=costo_total,
            recurrencia='unica',
            es_plantilla=False,
            notas=f'Registrado automáticamente desde módulo Compras.',
            creado_por=getattr(c, 'creado_por', None)
        )
        db.session.add(gasto)
        db.session.flush()
        # Partida doble automática
        _crear_asiento_auto(
            tipo='compra', subtipo='compra_materia',
            descripcion=f'Compra: {c.nombre_item} — {tipo_label}',
            monto=costo_total,
            cuenta_debe='Inventario materias primas',
            cuenta_haber='Cuentas por pagar proveedores',
            clasificacion='egreso',
            referencia=c.nro_factura or None,
            gasto_id=gasto.id,
            proveedor_id=c.proveedor_id
        )
    return c

def _make_xlsx(titulo, headers, rows):
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo
    # Cabecera empresa
    ws.merge_cells(f'A1:{chr(64+len(headers))}1')
    ws['A1'] = 'Evore CRM — ' + titulo
    ws['A1'].font = Font(bold=True, size=13, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor='5E72E4')
    ws['A1'].alignment = Alignment(horizontal='center')
    # Headers
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1A1F36')
        cell.alignment = Alignment(horizontal='center')
    # Datos
    for r_idx, row in enumerate(rows, 3):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    # Autofit columns
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def _procesar_orden_produccion(cot):
    """
    Al confirmar una cotización, por cada item:
    1. Verifica stock actual vs requerido y crea OrdenProduccion
    2. Reserva materias primas disponibles según BOM
    3. Si faltan materias, crea tareas comprar_materias + verificar_abono y notifica admins
    """
    admins = User.query.filter_by(rol='admin', activo=True).all()
    primer_admin_id = admins[0].id if admins else current_user.id

    for item in cot.items:
        if not item.producto_id: continue
        prod = db.session.get(Producto, item.producto_id)
        if not prod: continue
        cant_requerida = item.cantidad
        cant_en_stock  = min(prod.stock, cant_requerida)
        cant_producir  = max(0, cant_requerida - cant_en_stock)

        # Crear orden de producción
        orden = OrdenProduccion(
            cotizacion_id=cot.id,
            producto_id=prod.id,
            cantidad_total=cant_requerida,
            cantidad_stock=cant_en_stock,
            cantidad_producir=cant_producir,
            estado='en_produccion' if cant_producir > 0 else 'completado',
            notas=f'Cotización #{cot.numero or cot.id} — {cot.titulo}',
            creado_por=current_user.id
        )
        db.session.add(orden)
        db.session.flush()  # obtener orden.id para vincular reservas

        if cant_producir <= 0:
            continue  # ya hay suficiente stock, sin acción adicional

        # Verificar BOM
        receta = RecetaProducto.query.filter_by(producto_id=prod.id, activo=True).first()
        materias_faltantes = []
        if receta and receta.unidades_produce > 0:
            factor = cant_producir / receta.unidades_produce
            for ri in receta.items:
                mp = db.session.get(MateriaPrima, ri.materia_prima_id)
                if not mp: continue
                necesaria = ri.cantidad_por_unidad * factor
                disponible = mp.stock_disponible
                if disponible >= necesaria:
                    # Reservar automáticamente
                    mp.stock_disponible -= necesaria
                    mp.stock_reservado  += necesaria
                    db.session.add(ReservaProduccion(
                        materia_prima_id=mp.id,
                        cantidad=necesaria,
                        producto_id=prod.id,
                        orden_produccion_id=orden.id,
                        estado='reservado',
                        notas=f'Auto-reserva cot #{cot.numero or cot.id}',
                        creado_por=current_user.id
                    ))
                    # Verificar stock mínimo
                    from services.inventario import verificar_stock_minimo
                    verificar_stock_minimo(mp.id)
                else:
                    faltante = necesaria - disponible
                    materias_faltantes.append(
                        f'{mp.nombre}: falta {faltante:.3f} {mp.unidad} (disponible {disponible:.3f})'
                    )
                    # Reservar lo que haya
                    if disponible > 0:
                        mp.stock_disponible = 0
                        mp.stock_reservado += disponible
                        db.session.add(ReservaProduccion(
                            materia_prima_id=mp.id,
                            cantidad=disponible,
                            producto_id=prod.id,
                            orden_produccion_id=orden.id,
                            estado='reservado',
                            notas=f'Parcial cot #{cot.numero or cot.id}',
                            creado_por=current_user.id
                        ))
                    # Verificar stock mínimo (stock agotado o bajo)
                    from services.inventario import verificar_stock_minimo
                    verificar_stock_minimo(mp.id)

        if materias_faltantes:
            descripcion_falta = '\n'.join(materias_faltantes)
            desc_tareas = (f'Cotización: #{cot.numero or cot.id} — {cot.titulo}\n'
                           f'Producto: {prod.nombre} (x{cant_producir})\n\n'
                           f'Materiales faltantes:\n{descripcion_falta}')
            from datetime import timedelta
            venc = (datetime.utcnow() + timedelta(days=3)).date()

            t_compra = Tarea(
                titulo=f'Comprar materias — {prod.nombre} (cot #{cot.numero or cot.id})',
                descripcion=desc_tareas,
                estado='pendiente', prioridad='alta',
                asignado_a=primer_admin_id,
                creado_por=current_user.id,
                fecha_vencimiento=venc,
                cotizacion_id=cot.id,
                tarea_tipo='comprar_materias'
            )
            db.session.add(t_compra); db.session.flush()

            t_abono = Tarea(
                titulo=f'Verificar abono — {cot.titulo}',
                descripcion=(f'Confirmar recepción del anticipo antes de comprar materias.\n'
                             f'Cotización: #{cot.numero or cot.id}'),
                estado='pendiente', prioridad='alta',
                asignado_a=primer_admin_id,
                creado_por=current_user.id,
                fecha_vencimiento=venc,
                cotizacion_id=cot.id,
                tarea_tipo='verificar_abono',
                tarea_pareja_id=t_compra.id
            )
            db.session.add(t_abono); db.session.flush()
            # link inverso
            t_compra.tarea_pareja_id = t_abono.id

            # Notificar a todos los admins
            for adm in admins:
                _crear_notificacion(
                    adm.id, 'alerta_stock',
                    f'⚠️ Materiales insuficientes — {prod.nombre}',
                    f'Faltan materias para cotización #{cot.numero or cot.id}. '
                    f'Se crearon tareas de compra y abono.',
                    url_for('tareas')
                )

def _procesar_venta_produccion(venta):
    """
    Al guardar una venta, por cada item con producto:
    1. Compara stock actual vs cantidad requerida
    2. Si falta stock → crea OrdenProduccion (vinculada a esta venta)
    3. Reserva/descuenta materias primas disponibles según BOM inmediatamente
    4. Si faltan materias, crea tareas comprar_materias + verificar_abono
    El check de duplicado es POR VENTA+PRODUCTO (no global), así cada venta
    genera sus propias órdenes y las ventas posteriores no se bloquean.
    """
    try:
        admins = User.query.filter_by(rol='admin', activo=True).all()
        primer_admin_id = admins[0].id if admins else current_user.id
        from datetime import timedelta

        for item in venta.items:
            if not item.producto_id:
                continue
            prod = db.session.get(Producto, item.producto_id)
            if not prod:
                continue
            cant_requerida = item.cantidad
            cant_en_stock  = min(prod.stock or 0, cant_requerida)
            cant_producir  = max(0.0, cant_requerida - cant_en_stock)

            if cant_producir <= 0:
                continue  # stock suficiente, sin producción necesaria

            # Check duplicado POR VENTA+PRODUCTO (no bloquea otras ventas del mismo producto)
            existente = OrdenProduccion.query.filter(
                OrdenProduccion.venta_id == venta.id,
                OrdenProduccion.producto_id == prod.id,
                OrdenProduccion.estado != 'completado'
            ).first()
            if existente:
                continue

            orden = OrdenProduccion(
                venta_id=venta.id,
                producto_id=prod.id,
                cantidad_total=cant_requerida,
                cantidad_stock=cant_en_stock,
                cantidad_producir=cant_producir,
                estado='en_produccion',
                notas=f'Venta: {venta.titulo} — {prod.nombre} x{cant_requerida}',
                creado_por=current_user.id
            )
            db.session.add(orden)
            db.session.flush()  # obtener orden.id para vincular reservas

            # Descontar/reservar materias primas del BOM inmediatamente
            receta = RecetaProducto.query.filter_by(producto_id=prod.id, activo=True).first()
            materias_faltantes = []
            if receta and receta.unidades_produce > 0:
                factor = cant_producir / receta.unidades_produce
                for ri in receta.items:
                    mp = db.session.get(MateriaPrima, ri.materia_prima_id)
                    if not mp: continue
                    necesaria  = ri.cantidad_por_unidad * factor
                    disponible = mp.stock_disponible or 0
                    if disponible >= necesaria:
                        # Disponible: reservar completamente
                        mp.stock_disponible -= necesaria
                        mp.stock_reservado   = (mp.stock_reservado or 0) + necesaria
                        db.session.add(ReservaProduccion(
                            materia_prima_id=mp.id, cantidad=necesaria,
                            producto_id=prod.id, venta_id=venta.id,
                            orden_produccion_id=orden.id, estado='reservado',
                            notas=f'Auto-reserva venta: {venta.titulo}',
                            creado_por=current_user.id
                        ))
                        # Verificar stock mínimo
                        from services.inventario import verificar_stock_minimo
                        verificar_stock_minimo(mp.id)
                    else:
                        # Insuficiente: reservar lo que haya + anotar faltante
                        faltante = necesaria - disponible
                        materias_faltantes.append(
                            f'{mp.nombre}: falta {faltante:.3f} {mp.unidad} (disp. {disponible:.3f})'
                        )
                        if disponible > 0:
                            mp.stock_disponible = 0
                            mp.stock_reservado   = (mp.stock_reservado or 0) + disponible
                            db.session.add(ReservaProduccion(
                                materia_prima_id=mp.id, cantidad=disponible,
                                producto_id=prod.id, venta_id=venta.id,
                                orden_produccion_id=orden.id, estado='reservado',
                                notas=f'Parcial venta: {venta.titulo}',
                                creado_por=current_user.id
                            ))
                        # Verificar stock mínimo (stock agotado o bajo)
                        from services.inventario import verificar_stock_minimo
                        verificar_stock_minimo(mp.id)

            if materias_faltantes:
                desc_falta = '\n'.join(materias_faltantes)
                desc_t = (f'Venta: {venta.titulo}\n'
                          f'Producto: {prod.nombre} (x{cant_producir:.2f})\n\n'
                          f'Materiales faltantes:\n{desc_falta}')
                venc = (datetime.utcnow() + timedelta(days=3)).date()
                t_compra = Tarea(
                    titulo=f'Comprar materiales — {prod.nombre} (venta)',
                    descripcion=desc_t, estado='pendiente', prioridad='alta',
                    asignado_a=primer_admin_id, creado_por=current_user.id,
                    fecha_vencimiento=venc, tarea_tipo='comprar_materias'
                )
                db.session.add(t_compra); db.session.flush()
                t_abono = Tarea(
                    titulo=f'Verificar abono — compra {prod.nombre}',
                    descripcion=f'Confirmar anticipo antes de comprar materiales para {prod.nombre}.\n\n{desc_falta}',
                    estado='pendiente', prioridad='alta',
                    asignado_a=primer_admin_id, creado_por=current_user.id,
                    fecha_vencimiento=venc, tarea_tipo='verificar_abono',
                    tarea_pareja_id=t_compra.id
                )
                db.session.add(t_abono); db.session.flush()
                t_compra.tarea_pareja_id = t_abono.id
                orden.estado = 'pendiente_materiales'
                for adm in admins:
                    _crear_notificacion(
                        adm.id, 'alerta_stock',
                        f'⚠️ Materiales insuficientes — {prod.nombre}',
                        f'Venta "{venta.titulo}" requiere producción. Faltan materias.',
                        url_for('tareas')
                    )
    except Exception as ex:
        db.session.rollback()
        logging.warning(f'_procesar_venta_produccion error: {ex}')

def _modulos_user(user):
    if not user or not user.is_authenticated: return []
    rol_activo = _get_rol_activo(user)
    if rol_activo == 'admin': return _MODULOS_TODOS
    # Si el rol activo es diferente al principal, usar modulos del rol activo
    if rol_activo != user.rol:
        return _MODULOS_ROL.get(rol_activo, ['tareas','notas'])
    # Rol principal: intentar custom modules primero
    try:
        custom = json.loads(user.modulos_permitidos or '[]')
        if custom: return custom
    except Exception: pass
    return _MODULOS_ROL.get(user.rol, ['tareas','notas'])

def generar_csv_response(rows, headers, filename='export.csv'):
    """Genera una Response Flask con un CSV descargable."""
    import csv, io
    from flask import make_response
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    resp = make_response(output.getvalue())
    resp.headers['Content-Type'] = 'text/csv; charset=utf-8'
    resp.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return resp


def register_app_hooks(app):
    """Register template filters and context processors on the Flask app."""
    app.template_filter('cop')(cop)
    app.template_filter('moneda')(moneda)
    app.template_filter('moneda0')(moneda0)
    # Also register as Jinja globals so they can be called as functions: {{ moneda(value) }}
    app.jinja_env.globals['moneda'] = moneda
    app.jinja_env.globals['moneda0'] = moneda0
    app.jinja_env.globals['cop'] = cop
    app.context_processor(inject_globals)